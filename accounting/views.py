import math
import json
import os
import tempfile
import unicodedata
from django.shortcuts import render
from django.http import JsonResponse, FileResponse
from django.urls import include, path
from django.core.exceptions import PermissionDenied, ObjectDoesNotExist
from django.utils.datastructures import MultiValueDictKeyError
from django.db import models, IntegrityError, transaction
from django.db.models import  Max, Sum, Q, F, Subquery, OuterRef, Count
from django.db.models.functions import Coalesce
from django.core import serializers
from django.core.paginator import Paginator, EmptyPage
from django.conf import settings
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from andinasoft.models import (proyectos, empresas, 
                               cuentas_pagos, Profiles, asesores,clientes, Gtt,
                               Detalle_gtt, entidades_bancarias)
from andinasoft.shared_models import (Pagocomision, Recaudos_general, 
                                      formas_pago, RecaudosNoradicados, ventas_nuevas,
                                      Inmuebles, Adjudicacion)
from andina.decorators import group_perm_required, check_project, check_perms, check_groups
from andinasoft.create_pdf import GenerarPDF
from accounting.models import (distribucion_centros_costos, egresos_contable, egresos_banco, conciliaciones, 
                               Facturas, Pagos, gastos_caja, history_facturas, impuestos_legalizacion,info_interfaces,
                               Anticipos, otros_ingresos, parametros, reembolsos_caja, transferencias_companias,
                               docs_cuentas_oficinas, cuentas_intercompanias,
                               pago_detallado_relacionado, saldos_cuentas_tesoreria, info_facturas,
                               solicitud_anticipos, legalizacion_anticipos, conceptos_legalizacion,
                               reintegros_anticipos, Partners, Countries, States, Cities, archivo_contable,
                               PlinkMovement, WompiMovement)
from andinasoft.ajax_request import JSONRender
from andinasoft.utilities import get_text_from_file, pdf_gen, Utilidades
from andinasoft.handlers_functions import envio_notificacion
from accounting.models_alttum import Pagos_facturas, Anticipos_hotels as anticipos_hotels
from accounting import forms
from decimal import Decimal, InvalidOperation
from apis.nominapp.nominapp_api import period
from dateutil import relativedelta
import operator
import openpyxl
import datetime
import random
import string
import traceback
import csv
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.utils.dateparse import parse_date
from api_auth.decorators import api_token_auth
import re


def _parse_decimal(value, default=Decimal('0')):
    if value in (None, ''):
        return default
    try:
        cleaned = str(value).replace(',', '')
        return Decimal(cleaned)
    except (InvalidOperation, TypeError, ValueError):
        return default


def _parse_compact_date(value):
    if not value:
        return None
    value = str(value).strip()
    if not value:
        return None
    try:
        if len(value) == 8 and value.isdigit():
            return datetime.datetime.strptime(value, '%Y%m%d').date()
        return datetime.datetime.strptime(value, '%Y-%m-%d').date()
    except ValueError:
        return None


def _parse_datetime(value):
    if not value:
        return None
    try:
        return datetime.datetime.strptime(value, '%Y-%m-%d %H:%M:%S')
    except ValueError:
        try:
            return datetime.datetime.strptime(value, '%Y-%m-%dT%H:%M:%S')
        except ValueError:
            return None


def _media_url(file_field):
    if not file_field:
        return ''
    try:
        return file_field.url
    except Exception:
        return str(file_field)

def _normalize_account(value):
    if not value:
        return None
    return re.sub(r'[^A-Za-z0-9]', '', str(value)).lower() or None

def _get_account_by_number(account_number):
    normalized = _normalize_account(account_number)
    if not normalized:
        return None
    for cuenta in cuentas_pagos.objects.all():
        normalized_current = _normalize_account(cuenta.cuentabanco)
        if normalized_current == normalized:
            return cuenta
    return None

def _normalize_text_for_search(text):
    """
    Normaliza texto removiendo diacríticas y convirtiendo a minúsculas.
    Usado para búsquedas insensibles a acentos y mayúsculas.
    """
    if not text:
        return ''
    # Normalizar caracteres Unicode (NFD = descomponer caracteres con diacríticos)
    normalized = unicodedata.normalize('NFD', str(text))
    # Remover marcas diacríticas (categoría 'Mn' = Nonspacing Mark)
    without_accents = ''.join(char for char in normalized if unicodedata.category(char) != 'Mn')
    return without_accents.lower()

def _fuzzy_match(word1, word2, threshold=0.7):
    """
    Calcula si dos palabras son similares usando distancia de Levenshtein.
    Retorna True si la similitud es >= threshold (default 70%).
    """
    if not word1 or not word2:
        return False

    # Caso exacto (optimización)
    if word1 == word2:
        return True

    # Si una palabra está contenida en la otra (subcadena)
    if word1 in word2 or word2 in word1:
        # Calcular similitud basada en longitud de la subcadena
        min_len = min(len(word1), len(word2))
        max_len = max(len(word1), len(word2))
        similarity = min_len / max_len
        return similarity >= threshold

    # Calcular distancia de Levenshtein
    len1, len2 = len(word1), len(word2)

    # Crear matriz para programación dinámica
    dp = [[0] * (len2 + 1) for _ in range(len1 + 1)]

    # Inicializar primera fila y columna
    for i in range(len1 + 1):
        dp[i][0] = i
    for j in range(len2 + 1):
        dp[0][j] = j

    # Llenar matriz
    for i in range(1, len1 + 1):
        for j in range(1, len2 + 1):
            if word1[i-1] == word2[j-1]:
                dp[i][j] = dp[i-1][j-1]
            else:
                dp[i][j] = 1 + min(dp[i-1][j], dp[i][j-1], dp[i-1][j-1])

    # Calcular similitud
    max_len = max(len1, len2)
    distance = dp[len1][len2]
    similarity = 1 - (distance / max_len)

    return similarity >= threshold


def _parse_list_param(value):
    if not value:
        return []
    parts = [item.strip() for item in str(value).split(',')]
    return [item for item in parts if item]


def _calculate_lote_valor(lote):
    area = lote.areaprivada or Decimal('0')
    vr_m2 = lote.vrmetrocuadrado or Decimal('0')
    fac_via = lote.fac_valor_via_principal or Decimal('1')
    fac_area = lote.fac_valor_area_social or Decimal('1')
    fac_esq = lote.fac_valor_esquinero or Decimal('1')
    incrementos = fac_via * fac_area * fac_esq
    valor_lote = area * vr_m2 * incrementos
    return Utilidades().redondear_numero(numero=valor_lote, multiplo=1000000, redondeo='>')


def _normalize_lote_estado(value):
    if value is None:
        return None
    normalized = ' '.join(str(value).strip().lower().split())
    estados = {
        'libre': 'Libre',
        'bloqueado': 'Bloqueado',
        'sin liberar': 'Sin Liberar'
    }
    return estados.get(normalized)


def _get_cliente_nombre(cliente_id):
    if not cliente_id:
        return None
    try:
        return clientes.objects.get(pk=cliente_id).nombrecompleto
    except clientes.DoesNotExist:
        return None


@login_required
@group_perm_required(('andinasoft.view_facturas',),raise_exception=True)
def principal(request):
    context = {
       'form-int-egr':forms.form_interfaz_egreso,
       'form-int-notas':forms.form_interfaz_notas, 
       'proyectos':proyectos.objects.all().exclude(proyecto__icontains='Alttum')
    }
    
    return render(request,'accounting/principal.html',context)

@login_required
@group_perm_required(('accounting.view_egresos_banco',),raise_exception=True)
def movimientos(request):
    context={
        'empresas':empresas.objects.all(),
        'formmvto':forms.form_mvtos_concil()
    }
    return render(request,'accounting/movimientos.html',context)

@csrf_exempt
@api_token_auth
@require_http_methods(["POST"])
def api_bank_movements(request):
    if not request.user.is_authenticated or request.user.is_anonymous:
        return JsonResponse({'detail': 'Token inválido o no autenticado'}, status=401)
    try:
        payload = json.loads(request.body.decode('utf-8'))
    except (json.JSONDecodeError, UnicodeDecodeError):
        return JsonResponse({'detail':'JSON inválido'}, status=400)
    
    if not isinstance(payload, dict):
        return JsonResponse({'detail':'El cuerpo debe ser un objeto JSON'}, status=400)
    
    movimientos_payload = payload.get('movimientos')
    if movimientos_payload is None:
        return JsonResponse({'detail':'Debe enviar el campo "movimientos"'}, status=400)
    
    empresa_id = payload.get('empresa')
    cuenta_id = payload.get('cuenta')
    cuenta_numero = payload.get('cuenta_numero')
    
    cuenta_obj = None
    if cuenta_id:
        try:
            cuenta_obj = cuentas_pagos.objects.get(pk=cuenta_id)
        except cuentas_pagos.DoesNotExist:
            return JsonResponse({'detail':'Cuenta no encontrada'}, status=404)
    elif cuenta_numero:
        cuenta_obj = _get_account_by_number(cuenta_numero)
        if not cuenta_obj:
            return JsonResponse({'detail':'Cuenta no encontrada'}, status=404)
    else:
        return JsonResponse({'detail':'Debes enviar "cuenta" o "cuenta_numero"'}, status=400)
    empresa_obj = None
    if empresa_id:
        try:
            empresa_obj = empresas.objects.get(pk=empresa_id)
        except empresas.DoesNotExist:
            return JsonResponse({'detail':'Empresa no encontrada'}, status=404)
        if cuenta_obj.nit_empresa_id and cuenta_obj.nit_empresa_id != empresa_obj.pk:
            return JsonResponse({'detail':'La cuenta no pertenece a la empresa indicada'}, status=400)
    else:
        if cuenta_obj.nit_empresa_id:
            try:
                empresa_obj = empresas.objects.get(pk=cuenta_obj.nit_empresa_id)
            except empresas.DoesNotExist:
                return JsonResponse({'detail':'La empresa asociada a la cuenta no existe'}, status=400)
        else:
            return JsonResponse({'detail':'No se pudo determinar la empresa. Envía "empresa" o relaciona la cuenta a una empresa.'}, status=400)
    
    if not isinstance(movimientos_payload, list) or len(movimientos_payload) == 0:
        return JsonResponse({'detail':'El campo "movimientos" debe ser una lista con al menos un elemento'}, status=400)
    
    movimientos_limpios = []
    for index, movimiento in enumerate(movimientos_payload, start=1):
        if not isinstance(movimiento, dict):
            return JsonResponse({'detail':f'El movimiento #{index} no es un objeto válido'}, status=400)
        
        fecha = parse_date(str(movimiento.get('fecha','')))
        if fecha is None:
            return JsonResponse({'detail':f'La fecha del movimiento #{index} es inválida o no se envió (YYYY-MM-DD)'}, status=400)
        
        descripcion = movimiento.get('descripcion','').strip()
        if descripcion == '':
            return JsonResponse({'detail':f'La descripción del movimiento #{index} es obligatoria'}, status=400)
        
        valor_raw = movimiento.get('valor')
        try:
            valor = Decimal(str(valor_raw))
        except (InvalidOperation, TypeError):
            return JsonResponse({'detail':f'El valor del movimiento #{index} es inválido'}, status=400)
        if valor == 0:
            return JsonResponse({'detail':f'El valor del movimiento #{index} no puede ser cero'}, status=400)
        
        estado = movimiento.get('estado','SIN CONCILIAR').upper()
        if estado not in ('CONCILIADO','SIN CONCILIAR'):
            return JsonResponse({'detail':f'El estado del movimiento #{index} no es válido'}, status=400)
        
        pago_id = movimiento.get('pago_id')
        anticipo_id = movimiento.get('anticipo_id')
        transferencia_id = movimiento.get('transferencia_id')
        
        pago_obj = None
        anticipo_obj = None
        transferencia_obj = None
        
        if pago_id:
            try:
                pago_obj = Pagos.objects.get(pk=pago_id)
            except Pagos.DoesNotExist:
                return JsonResponse({'detail':f'El pago asociado del movimiento #{index} no existe'}, status=404)
        if anticipo_id:
            try:
                anticipo_obj = Anticipos.objects.get(pk=anticipo_id)
            except Anticipos.DoesNotExist:
                return JsonResponse({'detail':f'El anticipo asociado del movimiento #{index} no existe'}, status=404)
        if transferencia_id:
            try:
                transferencia_obj = transferencias_companias.objects.get(pk=transferencia_id)
            except transferencias_companias.DoesNotExist:
                return JsonResponse({'detail':f'La transferencia asociada del movimiento #{index} no existe'}, status=404)
        
        movimientos_limpios.append({
            'fecha': fecha,
            'descripcion': descripcion,
            'referencia': movimiento.get('referencia'),
            'valor': valor,
            'estado': estado,
            'pago': pago_obj,
            'anticipo': anticipo_obj,
            'transferencia': transferencia_obj,
            'external_id': movimiento.get('external_id')
        })
    
    registros = []
    with transaction.atomic():
        for item in movimientos_limpios:
            registro = egresos_banco.objects.create(
                empresa = empresa_obj,
                cuenta = cuenta_obj,
                fecha = item['fecha'],
                descripcion = item['descripcion'],
                referencia = item['referencia'],
                valor = item['valor'],
                estado = item['estado'],
                pago_asociado = item['pago'],
                anticipo_asociado = item['anticipo'],
                transferencia_asociada = item['transferencia']
            )
            registros.append({
                'id_mvto': registro.id_mvto,
                'external_id': item['external_id']
            })
    
    data = {
        'empresa': empresa_obj.pk,
        'cuenta': cuenta_obj.pk,
        'total_movimientos': len(registros),
        'movimientos': registros
    }
    return JsonResponse(data, status=201)


@api_token_auth
@require_http_methods(["GET"])
def api_lotes(request):
    if not request.user.is_authenticated or request.user.is_anonymous:
        return JsonResponse({'detail': 'Token inválido o no autenticado'}, status=401)
    proyecto = (request.GET.get('proyecto') or '').strip()
    if not proyecto:
        return JsonResponse({'detail': 'Debes enviar el parámetro "proyecto".'}, status=400)
    if not proyectos.objects.filter(proyecto__iexact=proyecto).exists():
        return JsonResponse({'detail': f'No encontramos el proyecto "{proyecto}".'}, status=404)
    if not check_project(request, proyecto, raise_exception=False):
        return JsonResponse({'detail': f'No tienes acceso al proyecto "{proyecto}".'}, status=403)

    manzanas = _parse_list_param(request.GET.get('manzana'))
    idinmueble = (request.GET.get('idinmueble') or '').strip()
    estado_raw = (request.GET.get('estado') or '').strip()
    estados = []
    if estado_raw:
        for item in estado_raw.split(','):
            normalized = _normalize_lote_estado(item)
            if normalized:
                estados.append(normalized)
    if estado_raw and not estados and not idinmueble:
        return JsonResponse({
            'detail': 'El estado enviado no es válido.',
            'estado_permitido': ['Libre', 'Bloqueado', 'Sin Liberar']
        }, status=400)
    try:
        inventario = Inmuebles.objects.using(proyecto).all()
        if idinmueble:
            inventario = inventario.filter(pk=idinmueble)
        else:
            if estados:
                estado_filter = Q()
                for estado in estados:
                    estado_filter |= Q(estado__iexact=estado)
                inventario = inventario.filter(estado_filter)
            else:
                inventario = inventario.filter(estado__iexact='Libre')
            if manzanas:
                inventario = inventario.filter(manzananumero__in=manzanas)
        data = []
        for lote in inventario:
            relacion = None
            if idinmueble:
                estado_lote = (lote.estado or '').strip().lower()
                if estado_lote == 'adjudicado':
                    adj = Adjudicacion.objects.using(proyecto).filter(idinmueble=lote.idinmueble).first()
                    if adj:
                        cliente_nombre = _get_cliente_nombre(adj.idtercero1)
                        relacion = {
                            'tipo': 'adjudicacion',
                            'referencia': adj.idadjudicacion,
                            'cliente': cliente_nombre
                        }
                elif estado_lote == 'reservado':
                    venta = (ventas_nuevas.objects.using(proyecto)
                             .filter(inmueble=lote.idinmueble)
                             .order_by('-fecha_contrato', '-id_venta')
                             .first())
                    if venta:
                        cliente_nombre = _get_cliente_nombre(venta.id_t1)
                        relacion = {
                            'tipo': 'reserva',
                            'referencia': venta.id_venta,
                            'cliente': cliente_nombre
                        }
            data.append({
                'idinmueble': lote.idinmueble,
                'estado': lote.estado,
                'manzana': lote.manzananumero,
                'lote': lote.lotenumero,
                'area_privada': lote.areaprivada,
                'precio_m2': lote.vrmetrocuadrado,
                'valor_lote': _calculate_lote_valor(lote),
                'motivo_bloqueo': lote.obsbloqueo,
                'usuario_bloqueo': lote.usuariobloquea,
                'relacion': relacion
            })
        return JsonResponse({'count': len(data), 'data': data}, status=200)
    except Exception as exc:
        return JsonResponse({'detail': 'Error al consultar lotes libres.', 'error': str(exc)}, status=500)


@csrf_exempt
@api_token_auth
@require_http_methods(["POST"])
def api_cambiar_estado_lote(request):
    if not request.user.is_authenticated or request.user.is_anonymous:
        return JsonResponse({'detail': 'Token inválido o no autenticado'}, status=401)
    payload = {}
    if request.body:
        try:
            payload = json.loads(request.body.decode('utf-8'))
        except (json.JSONDecodeError, UnicodeDecodeError):
            return JsonResponse({'detail': 'JSON inválido'}, status=400)
    proyecto = (payload.get('proyecto') if isinstance(payload, dict) else None) or request.POST.get('proyecto')
    inmueble_id = (payload.get('idinmueble') if isinstance(payload, dict) else None) or request.POST.get('idinmueble')
    estado_nuevo = (payload.get('estado') if isinstance(payload, dict) else None) or request.POST.get('estado')
    motivo_bloqueo = (payload.get('motivo_bloqueo') if isinstance(payload, dict) else None) or request.POST.get('motivo_bloqueo')
    if not proyecto:
        return JsonResponse({'detail': 'Debes enviar "proyecto".'}, status=400)
    if not inmueble_id:
        return JsonResponse({'detail': 'Debes enviar "idinmueble".'}, status=400)
    if not estado_nuevo:
        return JsonResponse({'detail': 'Debes enviar "estado".'}, status=400)
    if not proyectos.objects.filter(proyecto__iexact=proyecto).exists():
        return JsonResponse({'detail': f'No encontramos el proyecto "{proyecto}".'}, status=404)
    if not check_project(request, proyecto, raise_exception=False):
        return JsonResponse({'detail': f'No tienes acceso al proyecto "{proyecto}".'}, status=403)

    try:
        lote = Inmuebles.objects.using(proyecto).get(pk=inmueble_id)
    except Inmuebles.DoesNotExist:
        return JsonResponse({'detail': f'El lote "{inmueble_id}" no existe.'}, status=404)
    except Exception as exc:
        return JsonResponse({'detail': 'Error al buscar el lote.', 'error': str(exc)}, status=500)

    estado_actual = (lote.estado or '').strip()
    estado_nuevo_norm = _normalize_lote_estado(estado_nuevo)
    if not estado_nuevo_norm:
        return JsonResponse({
            'detail': 'El estado enviado no es válido.',
            'estado_permitido': ['Libre', 'Bloqueado', 'Sin Liberar']
        }, status=400)
    if estado_nuevo_norm == 'Bloqueado' and not (motivo_bloqueo or '').strip():
        return JsonResponse({'detail': 'Debes enviar "motivo_bloqueo" para bloquear un lote.'}, status=400)
    if estado_actual == estado_nuevo_norm:
        return JsonResponse({
            'detail': 'El lote ya se encuentra en el estado solicitado.',
            'estado_actual': estado_actual
        }, status=409)

    try:
        lote.estado = estado_nuevo_norm
        if estado_nuevo_norm == 'Bloqueado':
            lote.obsbloqueo = motivo_bloqueo
            lote.usuariobloquea = str(request.user)
            lote.fechadesbloque = datetime.datetime.today()
        else:
            lote.usuariobloquea = ''
            lote.obsbloqueo = ''
            lote.fechadesbloque = None
        lote.save(using=proyecto)
        return JsonResponse({
            'idinmueble': lote.idinmueble,
            'estado_anterior': estado_actual,
            'estado_actual': lote.estado
        }, status=200)
    except Exception as exc:
        return JsonResponse({'detail': 'Error al actualizar el estado del lote.', 'error': str(exc)}, status=500)


@api_token_auth
@require_http_methods(["GET"])
def api_bank_movements_list(request):
    if not request.user.is_authenticated or request.user.is_anonymous:
        return JsonResponse({'detail': 'Token inválido o no autenticado'}, status=401)
    cuenta_id = request.GET.get('cuenta')
    cuenta_numero = request.GET.get('cuenta_numero')
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    empresa_id = request.GET.get('empresa')
    estado = request.GET.get('estado')
    usado_agente = request.GET.get('usado_agente')
    valor_positivo = request.GET.get('valor_positivo')
    descripcion = request.GET.get('descripcion')
    page = request.GET.get('page', 1)
    page_size = request.GET.get('page_size', 100)
    
    if not any([cuenta_id, cuenta_numero, empresa_id]):
        return JsonResponse({
            'detail':'Debes enviar al menos "cuenta", "cuenta_numero" o "empresa"',
            'filters': {},
            'pagination': {'page': 1, 'page_size': 0, 'total_pages': 0, 'total_records': 0},
            'movimientos': []
        }, status=200)
    if not fecha_desde or not fecha_hasta:
        return JsonResponse({
            'detail':'Debes enviar "fecha_desde" y "fecha_hasta" (YYYY-MM-DD)',
            'filters': {},
            'pagination': {'page': 1, 'page_size': 0, 'total_pages': 0, 'total_records': 0},
            'movimientos': []
        }, status=200)

    fecha_desde_dt = parse_date(fecha_desde)
    fecha_hasta_dt = parse_date(fecha_hasta)
    if not fecha_desde_dt or not fecha_hasta_dt:
        return JsonResponse({
            'detail':'Formato de fecha inválido, usa YYYY-MM-DD',
            'filters': {},
            'pagination': {'page': 1, 'page_size': 0, 'total_pages': 0, 'total_records': 0},
            'movimientos': []
        }, status=200)
    if fecha_desde_dt > fecha_hasta_dt:
        return JsonResponse({
            'detail':'"fecha_desde" no puede ser mayor que "fecha_hasta"',
            'filters': {},
            'pagination': {'page': 1, 'page_size': 0, 'total_pages': 0, 'total_records': 0},
            'movimientos': []
        }, status=200)
    
    cuenta_obj = None
    if cuenta_id:
        try:
            # Validar que cuenta_id sea un entero válido
            cuenta_id_int = int(cuenta_id)
            cuenta_obj = cuentas_pagos.objects.get(pk=cuenta_id_int)
        except ValueError:
            return JsonResponse({
                'detail': f'El ID de cuenta "{cuenta_id}" no es válido. Debe ser un número entero.',
                'cuenta_enviada': cuenta_id,
                'filters': {},
                'pagination': {'page': 1, 'page_size': 0, 'total_pages': 0, 'total_records': 0},
                'movimientos': []
            }, status=200)
        except cuentas_pagos.DoesNotExist:
            return JsonResponse({
                'detail': f'No se encontró ninguna cuenta bancaria con el ID {cuenta_id}.',
                'cuenta_enviada': cuenta_id,
                'filters': {},
                'pagination': {'page': 1, 'page_size': 0, 'total_pages': 0, 'total_records': 0},
                'movimientos': []
            }, status=200)
    elif cuenta_numero:
        cuenta_obj = _get_account_by_number(cuenta_numero)
        if not cuenta_obj:
            return JsonResponse({
                'detail': f'No se encontró ninguna cuenta bancaria con el número "{cuenta_numero}".',
                'cuenta_enviada': cuenta_numero,
                'sugerencia': 'Verifica que el número de cuenta sea correcto y esté registrado en el sistema.',
                'filters': {},
                'pagination': {'page': 1, 'page_size': 0, 'total_pages': 0, 'total_records': 0},
                'movimientos': []
            }, status=200)
    
    empresa_obj = None
    if empresa_id:
        try:
            # Intentar buscar por pk (NIT) primero
            empresa_obj = empresas.objects.get(pk=empresa_id)
        except empresas.DoesNotExist:
            # Si no existe, intentar buscar por nombre
            try:
                empresa_obj = empresas.objects.get(nombre__iexact=empresa_id)
            except empresas.DoesNotExist:
                return JsonResponse({
                    'detail': f'No se encontró ninguna empresa con el NIT o nombre "{empresa_id}".',
                    'empresa_enviada': empresa_id,
                    'sugerencia': 'Verifica que el NIT o nombre de la empresa sea correcto.',
                    'filters': {},
                    'pagination': {'page': 1, 'page_size': 0, 'total_pages': 0, 'total_records': 0},
                    'movimientos': []
                }, status=200)
        if cuenta_obj and cuenta_obj.nit_empresa_id and cuenta_obj.nit_empresa_id != empresa_obj.pk:
            return JsonResponse({
                'detail': f'La cuenta bancaria {cuenta_obj.cuentabanco} no pertenece a la empresa {empresa_obj.nombre}.',
                'cuenta': cuenta_obj.pk,
                'cuenta_numero': cuenta_obj.cuentabanco,
                'empresa_cuenta': cuenta_obj.nit_empresa.nombre if cuenta_obj.nit_empresa else 'Sin empresa',
                'empresa_solicitada': empresa_obj.nombre,
                'filters': {},
                'pagination': {'page': 1, 'page_size': 0, 'total_pages': 0, 'total_records': 0},
                'movimientos': []
            }, status=200)
    
    queryset = egresos_banco.objects.filter(
        fecha__range = (fecha_desde_dt, fecha_hasta_dt)
    ).select_related('empresa','cuenta').order_by('fecha','pk')
    
    if cuenta_obj:
        queryset = queryset.filter(cuenta = cuenta_obj)
    if empresa_obj:
        queryset = queryset.filter(empresa = empresa_obj)
    if estado:
        estado = estado.upper()
        if estado not in ('CONCILIADO','SIN CONCILIAR'):
            return JsonResponse({'detail':'El estado debe ser CONCILIADO o SIN CONCILIAR'}, status=400)
        queryset = queryset.filter(estado = estado)
    if usado_agente is not None:
        if usado_agente.lower() in ('true', '1', 'yes'):
            queryset = queryset.filter(usado_agente=True)
        elif usado_agente.lower() in ('false', '0', 'no'):
            queryset = queryset.filter(usado_agente=False)
    if valor_positivo is not None:
        if valor_positivo.lower() in ('true', '1', 'yes'):
            queryset = queryset.filter(valor__gt=0)

    # Filtro por descripción con búsqueda difusa (fuzzy matching ~70% similitud)
    if descripcion:
        # Normalizar el término de búsqueda y dividir en palabras
        search_term_normalized = _normalize_text_for_search(descripcion)
        search_words = search_term_normalized.split()

        # Filtrar movimientos usando fuzzy matching
        matching_ids = []
        for mvto in queryset:
            if mvto.descripcion:
                desc_normalized = _normalize_text_for_search(mvto.descripcion)
                desc_words = desc_normalized.split()

                # Verificar que cada palabra de búsqueda tenga una coincidencia difusa en la descripción
                all_words_match = True
                for search_word in search_words:
                    # Buscar si alguna palabra de la descripción coincide con la palabra de búsqueda
                    word_found = any(_fuzzy_match(search_word, desc_word, threshold=0.7)
                                   for desc_word in desc_words)
                    if not word_found:
                        all_words_match = False
                        break

                if all_words_match:
                    matching_ids.append(mvto.id_mvto)

        queryset = queryset.filter(id_mvto__in=matching_ids)

    try:
        page = int(page)
        page_size = min(max(int(page_size), 1), 500)
    except (ValueError, TypeError):
        return JsonResponse({'detail':'"page" y "page_size" deben ser números enteros'}, status=400)
    
    paginator = Paginator(queryset, page_size)
    try:
        page_obj = paginator.page(page)
    except EmptyPage:
        return JsonResponse({'detail':'La página solicitada no existe'}, status=404)
    
    movimientos = []
    for mvto in page_obj.object_list:
        mvto_data = {
            'id_mvto': mvto.id_mvto,
            'empresa': mvto.empresa.pk,
            'cuenta': mvto.cuenta.pk,
            'cuenta_verbose': mvto.cuenta.cuentabanco,
            'fecha': mvto.fecha.isoformat(),
            'descripcion': mvto.descripcion,
            'referencia': mvto.referencia,
            'valor': float(mvto.valor),
            'estado': mvto.estado,
            'conciliacion_id': mvto.conciliacion_id,
            'pago_id': mvto.pago_asociado_id,
            'anticipo_id': mvto.anticipo_asociado_id,
            'transferencia_id': mvto.transferencia_asociada_id,
            'usado_agente': mvto.usado_agente,
            'fecha_uso_agente': mvto.fecha_uso_agente.isoformat() if mvto.fecha_uso_agente else None,
            'recibo_asociado_agente': mvto.recibo_asociado_agente,
            'proyecto_asociado_agente': mvto.proyecto_asociado_agente
        }

        # Agregar información relacionada para egresos (valor <= 0)
        if mvto.valor <= 0:
            mvto_data['info_relacionada'] = None

            # Si tiene pago asociado
            if mvto.pago_asociado:
                pago = mvto.pago_asociado
                radicado = pago.nroradicado
                mvto_data['info_relacionada'] = {
                    'tipo': 'pago',
                    'radicado': {
                        'nro': radicado.nroradicado,
                        'descripcion': radicado.descripcion,
                        'tercero': radicado.nombretercero,
                        'valor': radicado.valor,
                        'soporte_radicado_url': _media_url(radicado.soporte_radicado)
                    },
                    'pago': {
                        'valor': pago.valor,
                        'fecha_pago': pago.fecha_pago.isoformat(),
                        'soporte_pago_url': _media_url(pago.soporte_pago)
                    }
                }

            # Si tiene anticipo asociado
            elif mvto.anticipo_asociado:
                anticipo = mvto.anticipo_asociado
                mvto_data['info_relacionada'] = {
                    'tipo': 'anticipo',
                    'anticipo': {
                        'descripcion': anticipo.descripcion,
                        'tercero': anticipo.nombre_tercero,
                        'valor': anticipo.valor,
                        'fecha_pago': anticipo.fecha_pago.isoformat(),
                        'tipo_anticipo': anticipo.tipo_anticipo.descripcion if anticipo.tipo_anticipo else None,
                        'soporte_pago_url': _media_url(anticipo.soporte_pago)
                    }
                }

            # Si tiene transferencia asociada
            elif mvto.transferencia_asociada:
                transf = mvto.transferencia_asociada
                mvto_data['info_relacionada'] = {
                    'tipo': 'transferencia',
                    'transferencia': {
                        'fecha': transf.fecha.isoformat(),
                        'empresa_sale': transf.empresa_sale.nombre,
                        'cuenta_sale': transf.cuenta_sale.cuentabanco,
                        'empresa_entra': transf.empresa_entra.nombre,
                        'cuenta_entra': transf.cuenta_entra.cuentabanco,
                        'valor': transf.valor,
                        'soporte_pago_url': _media_url(transf.soporte_pago)
                    }
                }

        movimientos.append(mvto_data)
    
    data = {
        'filters': {
            'empresa': empresa_obj.pk if empresa_obj else None,
            'cuenta': cuenta_obj.pk if cuenta_obj else None,
            'fecha_desde': fecha_desde_dt.isoformat(),
            'fecha_hasta': fecha_hasta_dt.isoformat(),
            'estado': estado if estado else None,
            'usado_agente': usado_agente if usado_agente else None,
            'valor_positivo': valor_positivo if valor_positivo else None,
            'descripcion': descripcion if descripcion else None
        },
        'pagination': {
            'page': page,
            'page_size': page_size,
            'total_pages': paginator.num_pages,
            'total_records': paginator.count
        },
        'movimientos': movimientos
    }
    return JsonResponse(data, status=200)

@csrf_exempt
@api_token_auth
@login_required
@group_perm_required(('accounting.change_egresos_banco',),raise_exception=True)
@require_http_methods(["PATCH"])
def api_bank_movement_mark_used(request, movement_id):
    try:
        payload = json.loads(request.body.decode('utf-8'))
    except (json.JSONDecodeError, UnicodeDecodeError):
        return JsonResponse({'detail':'JSON inválido'}, status=400)

    try:
        movement = egresos_banco.objects.get(pk=movement_id)
    except egresos_banco.DoesNotExist:
        return JsonResponse({'detail':'Movimiento no encontrado', 'error_code':'NOT_FOUND'}, status=404)

    usado = payload.get('usado_agente')
    recibo = payload.get('recibo_asociado_agente')
    proyecto = payload.get('proyecto_asociado_agente')

    if usado is None:
        return JsonResponse({'detail':'Debes enviar "usado_agente"', 'error_code':'MISSING_USADO'}, status=400)

    movement.usado_agente = bool(usado)
    if movement.usado_agente:
        from django.utils import timezone
        movement.fecha_uso_agente = timezone.now()
        movement.recibo_asociado_agente = recibo if recibo else None
        movement.proyecto_asociado_agente = proyecto if proyecto else None
    else:
        movement.fecha_uso_agente = None
        movement.recibo_asociado_agente = None
        movement.proyecto_asociado_agente = None

    movement.save()

    return JsonResponse({
        'id_mvto': movement.pk,
        'usado_agente': movement.usado_agente,
        'fecha_uso_agente': movement.fecha_uso_agente.isoformat(sep=' ') if movement.fecha_uso_agente else None,
        'recibo_asociado_agente': movement.recibo_asociado_agente,
        'proyecto_asociado_agente': movement.proyecto_asociado_agente,
        'mensaje': 'Movimiento marcado como usado' if movement.usado_agente else 'Movimiento desmarcado'
    }, status=200)

@csrf_exempt
@api_token_auth
@login_required
@group_perm_required(('accounting.view_egresos_banco',),raise_exception=True)
@require_http_methods(["GET"])
def api_bank_movements_for_receipt(request):
    """
    Endpoint para obtener movimientos bancarios relacionados con una solicitud de recibo.
    Si la solicitud tiene recibo asociado, busca el movimiento por recibo+proyecto y usa esa cuenta.
    Si no tiene recibo, usa el proyecto para buscar la cuenta más común.
    Luego retorna movimientos con valor positivo en rango ±1 día de la fecha de pago.
    """
    if not request.user.is_authenticated or request.user.is_anonymous:
        return JsonResponse({'detail': 'Token inválido o no autenticado'}, status=401)

    proyecto = request.GET.get('proyecto')
    fecha_pago = request.GET.get('fecha_pago')
    recibo_asociado = request.GET.get('recibo_asociado')

    if not proyecto or not fecha_pago:
        return JsonResponse({'detail':'Debes enviar "proyecto" y "fecha_pago" (YYYY-MM-DD)'}, status=400)

    fecha_pago_dt = parse_date(fecha_pago)
    if not fecha_pago_dt:
        return JsonResponse({'detail':'Formato de fecha inválido, usa YYYY-MM-DD'}, status=400)

    # Calcular rango de fechas ±1 día
    from datetime import timedelta
    fecha_desde = fecha_pago_dt - timedelta(days=1)
    fecha_hasta = fecha_pago_dt + timedelta(days=1)

    cuenta_obj = None
    movimiento_asociado_id = None

    # Si tiene recibo asociado, buscar el movimiento por recibo
    # Primero intenta con recibo + proyecto, luego solo con recibo
    if recibo_asociado and recibo_asociado.strip():
        try:
            # Intentar primero con recibo + proyecto
            movimiento = egresos_banco.objects.filter(
                recibo_asociado_agente=recibo_asociado,
                proyecto_asociado_agente=proyecto
            ).first()

            # Si no encuentra, buscar solo por recibo
            if not movimiento:
                movimiento = egresos_banco.objects.filter(
                    recibo_asociado_agente=recibo_asociado
                ).first()

            if movimiento:
                cuenta_obj = movimiento.cuenta
                movimiento_asociado_id = movimiento.id_mvto
        except Exception:
            pass

    # Si no se encontró cuenta por recibo, buscar la cuenta más común del proyecto
    if not cuenta_obj:
        # Buscar movimientos recientes del proyecto para encontrar cuenta común
        movimientos_proyecto = egresos_banco.objects.filter(
            proyecto_asociado_agente=proyecto
        ).values('cuenta').annotate(
            count=Count('cuenta')
        ).order_by('-count').first()

        if movimientos_proyecto:
            try:
                cuenta_obj = cuentas_pagos.objects.get(pk=movimientos_proyecto['cuenta'])
            except cuentas_pagos.DoesNotExist:
                pass

    # Si aún no hay cuenta, retornar error
    if not cuenta_obj:
        return JsonResponse({
            'detail': 'No se encontró una cuenta bancaria asociada al proyecto',
            'movimientos': [],
            'cuenta': None,
            'movimiento_asociado_id': None
        }, status=200)

    # Buscar movimientos en la cuenta con valor positivo en el rango de fechas
    queryset = egresos_banco.objects.filter(
        cuenta=cuenta_obj,
        fecha__range=(fecha_desde, fecha_hasta),
        valor__gt=0
    ).select_related('empresa', 'cuenta').order_by('fecha', 'pk')

    movimientos = []
    for mvto in queryset:
        movimientos.append({
            'id_mvto': mvto.id_mvto,
            'empresa': mvto.empresa.pk,
            'cuenta': mvto.cuenta.pk,
            'fecha': mvto.fecha.isoformat(),
            'descripcion': mvto.descripcion,
            'referencia': mvto.referencia,
            'valor': float(mvto.valor),
            'estado': mvto.estado,
            'conciliacion_id': mvto.conciliacion_id,
            'pago_id': mvto.pago_asociado_id,
            'anticipo_id': mvto.anticipo_asociado_id,
            'transferencia_id': mvto.transferencia_asociada_id,
            'usado_agente': mvto.usado_agente,
            'fecha_uso_agente': mvto.fecha_uso_agente.isoformat(sep=' ') if mvto.fecha_uso_agente else None,
            'recibo_asociado_agente': mvto.recibo_asociado_agente,
            'proyecto_asociado_agente': mvto.proyecto_asociado_agente
        })

    return JsonResponse({
        'cuenta': cuenta_obj.pk,
        'cuenta_numero': cuenta_obj.cuentabanco,
        'movimiento_asociado_id': movimiento_asociado_id,
        'fecha_desde': fecha_desde.isoformat(),
        'fecha_hasta': fecha_hasta.isoformat(),
        'total_movimientos': len(movimientos),
        'movimientos': movimientos
    }, status=200)

@csrf_exempt
@api_token_auth
@login_required
@group_perm_required(('accounting.add_plinkmovement',),raise_exception=True)
@require_http_methods(["POST"])
def api_plink_movements(request):
    try:
        payload = json.loads(request.body.decode('utf-8'))
    except (json.JSONDecodeError, UnicodeDecodeError):
        return JsonResponse({'detail':'JSON inválido'}, status=400)
    empresa_id = payload.get('empresa')
    cuenta_payload = payload.get('cuenta')
    cuenta_normalizada_payload = _normalize_account(cuenta_payload)
    movimientos = payload.get('movimientos')
    if not empresa_id or not movimientos:
        return JsonResponse({'detail':'"empresa" y "movimientos" son obligatorios'}, status=400)
    try:
        empresa_obj = empresas.objects.get(pk=empresa_id)
    except empresas.DoesNotExist:
        return JsonResponse({'detail':'Empresa no encontrada'}, status=404)
    registros = []
    with transaction.atomic():
        for index, movimiento in enumerate(movimientos, start=1):
            fecha_transaccion = _parse_compact_date(movimiento.get('FECHA DE TRANSACCION'))
            if not fecha_transaccion:
                return JsonResponse({'detail':f'La fecha de transacción del movimiento #{index} es inválida'}, status=400)
            fecha_canje = _parse_compact_date(movimiento.get('FECHA DE CANJE'))
            fecha_compensacion = _parse_compact_date(movimiento.get('FECHA DE COMPENSACION'))
            cuenta_mov = movimiento.get('CUENTA DE CONSIGNACION') or cuenta_payload
            cuenta_norm = _normalize_account(cuenta_mov) or cuenta_normalizada_payload
            obj = PlinkMovement.objects.create(
                empresa = empresa_obj,
                nit = movimiento.get('NIT','').strip(),
                codigo_establecimiento = movimiento.get('CODIGO ESTABLECIMIENTO','').strip(),
                origen_compra = movimiento.get('ORIGEN DE LA COMPRA'),
                tipo_transaccion = movimiento.get('TIPO TRANSACCION'),
                franquicia = movimiento.get('FRANQUICIA'),
                identificador_red = movimiento.get('IDENTIFICADOR DE RED'),
                fecha_transaccion = fecha_transaccion,
                fecha_canje = fecha_canje,
                cuenta_consignacion = cuenta_mov,
                valor_compra = _parse_decimal(movimiento.get('VALOR COMPRA')),
                valor_propina = _parse_decimal(movimiento.get('VALOR PROPINA')),
                valor_iva = _parse_decimal(movimiento.get('VALOR IVA')),
                valor_impoconsumo = _parse_decimal(movimiento.get('VALOR IMPOCONSUMO')),
                valor_total = _parse_decimal(movimiento.get('VALOR TOTAL')),
                valor_comision = _parse_decimal(movimiento.get('VALOR COMISION')),
                valor_retefuente = _parse_decimal(movimiento.get('VALOR RETEFUENTE')),
                valor_rete_iva = _parse_decimal(movimiento.get('VALOR RETE IVA')),
                valor_rte_ica = _parse_decimal(movimiento.get('VALOR RTE ICA')),
                valor_provision = _parse_decimal(movimiento.get('VALOR PROVISION')),
                valor_neto = _parse_decimal(movimiento.get('VALOR NETO')),
                codigo_autorizacion = movimiento.get('CODIGO AUTORIZACION'),
                tipo_tarjeta = movimiento.get('TIPO TARJETA'),
                numero_terminal = movimiento.get('NO TERMINAL'),
                tarjeta = movimiento.get('TARJETA'),
                comision_porcentual = movimiento.get('COMISION PORCENTUAL'),
                comision_base = _parse_decimal(movimiento.get('COMISION BASE')),
                fecha_compensacion = fecha_compensacion,
                cuenta_normalizada = cuenta_norm,
            )
            registros.append({
                'id': obj.pk,
                'codigo_autorizacion': obj.codigo_autorizacion,
                'fecha_transaccion': obj.fecha_transaccion.isoformat()
            })
    return JsonResponse({
        'empresa': empresa_obj.pk,
        'total': len(registros),
        'movimientos': registros
    }, status=201)

@api_token_auth
@login_required
@group_perm_required(('accounting.view_plinkmovement',),raise_exception=True)
@require_http_methods(["GET"])
def api_plink_movements_list(request):
    empresa_id = request.GET.get('empresa')
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    usado_agente = request.GET.get('usado_agente')
    if not empresa_id or not fecha_desde or not fecha_hasta:
        return JsonResponse({'detail':'Debes enviar "empresa", "fecha_desde" y "fecha_hasta"'}, status=400)
    fecha_desde_dt = parse_date(fecha_desde)
    fecha_hasta_dt = parse_date(fecha_hasta)
    if not fecha_desde_dt or not fecha_hasta_dt:
        return JsonResponse({'detail':'Formato de fecha inválido'}, status=400)
    if fecha_desde_dt > fecha_hasta_dt:
        return JsonResponse({'detail':'"fecha_desde" no puede ser mayor que "fecha_hasta"'}, status=400)
    try:
        empresa_obj = empresas.objects.get(pk=empresa_id)
    except empresas.DoesNotExist:
        return JsonResponse({'detail':'Empresa no encontrada'}, status=404)
    queryset = PlinkMovement.objects.filter(
        empresa = empresa_obj,
        fecha_transaccion__range = (fecha_desde_dt, fecha_hasta_dt)
    )
    # Filtrar por usado_agente si se proporciona
    if usado_agente is not None:
        if usado_agente.lower() in ('true', '1'):
            queryset = queryset.filter(usado_agente=True)
        elif usado_agente.lower() in ('false', '0'):
            queryset = queryset.filter(usado_agente=False)
    queryset = queryset.order_by('fecha_transaccion','pk')
    data = []
    for mv in queryset:
        data.append({
            'id': mv.pk,
            'fecha_transaccion': mv.fecha_transaccion.isoformat(),
            'fecha_canje': mv.fecha_canje.isoformat() if mv.fecha_canje else None,
            'codigo_autorizacion': mv.codigo_autorizacion,
            'valor_total': float(mv.valor_total),
            'valor_neto': float(mv.valor_neto),
            'cuenta_consignacion': mv.cuenta_consignacion,
            'cuenta_normalizada': mv.cuenta_normalizada,
            'tipo_transaccion': mv.tipo_transaccion,
            'franquicia': mv.franquicia,
            'usado_agente': mv.usado_agente,
            'fecha_uso_agente': mv.fecha_uso_agente.isoformat(sep=' ') if mv.fecha_uso_agente else None,
            'recibo_asociado_agente': mv.recibo_asociado_agente,
        })
    return JsonResponse({
        'empresa': empresa_obj.pk,
        'total': queryset.count(),
        'movimientos': data
    }, status=200)

@csrf_exempt
@api_token_auth
@login_required
@group_perm_required(('accounting.add_wompimovement',),raise_exception=True)
@require_http_methods(["POST"])
def api_wompi_movements(request):
    try:
        payload = json.loads(request.body.decode('utf-8'))
    except (json.JSONDecodeError, UnicodeDecodeError):
        return JsonResponse({'detail':'JSON inválido'}, status=400)
    empresa_id = payload.get('empresa')
    cuenta_payload = payload.get('cuenta')
    cuenta_normalizada_payload = _normalize_account(cuenta_payload)
    movimientos = payload.get('movimientos')
    if not empresa_id or not movimientos:
        return JsonResponse({'detail':'"empresa" y "movimientos" son obligatorios'}, status=400)
    try:
        empresa_obj = empresas.objects.get(pk=empresa_id)
    except empresas.DoesNotExist:
        return JsonResponse({'detail':'Empresa no encontrada'}, status=404)
    registros = []
    with transaction.atomic():
        for index, movimiento in enumerate(movimientos, start=1):
            fecha = _parse_datetime(movimiento.get('fecha'))
            if not fecha:
                return JsonResponse({'detail':f'La fecha del movimiento #{index} es inválida'}, status=400)
            transaction_id = movimiento.get('id de la transaccion')
            if not transaction_id:
                return JsonResponse({'detail':f'El id de la transacción del movimiento #{index} es obligatorio'}, status=400)
            obj = WompiMovement.objects.create(
                empresa = empresa_obj,
                transaction_id = transaction_id,
                fecha = fecha,
                referencia = movimiento.get('referencia',''),
                monto = _parse_decimal(movimiento.get('monto')),
                iva = _parse_decimal(movimiento.get('iva')),
                impuesto_consumo = _parse_decimal(movimiento.get('impuesto al consumo')),
                moneda = movimiento.get('moneda','COP'),
                medio_pago = movimiento.get('medio de pago',''),
                email_pagador = movimiento.get('email del pagador'),
                nombre_pagador = movimiento.get('nombre del pagador'),
                telefono_pagador = movimiento.get('telefono del pagador'),
                id_conciliacion = movimiento.get('id conciliacion'),
                id_link_pago = movimiento.get('id link de pago'),
                documento_pagador = movimiento.get('documento del pagador'),
                tipo_documento_pagador = movimiento.get('tipo de documento del pagador'),
                referencia_1_nombre = movimiento.get('ref. 1 nombre'),
                referencia_1 = movimiento.get('ref. 1'),
                cuenta_normalizada = cuenta_normalizada_payload,
            )
            registros.append({
                'id': obj.pk,
                'transaction_id': obj.transaction_id,
                'fecha': obj.fecha.isoformat()
            })
    return JsonResponse({
        'empresa': empresa_obj.pk,
        'total': len(registros),
        'movimientos': registros
    }, status=201)

@api_token_auth
@login_required
@group_perm_required(('accounting.view_wompimovement',),raise_exception=True)
@require_http_methods(["GET"])
def api_wompi_movements_list(request):
    empresa_id = request.GET.get('empresa')
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    usado_agente = request.GET.get('usado_agente')
    if not empresa_id or not fecha_desde or not fecha_hasta:
        return JsonResponse({'detail':'Debes enviar "empresa", "fecha_desde" y "fecha_hasta"'}, status=400)
    fecha_desde_dt = parse_date(fecha_desde)
    fecha_hasta_dt = parse_date(fecha_hasta)
    if not fecha_desde_dt or not fecha_hasta_dt:
        return JsonResponse({'detail':'Formato de fecha inválido'}, status=400)
    if fecha_desde_dt > fecha_hasta_dt:
        return JsonResponse({'detail':'"fecha_desde" no puede ser mayor que "fecha_hasta"'}, status=400)
    try:
        empresa_obj = empresas.objects.get(pk=empresa_id)
    except empresas.DoesNotExist:
        return JsonResponse({'detail':'Empresa no encontrada'}, status=404)
    queryset = WompiMovement.objects.filter(
        empresa = empresa_obj,
        fecha__date__gte = fecha_desde_dt,
        fecha__date__lte = fecha_hasta_dt
    )
    # Filtrar por usado_agente si se proporciona
    if usado_agente is not None:
        if usado_agente.lower() in ('true', '1'):
            queryset = queryset.filter(usado_agente=True)
        elif usado_agente.lower() in ('false', '0'):
            queryset = queryset.filter(usado_agente=False)
    queryset = queryset.order_by('fecha','pk')
    data = []
    for mv in queryset:
        data.append({
            'id': mv.pk,
            'transaction_id': mv.transaction_id,
            'fecha': mv.fecha.isoformat(sep=' '),
            'referencia': mv.referencia,
            'monto': float(mv.monto),
            'medio_pago': mv.medio_pago,
            'email_pagador': mv.email_pagador,
            'id_conciliacion': mv.id_conciliacion,
            'cuenta_normalizada': mv.cuenta_normalizada,
            'usado_agente': mv.usado_agente,
            'fecha_uso_agente': mv.fecha_uso_agente.isoformat(sep=' ') if mv.fecha_uso_agente else None,
            'recibo_asociado_agente': mv.recibo_asociado_agente,
        })
    return JsonResponse({
        'empresa': empresa_obj.pk,
        'total': queryset.count(),
        'movimientos': data
    }, status=200)

@csrf_exempt
@api_token_auth
@login_required
@group_perm_required(('accounting.change_plinkmovement',),raise_exception=True)
@require_http_methods(["PATCH"])
def api_plink_movement_mark_used(request, movement_id):
    try:
        payload = json.loads(request.body.decode('utf-8'))
    except (json.JSONDecodeError, UnicodeDecodeError):
        return JsonResponse({'detail':'JSON inválido'}, status=400)

    try:
        movement = PlinkMovement.objects.get(pk=movement_id)
    except PlinkMovement.DoesNotExist:
        return JsonResponse({'detail':'Movimiento no encontrado', 'error_code':'NOT_FOUND'}, status=404)

    usado = payload.get('usado_agente')
    recibo = payload.get('recibo_asociado_agente')

    if usado is None:
        return JsonResponse({'detail':'Debes enviar "usado_agente"', 'error_code':'MISSING_USADO'}, status=400)

    movement.usado_agente = bool(usado)
    if movement.usado_agente:
        from django.utils import timezone
        movement.fecha_uso_agente = timezone.now()
        movement.recibo_asociado_agente = recibo if recibo else None
    else:
        movement.fecha_uso_agente = None
        movement.recibo_asociado_agente = None

    movement.save()

    return JsonResponse({
        'id': movement.pk,
        'usado_agente': movement.usado_agente,
        'fecha_uso_agente': movement.fecha_uso_agente.isoformat(sep=' ') if movement.fecha_uso_agente else None,
        'recibo_asociado_agente': movement.recibo_asociado_agente,
        'mensaje': 'Movimiento marcado como usado' if movement.usado_agente else 'Movimiento desmarcado'
    }, status=200)

@csrf_exempt
@api_token_auth
@login_required
@group_perm_required(('accounting.change_wompimovement',),raise_exception=True)
@require_http_methods(["PATCH"])
def api_wompi_movement_mark_used(request, movement_id):
    try:
        payload = json.loads(request.body.decode('utf-8'))
    except (json.JSONDecodeError, UnicodeDecodeError):
        return JsonResponse({'detail':'JSON inválido'}, status=400)

    try:
        movement = WompiMovement.objects.get(pk=movement_id)
    except WompiMovement.DoesNotExist:
        return JsonResponse({'detail':'Movimiento no encontrado', 'error_code':'NOT_FOUND'}, status=404)

    usado = payload.get('usado_agente')
    recibo = payload.get('recibo_asociado_agente')

    if usado is None:
        return JsonResponse({'detail':'Debes enviar "usado_agente"', 'error_code':'MISSING_USADO'}, status=400)

    movement.usado_agente = bool(usado)
    if movement.usado_agente:
        from django.utils import timezone
        movement.fecha_uso_agente = timezone.now()
        movement.recibo_asociado_agente = recibo if recibo else None
    else:
        movement.fecha_uso_agente = None
        movement.recibo_asociado_agente = None

    movement.save()

    return JsonResponse({
        'id': movement.pk,
        'usado_agente': movement.usado_agente,
        'fecha_uso_agente': movement.fecha_uso_agente.isoformat(sep=' ') if movement.fecha_uso_agente else None,
        'recibo_asociado_agente': movement.recibo_asociado_agente,
        'mensaje': 'Movimiento marcado como usado' if movement.usado_agente else 'Movimiento desmarcado'
    }, status=200)

@login_required
@group_perm_required(('accounting.view_planilla_movimiento',),raise_exception=True)
def planilla_movimientos(request):
    context={
        'form':forms.form_nueva_planilla
    }
    return render(request,'accounting/planilla_mvto.html',context)

@login_required
@group_perm_required(('accounting.add_conciliaciones',),raise_exception=True)
def view_conciliaciones(request):
    context = {
        'empresas':empresas.objects.all()
    }
    
    return render(request,'accounting/conciliaciones.html',context)

@login_required
@group_perm_required(('accounting.view_conciliaciones',),raise_exception=True)
def lista_conciliaciones(request):
    context = {
        'empresas':empresas.objects.all()
    }
    return render(request,'accounting/reconciliationslist.html',context)

@login_required
@group_perm_required(('accounting.add_facturas',),raise_exception=True)
def radicar_factura(request):
    context = {
        'form':forms.form_nuevo_radicado
    }
    if request.method == 'POST':
        form = forms.form_nuevo_radicado(request.POST,request.FILES)
        if form.is_valid():
            nro_factura = form.cleaned_data.get('nro_factura')
            empresa = form.cleaned_data.get('empresa')
            id_tercero = form.cleaned_data.get('id_tercero')
            nombre_tercero = form.cleaned_data.get('nombre_tercero')
            fecha_factura = form.cleaned_data.get('fecha_factura')
            fecha_vencimiento = form.cleaned_data.get('fecha_vencimiento')
            descripcion = form.cleaned_data.get('descripcion')
            valor = form.cleaned_data.get('valor')
            oficina = form.cleaned_data.get('oficina')
            soporte = request.FILES['soporte']
            
            radicado = Facturas.objects.create(
                empresa=empresa,nrofactura=nro_factura,
                fechafactura=fecha_factura,idtercero=id_tercero,
                fechavenc=fecha_vencimiento,nombretercero=nombre_tercero,
                oficina=oficina,valor=valor,soporte_radicado=soporte,
                descripcion=descripcion
            )
            action = f'Radicó la factura {nro_factura}'
            
            history_facturas.objects.create(
                factura=radicado,usuario=request.user,
                accion=action,
                ubicacion='Recepcion'
            )
            context.update({
                'topalert':{
                    'alert':True,
                    'title':'Listo!',
                    'message':f'Se radicó la factura con el consecutivo Nº {radicado.pk}',
                    'class':'alert-success'
                }
            })
            
            msj = f'''Hola :), <b>{request.user.get_full_name()}</b> ha radicado una nueva factura:<br>
                <ul>
                    <li>Id: {radicado.pk}</li>
                    <li>Empresa: {radicado.empresa.nombre}</li>
                    <li>Oficina: {radicado.oficina}</li>
                    <li>Tercero: {radicado.nombretercero}</li>
                    <li>Valor: ${radicado.valor:,}</li>
                </ul>'''
            
            envio_notificacion(msj,'Nueva factura radicada',['auxiliarcontable2@somosandina.co','davidcalao@somosandina.co'])
            
        else:
            context['form']=form
            
    return render(request,'accounting/radicar_factura.html',context)

@login_required
@group_perm_required(('accounting.change_facturas',),raise_exception=True)
def causar_factura(request):
    context = {
        'form_causar':forms.form_causar_rad,
        'form_rad_int':forms.form_rad_int,
    }
    facts_nr_contabilidad = Facturas.objects.filter(nrocausa__isnull=True)
    sin_recibir_mtr = 0
    sin_recibir_mde = 0
    for fact in facts_nr_contabilidad:
        ubicacion = history_facturas.objects.filter(factura=fact.pk).last().ubicacion
        if ubicacion == 'Recepcion':
            if fact.oficina == 'MONTERIA':
                sin_recibir_mtr += 1
            elif fact.oficina == 'MEDELLIN':
                sin_recibir_mde += 1
    if (sin_recibir_mtr + sin_recibir_mde)  >0:
        context['topalert']={
                'alert':True,
                'title':'Pendientes',
                'message':f'''Hay facturas radicadas sin recibir<br/>
                            <ul>
                                <li>Monteria: <strong>{sin_recibir_mtr}</strong></li>
                                <li>Medellin: <strong>{sin_recibir_mde}</strong></li>
                            </ul>''',
                'class':'alert-info'
                    }
    if request.method =='POST':
        if not request.is_ajax():
            if request.POST.get('btnRadInt'):
                id_tercero = request.POST.get('id_tercero_rad_int')
                nombre_tercero = request.POST.get('nombre_tercero_rad_int')
                oficina = request.POST.get('oficina_rad_int')
                empresa = request.POST.get('empresa_rad_int')
                tipo_rad = request.POST.get('tipo_radicado')
                causacion = request.POST.get('documento_causacion')
                cxp = request.POST.get('tipo_cxp')
                descripcion = request.POST.get('descripcion_rad_int')
                fecha = request.POST.get('fecha_rad_int')
                valor = request.POST.get('valor_neto_rad_int').replace(',','')
                soporte_causacion = request.FILES.get('soporte_causacion')
                proyecto_ascociado = request.POST.get('proyecto_ascociado')
                centro_costo = request.POST.get('centro_costo')
                
                nrofact = f'INT-{empresa}-{causacion}'
                obj_empresa = empresas.objects.get(pk=empresa)
                obj_cxp = info_interfaces.objects.get(pk=cxp)
                check_pagado = request.POST.get('check_pagado')
                try:
                    soporte = request.FILES['soporte_rad_int']
                except MultiValueDictKeyError:
                    soporte= None
                
                radicado = Facturas.objects.create(
                    empresa=obj_empresa,descripcion=descripcion,
                    fechafactura=fecha,nrofactura=nrofact,
                    idtercero=id_tercero,nombretercero=nombre_tercero,
                    valor=valor,pago_neto=valor,fechavenc=fecha,nrocausa=causacion,
                    cuenta_por_pagar=obj_cxp,fechacausa=fecha,origen='Interno',
                    oficina = oficina,soporte_radicado=soporte,secuencia_cxp=1,
                    soporte_causacion = soporte_causacion, proyecto_rel=proyecto_ascociado,
                    centro_costo = centro_costo
                )
                
                
                mensaje = f'Se creo el radicado interno con numero {radicado.pk}'
                action = f'Creó el radicado interno'
                ubicacion = 'Contabilidad'
                
                if check_pagado == 'on':
                    empresa_pago = request.POST.get('empresa_pago_rad')
                    empresa_obj = empresas.objects.get(pk=empresa_pago)
                    cuenta = request.POST.get('cuenta_pago_rad')
                    fecha_pago = request.POST.get('fecha_pago_rad')
                    try:
                        soporte_pago = request.FILES['soporte_pago_rad']
                    except MultiValueDictKeyError:
                        soporte_pago= None
                    obj_cuenta = cuentas_pagos.objects.get(pk=cuenta)
                    pago = Pagos.objects.create(empresa=empresa_obj,nroradicado=radicado,
                                                valor=valor,usuario=request.user,
                                                cuenta=obj_cuenta,fecha_pago=fecha_pago,
                                                soporte_pago=soporte_pago)
                    mensaje += ' y se registró pago correspondiente'
                    action += ' con pago asociado.'
                    ubicacion = 'Tesoreria'
                history_facturas.objects.create(
                    factura=radicado,usuario=request.user,
                    accion=action,
                    ubicacion=ubicacion
                )
                context['topalert'] = {
                        'alert':True,
                        'title':'Hecho!',
                        'message':mensaje,
                        'class':'alert-success'
                    }
                
                val = request.POST.get('valor_neto_rad_int')
                msj = f'''Hola :), <b>{request.user.get_full_name()}</b> ha causado una nueva factura:<br>
                    <ul>
                        <li>Id: {radicado.pk}</li>
                        <li>Empresa: {radicado.empresa.nombre}</li>
                        <li>Oficina: {radicado.oficina}</li>
                        <li>Tercero: {radicado.nombretercero}</li>
                        <li>Causacion: {radicado.nrocausa}</li>
                        <li>Valor: ${val}</li>
                        <li>Pago neto: ${val}</li>
                    </ul>'''
                
                
                asunto = f'Nuevo estado: Radicado {radicado.pk} CAUSADO'
                
                if radicado.oficina == 'MONTERIA':
                    
                    envio_notificacion(msj,asunto,['admin1@somosandina.co','recepcionmtr@somosandina.co'])
                
                else:
                    
                    envio_notificacion(msj,asunto,['operaciones2@somosandina.co','auxiliartesoreria@somosandina.co'])
                
                  
    return render(request,'accounting/causar_factura.html',context)

@login_required
@group_perm_required(('accounting.add_pagos',),raise_exception=True)
def pagar_factura(request):
    context = {
        'empresas':empresas.objects.all(),
        'form_facturas':forms.form_nuevo_pago,
        'form_anticipo':forms.form_anticipos_tesor,
        'form_transf':forms.form_transferencias,
        'formgtt':forms.form_asociar_gtt,
        'formcomis':forms.form_asociar_comision,
        'formnomina':forms.form_asociar_nomina,
        'formotrospag':forms.form_asociar_otros_pagos,
        'form_buscar_mvto': forms.form_buscar_mvto_pago,
    }
    sin_recibir_mde = 0
    sin_recibir_mtr = 0

    pendientes_qs = info_facturas.objects.filter(
        ubicacion='Contabilidad'
    ).exclude(
        Q(causacion__isnull=True) | Q(causacion__exact='')
    )
    pendientes_por_oficina = pendientes_qs.values('oficina').annotate(total=Count('radicado'))
    for row in pendientes_por_oficina:
        oficina = row.get('oficina')
        total = row.get('total') or 0
        if oficina == 'MONTERIA':
            sin_recibir_mtr = total
        elif oficina == 'MEDELLIN':
            sin_recibir_mde = total
    if (sin_recibir_mtr + sin_recibir_mde)>0:
        context['topalert']={                                                                                                                                                                                                                       
                'alert':True,
                'title':'Pendientes',
                'message':f'''Hay facturas causadas sin recibir<br/>
                            <ul>
                                <li>Monteria: <strong>{sin_recibir_mtr}</strong></li>
                                <li>Medellin: <strong>{sin_recibir_mde}</strong></li>
                            </ul>''',
                'class':'alert-info'
                    }
    return render(request,'accounting/facturas_tesoreria.html',context)

@login_required
@group_perm_required(('accounting.view_facturas',),raise_exception=True)
def lista_facturas(request):
    context = {
    }
    
    return render(request,'accounting/lista_facturas.html',context)

@login_required
@group_perm_required(('accounting.view_pagos',),raise_exception=True)
def lista_pagos(request):
    context = {
        'empresas' : empresas.objects.all()    
    }
    
    return render(request,'accounting/lista_pagos.html',context)

@login_required
@group_perm_required(('accounting.view_otros_ingresos',),raise_exception=True)
def view_otros_ingresos(request):
    context = {
        'form':forms.form_otros_ingresos
        
    }
    if request.method == 'POST':
        check_perms(request,('accounting.add_otros_ingresos',))
        if request.is_ajax():
            form = forms.form_otros_ingresos(request.POST)
            if form.is_valid():
                fecha = form.cleaned_data.get('fecha')
                id_tercero = form.cleaned_data.get('id_tercero')
                nombre_tercero = form.cleaned_data.get('nombre_tercero')
                concepto = form.cleaned_data.get('concepto')
                empresa = form.cleaned_data.get('empresa')
                cuenta = form.cleaned_data.get('cuenta')
                oficina = form.cleaned_data.get('oficina')
                valor = form.cleaned_data.get('valor').replace(',','')
                
                obj_empresa = empresa
                obj_cuenta = cuenta
                
                ingreso = otros_ingresos.objects.create(
                    empresa=obj_empresa,descripcion=concepto,valor=valor,
                    oficina=oficina,usuario=request.user,cuenta=obj_cuenta,id_tercero=id_tercero,
                    nombre_tercero=nombre_tercero,fecha_ing=fecha
                )
                consecutivo_ofic = otros_ingresos.objects.filter(pk__lte=ingreso.pk,
                        oficina=ingreso.oficina).count()
                filename = f'Recibo_ingreso_{ingreso.oficina}_{consecutivo_ofic}.pdf'
                ruta = settings.MEDIA_ROOT+'/tmp/pdf/'+filename
                GenerarPDF().reciboIngreso(ingreso,consecutivo_ofic,ruta)
                ruta_download = settings.MEDIA_URL + 'tmp/pdf/' + filename
                texto = f'Se creó el recibo de ingreso #{consecutivo_ofic}, puedes descargarlo <a href="{ruta_download}" class="font-weight-bold" target="_blank">Aqui</a>'
                data = {
                    'mensaje':{
                        'message':texto,
                    }
                }
            else:
                data = {
                    'mensaje':{
                        'message':form.errors,
                        
                    }
                }
            return JsonResponse(data)
    
    return render(request,'accounting/otros_ingresos.html',context)

#ajax
def ajax_print_interfaz_banco(request):
    if request.method == 'POST':
        desde = request.POST.get('fecha_desde')
        hasta = request.POST.get('fecha_hasta')
        empresa = request.POST.get('empresa')
        cuenta = request.POST.get('cuenta')
        proyecto = request.POST.get('proyecto')
        nits = {
            'Promotora Sandville':900993044,
            'Promotora Westville':901132949,
            'Status Comercializadora':901018375,
            'Quadrata Constructores':901004733,
            'Terranova Desarrolladora Turistica':900712229
        }
        tipo = request.POST.get('tipo')
        
        nit_empresa = nits.get(empresa)
        cuenta_debitar = int(cuenta.replace('-',""))
        book=openpyxl.load_workbook("resources/excel_formats/InterfazPAB.xlsx")
        sheet=book.active
        sheet.cell(2,1,nit_empresa)
        sheet.cell(2,2,220)
        letra=random.choice(string.ascii_uppercase)
        numero=random.randint(1,9)
        secuencia=f'{letra}{numero}'
        sheet.cell(2,3,"I")
        sheet.cell(2,4,secuencia)
        sheet.cell(2,5,cuenta_debitar)
        nom_proy=proyecto.replace(' ',"").upper()[:4]
        sheet.cell(2,6,'D')
        
        if tipo == 'comision':
            sheet.cell(2,7,'COMISIONES')
            stmt=f'CALL detalle_comisiones_fecha("{desde}","{hasta}")'
            comisiones=Pagocomision.objects.using(proyecto).raw(stmt)
            
            row=4
            for line in comisiones:
                datos_asesor = asesores.objects.get(pk=line.idgestor)
                sheet.cell(row,1,1)
                sheet.cell(row,2,int(line.idgestor))
                sheet.cell(row,3,line.nombre.upper()[:31])
                if datos_asesor.tipo_cuenta == 'Ahorros': tipo_cuenta=37
                elif datos_asesor.tipo_cuenta == 'Corriente': tipo_cuenta=40
                else: tipo_cuenta=""
                sheet.cell(row,4,tipo_cuenta)
                try: banco=entidades_bancarias.objects.get(banco=datos_asesor.banco).codigo
                except: banco=""
                sheet.cell(row,5,banco)
                try: sheet.cell(row,6,int(datos_asesor.cuenta))
                except:pass
                referencia = proyecto.replace(" ","").upper()
                ref=f'COMIS{referencia}'[:20]
                sheet.cell(row,9,ref)
                sheet.cell(row,11,line.pagoneto)
                row+=1
            
            nombre_doc=f'Interfaz_PAB_Comisiones_{proyecto}_{desde}_{hasta}.xlsx'
            
        elif tipo == 'gtt':
            sheet.cell(2,7,'GTT')
            gtts = Detalle_gtt.objects.filter(gtt__fecha_desde=desde,gtt__fecha_hasta=hasta,
                                              gtt__proyecto=proyecto)
            row=4
            for line in gtts:
                datos_asesor = line.asesor
                sheet.cell(row,1,1)
                sheet.cell(row,2,int(datos_asesor.pk))
                sheet.cell(row,3,datos_asesor.nombre.upper()[:31])
                if datos_asesor.tipo_cuenta == 'Ahorros': tipo_cuenta=37
                elif datos_asesor.tipo_cuenta == 'Corriente': tipo_cuenta=40
                else: tipo_cuenta=""
                sheet.cell(row,4,tipo_cuenta)
                try: banco=entidades_bancarias.objects.get(banco=datos_asesor.banco).codigo
                except: banco=""
                sheet.cell(row,5,banco)
                try: sheet.cell(row,6,int(datos_asesor.cuenta))
                except:pass
                referencia = proyecto.replace(" ","").upper()
                ref=f'GTT{referencia}'[:20]
                sheet.cell(row,9,ref)
                sheet.cell(row,11,line.valor)
                row+=1
            nombre_doc=f'Interfaz_PAB_GTT_{proyecto}_{desde}_{hasta}.xlsx'
            
        ruta=settings.MEDIA_ROOT+'/tmp/'+nombre_doc
        book.save(ruta)
        
        data = {
            'ruta': settings.MEDIA_URL + 'tmp/' + nombre_doc
        }
        
        return JsonResponse(data)

@login_required
def impr_interf_egresos(request):
    if request.method == 'POST':
        if request.is_ajax():
            try:
                empresa_egr = request.POST.get('empresa_egr')
                cuentas_egr = request.POST.getlist('cuentas_egr')
                egr_desde = request.POST.get('egr_desde')
                egr_hasta = request.POST.get('egr_hasta')
                consec_egr = int(request.POST.get('numero_inicial_egr'))
                list_mvtos = []
                for cuenta in cuentas_egr:
                    cuenta = int(cuenta)
                    documento = docs_cuentas_oficinas.objects.get(cuenta=cuenta).documento
                    
                    obj_transf = transferencias_companias.objects.filter(
                        cuenta_sale=cuenta,
                        fecha__gte=egr_desde,fecha__lte=egr_hasta,  
                        empresa_entra=F('empresa_sale')).order_by('empresa_sale')
                    for mvto in obj_transf:
                        list_mvtos.append({
                            'pk':mvto.pk,
                            'fecha':mvto.fecha,
                            'tipo_doc':documento.split('-')[0],
                            'codigo_doc':documento.split('-')[1],
                            'cuenta':mvto.cuenta_entra.nro_cuentacontable,
                            'natur_op':'D',
                            'valor_pago':mvto.valor,
                            'anno_pago':mvto.fecha.year,
                            'mes_pago':mvto.fecha.month,
                            'dia_pago':mvto.fecha.day,
                            'nit_pago':'',
                            'descrip_sec':f'TRANSFERENCIA DESDE CUENTA {mvto.cuenta_sale.cuentabanco}',
                            'tipo_comp_cruce':'',
                            'num_cruce':'',
                            'vencimiento_cruce':'',
                            'anno_cruce':'',
                            'mes_cruce':'',
                            'dia_cruce':'',
                        })
                        list_mvtos.append({
                            'pk':mvto.pk,
                            'fecha':mvto.fecha,
                            'tipo_doc':documento.split('-')[0],
                            'codigo_doc':documento.split('-')[1],
                            'cuenta':mvto.cuenta_sale.nro_cuentacontable,
                            'natur_op':'C',
                            'valor_pago':mvto.valor,
                            'anno_pago':mvto.fecha.year,
                            'mes_pago':mvto.fecha.month,
                            'dia_pago':mvto.fecha.day,
                            'nit_pago':'',
                            'descrip_sec':f'TRANSFERENCIA HACIA CUENTA {mvto.cuenta_entra.cuentabanco}',
                            'tipo_comp_cruce':'',
                            'num_cruce':'',
                            'vencimiento_cruce':'',
                            'anno_cruce':'',
                            'mes_cruce':'',
                            'dia_cruce':'',
                            'salta':True,
                        })
                    obj_anticipos = Anticipos.objects.filter(
                        fecha_pago__gte=egr_desde,fecha_pago__lte=egr_hasta,cuenta=cuenta
                    )
                    for mvto in obj_anticipos:
                        cuenta_anticipo = int(mvto.tipo_anticipo.cuenta_debito_1)
                        list_mvtos.append({
                            'pk':mvto.pk,
                            'fecha':mvto.fecha_pago,
                            'tipo_doc':documento.split('-')[0],
                            'codigo_doc':documento.split('-')[1],
                            'cuenta':cuenta_anticipo,
                            'natur_op':'D',
                            'valor_pago':mvto.valor,
                            'anno_pago':mvto.fecha_pago.year,
                            'mes_pago':mvto.fecha_pago.month,
                            'dia_pago':mvto.fecha_pago.day,
                            'nit_pago':mvto.id_tercero,
                            'descrip_sec':f'{mvto.descripcion.upper()[:26]}',
                            'tipo_comp_cruce':documento,
                            'num_cruce':'ANTICIPO',
                            'vencimiento_cruce':1,
                            'anno_cruce':mvto.fecha_pago.year,
                            'mes_cruce':mvto.fecha_pago.month,
                            'dia_cruce':mvto.fecha_pago.day,
                        })
                        list_mvtos.append({
                            'pk':mvto.pk,
                            'fecha':mvto.fecha_pago,
                            'tipo_doc':documento.split('-')[0],
                            'codigo_doc':documento.split('-')[1],
                            'cuenta':mvto.cuenta.nro_cuentacontable,
                            'natur_op':'C',
                            'valor_pago':mvto.valor,
                            'anno_pago':mvto.fecha_pago.year,
                            'mes_pago':mvto.fecha_pago.month,
                            'dia_pago':mvto.fecha_pago.day,
                            'nit_pago':mvto.id_tercero,
                            'descrip_sec':f'{mvto.descripcion.upper()[:26]}',
                            'tipo_comp_cruce':'',
                            'num_cruce':'',
                            'vencimiento_cruce':'',
                            'anno_cruce':'',
                            'mes_cruce':'',
                            'dia_cruce':'',
                            'salta':True
                        })
                    obj_transf = transferencias_companias.objects.filter(
                                cuenta_sale=cuenta,
                                fecha__gte=egr_desde,fecha__lte=egr_hasta,  
                            ).exclude(empresa_entra=F('empresa_sale')).order_by('empresa_sale')
                    for mvto in obj_transf:
                        cuenta_interc = cuentas_intercompanias.objects.get(
                            empresa_desde=mvto.empresa_sale,empresa_hacia=mvto.empresa_entra).cuenta_por_cobrar
                        list_mvtos.append({
                            'pk':mvto.pk,
                            'fecha':mvto.fecha,
                            'tipo_doc':documento.split('-')[0],
                            'codigo_doc':documento.split('-')[1],
                            'cuenta':cuenta_interc,
                            'natur_op':'D',
                            'valor_pago':mvto.valor,
                            'anno_pago':mvto.fecha.year,
                            'mes_pago':mvto.fecha.month,
                            'dia_pago':mvto.fecha.day,
                            'nit_pago':mvto.empresa_entra.pk,
                            'descrip_sec':f'TRANSFERENCIA HACIA {mvto.empresa_entra.nombre.upper()} CUENTA {mvto.cuenta_entra.cuentabanco}',
                            'tipo_comp_cruce':'',
                            'num_cruce':'',
                            'vencimiento_cruce':'',
                            'anno_cruce':'',
                            'mes_cruce':'',
                            'dia_cruce':'',
                        })
                        list_mvtos.append({
                            'pk':mvto.pk,
                            'fecha':mvto.fecha,
                            'tipo_doc':documento.split('-')[0],
                            'codigo_doc':documento.split('-')[1],
                            'cuenta':mvto.cuenta_sale.nro_cuentacontable,
                            'natur_op':'C',
                            'valor_pago':mvto.valor,
                            'anno_pago':mvto.fecha.year,
                            'mes_pago':mvto.fecha.month,
                            'dia_pago':mvto.fecha.day,
                            'nit_pago':'',
                            'descrip_sec':f'TRANSFERENCIA HACIA {mvto.empresa_entra.nombre.upper()} CUENTA {mvto.cuenta_entra.cuentabanco}',
                            'tipo_comp_cruce':'',
                            'num_cruce':'',
                            'vencimiento_cruce':'',
                            'anno_cruce':'',
                            'mes_cruce':'',
                            'dia_cruce':'',
                            'salta':True
                        })
                        
                    obj_pagos = Pagos.objects.filter(
                                fecha_pago__gte=egr_desde,fecha_pago__lte=egr_hasta,cuenta=cuenta
                            )
                    for mvto in obj_pagos:
                        cxp = mvto.nroradicado.cuenta_por_pagar.cuenta_credito_1
                        doc_cruce = mvto.nroradicado.nrocausa
                        doc_cruce = doc_cruce.split('-')
                        tipo_doc_cruce = f'{doc_cruce[0]}-{doc_cruce[1]}'
                        num_cruce = doc_cruce[2]
                        vencimient = mvto.nroradicado.secuencia_cxp
                        pagos_detallados = pago_detallado_relacionado.objects.filter(pago=mvto.pk)
                        if pagos_detallados.exists():
                            if mvto.nroradicado.empresa.pk != mvto.empresa.pk:
                                cuenta_interc = cuentas_intercompanias.objects.get(
                                    empresa_hacia=mvto.nroradicado.empresa.pk,
                                    empresa_desde=mvto.empresa.pk
                                ).cuenta_por_cobrar
                                list_mvtos.append({
                                    'pk':mvto.pk,
                                    'fecha':mvto.fecha_pago,
                                    'tipo_doc':documento.split('-')[0],
                                    'codigo_doc':documento.split('-')[1],
                                    'cuenta':cuenta_interc,
                                    'natur_op':'D',
                                    'valor_pago':mvto.valor,
                                    'anno_pago':mvto.fecha_pago.year,
                                    'mes_pago':mvto.fecha_pago.month,
                                    'dia_pago':mvto.fecha_pago.day,
                                    'nit_pago':mvto.nroradicado.empresa.pk,
                                    'descrip_sec':f'CXC PAGO FACT {mvto.nroradicado.nrofactura} CAUSACION {mvto.nroradicado.nrocausa} -{mvto.nroradicado.descripcion.upper()[:20]}',
                                    'tipo_comp_cruce':'',
                                    'num_cruce':'',
                                    'vencimiento_cruce':'',
                                    'anno_cruce':'',
                                    'mes_cruce':'',
                                    'dia_cruce':'',
                                })
                            for pago in pagos_detallados:
                                if mvto.nroradicado.empresa.pk == mvto.empresa.pk:
                                    list_mvtos.append({
                                        'pk':mvto.pk,
                                        'fecha':mvto.fecha_pago,
                                        'tipo_doc':documento.split('-')[0],
                                        'codigo_doc':documento.split('-')[1],
                                        'cuenta':int(cxp),
                                        'natur_op':'D',
                                        'valor_pago':pago.valor,
                                        'anno_pago':mvto.fecha_pago.year,
                                        'mes_pago':mvto.fecha_pago.month,
                                        'dia_pago':mvto.fecha_pago.day,
                                        'nit_pago':pago.id_tercero,
                                        'descrip_sec':f'PAGO {pago.nombre_tercero} {mvto.nroradicado.descripcion.upper()[:20]}',
                                        'tipo_comp_cruce':tipo_doc_cruce,
                                        'num_cruce':num_cruce,
                                        'vencimiento_cruce':pago.vencimiento,
                                        'anno_cruce':mvto.nroradicado.fechacausa.year,
                                        'mes_cruce':mvto.nroradicado.fechacausa.month,
                                        'dia_cruce':mvto.nroradicado.fechacausa.day,
                                    })
                                if pago.pk == pagos_detallados.last().pk: salta = True
                                else: salta = False
                                list_mvtos.append({
                                    'pk':mvto.pk,
                                    'fecha':mvto.fecha_pago,
                                    'tipo_doc':documento.split('-')[0],
                                    'codigo_doc':documento.split('-')[1],
                                    'cuenta':mvto.cuenta.nro_cuentacontable,
                                    'natur_op':'C',
                                    'valor_pago':pago.valor,
                                    'anno_pago':mvto.fecha_pago.year,
                                    'mes_pago':mvto.fecha_pago.month,
                                    'dia_pago':mvto.fecha_pago.day,
                                    'nit_pago':pago.id_tercero,
                                    'descrip_sec':f'PAGO {pago.nombre_tercero} {mvto.nroradicado.descripcion.upper()[:20]}',
                                    'tipo_comp_cruce':'',
                                    'num_cruce':'',
                                    'vencimiento_cruce':'',
                                    'anno_cruce':'',
                                    'mes_cruce':'',
                                    'dia_cruce':'',
                                    'salta':salta
                                })
                        else:
                            if mvto.nroradicado.empresa.pk == mvto.empresa.pk:
                                list_mvtos.append({
                                    'pk':mvto.pk,
                                    'fecha':mvto.fecha_pago,
                                    'tipo_doc':documento.split('-')[0],
                                    'codigo_doc':documento.split('-')[1],
                                    'cuenta':int(cxp),
                                    'natur_op':'D',
                                    'valor_pago':mvto.valor,
                                    'anno_pago':mvto.fecha_pago.year,
                                    'mes_pago':mvto.fecha_pago.month,
                                    'dia_pago':mvto.fecha_pago.day,
                                    'nit_pago':mvto.nroradicado.idtercero,
                                    'descrip_sec':f'PAGO FACT {mvto.nroradicado.nrofactura}-{mvto.nroradicado.descripcion.upper()[:20]}',
                                    'tipo_comp_cruce':tipo_doc_cruce,
                                    'num_cruce':num_cruce,
                                    'vencimiento_cruce':vencimient,
                                    'anno_cruce':mvto.nroradicado.fechacausa.year,
                                    'mes_cruce':mvto.nroradicado.fechacausa.month,
                                    'dia_cruce':mvto.nroradicado.fechacausa.day,
                                })
                            else:
                                cuenta_interc = cuentas_intercompanias.objects.get(
                                    empresa_hacia=mvto.nroradicado.empresa.pk,
                                    empresa_desde=mvto.empresa.pk
                                ).cuenta_por_cobrar
                                list_mvtos.append({
                                    'pk':mvto.pk,
                                    'fecha':mvto.fecha_pago,
                                    'tipo_doc':documento.split('-')[0],
                                    'codigo_doc':documento.split('-')[1],
                                    'cuenta':cuenta_interc,
                                    'natur_op':'D',
                                    'valor_pago':mvto.valor,
                                    'anno_pago':mvto.fecha_pago.year,
                                    'mes_pago':mvto.fecha_pago.month,
                                    'dia_pago':mvto.fecha_pago.day,
                                    'nit_pago':mvto.nroradicado.empresa.pk,
                                    'descrip_sec':f'CXC PAGO FACT {mvto.nroradicado.nrofactura} CAUSACION {mvto.nroradicado.nrocausa} -{mvto.nroradicado.descripcion.upper()[:20]}',
                                    'tipo_comp_cruce':'',
                                    'num_cruce':'',
                                    'vencimiento_cruce':'',
                                    'anno_cruce':'',
                                    'mes_cruce':'',
                                    'dia_cruce':'',
                                })
                            list_mvtos.append({
                                'pk':mvto.pk,
                                'fecha':mvto.fecha_pago,
                                'tipo_doc':documento.split('-')[0],
                                'codigo_doc':documento.split('-')[1],
                                'cuenta':mvto.cuenta.nro_cuentacontable,
                                'natur_op':'C',
                                'valor_pago':mvto.valor,
                                'anno_pago':mvto.fecha_pago.year,
                                'mes_pago':mvto.fecha_pago.month,
                                'dia_pago':mvto.fecha_pago.day,
                                'nit_pago':mvto.nroradicado.idtercero,
                                'descrip_sec':f'PAGO FACT {mvto.nroradicado.nrofactura}-{mvto.nroradicado.descripcion.upper()[:20]}',
                                'tipo_comp_cruce':'',
                                'num_cruce':'',
                                'vencimiento_cruce':'',
                                'anno_cruce':'',
                                'mes_cruce':'',
                                'dia_cruce':'',
                                'salta':True
                            })
                    
                list_mvtos_ord = sorted(list_mvtos,key=operator.itemgetter('fecha'))
                
                book=openpyxl.load_workbook("resources/excel_formats/InterfazSIIGO.xlsx")
                sheet=book.active
                row=6
                for pago in list_mvtos_ord:
                    sheet.cell(row,1,pago.get('tipo_doc'))
                    sheet.cell(row,2,pago.get('codigo_doc'))
                    sheet.cell(row,3,consec_egr)
                    sheet.cell(row,4,pago.get('cuenta'))
                    sheet.cell(row,5,pago.get('natur_op'))
                    sheet.cell(row,6,pago.get('valor_pago'))
                    sheet.cell(row,7,pago.get('anno_pago'))
                    sheet.cell(row,8,pago.get('mes_pago'))
                    sheet.cell(row,9,pago.get('dia_pago'))
                    sheet.cell(row,16,pago.get('nit_pago'))
                    sheet.cell(row,18,pago.get('descrip_sec').upper())
                    sheet.cell(row,58,pago.get('tipo_comp_cruce'))
                    if  pago.get('num_cruce') == 'ANTICIPO': 
                        sheet.cell(row,59,consec_egr)
                    else:
                        sheet.cell(row,59,pago.get('num_cruce'))
                    sheet.cell(row,60,pago.get('vencimiento_cruce'))
                    sheet.cell(row,61,pago.get('anno_cruce'))
                    sheet.cell(row,62,pago.get('mes_cruce'))
                    sheet.cell(row,63,pago.get('dia_cruce'))
                    row+=1
                    if pago.get('salta'): consec_egr += 1
                
                nombre_doc=f'Interfaz_egresos_{egr_desde}_{egr_hasta}.xlsx'
                ruta=settings.MEDIA_ROOT+'/tmp/xlsx/'+nombre_doc
                ruta_dw=settings.MEDIA_URL+'tmp/xlsx/'+nombre_doc
                book.save(ruta)
                
                data = {
                    'texto':f'''<ul>
                        <li>
                            Puedes descargar el archivo de egresos <strong><a href="{ruta_dw}" target="_blank">aquí</a></strong>
                        </li>
                    </ul>'''
                }   
            except:
                data={
                    'texto':traceback.format_exc()
                }
            return JsonResponse(data)

@login_required
def ajax_impr_int_notas(request):
   if request.method == 'POST':
       if request.is_ajax():
            empresa_egr = request.POST.get('empresa_notas')
            egr_desde = request.POST.get('fecha_nota_desde')
            egr_hasta = request.POST.get('fecha_nota_hasta')
            consec_egr = int(request.POST.get('numero_inicial_nota'))
            cuentas_egr = request.POST.getlist('cuentas_notas')
            oficina_egr = request.POST.get('oficina_notas')
            list_mvtos = [] 
            
            for cuenta in cuentas_egr:
                obj_intercomp = transferencias_companias.objects.filter(
                            cuenta_entra=cuenta,
                            fecha__gte=egr_desde,fecha__lte=egr_hasta,  
                    ).exclude(empresa_entra=F('empresa_sale')).order_by('empresa_sale')
                for mvto in obj_intercomp:
                    obj_inter_nota = cuentas_intercompanias.objects.get(
                            empresa_desde= mvto.empresa_sale.pk,
                            empresa_hacia = empresa_egr
                        )
                    doc_nota = obj_inter_nota.documento
                    cuenta_interc_nota =  obj_inter_nota.cuenta_por_pagar
                    
                    list_mvtos.append({
                        'pk':mvto.pk,
                        'fecha':mvto.fecha,
                        'tipo_doc':doc_nota.split('-')[0],
                        'codigo_doc':int(doc_nota.split('-')[1]),
                        'cuenta':mvto.cuenta_entra.nro_cuentacontable,
                        'natur_op':'D',
                        'valor_pago':mvto.valor,
                        'anno_pago':mvto.fecha.year,
                        'mes_pago':mvto.fecha.month,
                        'dia_pago':mvto.fecha.day,
                        'nit_pago':mvto.empresa_sale.Nit,
                        'descrip_sec':f'TRANSFERENCIA DESDE {mvto.empresa_sale.nombre.upper()} CUENTA {mvto.cuenta_sale.cuentabanco}',
                        'tipo_comp_cruce':'',
                        'num_cruce':'',
                        'vencimiento_cruce':'',
                        'anno_cruce':'',
                        'mes_cruce':'',
                        'dia_cruce':'',
                    })
                    list_mvtos.append({
                        'pk':mvto.pk,
                        'fecha':mvto.fecha,
                        'tipo_doc':doc_nota.split('-')[0],
                        'codigo_doc':int(doc_nota.split('-')[1]),
                        'cuenta':cuenta_interc_nota,
                        'natur_op':'C',
                        'valor_pago':mvto.valor,
                        'anno_pago':mvto.fecha.year,
                        'mes_pago':mvto.fecha.month,
                        'dia_pago':mvto.fecha.day,
                        'nit_pago':mvto.empresa_sale.Nit,
                        'descrip_sec':f'TRANSFERENCIA DESDE {mvto.empresa_sale.nombre.upper()} CUENTA {mvto.cuenta_sale.cuentabanco}',
                        'tipo_comp_cruce':'',
                        'num_cruce':'',
                        'vencimiento_cruce':'',
                        'anno_cruce':'',
                        'mes_cruce':'',
                        'dia_cruce':'',
                    })
                obj_pagos = Pagos.objects.filter(
                    nroradicado__empresa_id = empresa_egr,
                    nroradicado__oficina = oficina_egr,
                    fecha_pago__gte = egr_desde, fecha_pago__lte = egr_hasta
                ).exclude(empresa=F('nroradicado__empresa'))
                for mvto in obj_pagos:
                    obj_inter_nota = cuentas_intercompanias.objects.get(
                            empresa_desde= mvto.empresa.pk,
                            empresa_hacia = empresa_egr,
                        )
                    doc_nota = obj_inter_nota.documento
                    cuenta_interc_nota =  obj_inter_nota.cuenta_por_pagar
                    doc_cruce = mvto.nroradicado.nrocausa.split('-')
                    descrip = f'PAGO FACT {mvto.nroradicado.nrofactura} {mvto.nroradicado.nombretercero}-{mvto.nroradicado.descripcion}'.upper()
                    descrip_2 = f'CXP PAGO FACT {mvto.nroradicado.nrofactura}-{mvto.nroradicado.descripcion}'
                    list_mvtos.append({
                        'pk':mvto.pk,
                        'fecha':mvto.fecha_pago,
                        'tipo_doc':doc_nota.split('-')[0],
                        'codigo_doc':int(doc_nota.split('-')[1]),
                        'cuenta':int(mvto.nroradicado.cuenta_por_pagar.cuenta_credito_1),
                        'natur_op':'D',
                        'valor_pago':mvto.valor,
                        'anno_pago':mvto.fecha_pago.year,
                        'mes_pago':mvto.fecha_pago.month,
                        'dia_pago':mvto.fecha_pago.day,
                        'nit_pago':mvto.nroradicado.idtercero,
                        'descrip_sec':descrip.rstrip("\n"),
                        'tipo_comp_cruce':f'{doc_cruce[0]}-{doc_cruce[1]}',
                        'num_cruce':doc_cruce[2],
                        'vencimiento_cruce':mvto.nroradicado.secuencia_cxp,
                        'anno_cruce':mvto.nroradicado.fechacausa.year,
                        'mes_cruce':mvto.nroradicado.fechacausa.month,
                        'dia_cruce':mvto.nroradicado.fechacausa.day,
                    })
                    list_mvtos.append({
                        'pk':mvto.pk,
                        'fecha':mvto.fecha_pago,
                        'tipo_doc':doc_nota.split('-')[0],
                        'codigo_doc':int(doc_nota.split('-')[1]),
                        'cuenta':cuenta_interc_nota,
                        'natur_op':'C',
                        'valor_pago':mvto.valor,
                        'anno_pago':mvto.fecha_pago.year,
                        'mes_pago':mvto.fecha_pago.month,
                        'dia_pago':mvto.fecha_pago.day,
                        'nit_pago':mvto.empresa.Nit,
                        'descrip_sec':descrip_2.rstrip('\n'),
                        'tipo_comp_cruce':'',
                        'num_cruce':'',
                        'vencimiento_cruce':'',
                        'anno_cruce':'',
                        'mes_cruce':'',
                        'dia_cruce':'',
                    })
                
            list_mvtos_ord = sorted(list_mvtos,key=operator.itemgetter('fecha'))
            
            book=openpyxl.load_workbook("resources/excel_formats/InterfazSIIGO.xlsx")
            sheet=book.active
            row=6
            for pago in list_mvtos_ord:
                sheet.cell(row,1,pago.get('tipo_doc'))
                sheet.cell(row,2,pago.get('codigo_doc'))
                sheet.cell(row,3,consec_egr)
                sheet.cell(row,4,pago.get('cuenta'))
                sheet.cell(row,5,pago.get('natur_op'))
                sheet.cell(row,6,pago.get('valor_pago'))
                sheet.cell(row,7,pago.get('anno_pago'))
                sheet.cell(row,8,pago.get('mes_pago'))
                sheet.cell(row,9,pago.get('dia_pago'))
                sheet.cell(row,16,pago.get('nit_pago'))
                sheet.cell(row,18,pago.get('descrip_sec'))
                sheet.cell(row,58,pago.get('tipo_comp_cruce'))
                sheet.cell(row,59,pago.get('num_cruce'))
                sheet.cell(row,60,pago.get('vencimiento_cruce'))
                sheet.cell(row,61,pago.get('anno_cruce'))
                sheet.cell(row,62,pago.get('mes_cruce'))
                sheet.cell(row,63,pago.get('dia_cruce'))
                row+=1
                if pago.get('natur_op') == 'C': consec_egr += 1
            
            nombre_doc=f'Interfaz_notas_{empresa_egr}_{egr_desde}_{egr_hasta}.xlsx'
            ruta=settings.MEDIA_ROOT+'/tmp/xlsx/'+nombre_doc
            ruta_dw=settings.MEDIA_URL+'tmp/xlsx/'+nombre_doc
            book.save(ruta)
            
            data = {
                'texto':f'''<ul>
                    <li>
                        Puedes descargar el archivo de notas <strong><a href="{ruta_dw}" target="_blank">aquí</a></strong>
                    </li>
                </ul>'''
            }   
            return JsonResponse(data)

@login_required   
def ajax_impr_otros_ing(request):
    if request.method == 'POST':
        if request.is_ajax():
            pass

@login_required
def ajax_movimientos_cont(request):
    if request.is_ajax():
        if request.method == 'GET':
            check_perms(request,('accounting.view_egresos_contable',))
            cuenta = request.GET.get('cuenta')
            estado = request.GET.get('estado')
            conciliacion = request.GET.get('conciliacion')
            filter_pago_asociado = request.GET.get('pagoasociado')
            fecha_desde = request.GET.get('desde')
            fecha_hasta = request.GET.get('hasta')
            id_pago_asociado = request.GET.get('id_pago_asociado')
            id_ant_asociado = request.GET.get('id_ant_asociado')
            
            
            if id_pago_asociado:
                obj_pago = egresos_banco.objects.filter(pago_asociado=id_pago_asociado)
                conciliaciones_list = obj_pago.values('conciliacion').distinct()
                
                model = egresos_contable.objects.filter(fecha="1999-01-01")
                for c in list(conciliaciones_list):
                    if c.get('conciliacion') != None:
                        model = model | egresos_contable.objects.filter(conciliacion=c.get('conciliacion'))
                    
                obj_mvto = model
            elif id_ant_asociado:
                obj_pago = egresos_banco.objects.filter(pago_asociado=id_ant_asociado)
                conciliaciones_list = obj_pago.values('conciliacion').distinct()
                
                model = egresos_contable.objects.filter(fecha="1999-01-01")
                for c in list(conciliaciones_list):
                    if c.get('conciliacion') != None:
                        model = model | egresos_contable.objects.filter(conciliacion=c.get('conciliacion'))
                    
                obj_mvto = model
            else:
                
                obj_mvto = egresos_contable.objects.filter(
                    cuenta = cuenta
                )
            
            if estado:
                obj_mvto = obj_mvto.filter(estado=estado)
            if conciliacion:
                obj_mvto = obj_mvto.filter(conciliacion=conciliacion)
            if fecha_desde:
                obj_mvto = obj_mvto.filter(fecha__range = (fecha_desde, fecha_hasta)    )
            
            if filter_pago_asociado:
                obj_mvto = obj_mvto.filter(valor__lt=0)
            
            list_mvto = []
            i=0
            for line in obj_mvto:
                frmt_fecha = datetime.datetime.strftime(line.fecha,'%d/%m/%Y')
                if line.conciliacion == None: conciliac = ""
                else: conciliac = line.conciliacion.pk
                
                list_mvto.append({
                   'id':i,
                   'pk':line.pk,
                   'comprobante':line.comprobante,
                   'fecha':frmt_fecha,
                   'descripcion':line.descripcion,
                   'valor':f'{line.valor:,}',
                   'estado':line.estado,
                   'conciliacion':conciliac,
                   'clasificacion':line.tipo,
                   'soporte':_media_url(line.soporte_egreso),
                })
            data = {
                'data':list_mvto
            }
            return JsonResponse(data)

@login_required
def ajax_movimientos_banco(request):
    if request.is_ajax():
        if request.method == 'GET':
            list_mvto = []
            if check_perms(request,('accounting.view_egresos_banco',),raise_exception=False):
                
                cuenta = request.GET.get('cuenta')
                estado = request.GET.get('estado')
                conciliacion = request.GET.get('conciliacion')
                fecha = request.GET.get('fecha')
                tipo= request.GET.get('tipo')
                fecha_desde = request.GET.get('desde')
                fecha_hasta = request.GET.get('hasta')
                pago_asociado = request.GET.get('pago_asociado')
                ant_asociado = request.GET.get('ant_asociado')
                
                if pago_asociado:
                    obj_mvto = egresos_banco.objects.filter(pago_asociado=pago_asociado)
                elif ant_asociado:
                    obj_mvto = egresos_banco.objects.filter(anticipo_asociado=ant_asociado)
                else:
                    obj_mvto = egresos_banco.objects.filter(
                        cuenta = cuenta
                    )
                if estado:
                    obj_mvto = obj_mvto.filter(estado=estado)
                if conciliacion:
                    obj_mvto = obj_mvto.filter(conciliacion=conciliacion)
                if fecha:
                    obj_mvto = obj_mvto.filter(fecha=fecha)
                if fecha_desde:
                    obj_mvto = obj_mvto.filter(fecha__range=(fecha_desde,fecha_hasta))
                
                
                if tipo == 'sin_pago':
                    obj_mvto = obj_mvto.filter(
                            pago_asociado__isnull=True, anticipo_asociado__isnull=True,
                        transferencia_asociada__isnull=True,valor__lt=0
                                ).exclude(Q(descripcion__icontains='cobro iva pagos automaticos')|
                                Q(descripcion__icontains='impto gobierno 4x1000')|
                                Q(descripcion__icontains='impto gobierno 4x1000 cxc')|
                                Q(descripcion__icontains='iva comision rec cajero autom')|
                                Q(descripcion__icontains='comision recaudo caja')|
                                Q(descripcion__icontains='iva comis')|
                                Q(descripcion__icontains='(DIVIDIDO)')|
                                Q(descripcion__icontains='comision rec')|
                                Q(descripcion__icontains='comision pago a proveedores')|
                                Q(descripcion__icontains='COMISION PAGO A OTROS BANCOS')|
                                Q(descripcion__icontains='AJUSTE INTERESES SOBREG N DIAS')|
                                Q(descripcion__icontains='TRANSFERENCIA CTA SUC VIRTUAL')|
                                Q(descripcion__icontains='TRANSFERENCIA')|
                                Q(descripcion__icontains='PAGO SUC VIRT TC')|
                                Q(descripcion__icontains='TRASLADO A FONDO DE INVERSION')|
                                Q(descripcion__icontains='STATUS')|
                                Q(descripcion__icontains='QUADRATA')|
                                Q(descripcion__icontains='SANDVILLE')|
                                Q(descripcion__icontains='PROMOTORA SANDV')|
                                Q(descripcion__icontains='ANDINA')|
                                Q(descripcion__icontains='WESTV')|
                                Q(descripcion__icontains='RETIRO SUCURSAL CON TARJETA')|
                                Q(descripcion__icontains='MORA TARJETA')|
                                Q(descripcion__icontains='comision pago de nomina')
                            )
                elif tipo == 'transfer':
                    obj_mvto = obj_mvto.filter(
                                Q(descripcion__icontains='TRANSFERENCIA CTA SUC VIRTUAL')|
                                Q(descripcion__icontains='STATUS')|
                                Q(descripcion__icontains='QUADRATA')|
                                Q(descripcion__icontains='SANDVILLE')|
                                Q(descripcion__icontains='PROMOTORA SANDV')|
                                Q(descripcion__icontains='WESTV')|
                                Q(descripcion__icontains='MORA TARJETA')|
                                Q(descripcion__icontains='PAGO SUC VIRT TC')|
                                Q(descripcion__icontains='TRASLADO A FONDO DE INVERSION')|
                                Q(descripcion__icontains='RETIRO SUCURSAL CON TARJETA')|
                                Q(descripcion__icontains='RETIRO CORRESPONSAL')|
                                Q(descripcion__icontains='TRANSFERENCIA CTAS BANCOLOMBIA')|
                                Q(descripcion__icontains='TRANSFERENCIA')|
                                Q(descripcion__icontains='RETIRO CAJERO')|
                                Q(descripcion__icontains='DEBITO SUCURSAL BANC')|
                                Q(descripcion__icontains='ANDINA'),
                                pago_asociado__isnull=True, anticipo_asociado__isnull=True,
                            transferencia_asociada__isnull=True,valor__lt=0
                            )
                elif tipo == 'general':
                    obj_mvto = obj_mvto.filter(
                                valor__lt=0
                                ).exclude(Q(descripcion__icontains='cobro iva pagos automaticos')|
                                Q(descripcion__icontains='impto gobierno 4x1000')|
                                Q(descripcion__icontains='impto gobierno 4x1000 cxc')|
                                Q(descripcion__icontains='iva comision rec cajero autom')|
                                Q(descripcion__icontains='comision recaudo caja')|
                                Q(descripcion__icontains='iva comis')|
                                Q(descripcion__icontains='PAGO CXC DE CTA')|
                                Q(descripcion__icontains='INTERESES SOBREG')|
                                Q(descripcion__icontains='(DIVIDIDO)')|
                                Q(descripcion__icontains='comision rec')|
                                Q(descripcion__icontains='comision pago a proveedores')|
                                Q(descripcion__icontains='COMISION PAGO A OTROS BANCOS')|
                                Q(descripcion__icontains='AJUSTE INTERESES SOBREG N DIAS')|
                                Q(descripcion__icontains='comision pago de nomina')
                            )
                        
                    
                
                i=0
                for line in obj_mvto:
                    frmt_fecha = datetime.datetime.strftime(line.fecha,'%d/%m/%Y')
                    if line.conciliacion == None: conciliac = ""
                    else: conciliac = line.conciliacion.pk
                    objeto = {
                    'id':i,
                    'pk':line.pk,
                    'fecha':frmt_fecha,
                    'descripcion':line.descripcion,
                    'referencia':line.referencia,
                    'valor':line.valor,
                    'estado':line.estado,
                    'conciliacion':conciliac,
                    }
                    if line.tipo_pago()[0] != None:
                        objeto['tipo']=line.tipo_pago()[0]
                        objeto['pagorelacionado']=line.tipo_pago()[1].pk
                        objeto['soporte_pago']= _media_url(line.tipo_pago()[1].soporte_pago)
                        objeto['usuario']=line.tipo_pago()[1].usuario.username
                        if line.tipo_pago()[0] == 'Pago':
                            objeto['soporte_causacion']= _media_url(line.tipo_pago()[1].nroradicado.soporte_causacion)
                            objeto['soporte_factura']= _media_url(line.tipo_pago()[1].nroradicado.soporte_radicado)
                            
                        else :
                            objeto['soporte_causacion']= ""
                            objeto['soporte_factura']= ""
                            
                    else:
                        objeto['tipo']=""
                        objeto['pagorelacionado']=""
                        objeto['soporte_pago']= ""
                        objeto['usuario']=""
                        objeto['soporte_causacion']= ""
                        objeto['soporte_factura']= ""
                    list_mvto.append(objeto)
            data = {
                'data':list_mvto
            }
            return JsonResponse(data)

@login_required
def ajax_cuentas_empresas(request):
    if request.is_ajax():
        if request.method == 'GET':
            empresa = request.GET.get('empresa')
            group = request.GET.get('group')
            obj_cuentas = cuentas_pagos.objects.filter(nit_empresa=empresa, activo=True)
            if group == 'onlycash': obj_cuentas = obj_cuentas.filter(es_caja=True)
            
                
            json_cuentas = serializers.serialize('json',obj_cuentas)
            data={
                'cuentas':json_cuentas
            }
            return JsonResponse(data)


@login_required            
def ajax_cargar_movimiento_contable(request):
    if request.is_ajax():
        if request.method == 'POST':
            check_perms(request,('accounting.add_egresos_contable',))
            empresa = request.POST.get('empresa_carga')
            archivo_mvto = request.FILES.get('archivoCarga')
            
            soporte_egreso = request.FILES.get('soporte-egreso')
            
            if soporte_egreso:
                linea = request.POST.get('comprobante')
                obj_linea = egresos_contable.objects.get(pk=linea)
                comprobante = obj_linea.comprobante.split(' ')
                comprobante = f'{comprobante[0]} {comprobante[1]} {comprobante[2]}'
                
                c = egresos_contable.objects.filter(
                    comprobante__icontains = comprobante,
                    empresa = obj_linea.empresa.pk
                )
                
                for i in c:
                    i.soporte_egreso = soporte_egreso
                    i.save()
                    
                data = {
                    'title':'Andinasoft dice:',
                    'msj': 'Soporte cargado',
                    'class':'alert-success'
                }
                    
                return JsonResponse(data)
                
            
            try:
                ultima_fila = int(request.POST.get('ultima_fila'))
                obj_empresa = empresas.objects.get(pk=empresa)
                movimientos = cargar_gastos_contables(
                    empresa=obj_empresa,file=archivo_mvto,
                    ultima_linea=ultima_fila
                )
                if movimientos['cargadas']==ultima_fila-7:
                    data={
                        'title':'Hecho!',
                        'msj':f'Fueron cargadas todas las lineas ({ultima_fila-7})',
                        'class':'alert-success'
                    }
                elif 0<movimientos['cargadas']<ultima_fila-7:
                    data={
                        'title':'Alerta!',
                        'msj':f"Fueron cargadas {movimientos['cargadas']} lineas pero {movimientos['errores']} no pudieron ser cargadas pues ya existe(n) en la base de datos",
                        'class':'alert-warning',
                    }
                else:
                    data={
                        'title':'Oh Oh!',
                        'msj':f"No fue cargado el movimiento, todas las lineas existen en la base de datos",
                        'class':'alert-danger'
                    }
            except ObjectDoesNotExist:
                data={
                        'title':'Oh Oh!',
                        'msj':'Alguna de las cuentas en el movimiento que estas intentando cargar no corresponde a la empresa seleccionada',
                        'class':'alert-danger'
                    }
            except:
                data={
                        'title':'Se ha producido un error!',
                        'msj':traceback.format_exc(),
                        'class':'alert-danger'
                    }
            return JsonResponse(data)

@login_required
def ajax_eliminar_linea_mvto_contable(request):
    if request.is_ajax():
        if request.method == 'GET':
            check_perms(request,('accounting.delete_egresos_contable',))
            linea = request.GET.get('linea')
            obj_linea = egresos_contable.objects.get(pk=linea)
            obj_linea.delete()
            data = {
                'errors':0
            }
            return JsonResponse(data)

@login_required
def ajax_eliminar_linea_mvto_bco(request):
    if request.is_ajax():
        if request.method == 'GET':
            check_perms(request,('accounting.delete_egresos_banco',))
            linea = request.GET.get('linea')
            obj_linea = egresos_banco.objects.get(pk=linea)
            obj_linea.delete()
            data = {
                'errors':0
            }
            return JsonResponse(data)

@login_required
def ajax_cargar_movimiento_banco(request):
    if request.is_ajax():
        if request.method == 'POST':
            check_perms(request,('accounting.add_egresos_banco',))
            empresa = request.POST.get('empresa_carga_bco')
            archivo = request.FILES['archivoBco']
            confirm = request.POST.get('confirm')
            
            confirm = True if confirm == 'true' else False
            
            obj_empresa = empresas.objects.get(pk=empresa)
            carga = cargar_gastos_banco(empresa=obj_empresa,file=archivo, confirm_duplicated=confirm)
            
            
            if carga['duplicados']:
                data = {
                    'title':'Oh Oh!',
                    'msj':'Hay muchas lineas de banco que ya existen en la base de datos, ¿Estas seguro que deseas contiuar?',
                    'class':'alert-warning'
                }
            elif carga['errores'] > 0:
                data = {
                'title':'Oh Oh!',
                'msj':'Se produjo un error al cargar el movimiento de banco',
                'class':'alert-danger'
            }
            else:
                data = {
                'title':'Hecho!',
                'msj':'El movimiento bancario fue cargado con exito',
                'class':'alert-success'
            }
            return JsonResponse(data)

@login_required
def ajax_nueva_conciliacion(request):
    if request.method =='POST':
        if request.is_ajax():
            todo = request.POST.get('todo')
            
            if todo == 'dividir_mvto':
                line = request.POST.get('line')
                obj_line = egresos_banco.objects.get(pk=line)
                obj_concilacion = conciliaciones.objects.create(
                    empresa = obj_line.empresa,
                    cuenta_asociada=obj_line.cuenta,
                    fecha_crea=datetime.date.today(),
                    usuario_crea=request.user
                )
                
                credito_linea_act = egresos_banco.objects.create(
                    empresa = obj_line.empresa , cuenta = obj_line.cuenta,
                    fecha = obj_line.fecha,
                    descripcion ='AJUSTE POR DIVISION ' + obj_line.descripcion,
                    referencia =  'LINEA ' + line,
                    valor = obj_line.valor * -1,
                    estado = 'CONCILIADO',
                    conciliacion = obj_concilacion
                )
                
                nuevas_lineas = request.POST.getlist('nuevo_valor_linea')
                
                j = 1
                for i in nuevas_lineas:
                    descripcion = f'DIVISION ({j}/{len(nuevas_lineas)}) - {obj_line.descripcion}'
                    valor = float(i.replace(',','')) * -1
                    egresos_banco.objects.create(
                        empresa = obj_line.empresa, 
                        cuenta = obj_line.cuenta,
                        fecha = obj_line.fecha,
                        descripcion =descripcion,
                        referencia = f'LINEA {line}',
                        valor = valor,
                        estado = 'SIN CONCILIAR',
                    )
                    j+=1
                
                obj_line.descripcion = obj_line.descripcion + ' (DIVIDIDO)'
                obj_line.conciliacion = obj_concilacion
                obj_line.estado = 'CONCILIADO'
                obj_line.save()
                
                data = {
                    'title':'Hecho!',
                    'msj':f'Se dividio la linea {line} en {len(nuevas_lineas)} lineas',
                    'class':'alert-success'
                }
                
                return JsonResponse(data)
                    
                
                
                
            
            check_perms(request,('accounting.add_conciliaciones',))
            list_conc_cont = request.POST.getlist('conc-cont')
            list_conc_bco = request.POST.getlist('conc-bco')
            empresa = request.POST.get('empresa')
            cuenta = request.POST.get('cuenta')
            obj_concilacion = conciliaciones.objects.create(
                empresa = empresas.objects.get(pk=empresa),
                cuenta_asociada=cuentas_pagos.objects.get(pk=cuenta),
                fecha_crea=datetime.date.today(),
                usuario_crea=request.user
            )
            
            for i in list_conc_cont:
                obj_con_cont = egresos_contable.objects.get(pk=i)
                obj_con_cont.estado = 'CONCILIADO'
                obj_con_cont.conciliacion = obj_concilacion
                obj_con_cont.save()
            for i in list_conc_bco:
                obj_con_bco = egresos_banco.objects.get(pk=i)
                obj_con_bco.estado = 'CONCILIADO'
                obj_con_bco.conciliacion = obj_concilacion
                obj_con_bco.save()
            data={
                
            }
            return JsonResponse(data)

@login_required
def ajax_saldos_conciliacion(request):
    if request.method == 'GET':
        if request.is_ajax():
            cuenta = request.GET.get('cuenta')
            empresa = request.GET.get('empresa')
            
            obj_mvtos_cont = egresos_contable.objects.filter(empresa=empresa,cuenta=cuenta)
            saldo_final_cont = obj_mvtos_cont.aggregate(Sum('valor'))
            saldo_final_cont = saldo_final_cont['valor__sum']
            if saldo_final_cont == None:  saldo_final_cont=0
            
            no_conciliados_cont = obj_mvtos_cont.filter(estado='SIN CONCILIAR')
            total_nc_cont = no_conciliados_cont.aggregate(Sum('valor'))
            total_nc_cont = total_nc_cont['valor__sum']
            if total_nc_cont == None:  total_nc_cont=0
            
            obj_mvtos_bco = egresos_banco.objects.filter(empresa=empresa,cuenta=cuenta)
            saldo_final_bco = obj_mvtos_bco.aggregate(Sum('valor'))
            saldo_final_bco = saldo_final_bco['valor__sum']
            if saldo_final_bco == None:  saldo_final_bco=0
            
            no_conciliados_bco = obj_mvtos_bco.filter(estado='SIN CONCILIAR')
            total_nc_bco = no_conciliados_bco.aggregate(Sum('valor'))
            total_nc_bco = total_nc_bco['valor__sum']
            if total_nc_bco == None:  total_nc_bco=0
            
            verif_cont = saldo_final_cont + total_nc_bco
            verif_bco = saldo_final_bco + total_nc_cont
            
            data = {
                'saldo_libros':saldo_final_cont,
                'saldo_banco':saldo_final_bco,
                'verif_libros':verif_cont,
                'vefif_bancos':verif_bco
            }
            return JsonResponse(data)

@login_required
def ajax_imprimir_conciliacion(request):
    if request.method == 'GET':
    
        empresa = request.GET.get('empresa')
        cuenta = request.GET.get('cuenta')
        conciliacion = request.GET.get('conciliacion')
        
        fecha_max_cont = egresos_contable.objects.filter(
            empresa=empresa,cuenta=cuenta,conciliacion__lte=conciliacion
                    ).aggregate(Max('fecha'))['fecha__max']
        
        fecha_max_bco = egresos_banco.objects.filter(
            empresa=empresa,cuenta=cuenta,conciliacion__lte=conciliacion
                    ).aggregate(Max('fecha'))['fecha__max']
        
        fecha_max = max(fecha_max_bco,fecha_max_cont)
        
        obj_mvto_contable = egresos_contable.objects.filter(
            Q(conciliacion__gt=conciliacion)|Q(conciliacion__isnull=True),
            empresa=empresa,cuenta=cuenta,
            fecha__lte=fecha_max
        )
        
        obj_mvto_bco = egresos_banco.objects.filter(
            Q(conciliacion__gt=conciliacion)|Q(conciliacion__isnull=True),
            empresa=empresa,cuenta=cuenta,
            fecha__lte=fecha_max
        )
        
        saldo_cont = egresos_contable.objects.filter(
                    empresa =empresa,cuenta = cuenta,
                    fecha__lte=fecha_max
                ).aggregate(Sum('valor'))['valor__sum']
        saldo_banco = egresos_banco.objects.filter(
                    empresa =empresa,cuenta = cuenta,
                    fecha__lte=fecha_max
                ).aggregate(Sum('valor'))['valor__sum']
        cuenta_banco = cuentas_pagos.objects.get(pk=cuenta).cuentabanco
        filename = f'Conciliacion_{empresas.objects.get(pk=empresa).pk}_{cuenta_banco}_{conciliacion}.pdf'
        ruta = settings.MEDIA_ROOT +'/tmp/pdf/'+filename
        GenerarPDF().Conciliacion(
            empresa = empresas.objects.get(pk=empresa),
            cuenta_banco= cuentas_pagos.objects.get(pk=cuenta).cuentabanco,
            cuenta_cont=cuentas_pagos.objects.get(pk=cuenta).cuentacontable,
            saldo_banco = saldo_banco,
            saldo_cont=saldo_cont,
            fecha_conciliacion = fecha_max,
            mvto_contable = obj_mvto_contable,
            mvto_banco = obj_mvto_bco,
            ruta= ruta,
            user=request.user,
            conciliacion = conciliaciones.objects.get(pk=conciliacion)
        )
        
        data = {
            'filename':filename,
        }
        return JsonResponse(data)
            
@login_required
def ajax_lista_conciliaciones(request):
    if request.method == 'GET':
        if request.is_ajax():
            empresa = request.GET.get('empresa')
            cuenta = request.GET.get('cuenta')
            
            obj_conciliaciones = conciliaciones.objects.filter(empresa=empresa,cuenta_asociada=cuenta)
            list_mvto = []
            i=0
            for line in obj_conciliaciones:
                list_mvto.append({
                   'id':i,
                   'pk':line.pk,
                   'fecha':line.fecha_crea,
                   'usuario':line.usuario_crea.username
                })
            data = {
                'data':list_mvto
            }
            return JsonResponse(data)

@login_required
def ajax_borrar_conciliacion(request):
    if request.method == 'POST':
        if request.is_ajax():
            check_perms(request,('accounting.delete_conciliaciones',))
            conciliacion = request.POST.get('conciliacion')
            
            obj_conciliacion = conciliaciones.objects.get(pk=conciliacion)
            
            obj_mvto_cont = egresos_contable.objects.filter(conciliacion=conciliacion)
            obj_mvto_bco = egresos_banco.objects.filter(conciliacion=conciliacion)
            
            for mvto in obj_mvto_cont:
                mvto.estado = 'SIN CONCILIAR'
                mvto.conciliacion = None
                mvto.save()
            for mvto in obj_mvto_bco:
                mvto.estado = 'SIN CONCILIAR'
                mvto.conciliacion = None
                mvto.save()
            obj_conciliacion.delete()
            data = {}
            return JsonResponse(data)

@login_required  
def ajax_coincidir_tercero(request):
    if request.method == 'GET':
        if request.is_ajax():
            tercero = request.GET.get('tercero')
            nombre = Facturas.objects.filter(idtercero=tercero)
            nombre_ingresos = otros_ingresos.objects.filter(id_tercero=tercero)
            if nombre.exists():
                nombre = nombre.last().nombretercero
            elif nombre_ingresos.exists():
                nombre = nombre_ingresos.last().nombre_tercero
            else:
                nombre = ''
            data = {
                'nombre':nombre
            }
            return JsonResponse(data)

@login_required
def ajax_info_facturas(request):
    if request.method == 'GET':
        if request.is_ajax():
            list_mvto=[]
            oficina = request.GET.get('oficina')
            ubicacion = request.GET.get('ubicacion')
            causado = request.GET.get('causado')
            tesoreria = request.GET.get('tesoreria')
            
            
            
            if oficina:
                obj_facturas = info_facturas.objects.filter(
                    oficina=oficina,
                    ubicacion=ubicacion
                )
                if causado:
                    obj_facturas = obj_facturas.exclude(Q(causacion__isnull=True) | Q(causacion__exact=''))
                if tesoreria:
                    obj_facturas = obj_facturas.filter(saldo__gt=0)

                obj_facturas = obj_facturas.order_by('-radicado').values(
                    'radicado',
                    'nombretercero',
                    'fecharadicado',
                    'fechafactura',
                    'valor',
                    'empresa',
                    'causacion',
                    'pagoneto',
                    'saldo',
                    'descripcion',
                    'origen',
                )

                for idx, line in enumerate(obj_facturas):
                    list_mvto.append({
                        'id': idx,
                        'pk': line.get('radicado'),
                        'tercero': line.get('nombretercero'),
                        'fecha_rad': line.get('fecharadicado'),
                        'fecha_fact': line.get('fechafactura'),
                        'valor': f"{line.get('valor') or 0:,}",
                        'empresa': line.get('empresa'),
                        'causacion': line.get('causacion'),
                        'pagoneto': f"{line.get('pagoneto') or 0:,}",
                        'saldo': f"{line.get('saldo') or 0:,}",
                        'descripcion': line.get('descripcion'),
                        'tipo': line.get('origen'),
                    })
                                
            data = {
                'data':list_mvto
            }
            
            return JsonResponse(data)
               
@login_required 
def ajax_recibir_factura(request):
    if request.method == 'POST':
        if request.is_ajax():
            recibe = request.POST.get('recibe')
            if check_groups(request,(recibe,),raise_exception=False):
                radicado = request.POST.get('radicado')
                recibe = request.POST.get('recibe')
                
                obj_rad = Facturas.objects.get(pk=radicado)
                accion = f'Recibio la factura {obj_rad.nrofactura}'
                obj_history = history_facturas.objects.create(
                    factura=obj_rad,usuario=request.user,
                    accion=accion,ubicacion=recibe
                )
                data = {
                    'title':'Listo!',
                    'msj':'Se recibió la factura con exito',
                    'class':'alert-success',
                }
            else:
                data = {
                    'title':'oh oh!',
                    'msj':f'No tienes permisos para recibir una factura en {recibe}',
                    'class':'alert-success',
                }   
            return JsonResponse(data)

@login_required
def ajax_data_factura(request):
    if request.method == 'GET':
        if request.is_ajax():
            factura = request.GET.get('pk')

            obj_factura = Facturas.objects.filter(pk=factura).first()
            if not obj_factura:
                return JsonResponse({'detail': 'Factura no encontrada'}, status=404)

            data = {
                'pk': obj_factura.pk,
                'soporte_radicado': _media_url(obj_factura.soporte_radicado),
                'soporte_causacion': _media_url(obj_factura.soporte_causacion),
                'data': serializers.serialize('json', [obj_factura]),
            }
            return JsonResponse(data)

@login_required         
def ajax_registrar_causacion(request):
    if request.method == 'POST':
        if request.is_ajax():
            fecha=request.POST.get('fecha')
            nro_causacion=request.POST.get('nro_causacion')
            pagoneto=request.POST.get('pagoneto')
            radicado=request.POST.get('radicado')
            cuenta_cxp = request.POST.get('cuenta_por_pagar')
            secuencia_cxp = request.POST.get('secuencia_cxp')
            soporte = request.FILES.get('soporte')
            proyecto_ascociado = request.POST.get('proyecto_ascociado')
            centro_costo = request.POST.get('centro_costo')
            
            obj_radicado = Facturas.objects.get(pk=radicado)
            obj_radicado.nrocausa = nro_causacion
            obj_radicado.fechacausa = fecha
            obj_radicado.pago_neto = pagoneto.replace('.','')
            obj_radicado.cuenta_por_pagar = info_interfaces.objects.get(pk=cuenta_cxp)
            obj_radicado.secuencia_cxp = secuencia_cxp
            obj_radicado.soporte_causacion = soporte
            obj_radicado.proyecto_rel = proyecto_ascociado
            obj_radicado.centro_costo = centro_costo
            obj_radicado.save()
            
            history_facturas.objects.create(
                factura=obj_radicado,usuario=request.user,
                accion=f'Registró causacion con el numero {nro_causacion}',
                ubicacion='Contabilidad'
            )
            
            msj = f'''Hola :), <b>{request.user.get_full_name()}</b> ha causado una nueva factura:<br>
                <ul>
                    <li>Id: {obj_radicado.pk}</li>
                    <li>Empresa: {obj_radicado.empresa.nombre}</li>
                    <li>Oficina: {obj_radicado.oficina}</li>
                    <li>Tercero: {obj_radicado.nombretercero}</li>
                    <li>Causacion: {obj_radicado.nrocausa}</li>
                    <li>Valor: ${obj_radicado.valor:,}</li>
                    <li>Pago neto: ${pagoneto}</li>
                </ul>'''
            
            
            asunto = f'Nuevo estado: Radicado {obj_radicado.pk} CAUSADO'
            
            if obj_radicado.oficina == 'MONTERIA':
                envio_notificacion(msj,asunto,['admin1@somosandina.co','recepcionmtr@somosandina.co'])
            else:
                envio_notificacion(msj,asunto,['operaciones2@somosandina.co','auxiliartesoreria@somosandina.co'])
            
            
            data = {
                    'title':'Listo!',
                    'msj':'Se registró la causacion con exito',
                    'class':'alert-success',
                }
            return JsonResponse(data)

@login_required
def ajax_history_facturas(request):
    if request.method == 'GET':
        if request.is_ajax():
            factura = request.GET.get('pk')
            
            obj_history = history_facturas.objects.filter(factura=factura).order_by('-fecha')
            list_history = []
            for line in obj_history:
                avatar = Profiles.objects.get(user=line.usuario)
                list_history.append({
                    'imagen': avatar.avatar.image.url if avatar.avatar.image else '',
                    'usuario':line.usuario.first_name+ ' '+line.usuario.last_name,
                    'accion':line.accion,
                    'ubicacion':line.ubicacion,
                    'fecha':line.fecha
                })
            data ={
                'data':list_history
            }
            return JsonResponse(data)

@login_required
def ajax_asociar_pago_radicado(request):
    if request.method == 'POST':
        if request.is_ajax():
            radicado= request.POST.get('radicado')
            soporte = request.FILES.get('soporte')
            valor = request.POST.get('valor').replace(',','')
            fecha = request.POST.get('fecha')
            empresa = request.POST.get('empresa')
            cuenta = request.POST.get('cuenta')
            obj_fact = Facturas.objects.get(pk=radicado)
            obj_empresa = empresas.objects.get(pk=empresa)
            obj_cuenta = cuentas_pagos.objects.get(pk=cuenta)
            
            
            pago = Pagos.objects.create(
                nroradicado=obj_fact,valor=valor,usuario=request.user,
                fecha_pago=fecha,empresa=obj_empresa,cuenta=obj_cuenta,
                soporte_pago=soporte
            )
            history_facturas.objects.create(
                factura=obj_fact,usuario=request.user,
                accion=f'Registró un pago por {int(valor):,}',
                ubicacion='Tesoreria'
            )
            
            if cuenta != "9" and cuenta != "16" and cuenta != "25":
            
                lineas_banco = request.POST.get('lineas_banco')
                
                for i in lineas_banco.split(','):
                    obj_linea = egresos_banco.objects.get(pk=i)
                    obj_linea.pago_asociado = pago
                    obj_linea.save()
            
            data = {
                
            }
            
            return JsonResponse(data)

@login_required
def ajax_pagos_por_factura(request):
    if request.method == 'GET':
        if request.is_ajax():
            id_radicado = request.GET.get('id_radicado')
            obj_pagos = Pagos.objects.filter(nroradicado=id_radicado)
            
            list_pagos = []
            i=0
            for pago in obj_pagos:
                fecha = datetime.datetime.strftime(pago.fecha_pago,'%d/%m/%Y')
                list_pagos.append({
                    'id':i,
                    'pk':pago.pk,
                    'fecha':fecha,
                    'usuario':pago.usuario.username,
                    'valor':f'{pago.valor:,}',
                    'cuenta':pago.cuenta.cuentabanco,
                    'empresa':pago.empresa.nombre,
                    'soporte':_media_url(pago.soporte_pago)
                })
                i+=1
            data = {
                'data':list_pagos
            }
            return JsonResponse(data)

@login_required
def ajax_lista_facturas(request):
    if request.method == 'GET':
        if request.is_ajax():
            oficina = request.GET.get('oficina')
            draw = int(request.GET.get('draw', 1))
            start = int(request.GET.get('start', 0))
            length = int(request.GET.get('length', 10))
            if length <= 0:
                length = 10
            search_val = request.GET.get('search[value]')

            if not oficina:
                return JsonResponse({
                    "draw": draw,
                    "recordsTotal": 0,
                    "recordsFiltered": 0,
                    "data": []
                })

            if oficina == 'TODAS':
                base_qs = Facturas.objects.all()
            else:
                base_qs = Facturas.objects.filter(oficina=oficina)
            filtered_qs = base_qs

            if search_val:
                filtered_qs = filtered_qs.filter(
                    Q(nroradicado__icontains=search_val)|
                    Q(nombretercero__icontains=search_val)|
                    Q(nrocausa__icontains=search_val)|
                    Q(nrofactura__icontains=search_val)
                )

            total_records = base_qs.count()
            total_filtered = filtered_qs.count()

            # Aplicar anotaciones costosas solo a la página solicitada por DataTables.
            last_ubicacion_subquery = history_facturas.objects.filter(
                factura=OuterRef('pk')
            ).order_by('-fecha').values('ubicacion')[:1]

            defered_list = filtered_qs.order_by('-fecharadicado').select_related('empresa').annotate(
                total_pagado=Coalesce(Sum('pagos__valor'), 0, output_field=models.IntegerField()),
            ).annotate(
                saldo=Coalesce(
                    F('pago_neto') - F('total_pagado'),
                    F('pago_neto'),
                    output_field=models.IntegerField()
                ),
                ubicacion_actual=Subquery(last_ubicacion_subquery),
            )[start:start+length]
            
            i=0
            list_facturas = []
            for factura in defered_list:
                ubicacion = factura.ubicacion_actual or ''
                try: pagoneto = f'{factura.pago_neto:,}'
                except TypeError: pagoneto=''
                try:saldo = f'{factura.saldo:,}'
                except TypeError: saldo = 0
                list_facturas.append({
                    'id':i,
                    'pk':factura.pk,
                    'tercero':factura.nombretercero,
                    'factura':factura.nrofactura,
                    'fechafact':factura.fechafactura,
                    'fecharadicado':factura.fecharadicado,
                    'empresa':factura.empresa.nombre,
                    'valor':f'{factura.valor:,}',
                    'fechavencimiento':factura.fechavenc,
                    'causacion':factura.nrocausa,
                    'pagoneto':pagoneto,
                    'saldo':saldo,
                    'ubicacion':ubicacion,
                    'soporte':_media_url(factura.soporte_radicado),
                    'oficina':factura.oficina,
                    'soporte_causacion':_media_url(factura.soporte_causacion)
                })
                i+=1
            data={
                "draw": draw,
                "recordsTotal": total_records,
                "recordsFiltered": total_filtered,
                'data':list_facturas
                }
            return JsonResponse(data)

@login_required
def ajax_tipo_egreso(request):
    if request.method=='POST':
        if request.is_ajax():
            egreso = request.POST.get('egreso')
            tipo = request.POST.get('tipo')
            obj_egreso = egresos_contable.objects.get(pk=egreso)
            obj_egreso.tipo = tipo
            obj_egreso.save()
            
            data = {
                
            }
            return JsonResponse(data)

@login_required
def ajax_get_mvto_diario(request):
    if request.method == 'GET':
        if request.is_ajax():
            tipo = request.GET.get('tipo')
            fecha = request.GET.get('fecha')
            empresa = request.GET.get('empresa')
            cuentas = request.GET.get('cuentas')
            try: list_cuentas = cuentas.split(',')
            except AttributeError: list_cuentas = []
            obj_movimiento = []
            list_mvtos = []
            i=0
            lista_proyectos = proyectos.objects.exclude(proyecto__icontains='Alttum')
            if tipo == 'ingresos':
                for proyecto in lista_proyectos:
                    nombre_proy = proyecto.proyecto
                    for cuenta in list_cuentas:
                        cuenta = int(cuenta)
                        cuenta_pago_asociado = formas_pago.objects.using(nombre_proy).filter(cuenta_asociada=cuenta)
                        if cuenta_pago_asociado.exists():
                            obj_recaudos = Recaudos_general.objects.using(nombre_proy).filter(
                                fecha=fecha,formapago=cuenta_pago_asociado[0].descripcion
                            ).order_by('numrecibo')
                            for mvto in obj_recaudos:
                                cliente = clientes.objects.get(pk=mvto.idtercero)
                                list_mvtos.append({
                                    'id':i,
                                    'pk':mvto.pk,
                                    'fecha':mvto.fecha,
                                    'comprobante':mvto.numrecibo,
                                    'descripcion':cliente.nombrecompleto.upper(),
                                    'valor':f'{int(mvto.valor):,}',
                                    'cuenta':mvto.formapago,
                                    'proyecto':nombre_proy,
                                })
                                i+=1
                            obj_noradicados = RecaudosNoradicados.objects.using(nombre_proy).filter(
                                fecha=fecha,formapago=cuenta_pago_asociado[0].descripcion
                            ).order_by('recibo')
                            for mvto in obj_noradicados:
                                contrato = ventas_nuevas.objects.using(nombre_proy).get(pk=mvto.contrato)
                                cliente = clientes.objects.get(pk=contrato.id_t1)
                                list_mvtos.append({
                                    'id':i,
                                    'pk':mvto.pk,
                                    'fecha':mvto.fecha,
                                    'comprobante':mvto.recibo,
                                    'descripcion':cliente.nombrecompleto.upper() + ' (NO RADICADO)',
                                    'valor':f'{int(mvto.valor):,}',
                                    'cuenta':mvto.formapago,
                                    'proyecto':nombre_proy,
                                })
                                i+=1
                            
                            
            elif tipo == 'transferencias':
                for cuenta in list_cuentas:
                    cuenta = int(cuenta)
                    obj_transf = transferencias_companias.objects.filter(
                        Q(cuenta_entra=cuenta)|Q(cuenta_sale=cuenta),
                        fecha=fecha,
                        empresa_entra=F('empresa_sale')).order_by('empresa_sale')
                    for mvto in obj_transf:
                        check = True
                        for j in list_mvtos:
                            if j.get('pk','') == mvto.pk: 
                                check = False
                        if check:
                            list_mvtos.append({
                                'id':i,
                                'pk':mvto.pk,
                                'fecha':mvto.fecha,
                                'cuenta_entra':mvto.cuenta_entra.cuentabanco,
                                'cuenta_sale':mvto.cuenta_sale.cuentabanco,
                                'valor':f'{int(mvto.valor):,}',
                                'empresa':mvto.empresa_sale.nombre,
                            })
                            i+=1
                    
            elif tipo == 'otrosingresos':
                for cuenta in list_cuentas:
                    obj_otrosingresos = otros_ingresos.objects.filter(
                    empresa=empresa,cuenta=cuenta,fecha_ing=fecha
                )
                for mvto in obj_otrosingresos:
                    consecutivo = otros_ingresos.objects.filter(
                        pk__lte=mvto.pk,oficina=mvto.oficina
                    ).count()
                    oficinas={'MEDELLIN':'MDE','MONTERIA':'MTR'}
                    list_mvtos.append({
                        'pk':mvto.pk,
                        'fecha':mvto.fecha_ing,
                        'tercero':mvto.nombre_tercero,
                        'cuenta':mvto.cuenta.cuentabanco,
                        'valor':f'{int(mvto.valor):,}',
                        'consecutivo':f'{oficinas[mvto.oficina]}-{consecutivo}',
                    })
            
            
            elif tipo == 'intercomp':
                for cuenta in list_cuentas:
                    cuenta = int(cuenta)
                    obj_transf = transferencias_companias.objects.filter(
                        Q(cuenta_entra=cuenta)|Q(cuenta_sale=cuenta),
                        fecha=fecha,
                    ).exclude(empresa_entra=F('empresa_sale')).order_by('empresa_sale')
                    for mvto in obj_transf:
                        check = True
                        for j in list_mvtos:
                            if j.get('pk','') == mvto.pk: 
                                check = False
                        if check:
                            list_mvtos.append({
                                'id':i,
                                'pk':mvto.pk,
                                'fecha':mvto.fecha,
                                'empresa_entra':mvto.empresa_entra.nombre,
                                'empresa_sale':mvto.empresa_sale.nombre,
                                'cuenta_entra':mvto.cuenta_entra.cuentabanco,
                                'cuenta_sale':mvto.cuenta_sale.cuentabanco,
                                'valor':f'{int(mvto.valor):,}',
                                'empresa':mvto.empresa_sale.nombre,
                            })
                            i+=1
            
            elif tipo == 'anticipo':
                for cuenta in list_cuentas:
                    cuenta = int(cuenta)
                    obj_anticipos = Anticipos.objects.filter(
                        fecha_pago=fecha,cuenta=cuenta
                    )
                    for mvto in obj_anticipos:
                        check = True
                        list_mvtos.append({
                            'id':i,
                            'pk':mvto.pk,
                            'fecha':mvto.fecha_pago,
                            'tercero':mvto.nombre_tercero,
                            'descripcion':mvto.descripcion.upper(),
                            'cuenta':mvto.cuenta.cuentabanco,
                            'valor':f'{int(mvto.valor):,}',
                        })
                        i+=1
            
            elif tipo == 'pago':
                for cuenta in list_cuentas:
                    cuenta = int(cuenta)
                    obj_pagos = Pagos.objects.filter(
                        fecha_pago=fecha,cuenta=cuenta
                    )
                    for mvto in obj_pagos:
                        check = True
                        list_mvtos.append({
                            'id':i,
                            'pk':mvto.pk,
                            'radicado':mvto.nroradicado.pk,
                            'fecha':mvto.fecha_pago,
                            'tercero':mvto.nroradicado.nombretercero,
                            'descripcion':mvto.nroradicado.descripcion,
                            'cuenta':mvto.cuenta.cuentabanco,
                            'valor':f'{int(mvto.valor):,}',
                        })
                        i+=1
            
            
            data={
                'data':list_mvtos
            }
            return JsonResponse(data)

@login_required
def ajax_print_planilla(request):
    if request.method == 'GET':
        if request.is_ajax():
            fecha = request.GET.get('fecha')
            empresa = request.GET.get('empresa')
            cuentas = request.GET.get('cuentas')
            try: list_cuentas = cuentas.split(',')
            except AttributeError: list_cuentas = []
            
            list_efectivo = {
                'ingresos':0,
                'otros_ingresos':0,
                'transferencias_entra':0,
                'intercompañia_entra':0,
                'transferencias_sale':0,
                'intercompañia_sale':0,
                'anticipos':0,
                'pagos':0
            }
            
            try:
                #ingresos
                list_ingresos = {}
                list_mvtos = []
                lista_proyectos = proyectos.objects.exclude(proyecto__icontains='Alttum')
                for proyecto in lista_proyectos:
                    nombre_proy = proyecto.proyecto
                    for cuenta in list_cuentas:
                        cuenta = int(cuenta)
                        cuenta_pago_asociado = formas_pago.objects.using(nombre_proy).filter(cuenta_asociada=cuenta)
                        if cuenta_pago_asociado.exists():
                            obj_recaudos = Recaudos_general.objects.using(nombre_proy).filter(
                                fecha=fecha,formapago=cuenta_pago_asociado[0].descripcion
                            ).order_by('numrecibo')
                            for mvto in obj_recaudos:
                                cliente = clientes.objects.get(pk=mvto.idtercero)
                                if list_ingresos.get(nombre_proy):
                                    list_ingresos.get(nombre_proy).append(
                                        {
                                            'pk':mvto.pk,
                                            'fecha':mvto.fecha,
                                            'comprobante':mvto.numrecibo,
                                            'descripcion':cliente.nombrecompleto.upper(),
                                            'valor':f'{int(mvto.valor):,}',
                                            'cuenta':mvto.formapago,
                                        }
                                    )
                                else:
                                    list_ingresos.update({nombre_proy:[
                                        {
                                            'pk':mvto.pk,
                                            'fecha':mvto.fecha,
                                            'comprobante':mvto.numrecibo,
                                            'descripcion':cliente.nombrecompleto.upper(),
                                            'valor':f'{int(mvto.valor):,}',
                                            'cuenta':mvto.formapago,
                                        }]
                                    })
                                if 'EFECTIVO' in mvto.formapago.upper():
                                    list_efectivo['ingresos'] += mvto.valor
                                    
                            obj_noradicados = RecaudosNoradicados.objects.using(nombre_proy).filter(
                                    fecha=fecha,formapago=cuenta_pago_asociado[0].descripcion
                                ).order_by('recibo')
                            for mvto in obj_noradicados:
                                
                                contrato = ventas_nuevas.objects.using(nombre_proy).get(pk=mvto.contrato)
                                cliente = clientes.objects.get(pk=contrato.id_t1)
                                
                                if list_ingresos.get(nombre_proy):
                                    list_ingresos.get(nombre_proy).append(
                                        {
                                            'pk':mvto.pk,
                                            'fecha':mvto.fecha,
                                            'comprobante':mvto.recibo,
                                            'descripcion':cliente.nombrecompleto.upper(),
                                            'valor':f'{int(mvto.valor):,}',
                                            'cuenta':mvto.formapago,
                                        }
                                    )
                                else:
                                    list_ingresos.update({nombre_proy:[
                                        {
                                            'pk':mvto.pk,
                                            'fecha':mvto.fecha,
                                            'comprobante':mvto.recibo,
                                            'descripcion':cliente.nombrecompleto.upper(),
                                            'valor':f'{int(mvto.valor):,}',
                                            'cuenta':mvto.formapago,
                                        }]
                                    })
                                if 'EFECTIVO' in mvto.formapago.upper():
                                    list_efectivo['ingresos'] += mvto.valor
                                
                #otros ingresos
                list_otrosingresos = []
                for cuenta in list_cuentas:
                    cuenta = int(cuenta)
                    obj_otrosingresos = otros_ingresos.objects.filter(
                        empresa=empresa,cuenta=cuenta,fecha_ing=fecha
                    )
                    for mvto in obj_otrosingresos:
                        consecutivo = otros_ingresos.objects.filter(
                            pk__lte=mvto.pk,oficina=mvto.oficina
                        ).count()
                        oficinas={'MEDELLIN':'MDE','MONTERIA':'MTR'}
                        list_otrosingresos.append({
                            'pk':mvto.pk,
                            'fecha':mvto.fecha_ing,
                            'tercero':mvto.nombre_tercero,
                            'cuenta':mvto.cuenta.cuentabanco,
                            'valor':f'{int(mvto.valor):,}',
                            'consecutivo':f'{oficinas[mvto.oficina]}-{consecutivo}',
                        })
                        if 'EFECTIVO' in mvto.cuenta.cuentabanco.upper():
                            list_efectivo['otros_ingresos'] += mvto.valor
                #transferencias
                list_transfer = []
                for cuenta in list_cuentas:
                    cuenta = int(cuenta)
                    obj_transf = transferencias_companias.objects.filter(
                        Q(cuenta_entra=cuenta)|Q(cuenta_sale=cuenta),
                        fecha=fecha,
                        empresa_entra=F('empresa_sale')).order_by('empresa_sale')
                    for mvto in obj_transf:
                        check = True
                        for j in list_transfer:
                            if j.get('pk','') == mvto.pk: 
                                check = False
                        if check:
                            list_transfer.append({
                                'pk':mvto.pk,
                                'fecha':mvto.fecha,
                                'cuenta_entra':mvto.cuenta_entra.cuentabanco,
                                'cuenta_sale':mvto.cuenta_sale.cuentabanco,
                                'valor':f'{int(mvto.valor):,}',
                                'empresa':mvto.empresa_sale.nombre,
                                'tipo':'transferencias'
                            })
                            if 'EFECTIVO' in mvto.cuenta_entra.cuentabanco.upper():
                                list_efectivo['transferencias_entra'] += mvto.valor
                            elif 'EFECTIVO' in mvto.cuenta_sale.cuentabanco.upper():
                                list_efectivo['transferencias_sale'] += mvto.valor
                #intercompañia
                list_intercomp = []
                for cuenta in list_cuentas:
                    cuenta = int(cuenta)
                    obj_transf = transferencias_companias.objects.filter(
                        Q(cuenta_entra=cuenta)|Q(cuenta_sale=cuenta),
                        fecha=fecha,
                    ).exclude(empresa_entra=F('empresa_sale')).order_by('empresa_sale')
                    for mvto in obj_transf:
                        check = True
                        for j in list_intercomp:
                            if j.get('pk','') == mvto.pk: 
                                check = False
                        if check:
                            if empresa == mvto.empresa_entra.pk: tipo='entra'
                            else: tipo = 'sale'
                            list_intercomp.append({
                                'pk':mvto.pk,
                                'fecha':mvto.fecha,
                                'empresa_entra':mvto.empresa_entra.nombre,
                                'empresa_sale':mvto.empresa_sale.nombre,
                                'cuenta_entra':mvto.cuenta_entra.cuentabanco,
                                'cuenta_sale':mvto.cuenta_sale.cuentabanco,
                                'valor':f'{int(mvto.valor):,}',
                                'empresa':mvto.empresa_sale.nombre,
                                'tipo':tipo
                                })
                            if 'EFECTIVO' in mvto.cuenta_entra.cuentabanco.upper() and tipo=='entra':
                                list_efectivo['intercompañia_entra'] += mvto.valor
                            elif 'EFECTIVO' in mvto.cuenta_sale.cuentabanco.upper() and tipo=='sale':
                                list_efectivo['intercompañia_sale'] += mvto.valor
                #anticipos
                list_anticipos = []
                for cuenta in list_cuentas:
                    cuenta = int(cuenta)
                    obj_anticipos = Anticipos.objects.filter(
                        fecha_pago=fecha,cuenta=cuenta
                    )
                    for mvto in obj_anticipos:
                        check = True
                        list_anticipos.append({
                            'pk':mvto.pk,
                            'fecha':mvto.fecha_pago,
                            'tercero':mvto.nombre_tercero,
                            'descripcion':mvto.descripcion.upper(),
                            'cuenta':mvto.cuenta.cuentabanco,
                            'valor':f'{int(mvto.valor):,}',
                            'tipo':'anticipos'
                        })
                        if 'EFECTIVO' in mvto.cuenta.cuentabanco.upper():
                            list_efectivo['anticipos'] += mvto.valor
                #pagos
                list_pagos = []
                for cuenta in list_cuentas:
                    cuenta = int(cuenta)
                    obj_pagos = Pagos.objects.filter(
                        fecha_pago=fecha,cuenta=cuenta
                    )
                    for mvto in obj_pagos:
                        check = True
                        list_pagos.append({
                            'pk':mvto.pk,
                            'causacion':mvto.nroradicado.nrocausa,
                            'radicado':mvto.nroradicado.pk,
                            'fecha':mvto.fecha_pago,
                            'tercero':mvto.nroradicado.nombretercero,
                            'descripcion':mvto.nroradicado.descripcion,
                            'cuenta':mvto.cuenta.cuentabanco,
                            'valor':f'{int(mvto.valor):,}',
                            'tipo':'pagos'
                        })
                        if 'EFECTIVO' in mvto.cuenta.cuentabanco.upper():
                            list_efectivo['pagos'] += mvto.valor
                
                has_efectivo =False
                cuenta_efectivo=None
                obj_saldo_ini=None
                cuentas_seleccionadas = []
                for cuenta in list_cuentas:
                    nombre_cta = cuentas_pagos.objects.get(pk=cuenta).cuentabanco
                    cuentas_seleccionadas.append(nombre_cta)
                    if 'EFECTIVO' in nombre_cta.upper():
                        has_efectivo=True
                        cuenta_efectivo = cuentas_pagos.objects.get(pk=cuenta)
                        obj_saldo_ini=saldos_cuentas_tesoreria.objects.filter(cuenta=cuenta_efectivo.pk).order_by('-fecha')
                
                obj_empresa = empresas.objects.get(pk=empresa)
                filename = f'Planilla_Mvtos_{empresa}_{fecha}.pdf'
                GenerarPDF().planillaMovimientos(
                    empresa=obj_empresa,fecha =fecha,cuentas=cuentas_seleccionadas,
                    obj_ingresos=list_ingresos,
                    obj_otrosing = list_otrosingresos,
                    obj_transf = list_transfer,
                    obj_intercomp=list_intercomp,
                    obj_anticipos = list_anticipos,
                    obj_pagos = list_pagos,
                    usuario = request.user,
                    has_efectivo = has_efectivo,
                    cuenta_efectivo =cuenta_efectivo,
                    pagos_efectivo=list_efectivo,
                    obj_saldo_ini = obj_saldo_ini,
                    ruta = settings.MEDIA_ROOT+f'/tmp/pdf/{filename}'
                )
                ruta_dw = settings.MEDIA_URL+f'tmp/pdf/{filename}'
                data = {
                    'title':'Listo!',
                    'msj':f'Puedes descargar la planilla generada <b><a href="{ruta_dw}" target="_blank">Aqui</b>',
                    'class':'alert-success'
                }
            except:
                data = {
                    'title':'Oh Oh',
                    'msj':traceback.format_exc(),
                    'class':'alert-danger'
                }
            return JsonResponse(data)

@login_required
def ajax_lista_pagos(request):
    if request.method == 'GET':
        if request.is_ajax():
            oficina = request.GET.get('oficina')
            if oficina == 'TODAS':
                obj_pagos = Pagos.objects.all().order_by('-fecha_pago')
            else:
                obj_pagos = Pagos.objects.filter(nroradicado__oficina=oficina).order_by('-fecha_pago')
            
            draw = request.GET.get('draw')
            start = request.GET.get('start')
            length = request.GET.get('length')
            search_val = request.GET.get('search[value]')
            
            if search_val != "":
                obj_pagos = obj_pagos.filter(
                    Q(nroradicado__nombretercero__icontains = search_val)|
                    Q(nroradicado__nrofactura__icontains = search_val)|
                    Q(nroradicado__nrocausa__icontains = search_val)
                )
            if int(length) == -1:
                defered_list = obj_pagos
            else:
                defered_list = obj_pagos[int(start):int(start)+int(length)]
            i=0
            list_mvto=[]
            for line in defered_list:
                fecha_fact = datetime.datetime.strftime(line.nroradicado.fechafactura,'%d/%m/%Y')
                fecha_pago = datetime.datetime.strftime(line.fecha_pago,'%d/%m/%Y')
                try: sc = _media_url(line.nroradicado.soporte_causacion)
                except: sc = ""
                list_mvto.append({
                    'id':i,
                    'pk':line.pk,
                    'tercero':line.nroradicado.nombretercero,
                    'fecha_fact':fecha_fact,
                    'fecha_pago':fecha_pago,
                    'valor':f'{line.valor:,}',
                    'empresa':line.empresa.nombre,
                    'cuenta':line.cuenta.cuentabanco,
                    'causacion':line.nroradicado.nrocausa,
                    'usuario':line.usuario.username,
                    'factura':line.nroradicado.nrofactura,
                    'soporte':_media_url(line.soporte_pago),
                    'soporte_radicado':_media_url(line.nroradicado.soporte_radicado),
                    'soporte_causacion':sc,
                    'oficina':line.nroradicado.oficina,
                    'tipo_rad':line.nroradicado.origen,
                })
                i+=0
            data = {
                "draw": int(draw),
                "recordsTotal": obj_pagos.count(),
                "recordsFiltered": obj_pagos.count(),
                'data':list_mvto
            }
            return JsonResponse(data)

@login_required
def ajax_info_interfaces(request):
    if request.method == 'GET':
        if request.is_ajax():
            tipo = request.GET.get('tipo')
            empresa = request.GET.get('empresa')
            try: obj_empresa = empresas.objects.get(nombre=empresa)
            except: obj_empresa = empresas.objects.get(pk=empresa)
            
            obj_cxp = info_interfaces.objects.filter(
                empresa=obj_empresa.pk,descripcion__icontains=tipo
            )
            json_cxp = serializers.serialize('json',obj_cxp)
            data = {
                'data':json_cxp
            }
            return JsonResponse(data)

@login_required
def ajax_registrar_anticipo(request):
    if request.method == 'POST':
        if request.is_ajax():
            todo = request.POST.get('todo')
            
            if todo == 'onlylink':
                anticipo = request.POST.get('anticipo')
                lineas_banco = request.POST.get('lineas_banco[]')
                obj_anticipo = Anticipos.objects.get(pk=anticipo)
                for i in lineas_banco.split(','):
                    obj_linea = egresos_banco.objects.get(pk=i)
                    obj_linea.anticipo_asociado = obj_anticipo
                    obj_linea.save()
                    
                return JsonResponse({})
                
            empresa = request.POST.get('empresa_ant')
            cuenta = request.POST.get('cuenta_ant')
            fecha = request.POST.get('fecha_anticip')
            id_tercero = request.POST.get('id_tercero')
            nombre_tercero = request.POST.get('nombre_tercero')
            valor_anticipo = request.POST.get('valor_anticip').replace(',','')
            tipo = request.POST.get('tipo_anticipo')
            descripcion = request.POST.get('descripcion_ant')
            oficina = request.POST.get('oficina')
            soporte = request.FILES['soporte_pago']
            
            
            anticipo = Anticipos.objects.create(
                empresa=empresas.objects.get(pk=empresa),
                descripcion = descripcion,valor=valor_anticipo,usuario=request.user,
                cuenta=cuentas_pagos.objects.get(pk=cuenta),fecha_pago=fecha,
                id_tercero=id_tercero,nombre_tercero=nombre_tercero,
                tipo_anticipo = info_interfaces.objects.get(pk=tipo),
                soporte_pago=soporte,oficina=oficina
            )
            
            if cuenta != "9" and cuenta != "16" and cuenta != "25":
            
                lineas_banco = request.POST.get('lineas_banco')
                    
                for i in lineas_banco.split(','):
                    obj_linea = egresos_banco.objects.get(pk=i)
                    obj_linea.anticipo_asociado = anticipo
                    obj_linea.save()
            
                     
            return JsonResponse({})

@login_required
def ajax_get_comisiones(request):
    if request.method == 'GET':
        if request.is_ajax():
            fecha = request.GET.get('fecha')
            proyecto = request.GET.get('proyecto').title().replace('Del','del')
            stmt=f'CALL detalle_comisiones_fecha("{fecha}","{fecha}")'
            comisiones=Pagocomision.objects.using(proyecto).raw(stmt)

            obj_comis = []
            for line in comisiones:
                datos_asesor = asesores.objects.get(pk=line.idgestor)
                pagoneto = int(line.pagoneto)
                obj_comis.append({
                    'idgestor':line.idgestor,
                    'nombregestor':line.nombre.upper(),
                    'pagoneto':f'{pagoneto:,}'
                })
            data={
                'comisiones':obj_comis
            }
            return JsonResponse(data)

@login_required
def ajax_print_egreso(request):
    if request.method == 'GET':
        if request.is_ajax():
            anticipo = request.GET.get('anticipo')
            if anticipo:
                id_pago = request.GET.get('idpago')
                obj_pagos = Anticipos.objects.get(pk=id_pago)
                consecutivo = Anticipos.objects.filter(pk__lt=obj_pagos.pk,
                                                oficina=obj_pagos.oficina
                            ).count()+1
                filename = f'Anticipo_{obj_pagos.oficina}_{consecutivo}.pdf'
                ruta = settings.MEDIA_ROOT+'/tmp/pdf/'+filename
                GenerarPDF().reciboAnticipos(obj_pagos,consecutivo,ruta)
                
                data = {
                    'ruta':settings.MEDIA_URL+'/tmp/pdf/'+filename
                }
            else:
                id_pago = request.GET.get('idpago')
                obj_pagos = Pagos.objects.get(pk=id_pago)
                consecutivo = Pagos.objects.filter(pk__lt=obj_pagos.pk,
                                                nroradicado__oficina=obj_pagos.nroradicado.oficina
                            ).count()+1
                filename = f'Egreso_{obj_pagos.nroradicado.oficina}_{consecutivo}.pdf'
                ruta = settings.MEDIA_ROOT+'/tmp/pdf/'+filename
                GenerarPDF().reciboEgreso(obj_pagos,consecutivo,ruta)
                
                data = {
                    'ruta':settings.MEDIA_URL+'/tmp/pdf/'+filename
                }

            return JsonResponse(data)

@login_required
def ajax_lista_otros_ingresos(request):
    if request.method == 'GET':
        if request.is_ajax():
            obj_otros_ingresos = otros_ingresos.objects.all()
            list_ingresos = []
            i = 0
            for ingreso in obj_otros_ingresos:
                list_ingresos.append({
                    'id':i,
                    'pk':ingreso.pk,
                    'fecha':ingreso.fecha_ing,
                    'tercero':ingreso.nombre_tercero,
                    'concepto':ingreso.descripcion,
                    'empresa':ingreso.empresa.nombre,
                    'cuenta':ingreso.cuenta.cuentabanco,
                    'usuario':ingreso.usuario.username,
                    'oficina':ingreso.oficina,
                    'valor':f'{ingreso.valor:,}',
                })
                i += 1
            data = {
                'data':list_ingresos
            }
            return JsonResponse(data)

@login_required
def ajax_imprimir_otros_ingresos(request):
    if request.method == 'GET':
        if request.is_ajax():
            id_rcdo = request.GET.get('pk')
            ingreso = otros_ingresos.objects.get(pk=id_rcdo)
            consecutivo_ofic = otros_ingresos.objects.filter(pk__lte=ingreso.pk,
                    oficina=ingreso.oficina).count()
            filename = f'Recibo_ingreso_{ingreso.oficina}_{consecutivo_ofic}.pdf'
            ruta = settings.MEDIA_ROOT+'/tmp/pdf/'+filename
            GenerarPDF().reciboIngreso(ingreso,consecutivo_ofic,ruta)
            ruta_download = settings.MEDIA_URL + 'tmp/pdf/' + filename
            data = {
                'ruta':ruta_download
            }
            return JsonResponse(data)

@login_required
def ajax_transferencia(request):
    if request.method == 'POST':
        if request.is_ajax():
            
            is_tesor = True if check_groups(request,('Tesoreria',),raise_exception=False) else False
            if not is_tesor:
                data = {
                    'msj': 'No tienes permiso de realizar un pago.',
                    'class': 'alert-danger'
                }
                
                return JsonResponse(data)
            
            empresa_entra = request.POST.get('empresa_entra')
            cuenta_entra = request.POST.get('cuenta_entra')
            empresa_sale = request.POST.get('empresa_sale')
            cuenta_sale = request.POST.get('cuenta_sale')
            fecha = request.POST.get('fecha_transf')
            oficina = request.POST.get('oficina')
            valor = request.POST.get('valor_transf').replace(',','')
            soporte = request.FILES.get('soporte')
            id_reemb = request.POST.get('reembolso')
            
            cuenta_entra = cuentas_pagos.objects.get(pk=cuenta_entra)
            cuenta_sale = cuentas_pagos.objects.get(pk=cuenta_sale)
            
            if id_reemb is not None:
                oficina = cuenta_entra.oficina
            
            
            transferencia = transferencias_companias.objects.create(valor=valor,fecha=fecha,
                    empresa_entra=empresas.objects.get(pk=empresa_entra),
                    empresa_sale=empresas.objects.get(pk=empresa_sale),
                    cuenta_entra = cuenta_entra,
                    cuenta_sale = cuenta_sale,
                    oficina = oficina,usuario=request.user, soporte_pago = soporte
                )
            
            if not "Efectivo" in cuenta_sale.cuentabanco:
                lineas_banco = request.POST.get('lineas_banco')
                    
                for i in lineas_banco.split(','):
                    obj_linea = egresos_banco.objects.get(pk=i)
                    obj_linea.transferencia_asociada = transferencia
                    obj_linea.save()
            
            data = {
                    'msj': 'Se creó la transferencia con exito',
                    'class': 'alert-success'
                }
            
            
            if id_reemb is not None:
                reembolso = reembolsos_caja.objects.get(pk=id_reemb)
                reembolso.transferencia_asociada = transferencia
                reembolso.save()
            
            return JsonResponse(data)

@login_required
def ajax_lista_transf(request):
    if request.method == 'GET':
        if request.is_ajax():
            oficina = request.GET.get('oficina')
            draw = request.GET.get('draw')
            start = request.GET.get('start')
            length = request.GET.get('length')
            search_val = request.GET.get('search[value]')
            
            if oficina == 'TODAS':
                obj_transf = transferencias_companias.objects.all().order_by('-fecha')
            else:
                obj_transf = transferencias_companias.objects.filter(oficina=oficina).order_by('-fecha')
                
            if search_val != "":
                obj_transf = obj_transf.filter(
                    Q(cuenta_entra__cuentabanco__icontains = search_val)|
                    Q(cuenta_sale__cuentabanco__icontains = search_val)|
                    Q(empresa_entra__nombre__icontains = search_val)|
                    Q(empresa_sale__nombre__icontains = search_val)
                )
            defered_list = obj_transf[int(start):int(start)+int(length)]
            i=0
            list_transf = []
            for transf in defered_list:
                fecha_pago = datetime.datetime.strftime(transf.fecha,'%d/%m/%Y')
                list_transf.append({
                    'id':i,
                    'pk':transf.pk,
                    'fecha':fecha_pago,
                    'emp_sale':transf.empresa_sale.nombre,
                    'cta_sale':transf.cuenta_sale.cuentabanco,
                    'emp_entra':transf.empresa_entra.nombre,
                    'cta_entra':transf.cuenta_entra.cuentabanco,
                    'usuario':transf.usuario.username,
                    'oficina':transf.oficina,
                    'valor':f'{transf.valor:,}',
                    'soporte':_media_url(transf.soporte_pago)
                })
                i+=1
            data = {
                "draw": int(draw),
                "recordsTotal": obj_transf.count(),
                "recordsFiltered": obj_transf.count(),
                'data':list_transf
            }
            return JsonResponse(data)

@login_required
def ajax_lista_ant(request):
    if request.method == 'GET':
        if request.is_ajax():
            oficina = request.GET.get('oficina')
            draw = request.GET.get('draw')
            start = request.GET.get('start')
            length = request.GET.get('length')
            search_val = request.GET.get('search[value]')
            if oficina == 'TODAS':
                obj_ant = Anticipos.objects.all().order_by('-fecha_pago')
            else:
                obj_ant = Anticipos.objects.filter(oficina=oficina).order_by('-fecha_pago')
            
            if search_val != "":
                obj_ant = obj_ant.filter(
                    Q(cuenta__cuentabanco__icontains = search_val)|
                    Q(empresa__nombre__icontains = search_val)|
                    Q(nombre_tercero__icontains = search_val)|
                    Q(descripcion__icontains = search_val)
                )
            
            defered_list = obj_ant[int(start):int(start)+int(length)]
                        
            i=0
            list_transf = []
            for ant in defered_list:
                fecha_pago = datetime.datetime.strftime(ant.fecha_pago,'%d/%m/%Y')
                list_transf.append({
                    'id':i,
                    'pk':ant.pk,
                    'fecha':fecha_pago,
                    'fecha_nofmt':ant.fecha_pago,
                    'tercero':ant.nombre_tercero,
                    'empresa':ant.empresa.nombre,
                    'descripcion':ant.descripcion,
                    'cuenta':ant.cuenta.cuentabanco,
                    'pkcuenta':ant.cuenta.pk,
                    'usuario':ant.usuario.username,
                    'oficina':ant.oficina,
                    'valor':f'{ant.valor:,}',
                    'soporte':_media_url(ant.soporte_pago),
                    'islinked': ant.is_linked()
                })
                i+=1
            data = {
                "draw": int(draw),
                "recordsTotal": obj_ant.count(),
                "recordsFiltered": obj_ant.count(),
                'data':list_transf
            }
            return JsonResponse(data)

@login_required
def ajax_gtt_por_proy(request):
    if request.method == 'GET':
        if request.is_ajax():
            proyecto = request.GET.get('proyecto')
            tipo = request.GET.get('tipo')
            if tipo == 'all':
                obj_gtt = Gtt.objects.filter(proyecto=proyecto,estado='Aprobado'
                    ).order_by('-fecha_desde')
                json_gtt = serializers.serialize('json',obj_gtt)
                return JsonResponse({'data':json_gtt})
            elif tipo == 'detalle':
                gtt = request.GET.get('gtt')
                obj_detalle = Detalle_gtt.objects.filter(gtt=gtt)
                list_mvtos = []
                total = 0
                for mvto in obj_detalle:
                    list_mvtos.append({
                        'cc':mvto.asesor.pk,
                        'nombre':mvto.asesor.nombre,
                        'valor':mvto.valor
                    })
                    total+=mvto.valor
                total = f'{total:,}'
                return JsonResponse({'data':list_mvtos,'total':total})
                
    if request.method == 'POST':
        if request.is_ajax():
            empresa = request.POST.get('empresa_gtt')
            radicado = request.POST.get('radicado')
            valor = request.POST.get('valor_gtt').replace(',','')
            cuenta = request.POST.get('cuenta_gtt')
            fecha = request.POST.get('fecha_pago_gtt')
            id_gtt = request.POST.get('fecha_gtt')
            idterceros = request.POST.getlist('idTercero')
            nombreterceros = request.POST.getlist('nombretercero')
            vencimientos = request.POST.getlist('vencimientoGtt')
            valor_detalle = request.POST.getlist('valorGtt')
            soporte = request.FILES.get('soporte_gtt')
            
            obj_cuenta = cuentas_pagos.objects.get(pk=cuenta)
            obj_empresa = empresas.objects.get(pk=empresa)
            obj_radicado = Facturas.objects.get(pk=radicado)
            
            obj_pago = Pagos.objects.create(
                empresa=obj_empresa,nroradicado=obj_radicado,
                valor=valor,usuario=request.user,
                cuenta=obj_cuenta,fecha_pago=fecha,
                soporte_pago=soporte
            )
            
            for i in range(len(idterceros)):
                pago_detallado_relacionado.objects.create(
                    valor=valor_detalle[i],tipo='GTT',
                    id_tercero=idterceros[i],
                    nombre_tercero=nombreterceros[i],
                    pago=obj_pago,vencimiento=vencimientos[i]
                )
            
            history_facturas.objects.create(
                factura = obj_radicado,usuario= request.user,
                accion='Asoció pago detallado de GTT',ubicacion='Tesoreria'
            )
            
            lineas_banco = request.POST.get('lineas_banco')
                
            for i in lineas_banco.split(','):
                obj_linea = egresos_banco.objects.get(pk=i)
                obj_linea.pago_asociado = obj_pago
                obj_linea.save()
            
            data = {
                'msj':'listo'
            }
            return JsonResponse(data)

@login_required
def ajax_comis_por_proy(request):
    if request.method == 'GET':
        if request.is_ajax():
            proyecto = request.GET.get('proyecto')
            fecha = request.GET.get('fecha')
            stmt=f'CALL detalle_comisiones_fecha("{fecha}","{fecha}")'
            comisiones=Pagocomision.objects.using(proyecto).raw(stmt) 
            list_mvtos = []
            total = 0
            for mvto in comisiones:
                list_mvtos.append({
                    'cc':mvto.idgestor,
                    'nombre':mvto.nombre,
                    'valor':mvto.pagoneto
                })
                total+=mvto.pagoneto
            total = f'{total:,}'
            return JsonResponse({'data':list_mvtos,'total':total})
                
    if request.method == 'POST':
        if request.is_ajax():
            try:
                empresa = request.POST.get('empresa_comis')
                radicado = request.POST.get('radicado')
                valor = request.POST.get('valor_comis').replace(',','')
                valor = valor.replace('.00','')
                cuenta = request.POST.get('cuenta_comis')
                fecha = request.POST.get('fecha_pago_comis')
                id_gtt = request.POST.get('fecha_comis')
                idterceros = request.POST.getlist('idTercero')
                nombreterceros = request.POST.getlist('nombretercero')
                vencimientos = request.POST.getlist('vencimientocomis')
                valor_detalle = request.POST.getlist('valorComis')
                soporte = request.FILES.get('soporte_comis')
                
                obj_cuenta = cuentas_pagos.objects.get(pk=cuenta)
                obj_empresa = empresas.objects.get(pk=empresa)
                obj_radicado = Facturas.objects.get(pk=radicado)
                
                obj_pago = Pagos.objects.create(
                    empresa=obj_empresa,nroradicado=obj_radicado,
                    valor=valor,usuario=request.user,
                    cuenta=obj_cuenta,fecha_pago=fecha,soporte_pago=soporte
                )
                
                for i in range(len(idterceros)):
                    pago_detallado_relacionado.objects.create(
                        valor=valor_detalle[i],tipo='COMISIONES',
                        id_tercero=idterceros[i],
                        nombre_tercero=nombreterceros[i],
                        pago=obj_pago,vencimiento=vencimientos[i]
                    )
                
                history_facturas.objects.create(
                    factura = obj_radicado,usuario= request.user,
                    accion='Asoció pago detallado de comisiones',ubicacion='Tesoreria'
                )
                
                lineas_banco = request.POST.get('lineas_banco')
                
                for i in lineas_banco.split(','):
                    obj_linea = egresos_banco.objects.get(pk=i)
                    obj_linea.pago_asociado = obj_pago
                    obj_linea.save()
                
                data = {
                    'title':'Listo',
                    'msj':'Se registraron los pagos de forma correcta',
                    'class':'alert-success'
                }
            except:
                data = {
                    'title':'Oh Oh!',
                    'msj':traceback.format_exc(),
                    'class':'alert-danger'
                }
            
            return JsonResponse(data)

@login_required
def ajax_otros_detalle(request):
                
    if request.method == 'POST':
        if request.is_ajax():
            try:
                empresa = request.POST.get('empresa_otros')
                radicado = request.POST.get('radicado')
                valor = request.POST.get('valor_otros').replace(',','')
                valor = valor.replace('.00','')
                cuenta = request.POST.get('cuenta_otros')
                fecha = request.POST.get('fecha_pago_otros')
                id_gtt = request.POST.get('fecha_comis')
                idterceros = request.POST.getlist('idTercero')
                nombreterceros = request.POST.getlist('nombretercero')
                vencimientos = request.POST.getlist('vencimiento_otros')
                valor_detalle = request.POST.getlist('valor_det_otros')
                soporte = request.FILES.get('soporte_otros')
                
                obj_cuenta = cuentas_pagos.objects.get(pk=cuenta)
                obj_empresa = empresas.objects.get(pk=empresa)
                obj_radicado = Facturas.objects.get(pk=radicado)
                
                obj_pago = Pagos.objects.create(
                    empresa=obj_empresa,nroradicado=obj_radicado,
                    valor=valor,usuario=request.user,
                    cuenta=obj_cuenta,fecha_pago=fecha,soporte_pago=soporte
                )
                
                for i in range(len(idterceros)):
                    pago_detallado_relacionado.objects.create(
                        valor=valor_detalle[i].replace(',',''),tipo='OTROS',
                        id_tercero=idterceros[i],
                        nombre_tercero=nombreterceros[i],
                        pago=obj_pago,vencimiento=vencimientos[i]
                    )
                
                history_facturas.objects.create(
                    factura = obj_radicado,usuario= request.user,
                    accion='Asoció pago detallado',ubicacion='Tesoreria'
                )
                
                lineas_banco = request.POST.get('lineas_banco')
                
                for i in lineas_banco.split(','):
                    obj_linea = egresos_banco.objects.get(pk=i)
                    obj_linea.pago_asociado = obj_pago
                    obj_linea.save()
                
                data = {
                    'title':'Listo',
                    'msj':'Se registraron los pagos de forma correcta',
                    'class':'alert-success'
                }
            except:
                data = {
                    'title':'Oh Oh!',
                    'msj':traceback.format_exc(),
                    'class':'alert-danger'
                }
            
            return JsonResponse(data)
        
@login_required
def ajax_nomina_detalle(request):
    if request.method == 'POST':
        if request.is_ajax():
            try:
                empresa = request.POST.get('empresa_pago_nomina')
                radicado = request.POST.get('radicado')
                valor = request.POST.get('valor_nomina').replace(',', '')
                valor = valor.replace('.00', '')
                cuenta = request.POST.get('cuenta_nomina')
                fecha = request.POST.get('fecha_pago_nomina')
                idterceros = request.POST.getlist('idTercero')
                nombreterceros = request.POST.getlist('nombretercero')
                vencimientos = request.POST.getlist('vencimientonomina')
                valor_detalle = request.POST.getlist('valornomina')
                soporte = request.FILES.get('soporte_nomina')

                obj_cuenta = cuentas_pagos.objects.get(pk=cuenta)
                obj_empresa = empresas.objects.get(pk=empresa)
                obj_radicado = Facturas.objects.get(pk=radicado)

                obj_pago = Pagos.objects.create(
                    empresa=obj_empresa, nroradicado=obj_radicado,
                    valor=valor, usuario=request.user,
                    cuenta=obj_cuenta, fecha_pago=fecha, soporte_pago=soporte
                )

                a, b = 'áéíóúüñ', 'aeiouun'
                trans = str.maketrans(a, b)
                for i in range(len(idterceros)):
                    nombre = nombreterceros[i].translate(trans)
                    pago_detallado_relacionado.objects.create(
                        valor=valor_detalle[i].replace(',', ''), tipo='NOMINA',
                        id_tercero=idterceros[i],
                        nombre_tercero=nombre,
                        pago=obj_pago, vencimiento=vencimientos[i]
                    )

                history_facturas.objects.create(
                    factura=obj_radicado, usuario=request.user,
                    accion='Asoció pago detallado', ubicacion='Tesoreria'
                )
                # Asociar movimientos bancarios solo si lineas_banco no está vacío
                lineas_banco = request.POST.get('lineas_banco')
                if lineas_banco:
                    for i in lineas_banco.split(','):
                        if i.strip():
                            obj_linea = egresos_banco.objects.get(pk=i)
                            obj_linea.pago_asociado = obj_pago
                            obj_linea.save()

                data = {
                    'title': 'Listo',
                    'msj': 'Se registraron los pagos de forma correcta',
                    'class': 'alert-success'
                }
            except:
                data = {
                    'title': 'Oh Oh!',
                    'msj': traceback.format_exc(),
                    'class': 'alert-danger'
                }

            return JsonResponse(data)

@login_required
def ajax_saldos_iniciales(request):
    if request.is_ajax():
        if request.method == 'POST':
            cuenta = request.POST.get('cuenta')
            empresa = request.POST.get('empresa')
            fecha = request.POST.get('fecha')
            saldo_inicial = request.POST.get('saldoinicial').replace(',','')
            forma = request.POST.get('forma')
            tipo = request.POST.get('tipo')
            
            try:
                obj_cuenta = cuentas_pagos.objects.get(pk=cuenta)
                if tipo == 'cajamenor':
                    if obj_cuenta.usuario_responsable.pk != request.user.pk:
                        data = {
                            'title':'Oh oh!',
                            'msj':'Solo el usuario responsable de la caja puede realizar arqueos',
                            'class':'alert-danger'
                        }
                        
                        return JsonResponse(data)
                
                #Calculamos el saldo que viene
                valor = 0
                saldo_transfer = transferencias_companias.objects.filter(
                    cuenta_entra = cuenta,
                    fecha__lte = fecha,
                    fecha__gte = "2024-08-01"
                ).aggregate(total=Sum('valor')).get('total')
                
                if saldo_transfer != None: valor += saldo_transfer
                
                saldo_transfer_sale = transferencias_companias.objects.filter(
                    cuenta_sale = cuenta,
                    fecha__lte = fecha,
                    fecha__gte = "2024-08-01"
                ).aggregate(total=Sum('valor')).get('total')
                
                if saldo_transfer_sale != None: valor -= saldo_transfer_sale
                
                saldo_hoteles = Pagos_facturas.objects.using('Alttum').filter(
                    forma_pago__cuenta_andinasoft = cuenta,
                    fecha__lte = fecha,
                    fecha__gte = "2024-08-01"
                ).aggregate(total=Sum('valor')).get('total')
                
                if saldo_hoteles != None: valor += saldo_hoteles
                
                saldo_ant_hoteles = anticipos_hotels.objects.using('Alttum').filter(
                        forma_pago__cuenta_andinasoft = cuenta,
                        fecha__lte = fecha,
                        fecha__gte = "2024-08-01"
                ).aggregate(total=Sum('valor')).get('total')
                
                if saldo_ant_hoteles != None: valor += saldo_ant_hoteles
                
                saldo_gastos = gastos_caja.objects.filter(
                        forma_pago=cuenta,
                        fecha_gasto__lte = fecha,
                        fecha_gasto__gte = "2024-08-01"
                ).aggregate(total=Sum('valor')).get('total')
                
                if saldo_gastos != None: valor -= saldo_gastos
                
                saldos_cuentas_tesoreria.objects.create(
                    cuenta = obj_cuenta,
                    fecha = fecha,
                    saldo_inicial = saldo_inicial,
                    usuario = request.user,
                    forma = forma,
                    saldo_teorico = valor
                )
                
                data = {
                'title':'Listo!',
                'msj':'Se cargo el saldo inicial para la cuenta y fecha seleccionada',
                'class':'alert-success'
                }
                
            except IntegrityError:
                data = {
                    'title':'Oh Oh!',
                    'msj':'Ya existe saldo inicial cargado para la cuenta y fecha seleccionada',
                    'class':'alert-danger'
                }
            
            return JsonResponse(data)
        
        if request.method == 'GET':
            cuenta = request.GET.get('cuenta')
            obj_saldos = saldos_cuentas_tesoreria.objects.filter(
                cuenta=cuenta
            ).order_by('-fecha')
            
            list_mvtos=[]
            i=0
            for mvto in obj_saldos:
                fecha = f'{mvto.fecha.day:02d}/{mvto.fecha.month:02d}/{mvto.fecha.year}'
                valor = 0
                saldo_transfer = transferencias_companias.objects.filter(
                    cuenta_entra = cuenta,
                    fecha__lte = mvto.fecha,
                    fecha__gte = "2024-08-01"
                ).aggregate(total=Sum('valor')).get('total')
                
                if saldo_transfer != None: valor += saldo_transfer
                
                saldo_transfer_sale = transferencias_companias.objects.filter(
                    cuenta_sale = cuenta,
                    fecha__lte =  mvto.fecha,
                    fecha__gte = "2024-08-01"
                ).aggregate(total=Sum('valor')).get('total')
                
                if saldo_transfer_sale != None: valor -= saldo_transfer_sale
                
                saldo_hoteles = Pagos_facturas.objects.using('Alttum').filter(
                    forma_pago__cuenta_andinasoft = cuenta,
                    fecha__lte = mvto.fecha,
                    fecha__gte = "2024-08-01"
                ).aggregate(total=Sum('valor')).get('total')
                
                if saldo_hoteles != None: valor += saldo_hoteles
                
                saldo_ant_hoteles = anticipos_hotels.objects.using('Alttum').filter(
                        forma_pago__cuenta_andinasoft = cuenta,
                        fecha__lte = mvto.fecha,
                    fecha__gte = "2024-08-01"
                ).aggregate(total=Sum('valor')).get('total')
                
                if saldo_ant_hoteles != None: valor += saldo_ant_hoteles
                
                saldo_gastos = gastos_caja.objects.filter(
                        forma_pago=cuenta,
                        fecha_gasto__lte = mvto.fecha,
                        fecha_gasto__gte = "2024-08-01"
                ).aggregate(total=Sum('valor')).get('total')
                
                if saldo_gastos != None: valor -= saldo_gastos
                saldo_teorico = 0 if mvto.saldo_teorico is None else mvto.saldo_teorico
                list_mvtos.append({
                    'id':i,
                    'pk':mvto.pk,
                    'fecha':fecha,
                    'saldo':f'{mvto.saldo_inicial:,}',
                    'usuario':mvto.usuario.username,
                    'saldoteorico':saldo_teorico,
                    'saldoteorhoy':valor,
                    'diferencia_guardad':mvto.saldo_inicial - saldo_teorico,
                    'diferencia_hoy':mvto.saldo_inicial - valor
                })
                i+=1
            data = {
                'data':list_mvtos
            }
            
            return JsonResponse(data)
        
def ajax_reclasificar_nomina(request):
    if request.method == 'POST':
        tipo = request.POST.get('tipo')
        if tipo == 'reclasificacion_cc':
            nomina = request.FILES.get('archivonomina')
            ultima_fila = request.POST.get('ultimafila')
            consecutivo = request.POST.get('consecutivo_nomina')
            if nomina != None:
                documento=openpyxl.load_workbook(nomina)
                hoja=documento.get_sheet_names()
                sheet_read=documento.get_sheet_by_name(hoja[0])
                
                book=openpyxl.load_workbook("resources/excel_formats/InterfazSIIGO.xlsx")
                sheet_write=book.active
                
                row=6
                
                distribuciones_cc = distribucion_centros_costos.objects.all()
                
                for fila in range(6,int(ultima_fila)+1):
                    cuenta=sheet_read[f'D{fila}'].value
                    tercero = sheet_read[f'P{fila}'].value
                    verif_tercero = distribuciones_cc.filter(cedula=tercero)
                    valor = sheet_read[f'F{fila}'].value.replace('.','').replace(',','.')
                    if verif_tercero.exists() and cuenta[0] == '5':
                        for distrib in verif_tercero:                            
                            valor_distribuido = float(valor) * distrib.porcentaje
                            centro_distribuido = distrib.centro
                            subc_distribuido = distrib.subcentro
                            
                            sheet_write.cell(row,1,sheet_read[f'A{fila}'].value)
                            sheet_write.cell(row,2,sheet_read[f'B{fila}'].value)
                            sheet_write.cell(row,3,consecutivo)
                            sheet_write.cell(row,4,sheet_read[f'D{fila}'].value)
                            sheet_write.cell(row,5,sheet_read[f'E{fila}'].value)
                            sheet_write.cell(row,6,valor_distribuido)
                            sheet_write.cell(row,7,sheet_read[f'G{fila}'].value)
                            sheet_write.cell(row,8,sheet_read[f'H{fila}'].value)
                            sheet_write.cell(row,9,sheet_read[f'I{fila}'].value)
                            sheet_write.cell(row,10,sheet_read[f'J{fila}'].value)
                            sheet_write.cell(row,11,sheet_read[f'K{fila}'].value)
                            sheet_write.cell(row,13,row-5)
                            sheet_write.cell(row,14,centro_distribuido)
                            sheet_write.cell(row,15,subc_distribuido)
                            sheet_write.cell(row,16,sheet_read[f'P{fila}'].value)
                            sheet_write.cell(row,17,sheet_read[f'R{fila}'].value)
                            row += 1
                    else:
                        sheet_write.cell(row,1,sheet_read[f'A{fila}'].value)
                        sheet_write.cell(row,2,sheet_read[f'B{fila}'].value)
                        sheet_write.cell(row,3,sheet_read[f'C{fila}'].value)
                        sheet_write.cell(row,4,sheet_read[f'D{fila}'].value)
                        sheet_write.cell(row,5,sheet_read[f'E{fila}'].value)
                        sheet_write.cell(row,6,valor)
                        sheet_write.cell(row,7,sheet_read[f'G{fila}'].value)
                        sheet_write.cell(row,8,sheet_read[f'H{fila}'].value)
                        sheet_write.cell(row,9,sheet_read[f'I{fila}'].value)
                        sheet_write.cell(row,10,sheet_read[f'J{fila}'].value)
                        sheet_write.cell(row,11,sheet_read[f'K{fila}'].value)
                        sheet_write.cell(row,13,row-5)
                        sheet_write.cell(row,14,sheet_read[f'N{fila}'].value)
                        sheet_write.cell(row,15,sheet_read[f'O{fila}'].value)
                        sheet_write.cell(row,16,sheet_read[f'P{fila}'].value)
                        sheet_write.cell(row,17,sheet_read[f'R{fila}'].value)
                        row += 1
                        
                nombre_doc=f'Nomina_con_CC_reclasificados.xlsx'
                ruta=settings.MEDIA_ROOT+'/tmp/xlsx/'+nombre_doc
                ruta_dw=settings.MEDIA_URL+'tmp/xlsx/'+nombre_doc
                book.save(ruta)
                
                data = {
                    'texto':f'''<ul>
                        <li>
                            Puedes descargar el archivo de egresos <strong><a href="{ruta_dw}" target="_blank">aquí</a></strong>
                        </li>
                    </ul>'''
                }
                
                return JsonResponse(data)


def ajax_nomina_asociada(request):
    if request.method == 'GET':
        year = int(request.GET.get('year'))
        month = int(request.GET.get('month'))
        quincena = int(request.GET.get('quincena'))
        empresa = int(request.GET.get('empresa'))
        
        periodo = period(empresa,year,month,quincena)
        quincena = periodo.get_period_values()
        
        data = {
            'data':quincena
        }
        return JsonResponse(data)


#accounts handler functions
def cargar_gastos_contables(empresa:"Debe ser una instancia del modelo empresa",
                            file:"Archivo en xlsx local",
                            ultima_linea:int):
    """
   Esta funcion carga el movimiento de las cuentas de disponible para su uso en la planilla diaria, conciliacion y arqueo de caja
    """
    file_server=settings.DIR_EXPORT+'movements_siigo_file.xlsx'
    with open(file_server,'wb+') as destination:
        for chunk in file.chunks():
            destination.write(chunk)
    documento=openpyxl.load_workbook(file_server)
    hoja=documento.get_sheet_names()
    sheet=documento.get_sheet_by_name(hoja[0])
    errores=0
    lineas_cargadas=0
    for fila in range(8,ultima_linea+1):
        manage=sheet[f'E{fila}'].value
        if manage=='  000 00000000000 000' or manage=='' or manage==None:
            pass
        else:
            cuenta=sheet[f'B{fila}'].value
            descrip_cta=sheet[f'C{fila}'].value
            comprobante=sheet[f'E{fila}'].value
            fecha=sheet[f'F{fila}'].value
            descripcion=sheet[f'I{fila}'].value
            debito=sheet[f'M{fila}'].value
            credito=sheet[f'N{fila}'].value
            if debito=='                 ' or debito==None:
                debito=0
            else:
                debito=(debito)
            if credito=='                 ' or credito==None:
                credito=0
            else:
                credito=(credito)
            valor=Decimal(debito-credito)
            
            try:
                obj_cuenta = cuentas_pagos.objects.get(nit_empresa=empresa.pk,cuentacontable=descrip_cta)
                egresos_contable.objects.create(
                    empresa = empresa,
                    fecha=fecha,
                    cuenta = obj_cuenta,
                    comprobante=comprobante,
                    descripcion=descripcion,
                    valor=valor,
                )
                lineas_cargadas+=1
            except IntegrityError:
                if errores==0:
                    file = open(settings.DIR_EXPORT+'comprobantes_no_cargados.txt','w')
                    file.write('Los siguientes comprobantes ya estan cargados en la base de datos para esta empresa:'+'\n')
                msj = f'{comprobante}||{fecha}||{valor}\n'
                file.write(msj)
                errores+=1
    if errores>0:
        file.close()
    response = {
        'cargadas':lineas_cargadas,
        'errores':errores,
    }
    return response




def cargar_gastos_banco(empresa:"Debe ser una instancia del modelo empresa",
                        file:"Archivo en csv local",
                        dia_mvto=None,
                        cuenta=None,
                        confirm_duplicated = False):
    """
   Esta funcion carga el movimiento de las cuentas de disponible para su uso en la planilla diaria, conciliacion y arqueo de caja
    """
    
    file_server=settings.DIR_EXPORT+'movements_banco_file.csv'
    with open(file_server,'wb+') as destination:
        for chunk in file.chunks():
            destination.write(chunk)
    file.close()
    i=1
    errores=0
    lineas_cargadas=0
    
    with open(file_server,errors='ignore',newline="") as csv_file:
        reader = csv.reader(csv_file)
        total_rows = 0
        movement_exists_prevent = 0
        lines_objects = []
        for row in reader:
            try:
                total_rows += 1
                obj_cuenta = cuentas_pagos.objects.filter(nit_empresa=empresa.pk,cuentabanco=row[0])
                fecha = f'{row[1][:4]}-{row[1][4:6]}-{row[1][6:]}'
                descripcion = row[2]
                valor = row[4]
                
                
                
                if obj_cuenta.exists():
                    obj_cuenta = obj_cuenta[0]
                    check_line = egresos_banco.objects.filter(
                        empresa = empresa,
                        fecha = fecha,
                        cuenta = obj_cuenta,
                        descripcion = descripcion,
                        valor = valor
                    )
                    if check_line.exists():
                        movement_exists_prevent += 1
                    
                    lines_objects.append(egresos_banco(
                        empresa = empresa,
                        fecha=fecha,
                        cuenta = obj_cuenta,
                        referencia=row[3],
                        descripcion=row[2],
                        valor=row[4],
                    ))
                    lineas_cargadas += 1
                else:
                    file_errors = open(settings.DIR_EXPORT+'comprobantes_no_cargados_banco.txt','w')
                    msj = f'Linea:{i}||La fecha y/o la cuenta seleccionada no coinciden con la linea a cargar\n'
            except:
                if errores==0:
                    file_errors = open(settings.DIR_EXPORT+'comprobantes_no_cargados_banco.txt','w')
                    file_errors.write('Los siguientes filas no pudieron ser cargadas'+'\n')
                msj = f'Linea:{i}||{traceback.format_exc()}\n'
                file_errors.write(msj)
                errores+=1
            i+=1
            
            
    file.close()
    
    duplicated = False
    if movement_exists_prevent > 2 and not confirm_duplicated:
        duplicated = True
    else:
        egresos_banco.objects.bulk_create(
                lines_objects
            )
    if errores>0:
        file_errors.close()
    response = {
        'cargadas':lineas_cargadas,
        'errores':errores,
        'duplicados':duplicated,
    }
    return response

#Anticipos

@login_required
def solicitud_de_anticipos(request):
    context = {
        'form_solicitar_ant':forms.form_solicitar_anticipos,
        'form_anticipo':forms.form_anticipos_tesor,
        'form_leg_anticipo':forms.form_legalizar_anticipo,
        'form_partners':forms.form_partners,
        'form_taxes':forms.form_taxes,
    }
    
    is_tesorero = check_perms(request,('accounting.add_anticipos',), raise_exception =  False)
    is_contabilidad = check_perms(request,('accounting.change_solicitud_anticipos',), raise_exception =  False)
    
    if request.is_ajax():
        if request.method == 'GET':
            if request.GET.get('datatable'):
                obj_anticipos_solic = solicitud_anticipos.objects.all().order_by('-fecha')
                if is_tesorero == False and is_contabilidad == False: 
                    obj_anticipos_solic = obj_anticipos_solic.filter(
                        Q(usuario_solicita = request.user.pk)|
                        Q(quien_aprueba = request.user.pk)
                    )
                
                data = {
                    'data':JSONRender(obj_anticipos_solic,
                                      query_functions=('has_reembolsos','data_user_solicita','recibo_reembolso')
                                      ).render()
                }
                
                return JsonResponse(data)
            if request.GET.get('history'):
                anticipo = request.GET.get('anticipo')
                
                
                solicitud = solicitud_anticipos.objects.get(pk=anticipo)
                avatar_solicita = solicitud.usuario_solicita.profiles.avatar.image.url if solicitud.usuario_solicita.profiles.avatar.image else ''
                body = f'''
                    <div class="media border-bottom border-gray pt-2">
                        <img src="{avatar_solicita}" class="mr-3 rounded-circle" width="60" height="60" alt="...">
                        <div class="media-body">
                        <h6 class="mt-0">{solicitud.usuario_solicita.get_full_name()} solicitó anticipo el {solicitud.fecha}</h6>
                        <p class="small"><b>Descripción:</b> {solicitud.descripcion}<br>
                        <b>Oficina:</b> {solicitud.oficina}</br>
                        <b>Empresa:</b> {solicitud.empresa.nombre}</br>
                        <b>Valor:</b>$ {solicitud.valor:,}</p>
                        </div>
                    </div>
                '''
                if solicitud.usuario_aprueba:
                    avatar_aprueba = solicitud.usuario_aprueba.profiles.avatar.image.url if solicitud.usuario_aprueba.profiles.avatar.image else ''
                    body += f'''
                    <div class="media border-bottom border-gray pt-2">
                        <img src="{avatar_aprueba}" class="mr-3 rounded-circle" width="60" height="60" alt="...">
                        <div class="media-body">
                        <h6 class="mt-0">{solicitud.usuario_aprueba.get_full_name()} aprobó el anticipo</h6>
                        </div>
                    </div>
                '''
                if solicitud.pago_anticipo:
                    obj_pago = solicitud.pago_anticipo
                    avatar_pago = solicitud.pago_anticipo.usuario.profiles.avatar.image.url if solicitud.pago_anticipo.usuario.profiles.avatar.image else ''
                    body += f'''
                    <div class="media border-bottom border-gray pt-2">
                        <img src="{avatar_pago}" class="mr-3 rounded-circle" width="60" height="60" alt="...">
                        <div class="media-body">
                        <h6 class="mt-0">{solicitud.pago_anticipo.usuario.get_full_name()} pagó el anticipo el {solicitud.pago_anticipo.fecha_pago}</h6>
                        </div>
                    </div>
                '''
                if solicitud.has_legalizacion():
                    obj_leg = legalizacion_anticipos.objects.filter(pk=solicitud.pk)
                    avatar_carga = obj_leg[0].usuario_carga.profiles.avatar.image.url if obj_leg[0].usuario_carga.profiles.avatar.image else ''
                    body += f'''
                    <div class="media border-bottom border-gray pt-2">
                        <img src="{avatar_carga}" class="mr-3 rounded-circle" width="60" height="60" alt="...">
                        <div class="media-body">
                        <h6 class="mt-0">{obj_leg[0].usuario_carga.get_full_name()} legalizó anticipo el {obj_leg[0].fecha_legalizacion}</h6>
                        </div>
                    </div>
                '''
                    if obj_leg[0].usuario_aprueba:
                        avatar_leg_aprueba = obj_leg[0].usuario_aprueba.profiles.avatar.image.url if obj_leg[0].usuario_aprueba.profiles.avatar.image else ''
                        body += f'''
                    <div class="media border-bottom border-gray pt-2">
                        <img src="{avatar_leg_aprueba}" class="mr-3 rounded-circle" width="60" height="60" alt="...">
                        <div class="media-body">
                        <h6 class="mt-0">{obj_leg[0].usuario_aprueba.get_full_name()} aprobó legalización de anticipo el {obj_leg[0].fecha_aprobacion}</h6>
                        </div>
                    </div>
                '''
                    
                body += '</div>'
                data = {
                    'body':body
                }
                
                return JsonResponse(data)
            
        if request.method == 'POST':
            to_do = request.POST.get('to_do')

            if to_do == 'approve':
                id_solicitud = request.POST.get('id_solicitud')
                
                obj_solicitud = solicitud_anticipos.objects.get(pk=id_solicitud)
                
                if obj_solicitud.quien_aprueba == request.user or request.user.is_superuser:
                    obj_solicitud.estado = 'aprobado'
                    obj_solicitud.usuario_aprueba = request.user
                    obj_solicitud.save()
                    
                    data = {
                        'class': 'alert-success',
                        'msj': 'La solicitud de anticipo fue aprobada con exito'
                    }
                    
                    msj = f'Hola <b>:)</b>, <b>{request.user.get_full_name()}</b> aprobó la solicitud de anticipo de <b>{obj_solicitud.usuario_solicita.get_full_name()}</b> por <b>${obj_solicitud.valor:,}</b> '
            
                    envio_notificacion(msj,'Anticipo aprobado',['operaciones2@somosandina.co','jorgeavila@somosandina.co'])
                else:
                    data = {
                        'class':'alert-danger',
                        'msj': 'No tienes permisos para aprobar un anticipo'
                    }
                    
                return JsonResponse(data)
            
            elif to_do == 'payadvance':
                if not check_perms(request,('accounting.change_anticipos',), raise_exception=False):
                    msj = 'No tienes permisos para registrar el pago de un anticipo'
                    clss = 'alert-danger'
                else:
                
                    empresa = request.POST.get('empresa_ant')
                    cuenta = request.POST.get('cuenta_ant')
                    fecha = request.POST.get('fecha_anticip')
                    id_tercero = request.POST.get('id_tercero')
                    nombre_tercero = request.POST.get('nombre_tercero')
                    valor_anticipo = request.POST.get('valor_anticip').replace(',','')
                    tipo = request.POST.get('tipo_anticipo')
                    descripcion = request.POST.get('descripcion_ant')
                    oficina = request.POST.get('oficina')
                    soporte = request.FILES.get('soporte_pago')
                    
                    
                    obj_pago = Anticipos.objects.create(
                        empresa=empresas.objects.get(pk=empresa),
                        descripcion = descripcion,valor=valor_anticipo,usuario=request.user,
                        cuenta=cuentas_pagos.objects.get(pk=cuenta),fecha_pago=fecha,
                        id_tercero=id_tercero,nombre_tercero=nombre_tercero,
                        tipo_anticipo = info_interfaces.objects.get(pk=tipo),
                        soporte_pago=soporte,oficina=oficina
                    )
                    
                    id_solicitud = request.POST.get('id_solicitud')
                    
                    obj_solicitud = solicitud_anticipos.objects.get(pk=id_solicitud)
                    
                    
                    obj_solicitud.estado = 'pagado'
                    obj_solicitud.pago_anticipo = obj_pago
                    obj_solicitud.save()

                    msj = 'Pago de anticipo registrado sin problemas'
                    clss = 'alert-success'
                
                data = {
                    'msj': msj,
                    'class': clss
                }
                    
                return JsonResponse(data)
            
            
    else:
        if request.method == 'POST':
            empresa_solicita = request.POST.get('empresa_solicita')
            descripcion = request.POST.get('descripcion')
            valor = request.POST.get('valor')
            oficina = request.POST.get('oficina')
            quien_aprueba = request.POST.get('quien_aprueba')
            
            solicitud_anticipos.objects.create(
                empresa = empresas.objects.get(pk=empresa_solicita),
                usuario_solicita = request.user,
                descripcion = descripcion, 
                valor = valor.replace(',',''),
                oficina = oficina,
                quien_aprueba = User.objects.get(pk=quien_aprueba)
            )
            
            email_user_aprueba = User.objects.get(pk=quien_aprueba).email
            
            msj = f'Hola :), <b>{request.user.get_full_name()}</b> ha solicitado un nuevo anticipo por <b>${valor}</b>.'
            
            envio_notificacion(msj,'Nuevo anticipo solicitado',['operaciones2@somosandina.co','jorgeavila@somosandina.co',email_user_aprueba])
        
            
    return render(request,'solicitar_anticipo.html', context)

@login_required
def legalizaciones(request):
    if request.method == 'GET':
        if request.is_ajax():
            todo = request.GET.get('todo')
            
            
            if todo == 'datatable':
                anticipo = request.GET.get('id_anticipo')
                obj_legalizaciones = legalizacion_anticipos.objects.filter(
                    anticipo = anticipo,
                )
                
                data = {
                    'data':JSONRender(obj_legalizaciones, query_functions=['nombre_tercero']).render()
                }
                
                return JsonResponse(data)
            
            elif todo == 'print_leg':
                anticipo = request.GET.get('anticipo')
                obj_anticipo = solicitud_anticipos.objects.get(pk=anticipo)
                obj_legalizaciones = legalizacion_anticipos.objects.filter(anticipo=anticipo
                                                                           ).order_by('fecha_gasto')
                
                context = {
                    'anticipo':obj_anticipo,
                    'legalizaciones':obj_legalizaciones,
                    'user':request.user,
                    'now':datetime.datetime.now(),
                }
                filename = f'Legalizacion_anticipo_id_{anticipo}_{obj_anticipo.usuario_solicita.username}.pdf'.replace(' ','_')
                
                pdf = pdf_gen(f'pdf/legalizacion_anticipo.html',context,filename)
                
                file = pdf.get('url')
                
                msj = f'Puedes descargar el archivo <a href="{file}" target="_blank"><b>Aquí</b></a>'
                bg = 'alert-success'
                
                data = {
                    'msj': msj,
                    'bg': bg
                }
                
                return JsonResponse(data)
        
            
    elif request.method == 'POST':
        if request.is_ajax():
            
            todo = request.POST.get('to_do')
            
            if todo == 'legalice':
                            
                fecha = request.POST.get('fecha')
                descripcion = request.POST.get('descripcion')
                nit_tercero = request.POST.get('nit_tercero')
                valor = request.POST.get('valor').replace(',','')
                val_fmt = request.POST.get('valor')
                soporte = request.FILES.get('soporte')
                concepto = request.POST.get('concepto')    
                anticipo = request.POST.get('id_solicitud')
                
                obj_anticipo = solicitud_anticipos.objects.get(pk = anticipo)
                obj_tercero = Partners.objects.get(pk=nit_tercero)
                
                
                obj_legal = legalizacion_anticipos.objects.create(
                    anticipo = obj_anticipo,
                    concepto = conceptos_legalizacion.objects.get(pk=concepto),
                    fecha_gasto = fecha, descripcion = descripcion,
                    tercero = obj_tercero,
                    valor = valor, soporte = soporte, usuario_carga = request.user
                )    
                
                msj = f'Hola <b>:)</b>, <b>{request.user.get_full_name()}</b> registró un gasto por ${val_fmt} a {obj_tercero.nombre_completo()} sobre un anticipo.'
            
                envio_notificacion(msj,'Nuevo gasto registrado',['auxiliarcontable2@somosandina.co',])
                
                return JsonResponse({'status':'ok'})
            
            elif todo == 'delete':
                legalizacion = request.POST.get('leg')
                
                obj_leg = legalizacion_anticipos.objects.get(pk=legalizacion)
                
                if obj_leg.anticipo.usuario_solicita == request.user or \
                    request.user.is_superuser:
                    obj_leg.delete()

                    msj = 'Se eliminó el gasto seleccionado'
                    bg = 'alert-success'
                else:
                    msj = 'Solo puedes eliminar legalizaciones que tu hayas realizado'
                    bg = 'alert-danger'
                
                data = {
                    'msj': msj,
                    'bg': bg
                }
                
                return JsonResponse(data)
            elif todo == 'approve_tesoreria':
                
                anticipo = request.POST.get('anticipo')
                cuenta = request.POST.get('cuenta')
                valor = request.POST.get('valor')
                
                if not check_perms(request,('accounting.add_otros_ingresos',),raise_exception=False):
                    data = {
                        'msj':'No tienes permisos para verificar reintegros a empresa',
                        'class':'alert-danger'
                    }
                    
                    return JsonResponse(data)    

                    
                obj_anticipo = solicitud_anticipos.objects.get(pk=anticipo)
                obj_cuenta = cuentas_pagos.objects.get(pk=cuenta)
                descrip = f'REINTEGRO DE ANTICIPO ({obj_anticipo.pk}): {obj_anticipo.descripcion}'
                
                ingreso = otros_ingresos.objects.create(
                    fecha_ing = datetime.date.today(),
                    id_tercero = obj_anticipo.usuario_solicita.profiles.identificacion,
                    nombre_tercero = obj_anticipo.usuario_solicita.get_full_name(),
                    descripcion = descrip, valor = valor, usuario = request.user,
                    empresa = obj_cuenta.nit_empresa, cuenta = obj_cuenta, oficina = obj_anticipo.oficina,
                )
                
                obj_reint = reintegros_anticipos.objects.create(
                    id_anticipo = solicitud_anticipos.objects.get(pk=anticipo),
                    id_otros_ingresos = ingreso,
                    tipo = 'D',
                    valor = valor                    
                )
                
                consecutivo_ofic = otros_ingresos.objects.filter(pk__lte=ingreso.pk,
                        oficina=ingreso.oficina).count()
                filename = f'Recibo_ingreso_{ingreso.oficina}_{consecutivo_ofic}.pdf'
                ruta = settings.MEDIA_ROOT+'/tmp/pdf/'+filename
                GenerarPDF().reciboIngreso(ingreso,consecutivo_ofic,ruta)
                ruta_download = settings.MEDIA_URL + 'tmp/pdf/' + filename
                
                data = {
                    'ruta':ruta_download,
                    'msj':f'Se genero el recibo de otros ingresos, puedes descargarlo <a target="_blank" href="{ruta_download}"><b>aquí</b></a>',
                    'class':'alert-success'
                }
                
                return JsonResponse(data) 
            
            elif todo == 'finish_leg':
                anticipo = request.POST.get('anticipo')
                obj_anticipo = solicitud_anticipos.objects.get(pk=anticipo)
                
                obj_anticipo.estado = 'prelegalizado'
                obj_anticipo.save()
                
                data = {
                    'msj':'El anticipo seleccionado fué cambiado a estado prelegalizado.',
                    'class':'alert-success'
                }
                
                msj = f'Hola <b>:)</b>, <b>{request.user.get_full_name()}</b> prelegalizó un anticipo de fecha {obj_anticipo.fecha}.'
            
                envio_notificacion(msj,'Nuevo gasto registrado',['auxiliarcontable2@somosandina.co',])
                
                return JsonResponse(data)
                
            elif todo == 'change_support_leg':
                legalizacion =  request.POST.get('legalizacion')
                file = request.FILES.get('nuevo_soporte')
                obj_legalizacion = legalizacion_anticipos.objects.get(pk = legalizacion)
                
                obj_legalizacion.soporte = file
                obj_legalizacion.save()
                
                data = {
                    'msj':'Se cambió el soporte de la legalización seleccionada.',
                    'class':'alert-success'
                }
                
                return JsonResponse(data)
            
            elif todo == 'change_concept':
                legalizacion =  request.POST.get('id_leg')
                concepto = request.POST.get('new_concept')
                obj_legalizacion = legalizacion_anticipos.objects.get(pk = legalizacion)
                obj_concept = conceptos_legalizacion.objects.get(pk=concepto)
                
                
                obj_legalizacion.concepto = obj_concept
                obj_legalizacion.save()
                
                data = {
                    'msj':'Se cambió el concepto de la legalización seleccionada.',
                    'class':'alert-success'
                }
                
                return JsonResponse(data)
            
            elif todo == 'add_taxes':
                legalizacion =  request.POST.get('id_leg')
                obj_legalizacion = legalizacion_anticipos.objects.get(pk = legalizacion)
                
                tipo_iva = request.POST.get('tipo_iva')
                valor_iva = request.POST.get('valor_iva').replace(',','')
                tipo_rte = request.POST.get('tipo_rte')
                valor_rte =  request.POST.get('valor_rte').replace(',','')
                rte_asumida = request.POST.get('rte_asumida')
                
                if int(valor_iva) > 0:
                    obj_legalizacion.cuenta_iva = impuestos_legalizacion.objects.get(pk=tipo_iva)
                    obj_legalizacion.valor_iva = valor_iva
                
                if int(valor_rte) > 0:
                    obj_legalizacion.cuenta_rte = impuestos_legalizacion.objects.get(pk=tipo_rte)
                    obj_legalizacion.valor_rte = valor_rte
                    
                    rte_asumida = True if rte_asumida == 'on' else False
                
                    obj_legalizacion.rte_asumida = rte_asumida
                
                obj_legalizacion.save()
                
                data = {
                    
                }
                
                return JsonResponse(data)
            
            elif todo == 'approve_contabilidad':
                if not check_perms(request,('accounting.change_solicitud_anticipos',),raise_exception=False):
                    data = {
                        'msj':'No tienes permisos para verificar legalizaciones',
                        'class':'alert-danger'
                    }
                    
                    return JsonResponse(data)   
                
                anticipo = request.POST.get('anticipo')
                obj_anticipo = solicitud_anticipos.objects.get(pk=anticipo)
                obj_legalizaciones = legalizacion_anticipos.objects.filter(anticipo = anticipo)
                
                for i in obj_legalizaciones:
                    i.usuario_aprueba = request.user
                    i.fecha_aprobacion = datetime.date.today()
                    i.save()
                
                obj_anticipo.estado = 'legalizado'
                obj_anticipo.save()
                
                
                data = {
                    'msj':'Se aprobó la legalización sin problemas.',
                    'class':'alert-success'
                }
                
                return JsonResponse(data)

            elif todo == 'delete':
                gasto = request.POST.get('gasto')
                
                obj_gasto = gastos_caja.objects.get(pk=gasto)
                
                if obj_gasto.usuario_carga == request.user or \
                    request.user.is_superuser:
                    obj_gasto.delete()

                    msj = 'Se eliminó el gasto seleccionado'
                    bg = 'alert-success'
                else:
                    msj = 'Solo puedes eliminar legalizaciones que tu hayas realizado'
                    bg = 'alert-danger'
                
                data = {
                    'msj': msj,
                    'bg': bg
                }
                
                return JsonResponse(data)
            
        
@login_required
def partners(request):
    if request.is_ajax():
        if request.method == 'POST':
            todo = request.POST.get('todo')
            if todo == 'create':
                form = request.POST
                
                tipo_id = form.get('document_type')
                id_numero = form.get('idTercero')
                nombres = form.get('nombres').upper()
                apellidos = form.get('apellidos').upper()
                telefono = form.get('telefono')
                pais = Countries.objects.get(pk=form.get('pais'))
                estado = States.objects.get(pk=form.get('estado'))
                ciudad = Cities.objects.get(pk=form.get('ciudad'))
                direccion = form.get('direccion')
                email = form.get('email')
                responsabilidad_fiscal = form.get('responsabilidad_fiscal')
                soporte = request.FILES.get('soporte_id')
                a,b = 'áéíóúüñ','aeiouun'
                trans = str.maketrans(a,b)
                soporte.name = soporte.name.translate(trans)
                
                if not soporte.name.endswith('.pdf'):
                    
                    data = {
                        'msj':'El soporte debe ser un archivo en formato PDF.',
                        'class':'alert-danger',
                        'status_response':'error'
                    }
                    
                    return JsonResponse(data)
                
                try:
                    partner = Partners.objects.create(
                        document_type = tipo_id, idTercero = id_numero, nombres = nombres, apellidos = apellidos,
                        telefono = telefono, pais = pais, estado = estado, ciudad = ciudad, direccion = direccion,
                        email = email, responsabilidad_fiscal = responsabilidad_fiscal, soporte_identificacion = soporte
                    )
                    
                    data = {
                        'msj':'El tercero fue creado',
                        'class':'alert-success',
                        'partner':{
                            'idTercero':id_numero,
                            'full_name': f'{nombres} {apellidos}'
                        },
                        'status_response':'success'
                    }
                except IntegrityError:
                    data = {
                        'msj':'El tercero ya existe',
                        'class':'alert-danger',
                        'status_response':'error'
                    }
                    
                return JsonResponse(data)                

@login_required
def conceptos(request):
    if request.is_ajax():
        if request.method == 'GET':
            empresa = request.GET.get('empresa')
            obj_conceptos = conceptos_legalizacion.objects.filter(activo=True
                                        ).values('pk','descripcion')
            data = {
                'data':list(obj_conceptos)
            }
            
            return JsonResponse(data)
                       

@login_required
def historical_data(request):
    
    context = {
        'form': forms.form_historic_accounting_data(),
    }
    
    if request.method == 'GET':
        if request.is_ajax():
            draw = request.GET.get('draw')
            start = request.GET.get('start')
            length = request.GET.get('length')
            search_val = request.GET.get('search[value]')
            
            empresa = request.GET.get('empresa')
            documento = request.GET.get('tipo_doc')
            consecutivo = request.GET.get('consecutivo')
            
            obj_archivos = archivo_contable.objects.all().order_by('-fecha_doc')
            total_records =  obj_archivos.count()
            
            
            
            if empresa!="" and empresa!=None:
                obj_archivos = obj_archivos.filter(
                    empresa = empresa
                )
            if documento!="" and documento!=None:
                obj_archivos = obj_archivos.filter(
                    tipo_doc = documento,
                )
            if consecutivo!="" and consecutivo!=None:
                obj_archivos = obj_archivos.filter(
                    consecutivo = consecutivo
                )            
                
            if search_val:
                obj_archivos = obj_archivos.filter(ocr_text__icontains = search_val)
            
            list_archivos= JSONRender(obj_archivos).render()
            defered_list = list_archivos[int(start):int(start)+int(length)]
            
            data = {
                "draw": int(draw),
                "recordsTotal": total_records,
                "recordsFiltered": obj_archivos.count(),
                'data':defered_list
            }
            return JsonResponse(data)
    
    if request.method == 'POST':
        
        form = forms.form_historic_accounting_data(request.POST, request.FILES)
        
        if form.is_valid():
            empresa = form.cleaned_data.get('empresa')
            tipo_doc = form.cleaned_data.get('tipo_doc')
            consecutivo_doc = form.cleaned_data.get('consecutivo_doc')
            fecha_doc = form.cleaned_data.get('fecha_doc')
            doc_file = form.cleaned_data.get('doc_file')
            
            
            instance = archivo_contable.objects.create(
                empresa = empresa, tipo_doc = tipo_doc.upper(),
                consecutivo = consecutivo_doc, fecha_doc = fecha_doc,
                document = doc_file, user_carga = request.user
            )
            
            temp_file_path = None
            try:
                try:
                    file_path = instance.document.path
                except Exception:
                    _, suffix = os.path.splitext(str(instance.document.name))
                    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix or '.tmp') as temp_file:
                        with instance.document.open('rb') as stored_file:
                            for chunk in iter(lambda: stored_file.read(1024 * 1024), b''):
                                if not chunk:
                                    break
                                temp_file.write(chunk)
                        temp_file_path = temp_file.name
                    file_path = temp_file_path

                ocr_text = get_text_from_file(file_path)
            finally:
                if temp_file_path and os.path.exists(temp_file_path):
                    os.remove(temp_file_path)
            
            instance.ocr_text = ocr_text
            instance.save()
    
    return render(request, 'historical_data.html',context)

#Cajas efectivo
@login_required
@group_perm_required(('accounting.view_gastos_caja',),raise_exception=True)
def cajas_efectivo(request):
    a,b = 'áéíóúüñÑ','aeiouunn'
    trans = str.maketrans(a,b)
    
    if request.is_ajax():
        if request.method == 'GET':
            todo = request.GET.get('todo')
            
            if todo == 'datatable':
                caja = request.GET.get('caja')
                fecha_desde = request.GET.get('fecha_desde')
                fecha_hasta = request.GET.get('fecha_hasta')
                solo_pendientes = request.GET.get('sp')
                legalizar = request.GET.get('legalizar')
                id_reemb = request.GET.get('id_reemb')
                include_incomes = request.GET.get('swinc')
                
                movements = []
                
                if caja:
                    #outcomes
                    if legalizar == 'true':
                        obj_gastos = gastos_caja.objects.filter(estado='Aprobado',reembolso__isnull=True,
                                                                forma_pago=caja)
                    elif id_reemb is not None:
                         obj_gastos = gastos_caja.objects.filter(reembolso=id_reemb)
                    else:
                        obj_gastos = gastos_caja.objects.filter(forma_pago=caja,
                                                fecha_gasto__range = (fecha_desde,fecha_hasta))
                    
                    if solo_pendientes == 'true':
                        obj_gastos = obj_gastos.filter(estado='Pendiente')
                        
                    for i in obj_gastos:
                        id_iva = i.cuenta_iva.id if i.cuenta_iva is not None else None
                        vr_iva = 0 if i.valor_iva is None else i.valor_iva
                        id_rte = i.cuenta_rte.id if i.cuenta_rte is not None else None
                        vr_rte = 0 if i.valor_rte is None else i.valor_rte
                        subtotal = i.valor - vr_iva + vr_rte
                        if i.rte_asumida is True: subtotal -= vr_rte
                        aprob = "" if i.usuario_aprueba is None else i.usuario_aprueba.username
                        reemb = 'Si' if i.has_reemb() else 'No'
                        movements.append({
                            'pk':i.pk,
                            'fecha': datetime.datetime.strftime(i.fecha_gasto,"%d/%m/%Y"),
                            'tercero': i.tercero.nombre_completo(),
                            'nit_tercero':i.tercero.idTercero,
                            'descripcion': i.descripcion.capitalize(),
                            'valor': i.valor * -1,
                            'legalizado':'',
                            'soporte': _media_url(i.soporte),
                            'tipo':'gasto',
                            'estado':i.estado,
                            'cuenta_iva':{
                                'id':id_iva,
                            },
                            'valor_iva':i.valor_iva,
                            'cuenta_rte':{
                                'id':id_rte,
                            },
                            'valor_rte':i.valor_rte,
                            'rte_asumida':i.rte_asumida,
                            'usuario_aprueba':aprob,
                            'legalizado':reemb,
                            'iva':vr_iva,
                            'rte': vr_rte,
                            'subtotal':subtotal,
                            'concepto':i.concepto.descripcion,
                            'soporte_doc_prov': _media_url(i.tercero.soporte_identificacion),
                        })
                        
                    if include_incomes == 'true':
                        obj_gastos = obj_gastos.filter(estado='Pendiente')
                    
                        #Calculamos el saldo que viene
                        valor = 0
                        saldo_transfer = transferencias_companias.objects.filter(
                            cuenta_entra = caja,
                            fecha__lt = fecha_desde,
                            fecha__gte = "2024-08-01"
                        ).aggregate(total=Sum('valor')).get('total')
                        
                        if saldo_transfer != None: valor += saldo_transfer
                        
                        saldo_transfer_sale = transferencias_companias.objects.filter(
                            cuenta_sale = caja,
                            fecha__lt = fecha_desde,
                            fecha__gte = "2024-08-01"
                        ).aggregate(total=Sum('valor')).get('total')
                        
                        if saldo_transfer_sale != None: valor -= saldo_transfer_sale
                        
                        saldo_hoteles = Pagos_facturas.objects.using('Alttum').filter(
                            forma_pago__cuenta_andinasoft = caja,
                            fecha__lt = fecha_desde,
                            fecha__gte = "2024-08-01"
                        ).aggregate(total=Sum('valor')).get('total')
                        
                        if saldo_hoteles != None: valor += saldo_hoteles
                        
                        saldo_ant_hoteles = anticipos_hotels.objects.using('Alttum').filter(
                                forma_pago__cuenta_andinasoft = caja,
                                fecha__lt = fecha_desde,
                                fecha__gte = "2024-08-01"
                        ).aggregate(total=Sum('valor')).get('total')
                        
                        if saldo_ant_hoteles != None: valor += saldo_ant_hoteles
                        
                        saldo_gastos = gastos_caja.objects.filter(
                                forma_pago=caja,
                                fecha_gasto__lt = fecha_desde,
                                fecha_gasto__gte = "2024-08-01"
                        ).aggregate(total=Sum('valor')).get('total')
                        
                        if saldo_gastos != None: valor -= saldo_gastos
                        
                        dia_saldo = datetime.datetime.strptime(fecha_desde,"%Y-%m-%d") - relativedelta.relativedelta(days=1)
                        movements.append({
                                'pk':0,
                                'fecha': datetime.datetime.strftime(dia_saldo,"%d/%m/%Y"),
                                'tercero': 'SALDO INICIAL',
                                'nit_tercero':'INICIO DE DIA',
                                'descripcion': 'SALDO A INICIO DE DIA',
                                'valor':valor,
                                'legalizado':'',
                                'soporte': '',
                                'tipo':'saldo',
                                'estado':'',
                                'iva':None,
                                'rte': None,
                                'subtotal':'',
                                'concepto':'SALDOS'
                            })
                        
                        obj_transferencias = transferencias_companias.objects.filter(
                            cuenta_entra = caja,
                            fecha__range = (fecha_desde,fecha_hasta)
                        )
                        
                        for i in obj_transferencias:
                            movements.append({
                                'pk':i.pk,
                                'fecha': datetime.datetime.strftime(i.fecha,"%d/%m/%Y"),
                                'tercero': i.empresa_sale.nombre,
                                'nit_tercero':i.empresa_sale.pk,
                                'descripcion': f'Transferencia desde cuenta {i.cuenta_sale.cuentabanco}',
                                'valor': i.valor,
                                'legalizado':'',
                                'soporte': '',
                                'estado':'',
                                'tipo':'ingreso',
                                'iva':None,
                                'rte': None,
                                'subtotal':'',
                                'concepto':'TRANSFERENCIAS'
                            })
                        
                        obj_transferencias_sale = transferencias_companias.objects.filter(
                            cuenta_sale = caja,
                            fecha__range = (fecha_desde,fecha_hasta)
                        )
                        
                        for i in obj_transferencias_sale:
                            movements.append({
                                'pk':i.pk,
                                'fecha': datetime.datetime.strftime(i.fecha,"%d/%m/%Y"),
                                'tercero': i.empresa_sale.nombre,
                                'nit_tercero':i.empresa_sale.pk,
                                'descripcion': f'Transferencia hacia cuenta {i.cuenta_entra.cuentabanco}',
                                'valor': i.valor * -1,
                                'legalizado':'',
                                'soporte': '',
                                'estado':'',
                                'tipo':'ingreso',
                                'iva':None,
                                'rte': None,
                                'subtotal':'',
                                'concepto':'TRANSFERENCIAS'
                            })
                        
                        
                        obj_incomes_hotel = Pagos_facturas.objects.using('Alttum').filter(
                            forma_pago__cuenta_andinasoft = caja,
                            fecha__range = (fecha_desde,fecha_hasta)
                        )
                        
                        for i in obj_incomes_hotel:
                            movements.append({
                                'pk':i.pk,
                                'fecha': datetime.datetime.strftime(i.fecha,"%d/%m/%Y"),
                                'tercero': i.factura.cliente(),
                                'nit_tercero':i.factura.nit_cliente(),
                                'descripcion': f'Pago factura N° {i.factura.pk}',
                                'valor': i.valor,
                                'legalizado':'',
                                'soporte': '',
                                'tipo':'ingreso',
                                'estado':'',
                                'iva':None,
                                'rte': None,
                                'subtotal':'',
                                'concepto':'INGRESOS'
                                
                            })
                        
                        obj_anticipos_hotel = anticipos_hotels.objects.using('Alttum').filter(
                                forma_pago__cuenta_andinasoft = caja,
                                fecha__range = (fecha_desde,fecha_hasta)
                        )
                        
                        for i in obj_anticipos_hotel:
                            movements.append({
                                'pk':i.pk,
                                'fecha': datetime.datetime.strftime(i.fecha,"%d/%m/%Y"),
                                'tercero': i.cliente(),
                                'nit_tercero':i.nit_cliente(),
                                'descripcion': f'Anticipo {i.tipo()[0]} {i.pk}',
                                'valor': i.valor,
                                'legalizado':'',
                                'soporte': '',
                                'tipo':'ingreso',
                                'estado':'',
                                'iva':None,
                                'rte': None,
                                'subtotal':'',
                                'concepto':'INGRESOS'
                            })
                        
                        
                    
                data = {
                    'data':movements
                }
                
                
                return JsonResponse(data)
            
            elif todo == 'datatable_reembolsos':
                
                caja = request.GET.get('caja')
                fecha_desde = request.GET.get('fecha_desde')
                fecha_hasta = request.GET.get('fecha_hasta')
                reembolsos = reembolsos_caja.objects.filter(
                    fecha_solicita__range = (fecha_desde,fecha_hasta),
                    caja = caja
                )
                
                data = {
                    'data': JSONRender(reembolsos).render()
                }
                
                return JsonResponse(data)
             
            elif todo == 'interf_reemb':
                id_reemb = request.GET.get('id_reemb')
                
                reembolso = reembolsos_caja.objects.get(pk=id_reemb)
                gastos = gastos_caja.objects.filter(reembolso=id_reemb)
                
                rte_asumida = impuestos_legalizacion.objects.get(descripcion__icontains ="retefte asumida")
                ajustes = impuestos_legalizacion.objects.get(descripcion__icontains ="ajuste al peso")
                
                
                book = openpyxl.load_workbook("resources/excel_formats/InterfazSIIGO.xlsx")
                sheet=book.active
                row=6
                for i in gastos:
                    vr_iva = 0 if i.valor_iva is None else i.valor_iva
                    vr_rte = 0 if i.valor_rte is None else i.valor_rte
                    
                    if i.forma_pago.empresa == "Promotora Sandville":
                        cuenta = i.concepto.cuenta_andina
                        cta_iva = i.cuenta_iva.cuenta_andina if vr_iva != 0 else ""
                        cta_rte = i.cuenta_rte.cuenta_andina if vr_rte != 0 else ""
                        cta_rte_asum = rte_asumida.cuenta_andina
                        cta_aprox = ajustes.cuenta_andina
                        
                    elif i.forma_pago.empresa == "Status Comercializadora":
                        cuenta = i.concepto.cuenta_status
                        cta_iva = i.cuenta_iva.cuenta_status if vr_iva != 0 else ""
                        cta_rte = i.cuenta_rte.cuenta_status if vr_rte != 0 else ""
                        cta_rte_asum = rte_asumida.cuenta_status
                        cta_aprox = ajustes.cuenta_status
                        
                    elif i.forma_pago.empresa == "Quadrata Constructores":
                        cuenta = i.concepto.cuenta_quadrata
                        cta_iva = i.cuenta_iva.cuenta_quadrata if vr_iva != 0 else ""
                        cta_rte = i.cuenta_rte.cuenta_quadrata if vr_rte != 0 else ""
                        cta_rte_asum = rte_asumida.cuenta_quadrata
                        cta_aprox = ajustes.cuenta_quadrata
                        
                    else:
                        cuenta = ""
                        cta_iva = ""
                        cta_rte = ""
                        cta_rte_asum = ""
                        cta_aprox = ""
                    sbt = i.valor -vr_iva + vr_rte
                    if i.rte_asumida: sbt -= vr_rte
                    sheet.cell(row,4,cuenta)
                    sheet.cell(row,5,"D")
                    sheet.cell(row,6,sbt)
                    sheet.cell(row,7,i.fecha_gasto.year)
                    sheet.cell(row,8,i.fecha_gasto.month)
                    sheet.cell(row,9,i.fecha_gasto.day)
                    sheet.cell(row,16,i.tercero.pk)
                    sheet.cell(row,18,i.descripcion.upper())
                    row+=1
                    
                    if cta_iva != "":
                        sheet.cell(row,4,cta_iva)
                        sheet.cell(row,5,"D")
                        sheet.cell(row,6, vr_iva)
                        sheet.cell(row,7, i.fecha_gasto.year)
                        sheet.cell(row,8, i.fecha_gasto.month)
                        sheet.cell(row,9, i.fecha_gasto.day)
                        sheet.cell(row,16, i.tercero.pk)
                        sheet.cell(row,18, "IVA - "+i.descripcion.upper())
                        row+=1

                    if cta_rte != "":
                        sheet.cell(row,4,cta_rte)
                        sheet.cell(row,5,"C")
                        sheet.cell(row,6, vr_rte)
                        sheet.cell(row,7, i.fecha_gasto.year)
                        sheet.cell(row,8, i.fecha_gasto.month)
                        sheet.cell(row,9, i.fecha_gasto.day)
                        sheet.cell(row,16, i.tercero.pk)
                        sheet.cell(row,18, "RETEFTE - "+i.descripcion.upper())
                        row+=1
                    
                        if i.rte_asumida is True:
                            sheet.cell(row,4,cta_rte_asum)
                            sheet.cell(row,5,"D")
                            sheet.cell(row,6, vr_rte)
                            sheet.cell(row,7, i.fecha_gasto.year)
                            sheet.cell(row,8, i.fecha_gasto.month)
                            sheet.cell(row,9, i.fecha_gasto.day)
                            sheet.cell(row,16, i.tercero.pk)
                            sheet.cell(row,18, "RETEFTE ASUMIDA- "+i.descripcion.upper())
                            row+=1
                            
                responsable = Profiles.objects.get(user=reembolso.caja.usuario_responsable.pk)
                valor = reembolso.valor
                valor_aprox = valor + (10000 - valor % 10000) % 10000
                vr_ajuste = valor_aprox - valor
                
                if vr_ajuste > 0:
                    sheet.cell(row,4,cta_aprox)
                    sheet.cell(row,5,"D")
                    sheet.cell(row,6, vr_ajuste)
                    sheet.cell(row,7, reembolso.fecha_aprueba.year)
                    sheet.cell(row,8, reembolso.fecha_aprueba.month)
                    sheet.cell(row,9, reembolso.fecha_aprueba.day)
                    sheet.cell(row,16, responsable.identificacion)
                    sheet.cell(row,18, "AJUSTE POR APROXIMACION PARA RETIRO EN CAJERO")
                    row+=1
                
                
                sheet.cell(row,4,reembolso.caja.nro_cuentacontable)
                sheet.cell(row,5,"C")
                sheet.cell(row,6,valor_aprox)
                sheet.cell(row,7,reembolso.fecha_aprueba.year)
                sheet.cell(row,8,reembolso.fecha_aprueba.month)
                sheet.cell(row,9,reembolso.fecha_aprueba.day)
                sheet.cell(row,16,responsable.identificacion)
                sheet.cell(row,18,F"REEMBOLSO CAJA {reembolso.caja.usuario_responsable.get_full_name().upper()}")
                    
                nombre_doc=f'Interfaz_legalizacion_caja_{reembolso.caja.usuario_responsable.username}.xlsx'
                ruta=settings.MEDIA_ROOT+'/tmp/xlsx/'+nombre_doc
                ruta_dw=settings.MEDIA_URL+'tmp/xlsx/'+nombre_doc
                book.save(ruta)
                
                data = {
                    'class': 'alert-success',
                    'msj':f'Puedes descargar el archivo de legalizacion <strong><a href="{ruta_dw}" target="_blank">aquí</a></strong>'
                }   
                
                return JsonResponse(data)
            
            elif todo == 'print_reemb':
                id_reemb = request.GET.get('id_reemb')
                reembolso = reembolsos_caja.objects.get(pk=id_reemb)
                gastos = gastos_caja.objects.filter(reembolso=id_reemb).order_by('fecha_gasto')
                
                context = {
                    'reembolso':reembolso,
                    'gastos_reembolso':gastos,
                    'user_aprueba': gastos.first().usuario_aprueba,
                    'user':request.user,
                    'now':datetime.datetime.now(),
                }
                filename = f'Reembolso_caja_{reembolso.caja.usuario_responsable.get_full_name().upper()}.pdf'.replace(' ','_')
                filename = filename.translate(trans)
                
                
                pdf = pdf_gen(f'pdf/reembolso_caja.html',context,filename)
                
                file = pdf.get('url')
                
                msj = f'Puedes descargar el archivo <a href="{file}" target="_blank"><b>Aquí</b></a>'
                bg = 'alert-success'
                
                data = {
                    'msj': msj,
                    'class': bg
                }
                
                return JsonResponse(data)
            
        elif request.method == 'POST':
            todo = request.POST.get('to_do')
            
            if todo == 'nuevogasto':
                id_caja = request.POST.get('id_caja')
                cuenta = cuentas_pagos.objects.get(pk=id_caja)
                owner = True if cuenta.usuario_responsable.pk == request.user.pk else False
                if not owner:
                    data = {
                        'msj': 'No eres el responsable de esta caja, no puedes registrar gastos.',
                        'class': 'alert-danger'
                    }
                    
                    return JsonResponse(data)
                
                fecha = request.POST.get('fecha')
                descripcion = request.POST.get('descripcion')
                nit_tercero = request.POST.get('nit_tercero')
                valor = request.POST.get('valor')
                soporte = request.FILES.get('soporte')
                concepto = request.POST.get('concepto')
                
                if not soporte.name.endswith('.pdf'):
                    
                    data = {
                        'msj':'El soporte debe ser un archivo en formato PDF.',
                        'class':'alert-danger'
                    }
                    
                    return JsonResponse(data)
                
                                
                obj_concepto = conceptos_legalizacion.objects.get(pk=concepto)
                
                partner = Partners.objects.get(pk=nit_tercero)
                
                soporte.name = soporte.name.translate(trans)
                
                if not soporte.name.endswith('.pdf'):
                    data = {
                        'msj': 'El soporte a cargar debe ser un PDF, vuelve a intentarlo',
                        'class': 'alert-danger'
                    }
                
                    return JsonResponse(data)
                
                gasto = gastos_caja.objects.create(
                    fecha_gasto = fecha, concepto = obj_concepto,
                    descripcion = descripcion.upper(),  tercero = partner,
                    valor = valor.replace(',',''), soporte = soporte,
                    usuario_carga = request.user, forma_pago = cuenta
                )
                
                data = {
                    'msj': 'El gasto fue registrado',
                    'class': 'alert-success'
                }
                
                return JsonResponse(data)
            
            elif todo == 'change_concept':
                                
                gasto =  request.POST.get('id_gasto')
                concepto = request.POST.get('new_concept')
                obj_gasto = gastos_caja.objects.get(pk = gasto)
                obj_concept = conceptos_legalizacion.objects.get(pk=concepto)
                
                owner = True if obj_gasto.forma_pago.usuario_responsable.pk == request.user.pk else False
                is_cont = True if check_groups(request,('Contabilidad',),raise_exception=False) else False
                if not owner and not is_cont:
                    data = {
                        'msj': 'No tienes permiso de realizar cambios sobre un gasto.',
                        'class': 'alert-danger'
                    }
                    
                    return JsonResponse(data)
                
                
                obj_gasto.concepto = obj_concept
                
                if obj_gasto.estado == 'Devuelto':
                    obj_gasto.estado = 'Reembolso'
                
                obj_gasto.save()
                
                data = {
                    'msj':'Se cambió el concepto del gasto seleccionado.',
                    'class':'alert-success'
                }
                
                return JsonResponse(data)
    
            elif todo == 'change_support_gasto':
                gasto =  request.POST.get('gasto')
                file = request.FILES.get('nuevo_soporte')
                obj_gasto = gastos_caja.objects.get(pk = gasto)
                
                owner = True if obj_gasto.forma_pago.usuario_responsable.pk == request.user.pk else False
                is_cont = True if check_groups(request,('Contabilidad',),raise_exception=False) else False
                if not owner and not is_cont:
                    data = {
                        'msj': 'No tienes permiso de realizar cambios sobre un gasto.',
                        'class': 'alert-danger'
                    }
                    
                    return JsonResponse(data)
                
                if not file.name.endswith('.pdf'):
                    
                    data = {
                        'msj':'El soporte debe ser un archivo en formato PDF.',
                        'class':'alert-danger'
                    }
                    
                    return JsonResponse(data)
                
                obj_gasto.soporte = file
                
                
                if obj_gasto.estado == 'Devuelto':
                    obj_gasto.estado = 'Reembolso'
                
                obj_gasto.save()
                
                data = {
                    'msj':'Se cambió el soporte del gasto seleccionado.',
                    'class':'alert-success'
                }
                
                return JsonResponse(data)
            
            elif todo == 'add_taxes':
                gasto =  request.POST.get('id_gasto')
                obj_gastos = gastos_caja.objects.get(pk = gasto)
                
                owner = True if obj_gastos.forma_pago.usuario_responsable.pk == request.user.pk else False
                is_cont = True if check_groups(request,('Contabilidad',),raise_exception=False) else False
                if not owner and not is_cont:
                    data = {
                        'msj': 'No tienes permiso de realizar cambios sobre un gasto.',
                        'class': 'alert-danger'
                    }
                    
                    return JsonResponse(data)
                
                tipo_iva = request.POST.get('tipo_iva')
                valor_iva = request.POST.get('valor_iva').replace(',','')
                tipo_rte = request.POST.get('tipo_rte')
                valor_rte =  request.POST.get('valor_rte').replace(',','')
                rte_asumida = request.POST.get('rte_asumida')
                
                if int(valor_iva) >= 0:
                    cuenta_iva = None if tipo_iva == "" or tipo_iva == None else \
                                    impuestos_legalizacion.objects.get(pk=tipo_iva)
                    obj_gastos.cuenta_iva = cuenta_iva
                    obj_gastos.valor_iva = valor_iva
                
                if int(valor_rte) >= 0:
                    cuenta_rte = None if tipo_rte == "" or tipo_rte == None else \
                                    impuestos_legalizacion.objects.get(pk=tipo_rte)
                    obj_gastos.cuenta_rte = cuenta_rte
                    obj_gastos.valor_rte = valor_rte
                    
                    rte_asumida = True if rte_asumida == 'on' else False
                
                    obj_gastos.rte_asumida = rte_asumida
                
                if obj_gastos.estado == 'Devuelto':
                    obj_gastos.estado = 'Reembolso'
                
                obj_gastos.save()
                
                data = {
                    
                }
                
                return JsonResponse(data)
            
            elif todo == 'approve':
                
                gasto = request.POST.get('gasto')
                
                obj_gasto = gastos_caja.objects.get(pk=gasto)
                
                if obj_gasto.forma_pago.usuario_aprobador.pk != request.user.pk \
                    and not request.user.is_superuser:
                    data = {
                        'msj':'No tienes permisos para aprobar este gasto',
                        'class':'alert-danger'
                    }
                
                else:
                    obj_gasto.estado = 'Aprobado'
                    obj_gasto.usuario_aprueba = request.user
                    obj_gasto.fecha_aprobacion = datetime.date.today()
                    obj_gasto.save()
                    
                    data = {
                        'msj':'Se aprobó el gasto seleccionado',
                        'class':'alert-success'
                    }
                
                return JsonResponse(data)
                
            elif todo == 'solicitar_reembolso':
                
                caja = request.POST.get('caja')
                
                gastos_sin_legalizar = gastos_caja.objects.filter(forma_pago=caja,
                                                                  estado = 'Aprobado',
                                                                  reembolso__isnull=True)
                
                if gastos_sin_legalizar.exists():
                    
                    valor = 0
                    obj_caja = cuentas_pagos.objects.get(pk=caja)
                    
                    owner = True if obj_caja.usuario_responsable.pk == request.user.pk else False
                    if not owner:
                        data = {
                            'msj': 'No eres el responsable de esta caja, no puedes solicitar reembolsos.',
                            'class': 'alert-danger'
                        }
                        
                        return JsonResponse(data)
                        
                    reembolso = reembolsos_caja.objects.create(usuario_solicita=request.user,
                                                            valor = valor,
                                                            caja=obj_caja)
                    
                    for i in gastos_sin_legalizar:
                        i.reembolso = reembolso
                        i.estado = 'Reembolso'
                        i.save()
                        valor += i.valor
                        
                    reembolso.valor = valor
                    reembolso.save()
                    
                    data = {
                            'msj':'Se solcitó el reembolso',
                            'class':'alert-success'
                        }
                    
                    try:
                        msj = f'''Hola :), <b>{request.user.get_full_name()}</b> ha solicitado un nuevo reembolso por {valor:,}<br>'''
                    
                        envio_notificacion(msj,f'Nueva reembolso solicitado',['auxiliarcontable2@somosandina.co','davidcalao@somosandina.co'])
                        
                    except:
                        pass
                    
                else:
                    data = {
                            'msj':'No hay gastos pendientes por legalizar.',
                            'class':'alert-danger'
                        }
                
                return JsonResponse(data)
                
            elif todo == 'approve_leg':
                line = request.POST.get('id_reemb')
                
                is_cont = True if check_groups(request,('Contabilidad',),raise_exception=False) else False
                if not is_cont:
                    data = {
                        'msj': 'No tienes permiso de aprobar un reembolso.',
                        'class': 'alert-danger'
                    }
                    
                    return JsonResponse(data)
                
                reemb = reembolsos_caja.objects.get(pk=line)
                reemb.usuario_aprueba = request.user
                reemb.fecha_aprueba = datetime.date.today()
                reemb.save()
                
                data = {
                        'msj':'Se aprobó el gasto seleccionado',
                        'class':'alert-success'
                    }
                
                try:
                    msj = f'''Hola :), <b>{request.user.get_full_name()}</b> ha verificado una reembolso de caja a {reemb.caja.usuario_responsable.get_full_name()} por {reemb.valor:,}
                            y se encuentra disponible para su pago.<br>'''
                    
                    
                    if reemb.caja.oficina == 'MONTERIA':
                        pagadores = [
                            'admin1@somosandina.co',
                            'recepcionmtr@somosandina.co',
                            'tatianamontes@somosandina.co',
                            reemb.usuario_responsable.email,
                            reemb.usuario_aprobador.email
                        ]
                        emails = list(set(pagadores))
                    else:
                        pagadores = [
                            'operaciones2@somosandina.co',
                            'jorgeavila@somosandina.co',
                            reemb.usuario_responsable.email,
                            reemb.usuario_aprobador.email
                        ]
                        emails = list(set(pagadores))
                        
                    envio_notificacion(msj,f'Reembolso de caja verificado',emails)
                    
                except:
                    pass
                
                return JsonResponse(data)
            
            elif todo == 'leg_reemb':
                
                is_cont = True if check_groups(request,('Contabilidad',),raise_exception=False) else False
                if not is_cont:
                    data = {
                        'msj': 'No tienes permiso de registrar una legalización.',
                        'class': 'alert-danger'
                    }
                    
                    return JsonResponse(data)
                
                id_leg = request.POST.get('id_leg')
                causacion = request.POST.get('causacion_leg_reemb')
                valor_reemb = request.POST.get('valor_reemb')
                soporte = request.FILES.get('soporte_leg_reemb')
                
                soporte.name = soporte.name.translate(trans)
                
                if not soporte.name.endswith('.pdf'):
                    
                    data = {
                        'msj':'El soporte debe ser un archivo en formato PDF.',
                        'class':'alert-danger'
                    }
                    
                    return JsonResponse(data)
                
                reembolso = reembolsos_caja.objects.get(pk=id_leg)
                reembolso.doc_legalizacion = causacion
                reembolso.soporte_legalizacion = soporte
                reembolso.valor_a_reembolsar = valor_reemb.replace(',','')
                reembolso.save()
                
                gastos = gastos_caja.objects.filter(reembolso=reembolso.pk)
                
                for i in gastos:
                    i.estado = 'Legalizado'
                    i.save()
                
                data = {
                        'msj':'Se registró la legalización del reembolso seleccionado',
                        'class':'alert-success'
                    }
                
                return JsonResponse(data)

            elif todo == 'delete':
                legalizacion = request.POST.get('gasto')
                
                obj_leg = gastos_caja.objects.get(pk=legalizacion)
                owner = True if obj_leg.forma_pago.usuario_responsable.pk == request.user.pk else False
                if not owner:
                    data = {
                        'msj': 'No eres el responsable de esta caja, no puedes eliminar gastos.',
                        'class': 'alert-danger'
                    }
                    
                    return JsonResponse(data)
                
                obj_leg.delete()

                msj = 'Se eliminó el gasto seleccionado'
                bg = 'alert-success'
                
                
                data = {
                    'msj': msj,
                    'bg': bg
                }
                
                return JsonResponse(data)
            
            elif todo == 'return_to_user':
                gasto = request.POST.get('gasto')
                obj_gasto = gastos_caja.objects.get(pk=gasto)
                obj_gasto.estado = 'Devuelto' 
                
                obj_gasto.save()
                
                data = {
                        'msj':'Se regresó el gasto seleccionado al usuario responsable',
                        'class':'alert-success'
                    }
                
                return JsonResponse(data)
            
    superuser = request.user.is_superuser
    contabilidad = check_groups(request,['Contabilidad',],raise_exception=False)
    tesoreria = check_groups(request,['Tesoreria',],raise_exception=False)
    
    usertype = None
    
    cajas_asignadas = cuentas_pagos.objects.filter(
        usuario_responsable = request.user.pk)
    
    cajas_aprobar = cuentas_pagos.objects.filter(
        usuario_aprobador = request.user.pk)
    
    if cajas_asignadas.exists(): usertype = 'responsable'
    if cajas_aprobar.exists(): usertype = 'aprobador'       

    if tesoreria: usertype = 'tesoreria'
    elif contabilidad: usertype = 'contabilidad'
       
    if superuser: usertype = 'superuser'
    
    
    context = {
        'formCaja': forms.form_cajas(user=request.user.pk,user_type=usertype),
        'form_leg_anticipo':forms.form_legalizar_anticipo,
        'form_partners':forms.form_partners,
        'form_taxes':forms.form_taxes,
        'form_transf':forms.form_transferencias,
        'form_reg_legaliz': forms.form_reg_legaliz,
        'abrir_mes_anterior':parametros.objects.get(descripcion='abrir_mes_anterior')
    }
    
    """ lista = [78725,78730,78706,78707,78709,78714,78708,78715,78722,78724,78723,78720,78705,78735,78713,78728,78718,78737,78733,78727,78736,78726,78712,78734,78704,78721,78716,78729,78731,78710,78703,78719,78748,78741,78747,78742,78744,78738,78745,78743,78739,78740,78746,78770,78796,78786,78775,78755,78769,78790,78764,78792,78760,78779,78773,78791,78753,78758,78788,78780,78766,78793,78782,78771,78787,78762,78757,78778,78772,78777,78761,78781,78799,78797,78756,78765,78776,78783,78794,78785,77048,78759,78763,78767,77871,78774,78795,77602,78083,78077,78043,78066,78068,78064,78070,78055,78088,78087,78065,78082,78049,78086,78051,78034,78050,78072,77642,78067,78059,78045,78061,78078,78073,78058,78041,78084,78069,78048,78053,78044,78042,78054,78071,78038,78040,78052,78056,78035,78089,78079,78037,78075,78081,78033,78060,78076,78085,78047,78080,78074,78306,78247,78218,78205,78305,78219,78242,78217,78151,78126,78315,78204,77820,78187,78222,78276,78221,78181,78202,78201,78149,78150,77713,78284,78285,78203,78243,78233,78195,78317,78192,78177,78178,78209,77888,78180,78297,78230,78318,78208,78234,78168,78138,78193,78212,78277,78278,78146,78173,78307,78288,78179,78100,78105,78125,78140,78141,78142,78143,78144,78145,78154,78157,78169,78170,78171,78172,78174,78175,78186,78196,78197,78198,78199,78213,78214,78235,78236,78237,78238,78239,78251,78252,78260,78261,78262,78263,78264,78265,78272,78273,78274,78275,78283,78296,78298,78299,78300,78301,78302,78303,78312,78148,78223,78282,78194,78279,78281,78165,78132,78159,78309,78090,78103,78123,78131,78133,78134,78135,78136,78137,78152,78155,78158,78160,78161,78162,78163,78164,78183,78188,78189,78190,78191,78206,78207,78224,78225,78226,78227,78228,78248,78249,78253,78254,78255,78256,78257,78258,78266,78267,78268,78269,78280,78286,78287,78289,78290,78291,78292,78293,78229,78220,78316,78184,78294,78124,78210,78259,78106,78271,78156,78250,78310,78104,78153,78270,78176,78314,78182,78139,78313,78244,78232,78185,78295,78231,78216,78240,78311,78246,78215,78304,78211,78245,78200,78614,78547,78425,78560,78418,78643,78451,78346,78608,78631,78594,78429,78518,78552,78544,78639,78484,78469,78483,78446,78565,78460,78617,78432,78448,78509,78574,77940,78597,78499,78593,78615,78611,78581,78582,78416,78506,78433,78450,78546,78550,78584,78634,78488,78489,78470,78492,78596,78598,78424,78491,78503,78573,78575,78512,78515,78494,78495,78507,78342,78519,78536,78500,78595,78465,78517,78505,78479,78486,78490,78496,78504,78487,78449,78498,78516,78501,78514,78481,78480,78497,78502,78508,78511,78513,78635,78636,78482,78510,78609,78545,78553,78637,78579,78468,78606,78431,78434,78435,78610,78640,78633,78473,78638,78641,78532,78447,78627,78628,78632,77938,78612,78430,77915,77944,77956,78442,78629,78616,78485,78620,78556,78583,78414,78645,78589,78426,78466,78343,78567,78458,78445,78462,78417,78467,78537,78576,78548,78472,78599,78471,78534,78551,78549,78423,78522,78523,78524,78535,78649,78520,78539,78571,78619,78323,78461,78538,78540,78541,78542,78543,78572,78564,78525,78526,78528,78422,78563,78527,78529,78530,78562,78588,78642,78602,78437,78586,78452,78441,76985,78345,78439,78454,78455,78475,78521,78533,78569,78618,78648,78621,78436,78444,78463,78622,78623,78651,78568,78646,78443,78585,78413,78570,78459,78344,78347,78474,78478,78561,78590,78600,78601,78613,78592,78604,78457,78559,78464,78456,78578,78652,78554,78415,78438,78477,78558,78626,78577,78647,78591,78453,78650,78440,78644,78566,78624,78476,78557,78531,78587,78625,78428,78603,78427,78555,77736,78669,78684,78679,78676,78681,78672,78692,78673,78677,78675,78691,78685,78686,78693,78662,78661,78671,78670,78680,78687,78678,78690,78683,78682,78663]

    for i in lista:
        e = egresos_banco.objects.get(pk=i)
        if e.estado == 'SIN CONCILIAR':
            e.delete()
            print(f'{i} se eliminó')
        else:
            print(f'{i} NO se eliminó')
     """
          
    return render(request, 'cajas_efectivo.html', context)


# ==============================================================================
# UPLOAD MOVEMENTS - Carga masiva de movimientos bancarios
# ==============================================================================

@login_required
@group_perm_required(('accounting.add_egresos_banco',), raise_exception=True)
def upload_movements(request):
    """
    Vista para cargar movimientos bancarios de forma masiva
    Permite seleccionar empresa y ver resumen de últimos 5 días
    """
    lista_empresas = empresas.objects.all().order_by('nombre')

    context = {
        'empresas': lista_empresas
    }
    return render(request, 'accounting/upload_movements.html', context)


@login_required
@group_perm_required(('accounting.view_egresos_banco',), raise_exception=True)
def api_movements_summary(request):
    """
    API para obtener resumen de movimientos de últimos 5 días

    Query params:
        - empresa: ID de la empresa (requerido)

    Response:
        {
            "empresa": 1,
            "empresa_nombre": "ANDINA SOFT S.A.S",
            "resumen": [
                {"fecha": "2026-01-02", "banco": 45, "wompi": 23, "plink": 12, "total": 80},
                ...
            ]
        }
    """
    import requests
    from datetime import timedelta
    from django.utils import timezone

    empresa_id = request.GET.get('empresa')

    if not empresa_id:
        return JsonResponse({
            'detail': 'Parámetro empresa es requerido'
        }, status=400)

    try:
        empresa = empresas.objects.get(pk=empresa_id)
    except empresas.DoesNotExist:
        return JsonResponse({
            'detail': f'Empresa {empresa_id} no encontrada'
        }, status=404)

    # Calcular últimos 5 días
    hoy = timezone.now().date()

    # Consultar Wompi y Plink para últimos 3 días (una sola llamada cada uno)
    wompi_data = _get_wompi_count_from_n8n(empresa_id)
    plink_data = _get_plink_count_from_n8n(empresa_id)

    resumen = []

    for i in range(5):
        fecha = hoy - timedelta(days=i)
        fecha_str = fecha.isoformat()

        # 1. BANCO: Consulta directa a Django model
        count_banco = egresos_banco.objects.filter(
            empresa=empresa,
            fecha=fecha
        ).count()

        # 2. WOMPI: Buscar en datos retornados de n8n
        count_wompi = wompi_data.get(fecha_str, 0)

        # 3. PLINK: Buscar en datos retornados de n8n
        count_plink = plink_data.get(fecha_str, 0)

        resumen.append({
            'fecha': fecha_str,
            'banco': count_banco,
            'wompi': count_wompi,
            'plink': count_plink,
            'total': count_banco + count_wompi + count_plink
        })

    return JsonResponse({
        'empresa': empresa.pk,
        'empresa_nombre': empresa.nombre,
        'resumen': resumen
    })


def _get_wompi_count_from_n8n(empresa_id):
    """
    Helper: Consulta cantidad de movimientos Wompi vía n8n
    Retorna dict con últimos 3 días: {"2026-01-02": 23, "2026-01-01": 15, ...}
    """
    import requests
    try:
        response = requests.get(
            settings.N8N_WEBHOOK_WOMPI_COUNT,
            params={'empresa_id': empresa_id},
            timeout=10
        )

        if response.status_code == 200:
            data = response.json()
            # Espera: {"movimientos": [{"fecha": "2026-01-02", "count": 23}, ...]}
            # o formato simplificado: {"2026-01-02": 23, "2026-01-01": 15}
            if 'movimientos' in data:
                return {item['fecha']: item['count'] for item in data['movimientos']}
            else:
                return data
        else:
            return {}

    except Exception as e:
        print(f"Error consultando Wompi count: {e}")
        return {}


def _get_plink_count_from_n8n(empresa_id):
    """
    Helper: Consulta cantidad de movimientos Plink vía n8n
    Retorna dict con últimos 3 días: {"2026-01-02": 12, "2026-01-01": 8, ...}
    """
    import requests
    try:
        response = requests.get(
            settings.N8N_WEBHOOK_PLINK_COUNT,
            params={'empresa_id': empresa_id},
            timeout=10
        )

        if response.status_code == 200:
            data = response.json()
            # Espera: {"movimientos": [{"fecha": "2026-01-02", "count": 12}, ...]}
            # o formato simplificado: {"2026-01-02": 12, "2026-01-01": 8}
            if 'movimientos' in data:
                return {item['fecha']: item['count'] for item in data['movimientos']}
            else:
                return data
        else:
            return {}

    except Exception as e:
        print(f"Error consultando Plink count: {e}")
        return {}


@login_required
@group_perm_required(('accounting.view_egresos_banco',), raise_exception=True)
def api_check_upload_status(request):
    """
    Endpoint para consultar el estado de una carga de movimientos
    GET /accounting/api/check-upload-status?job_id=<id>
    """
    from accounting.models import upload_movements_job

    if request.method != 'GET':
        return JsonResponse({'detail': 'Método no permitido'}, status=405)

    job_id = request.GET.get('job_id')
    if not job_id:
        return JsonResponse({'detail': 'job_id es requerido'}, status=400)

    try:
        job = upload_movements_job.objects.get(pk=job_id)

        # Verificar permisos: solo el usuario que creó el job o usuarios con permisos globales pueden verlo
        if job.usuario != request.user:
            perm = check_perms(request, ('andinasoft.add_recaudos_general',), raise_exception=False)
            if not perm:
                return JsonResponse({'detail': 'No tienes permiso para ver este trabajo'}, status=403)

        response_data = {
            'job_id': job.id,
            'status': job.status,
            'empresa': job.empresa.pk,
            'created_at': job.created_at.isoformat(),
            'updated_at': job.updated_at.isoformat(),
        }

        # Si está completado o con error, incluir detalles
        if job.status in ['completed', 'partial', 'failed']:
            response_data['completed_at'] = job.completed_at.isoformat() if job.completed_at else None
            response_data['mensaje'] = job.mensaje
            response_data['detalles'] = {
                'banco': job.movimientos_banco,
                'wompi': job.movimientos_wompi,
                'plink': job.movimientos_plink,
            }

            if job.movimientos_rechazados:
                response_data['movimientos_rechazados'] = job.movimientos_rechazados
                response_data['archivo_rechazados'] = job.archivo_rechazados
                response_data['nombre_archivo'] = job.nombre_archivo_rechazados

            if job.error_detail:
                response_data['error'] = job.error_detail

        return JsonResponse(response_data)

    except upload_movements_job.DoesNotExist:
        return JsonResponse({'detail': 'Trabajo no encontrado'}, status=404)


@csrf_exempt
def api_upload_callback(request):
    """
    Endpoint de callback para que n8n notifique el resultado del procesamiento
    POST /accounting/api/upload-callback

    Body esperado:
    {
        "job_id": 123,
        "success": true/false,
        "message": "...",
        "detalles": {
            "banco": 45,
            "wompi": 23,
            "plink": 12
        },
        "archivo_rechazados": "<base64>",  // opcional
        "nombre_archivo": "rechazados.xlsx"  // opcional
    }
    """
    from accounting.models import upload_movements_job
    from django.utils import timezone
    import json

    if request.method != 'POST':
        return JsonResponse({'detail': 'Método no permitido'}, status=405)

    try:
        data = json.loads(request.body.decode('utf-8'))
    except (json.JSONDecodeError, UnicodeDecodeError):
        return JsonResponse({'detail': 'JSON inválido'}, status=400)

    job_id = data.get('job_id')
    if not job_id:
        return JsonResponse({'detail': 'job_id es requerido'}, status=400)

    try:
        job = upload_movements_job.objects.get(pk=job_id)
    except upload_movements_job.DoesNotExist:
        return JsonResponse({'detail': 'Trabajo no encontrado'}, status=404)

    # Actualizar el job con los resultados
    success = data.get('success', False)
    detalles = data.get('detalles', {})

    if success:
        job.status = 'completed'
        job.mensaje = data.get('message', 'Movimientos cargados exitosamente')
        job.movimientos_banco = detalles.get('banco')
        job.movimientos_wompi = detalles.get('wompi')
        job.movimientos_plink = detalles.get('plink')
    else:
        # Verificar si hay rechazos parciales
        rechazados = data.get('movimientos_rechazados', 0)
        if rechazados > 0:
            job.status = 'partial'
            job.mensaje = data.get('message', f'{rechazados} movimientos fueron rechazados')
            job.movimientos_rechazados = rechazados
            job.archivo_rechazados = data.get('archivo_rechazados')
            job.nombre_archivo_rechazados = data.get('nombre_archivo')
            job.movimientos_banco = detalles.get('banco')
            job.movimientos_wompi = detalles.get('wompi')
            job.movimientos_plink = detalles.get('plink')
        else:
            job.status = 'failed'
            job.mensaje = data.get('message', 'Error al procesar movimientos')
            job.error_detail = data.get('error')

    job.completed_at = timezone.now()
    job.save()

    return JsonResponse({
        'status': 'ok',
        'job_id': job.id,
        'updated_status': job.status
    })


@login_required
@group_perm_required(('accounting.add_egresos_banco',), raise_exception=True)
def api_upload_movements(request):
    """
    API para enviar archivos de movimientos al webhook de n8n

    FormData:
        - empresa: ID de la empresa (requerido)
        - archivo_banco: archivo Excel/CSV (opcional)
        - archivo_wompi: archivo Excel/CSV (opcional)
        - archivo_plink: archivo Excel/CSV (opcional)

    Response exitoso:
        {
            "success": true,
            "message": "Movimientos cargados exitosamente",
            "detalles": {"banco": 45, "wompi": 23, "plink": 12}
        }

    Response con rechazados:
        {
            "success": false,
            "message": "Algunos movimientos fueron rechazados",
            "archivo_rechazados": "<base64>",
            "nombre_archivo": "rechazados_2026-01-02.xlsx"
        }
    """
    import requests
    from accounting.models import upload_movements_job

    if request.method != 'POST':
        return JsonResponse({
            'detail': 'Método no permitido'
        }, status=405)

    empresa_id = request.POST.get('empresa')

    if not empresa_id:
        return JsonResponse({
            'detail': 'Parámetro empresa es requerido'
        }, status=400)

    # Validar empresa
    try:
        empresa = empresas.objects.get(pk=empresa_id)
    except empresas.DoesNotExist:
        return JsonResponse({
            'detail': 'Empresa no encontrada'
        }, status=404)

    # Validar que al menos un archivo fue enviado
    tiene_banco = 'archivo_banco' in request.FILES
    tiene_wompi = 'archivo_wompi' in request.FILES
    tiene_plink = 'archivo_plink' in request.FILES

    if not any([tiene_banco, tiene_wompi, tiene_plink]):
        return JsonResponse({
            'detail': 'Debes enviar al menos un archivo de movimientos'
        }, status=400)

    # Crear job de procesamiento
    job = upload_movements_job.objects.create(
        empresa=empresa,
        usuario=request.user,
        status='processing',
        tiene_banco=tiene_banco,
        tiene_wompi=tiene_wompi,
        tiene_plink=tiene_plink
    )

    # Preparar archivos para enviar a n8n
    archivos = {}
    if tiene_banco:
        archivos['Movimiento_Banco'] = request.FILES['archivo_banco']
    if tiene_wompi:
        archivos['Movimiento_Wompi'] = request.FILES['archivo_wompi']
    if tiene_plink:
        archivos['Movimiento_Plink'] = request.FILES['archivo_plink']

    # Enviar a n8n webhook
    webhook_url = settings.N8N_WEBHOOK_UPLOAD_MOVEMENTS

    try:
        # Preparar files para requests con nombres específicos para n8n
        files = {
            key: (file.name, file.read(), file.content_type)
            for key, file in archivos.items()
        }

        # Construir callback URL
        callback_url = request.build_absolute_uri('/accounting/api/upload-callback')

        data = {
            'empresa': empresa_id,
            'job_id': job.id,
            'callback_url': callback_url
        }

        # Llamar webhook n8n sin esperar respuesta (timeout corto)
        response = requests.post(
            webhook_url,
            files=files,
            data=data,
            timeout=10  # Solo 10 segundos para iniciar el proceso
        )

        # Si n8n responde inmediatamente, procesarlo
        if response.status_code == 200:
            try:
                result = response.json()
                # Si n8n ya retornó el resultado completo (modo síncrono)
                if 'success' in result:
                    from django.utils import timezone
                    if result.get('success'):
                        job.status = 'completed'
                        job.mensaje = result.get('message', 'Movimientos cargados exitosamente')
                        detalles = result.get('detalles', {})
                        job.movimientos_banco = detalles.get('banco')
                        job.movimientos_wompi = detalles.get('wompi')
                        job.movimientos_plink = detalles.get('plink')
                    else:
                        rechazados = result.get('movimientos_rechazados', 0)
                        if rechazados > 0:
                            job.status = 'partial'
                            job.mensaje = result.get('message')
                            job.movimientos_rechazados = rechazados
                            job.archivo_rechazados = result.get('archivo_rechazados')
                            job.nombre_archivo_rechazados = result.get('nombre_archivo')
                        else:
                            job.status = 'failed'
                            job.mensaje = result.get('message', 'Error al procesar movimientos')
                    job.completed_at = timezone.now()
                    job.save()
            except (ValueError, KeyError):
                pass

        # Retornar inmediatamente con el job_id
        return JsonResponse({
            'job_id': job.id,
            'status': job.status,
            'message': 'Los archivos se están procesando. Consulta el estado con el job_id.'
        })

    except requests.exceptions.Timeout:
        # Timeout no es error, n8n está procesando
        return JsonResponse({
            'job_id': job.id,
            'status': 'processing',
            'message': 'Los archivos se están procesando en segundo plano.'
        })

    except requests.exceptions.RequestException as e:
        from django.utils import timezone
        job.status = 'failed'
        job.mensaje = 'Error al comunicarse con el servicio de procesamiento'
        job.error_detail = f'Webhook: {webhook_url} | Error: {str(e)}'
        job.completed_at = timezone.now()
        job.save()
        return JsonResponse({
            'job_id': job.id,
            'status': 'failed',
            'error': f'No fue posible conectar con n8n ({webhook_url})',
            'error_detail': str(e)
        }, status=502)

    except Exception as e:
        from django.utils import timezone
        job.status = 'failed'
        job.mensaje = 'Error inesperado'
        job.error_detail = str(e)
        job.completed_at = timezone.now()
        job.save()
        return JsonResponse({
            'job_id': job.id,
            'status': 'failed',
            'error': str(e)
        }, status=500)


urls = [
    path('principal',principal),
    path('movements',movimientos),
    path('upload-movements',upload_movements),
    path('api/movements-summary',api_movements_summary),
    path('api/upload-movements',api_upload_movements),
    path('api/upload-callback',api_upload_callback),
    path('api/check-upload-status',api_check_upload_status),
    path('api/bank-movements',api_bank_movements),
    path('api/lotes',api_lotes),
    path('api/lotes/estado',api_cambiar_estado_lote),
    path('api/bank-movements/list',api_bank_movements_list),
    path('api/bank-movements/for-receipt',api_bank_movements_for_receipt),
    path('api/bank-movements/<int:movement_id>/mark-used',api_bank_movement_mark_used),
    path('api/plink-movements',api_plink_movements),
    path('api/plink-movements/list',api_plink_movements_list),
    path('api/plink-movements/<int:movement_id>/mark-used',api_plink_movement_mark_used),
    path('api/wompi-movements',api_wompi_movements),
    path('api/wompi-movements/list',api_wompi_movements_list),
    path('api/wompi-movements/<int:movement_id>/mark-used',api_wompi_movement_mark_used),
    path('movement_sheet',planilla_movimientos),
    path('reconciliations',view_conciliaciones),
    path('reconciliationslist',lista_conciliaciones),
    path('radicarfactura',radicar_factura),
    path('causarfactura',causar_factura),
    path('pagarfactura',pagar_factura),
    path('listafacturas',lista_facturas),
    path('listapagos',lista_pagos),
    path('otherincomes',view_otros_ingresos),
    path('solicitaranticipos',solicitud_de_anticipos),
    path('legalizaciones',legalizaciones),
    path('partners',partners),
    path('conceptos',conceptos),
    path('archivocontable',historical_data),
    path('cajasefectivo',cajas_efectivo),
    path('ajax/comisionasociada',ajax_comis_por_proy),
    path('ajax/nominaasociada',ajax_nomina_asociada),
    path('ajax/gttasociado',ajax_gtt_por_proy),
    path('ajax/interfazEgresos',impr_interf_egresos),
    path('ajax/interfacesnotas',ajax_impr_int_notas),
    path('ajax/accountsmovements',ajax_movimientos_cont),
    path('ajax/companies_accounts',ajax_cuentas_empresas),
    path('ajax/load_account_movements',ajax_cargar_movimiento_contable),
    path('ajax/load_bancary_movements',ajax_cargar_movimiento_banco),
    path('ajax/delete_movements',ajax_eliminar_linea_mvto_contable),
    path('ajax/delete_bancarymovements',ajax_eliminar_linea_mvto_bco),
    path('ajax/bancarymovements',ajax_movimientos_banco),
    path('ajax/reconciliate',ajax_nueva_conciliacion),
    path('ajax/accountingfinalbalance',ajax_saldos_conciliacion),
    path('ajax/printreconciliate',ajax_imprimir_conciliacion),
    path('ajax/reconciliationsdetail',ajax_lista_conciliaciones),
    path('ajax/deletereconciliation',ajax_borrar_conciliacion),
    path('ajax/matchsupplier',ajax_coincidir_tercero),
    path('ajax/infofacturas',ajax_info_facturas),
    path('ajax/recibir',ajax_recibir_factura),
    path('ajax/datafactura',ajax_data_factura),
    path('ajax/causarfactura',ajax_registrar_causacion),
    path('ajax/gethistory',ajax_history_facturas),
    path('ajax/payfactura',ajax_asociar_pago_radicado),
    path('ajax/getpaymentsfact',ajax_pagos_por_factura),
    path('ajax/listafacturas',ajax_lista_facturas),
    path('ajax/paymenttype',ajax_tipo_egreso),
    path('ajax/getmvtosplanilla',ajax_get_mvto_diario),
    path('ajax/printplanilla',ajax_print_planilla),
    path('ajax/listapagos',ajax_lista_pagos),
    path('ajax/getinfointerfaz',ajax_info_interfaces),
    path('ajax/registraranticipo',ajax_registrar_anticipo),
    path('ajax/getcomisiones',ajax_get_comisiones),
    path('ajax/printegreso',ajax_print_egreso),
    path('ajax/listanotherincomes',ajax_lista_otros_ingresos),
    path('ajax/printotherincomes',ajax_imprimir_otros_ingresos),
    path('ajax/newtransfer',ajax_transferencia),
    path('ajax/transferlist',ajax_lista_transf),
    path('ajax/anticiposlista',ajax_lista_ant),
    path('ajax/saldosiniciales',ajax_saldos_iniciales),
    path('ajax/otrospagosdetallados',ajax_otros_detalle),
    path('ajax/nominadetallada',ajax_nomina_detalle),
    path('ajax/interfazbanco',ajax_print_interfaz_banco),
    path('ajax/reclasificarnomina',ajax_reclasificar_nomina),
]

        
    
