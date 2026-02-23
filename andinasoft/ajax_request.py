from copy import copy
from django.http import HttpResponseRedirect, JsonResponse, HttpResponse, FileResponse
from django.shortcuts import render, redirect
from django.core import serializers
from django.conf import settings
from django.urls import path, include
from django.db.models import Avg, Max, Min, Sum, Q
from openpyxl.styles import Font
from andinasoft.utilities import Utilidades
from andinasoft.shared_models import (Inmuebles, Adjudicacion, Cargos_comisiones, CargosFijos, Vista_Adjudicacion,
                                      titulares_por_adj, ventas_nuevas, AsignacionComisiones)
from andinasoft.models import clientes, sidebar_pinned, asesores, empresas, cuentas_pagos
from accounting.models import Countries, States, Cities
from apis.dlocal.api import Dlocal
import openpyxl
from django.db.models.fields.related import ForeignKey, ManyToManyField
from django.db.models.fields.files import FileField, ImageField
from django.contrib.auth.decorators import login_required
import datetime
from docx import Document


def data_inmueble(request,proyecto):
    
    if request.method == 'GET':
        if request.is_ajax():
            inmueble = request.GET.get('inmueble')
            inmueble = inmueble.split('>')[1].split('<')[0]
            obj_inmueble = Inmuebles.objects.using(proyecto).filter(idinmueble=inmueble)
            json_inmueble = serializers.serialize('json',obj_inmueble)
            data = {
                'data_inmueble':json_inmueble,
            }
            if obj_inmueble[0].estado == 'Adjudicado' or obj_inmueble[0].estado == 'Reservado':
                obj_adj = Adjudicacion.objects.using(proyecto).filter(idinmueble=inmueble).exclude(estado='Desistido')
                if obj_adj.exists():
                    valor=obj_adj[0].pk
                    tipo='Adjudicacion'
                    titular = titulares_por_adj.objects.using(proyecto).get(adj=valor).titular1
                    url = f"/adjudicaciones/{proyecto}/{valor}"
                else: 
                    obj_reserva = ventas_nuevas.objects.using(proyecto).filter(inmueble=inmueble).exclude(estado='Adjudicado').exclude(estado='Anulado')
                    if obj_reserva.exists():
                        valor = obj_reserva[0].pk
                        tipo='Reserva'
                        titular = clientes.objects.get(pk=obj_reserva[0].id_t1).nombrecompleto
                        if obj_reserva[0].estado=='Pendiente':
                            url = f"/comercial/acciones_venta/{proyecto}/{valor}"
                        elif obj_reserva[0].estado=='Aprobado':
                            url = f"/operaciones/adjudicar_contrato/{proyecto}/{valor}"
                        else: url = '#'
                data['infopropietario']={
                    'nombretitular':titular,
                    'tipo':tipo,
                    'id':valor,
                    'url':url
                }
                    
            return JsonResponse(data)

def spanish_datatables(request):
    data ={
	"sProcessing":     "Procesando...",
	"sLengthMenu":     "Mostrar _MENU_ registros",
	"sZeroRecords":    "No se encontraron resultados",
	"sEmptyTable":     "Ningún dato disponible en esta tabla",
	"sInfo":           "Mostrando registros del _START_ al _END_ de un total de _TOTAL_ registros",
	"sInfoEmpty":      "Mostrando registros del 0 al 0 de un total de 0 registros",
	"sInfoFiltered":   "(filtrado de un total de _MAX_ registros)",
	"sInfoPostFix":    "",
	"sSearch":         "Buscar:",
	"sUrl":            "",
	"sInfoThousands":  ",",
	"sLoadingRecords": "Cargando...",
	"oPaginate": {
		"sFirst":    "Primero",
		"sLast":     "Último",
		"sNext":     "Siguiente",
		"sPrevious": "Anterior"
	},
	"oAria": {
		"sSortAscending":  ": Activar para ordenar la columna de manera ascendente",
		"sSortDescending": ": Activar para ordenar la columna de manera descendente"
	},
    "searchBuilder": {
        "add":'Nuevo filtro',
        "clearAll":'Borrar filtro',
        "deleteTitle":'Borrar',
        "data":'Columna',
        "logicAnd":'Y',
        "logicOr":'O',
        "condition": 'Condición',
        "value": 'Valor',
        "title": 'Filtro Avanzado',
    },
    'buttons': {
                'colvis': 'Columnas',
                'copy': '<i class="fa fa-files-o"></i>',
                'excel':'<i class="fa fa-file-excel-o"></i>',
                'pdf':'<i class="fa fa-file-pdf-o"></i>',
                'pageLength':{
                    '_': 'Ver %d',
                    -1: 'Ver todo'
                },
                'copyTitle': 'Copiado al portapapeles',
                'copySuccess': {
                    '_': '%d lineas copiadas',
                    1: '1 lineas copiadas'
                }
            }
}
    return JsonResponse(data)

def spansh_search_conditions(request):
    data ={
        'conditions':{
            "num": {
                '=': {
                    "conditionName": 'Igual a'
                },
                '!=': {
                    "conditionName": 'Diferente de'
                },
                '!null': {
                    "conditionName": 'No esta vacio'
                },
                '<': {
                    "conditionName": 'Menor que'
                },
                '<=': {
                    "conditionName": 'Menor o igual que'
                },
                '>': {
                    "conditionName": 'Mayor que'
                },
                '>=': {
                    "conditionName": 'Mayor o igual que'
                },
                'null': {
                    "conditionName": 'Vacio'
                },
                'between': {
                    "conditionName": 'Entre'
                },
                '!between': {
                    "conditionName": 'No esta entre'
                }
            },
            "string": {
                '=': {
                    "conditionName": 'Igual a'
                },
                '!=': {
                    "conditionName": 'Diferente de'
                },
                '!null': {
                    "conditionName": 'No esta vacio'
                },
                'null': {
                    "conditionName": 'Vacio'
                },
                'starts': {
                    "conditionName": 'Empieza con'
                },
                'ends': {
                    "conditionName": 'Termina en'
                },
                'contains': {
                    "conditionName": 'Contiene'
                },
            }
        }
    }
    return JsonResponse(data)

def sb_pinned(request):
    if request.method == 'GET':
        if request.is_ajax():
            pinned = request.GET.get('pinned')
            minimal = request.GET.get('minimal')
            if pinned:
                obj_pinned = sidebar_pinned.objects.filter(usuario=request.user)
                if obj_pinned.exists():
                    sb_user = sidebar_pinned.objects.get(usuario=request.user)
                    sb_user.pinned=pinned
                    sb_user.save()
                else:
                    sidebar_pinned.objects.create(
                        usuario=request.user,
                        pinned=pinned
                    )
                return JsonResponse({'errors':False})
            if minimal:
                is_min=0
                obj_sb=sidebar_pinned.objects.filter(usuario=request.user)
                if obj_sb.exists():
                    is_min = obj_sb[0].pinned
                return JsonResponse({'sb_status':is_min})

@login_required
def base_propietarios(request):
    if request.method =='GET':
        proyecto = request.GET.get('proyecto')
        book=openpyxl.Workbook()
        sheet=book.active
        encabezados=['Nombre','Apellido','Nombre Completo','Adjudicacion','Titular','Tipo','Ciudad','Direccion','Celular','Email','Proyecto','Fecha contrato','id']
        sheet.append(encabezados)
        i=2
        obj_adjs = Adjudicacion.objects.using(proyecto).exclude(estado='Desistido')
        base_clientes = []
        for adj in obj_adjs:
            tit1 = clientes.objects.filter(pk=adj.idtercero1)
            if tit1.exists():
                sheet.cell(i,1,tit1[0].nombres)
                sheet.cell(i,2,tit1[0].apellidos)
                sheet.cell(i,3,tit1[0].nombrecompleto)
                sheet.cell(i,4,adj.pk)
                sheet.cell(i,5,'Titular 1')
                sheet.cell(i,6,adj.origenventa)
                sheet.cell(i,7,tit1[0].ciudad)
                sheet.cell(i,8,tit1[0].domicilio)
                sheet.cell(i,9,tit1[0].celular1)
                sheet.cell(i,10,tit1[0].email)
                sheet.cell(i,11,proyecto)
                sheet.cell(i,12,adj.fechacontrato)
                sheet.cell(i,13,tit1[0].pk)
                
                i+=1
            tit2 = clientes.objects.filter(pk=adj.idtercero2)
            if tit2.exists() and tit2[0].pk!="":
                sheet.cell(i,1,tit2[0].nombres)
                sheet.cell(i,2,tit2[0].apellidos)
                sheet.cell(i,3,tit2[0].nombrecompleto)
                sheet.cell(i,4,adj.pk)
                sheet.cell(i,5,'Titular 2')
                sheet.cell(i,6,adj.origenventa)
                sheet.cell(i,7,tit2[0].ciudad)
                sheet.cell(i,8,tit2[0].domicilio)
                sheet.cell(i,9,tit2[0].celular1)
                sheet.cell(i,10,tit2[0].email)
                sheet.cell(i,11,proyecto)
                sheet.cell(i,12,adj.fechacontrato)
                sheet.cell(i,13,tit2[0].pk)
                i+=1
            tit3 = clientes.objects.filter(pk=adj.idtercero3)
            if tit3.exists() and tit3[0].pk!="":
                sheet.cell(i,1,tit3[0].nombres)
                sheet.cell(i,2,tit3[0].apellidos)
                sheet.cell(i,3,tit3[0].nombrecompleto)
                sheet.cell(i,4,adj.pk)
                sheet.cell(i,5,'Titular 3')
                sheet.cell(i,6,adj.origenventa)
                sheet.cell(i,7,tit3[0].ciudad)
                sheet.cell(i,8,tit3[0].domicilio)
                sheet.cell(i,9,tit3[0].celular1)
                sheet.cell(i,10,tit3[0].email)
                sheet.cell(i,11,proyecto)
                sheet.cell(i,12,adj.fechacontrato)
                sheet.cell(i,13,tit3[0].pk)
                i+=1
            tit4 = clientes.objects.filter(pk=adj.idtercero4)
            if tit4.exists() and (tit4[0].pk!=""):
                sheet.cell(i,1,tit4[0].nombres)
                sheet.cell(i,2,tit4[0].apellidos)
                sheet.cell(i,3,tit4[0].nombrecompleto)
                sheet.cell(i,4,adj.pk)
                sheet.cell(i,5,'Titular 4')
                sheet.cell(i,6,adj.origenventa)
                sheet.cell(i,7,tit4[0].ciudad)
                sheet.cell(i,8,tit4[0].domicilio)
                sheet.cell(i,9,tit4[0].celular1)
                sheet.cell(i,10,tit4[0].email)
                sheet.cell(i,11,proyecto)
                sheet.cell(i,12,adj.fechacontrato)
                sheet.cell(i,13,tit4[0].pk)
                i+=1
        filename = f'Baseclientesactivos_{proyecto}.xlsx'
        ruta=settings.MEDIA_ROOT+'/tmp/pdf/'+filename
        book.save(ruta)
        """ data = {
            'file':'/media/tmp/pdf/'+filename    
        }
        return JsonResponse(data) """
        
        return FileResponse(open(ruta,'rb'),as_attachment=True,filename=filename)

@login_required
def cargos_comisiones(request):
    if request.method == 'GET':
        tipo = request.GET.get('tipo')
        proyecto = request.GET.get('proyecto')
        if tipo == 'cargos':
            cargo = request.GET.get('cargo')
            obj_cargo = Cargos_comisiones.objects.using(proyecto).filter(pk=cargo)
            data = {
                'data':JSONRender(obj_cargo).render()
            }
            return JsonResponse(data)
        if tipo == 'escala_asignada':
            tipo_escala = request.GET.get('tipo_escala')
            contrato = request.GET.get('venta')
            if tipo_escala == 'variable':
                obj_escala = AsignacionComisiones.objects.using(proyecto).filter(idadjudicacion=contrato
                            ).exclude(Q(idcargo=10)|Q(idcargo=11)|Q(idcargo=111)|Q(idcargo=12)|Q(idcargo=40))
                data = []
                for carg in obj_escala:
                    data.append({
                        'gestor':{
                            'cedula':carg.idgestor,
                            'nombre':asesores.objects.get(pk=carg.idgestor).nombre
                        },
                        'cargo':{
                            'id':carg.idcargo,
                            'nombre':Cargos_comisiones.objects.using(proyecto).get(pk=carg.idcargo).nombrecargo
                        },
                        'comision':carg.comision
                    })
                
                data_to_show = {
                    'data':data
                }
                
                return JsonResponse(data_to_show)
                
                
    if request.method == 'POST':
        venta = request.POST.get('venta')
        proyecto = request.POST.get('proyecto')
        
        cargos = request.POST.getlist('cargo_comis')
        gestores = request.POST.getlist('id_asesor_comis')
        tasa = request.POST.getlist('comision')
        
        
        escala_asignada = AsignacionComisiones.objects.using(proyecto).filter(idadjudicacion=venta)
        print(escala_asignada)
        if escala_asignada.exists():
            for cargo in escala_asignada:
                cargo.delete()
        
        for i in range(len(cargos)):
            cargo = cargos[i].split(' - ')[0]
            gestor = gestores[i].split(' - ')[0]
            
            AsignacionComisiones.objects.using(proyecto).create(
                id_comision=f'{cargo}-{venta}',
                usuario = 'Activo',idadjudicacion = venta,
                comision = tasa[i], idgestor = gestores[i],
                idcargo = cargo
            )
        
        cargos_fijos=CargosFijos.objects.using(proyecto).all()
        for fijo in cargos_fijos:
            AsignacionComisiones.objects.using(proyecto).create(id_comision=f'{fijo.idcargo}-{venta}',
                                            idadjudicacion=venta,
                                            fecha=datetime.datetime.today(),
                                            idgestor=fijo.cc_fija,
                                            idcargo=fijo.idcargo,
                                            comision=fijo.porc_fijo,
                                            usuario='Activo',
                                            )
            
        data = {
            'title':'Listo',
            'msj':'La escala fue registrada con exito',
            'class':'toast-success'
        }
        
        return JsonResponse(data)

@login_required
def info_asesores(request):
    if request.method == 'GET':
        obj_asesores = asesores.objects.filter(estado='Activo')
        data = {
            'data':JSONRender(obj_asesores).render()
        }
        return JsonResponse(data)

def ajax_info_pais(request):
    if request.is_ajax():
        if request.method == 'GET':
            tipo = request.GET.get('tipo')
            paises = []
            estados = []
            ciudades = []
            if tipo == 'paises':
                """ file_paises = open(settings.STATIC_ROOT+'/json/countries/countries.json',encoding="utf8")
                json_file = json.loads(file_paises.read().encode().decode('utf-8-sig')) """
                
                obj_countries = Countries.objects.all().order_by('country_name')
                
                paises = serializers.serialize('json',obj_countries)
                    
            if tipo == 'estados':
                pais = request.GET.get('pais')
                """ file_estados = open(settings.STATIC_ROOT+'/json/countries/states.json',encoding="utf8")
                json_file = json.loads(file_estados.read().encode().decode('utf-8-sig')) """
                
                obj_estados = States.objects.filter(country= pais).order_by('state_name')
                
                estados = serializers.serialize('json',obj_estados)
            
            if tipo == 'ciudades':
                estado = request.GET.get('estado')
                """file_ciudades = open(settings.STATIC_ROOT+'/json/countries/cities.json',encoding="utf8")
                json_file = json.loads(file_ciudades.read().encode().decode('utf-8-sig'))
                
                for ciudad in json_file['cities']:
                    if ciudad.get('id_state')==int(estado):
                        ciudades.append((ciudad.get('id'),ciudad.get('name'))) """
                
                obj_ciudades = Cities.objects.filter(state=estado).order_by('city_name')
                
                ciudades = serializers.serialize('json',obj_ciudades)
                        
            
            data = {
                'paises':paises,
                'estados':estados,
                'ciudades':ciudades,
            }
            
            return JsonResponse(data)

def ajax_estructura_linderos(request):

    if request.method == 'POST':
        file = request.FILES.get('linderos')
        
        documento=openpyxl.load_workbook(file)
        hojas=documento.get_sheet_names()
        sheet=documento.get_sheet_by_name(hojas[0])
        
        row = 2
        empty = False
        
        filename = 'linderos_texto.docx'
        document = Document()
        
        while empty is False:
            manzana = sheet.cell(row,1).value
            
            if manzana is None or manzana == '':
                empty = True
                continue
            
            lote = sheet.cell(row,2).value
            area = sheet.cell(row,3).value
            norte = sheet.cell(row,4).value
            tipo_norte = sheet.cell(row,5).value
            colindante_norte = sheet.cell(row,6).value
            sur = sheet.cell(row,7).value
            tipo_sur = sheet.cell(row,8).value
            colindante_sur = sheet.cell(row,9).value
            este = sheet.cell(row,10).value
            tipo_este = sheet.cell(row,11).value
            colindante_este = sheet.cell(row,12).value
            oeste = sheet.cell(row,13).value
            tipo_oeste = sheet.cell(row,14).value
            colindante_oeste = sheet.cell(row,15).value
            coeficiente = sheet.cell(row,16).value
            
            tipo_colindante = lambda x: "regular" if x == "R" else "irregular"
            
            def colindantes(x:str):
                x = x.replace("LT","Lote "
                    ).replace("Víainterna","Vía interna"
                    ).replace("VíaInterna","Vía interna"
                    ).replace("Viainterna","Vía interna"
                    ).replace("ZonaVerde","Zona Verde"
                    ).replace("  "," ")
                return x
            
            p = Utilidades()
            m_text = p.numeros_letras(manzana,'Numero').upper()
            l_text = p.numeros_letras(lote,'Numero').upper()
            texto = f'MANZANA {m_text} ({manzana}) LOTE {l_text} ({lote}): Lote de terreno, con un área de {round(area,2)} m2, comprendido dentro de los siguientes linderos: '
            texto += f'Por el NORTE, en línea {tipo_colindante(tipo_norte)}, con {colindantes(colindante_norte)}, en una longitud de {round(norte,2)} metros. '
            texto += f'Por el SUR, en línea {tipo_colindante(tipo_sur)}, con {colindantes(colindante_sur)}, en una longitud de {round(sur,2)} metros. '
            texto += f'Por el ESTE, en línea {tipo_colindante(tipo_este)}, con {colindantes(colindante_este)}, en una longitud de {round(este,2)} metros. '
            texto += f'Por el OESTE, en línea {tipo_colindante(tipo_oeste)}, con {colindantes(colindante_oeste)}, en una longitud de {round(oeste,2)} metros. '
            texto += f'Este lote tiene un coeficiente de copropiedad de {coeficiente}.'
            
            
            document.add_paragraph(texto, style="List Number")
            
            row += 1
        
        document.save('static_media/tmp/'+filename)
        
        txt = f'¡Listo! Se está descargando el archivo'
        
        data = {
            'status': 'success',
            'message':txt,
            'href': f'/media/tmp/{filename}'
        }
        
        return JsonResponse(data)

def ajax_revision_ds(request):
    if request.method == 'POST':
        if request.is_ajax:
            file_ds = request.FILES.get('file_dian')
            file_siigo = request.FILES.get('file_siigo')
            
            terceros = {}
            
            wb_ds = openpyxl.load_workbook(file_ds)
            sheets_ds=wb_ds.get_sheet_names()
            sheet_ds=wb_ds.get_sheet_by_name(sheets_ds[0])
            
            wb_siigo = openpyxl.load_workbook(file_siigo)
            sheets_siigo=wb_siigo.get_sheet_names()
            sheet_siigo=wb_siigo.get_sheet_by_name(sheets_siigo[0])
            
            
            row = 8
            blank = False
            while blank is False:
                control = sheet_siigo.cell(row,1).value
                if control is None or control == '':
                    blank = True
                    break
                
                tercero = sheet_siigo.cell(row,7).value
                if tercero != '' and tercero != "             " and tercero != None:
                    tercero = int(tercero)
                    valor = sheet_siigo.cell(row,13).value
                    
                    if (tercero > 999999999 or tercero < 800000000) and valor != "                 ":
                        if tercero not in terceros:
                            terceros[tercero] = {
                                'tercero': sheet_siigo.cell(row,8).value,
                                'siigo':valor,
                                'dian':0,
                                'rows_siigo':[row,],
                                'rows_dian':[]
                            }
                        else:
                            terceros[tercero]['siigo'] += valor
                            terceros[tercero]['rows_siigo'].append(row)
                
                row += 1
                
            row = 2
            blank = False
            while blank is False:
                control = sheet_ds.cell(row,1).value
                if control is None or control == '':
                    blank = True
                    break
                if control == 'Documento soporte con no obligados':
                    tercero = sheet_ds.cell(row,9).value
                    tercero = int(tercero)
                    valor = sheet_ds.cell(row,14).value
                    
                    if (tercero > 999999999 or tercero < 800000000):
                        if tercero not in terceros:
                            terceros[tercero] = {
                                'tercero': sheet_ds.cell(row,10).value,
                                'siigo':0,
                                'dian':valor,
                                'rows_siigo':[],
                                'rows_dian':[row,]
                            }
                        else:
                            terceros[tercero]['dian'] += valor
                            terceros[tercero]['rows_dian'].append(row)
                
                row += 1
                
            wb = openpyxl.Workbook()
            sheet = wb.active
            
            sheet.append([
                'NIT','Nombre','SIIGO','DIAN','Diferencia','Observacion'
            ])
            
            sheet.column_dimensions['C'].number_format = 'Comma'
            sheet.column_dimensions['D'].number_format = 'Comma'
            sheet.column_dimensions['E'].number_format = 'Comma'
            
            sheet.row_dimensions[1].font = Font(bold=True)
            
            for tercero, tercero_data in terceros.items():
                diferencia = tercero_data['dian'] - tercero_data['siigo']
                
                if diferencia <0:
                    obs = 'Falta realizar 1 o mas documentos soportes en la DIAN'
                elif diferencia > 0:
                    obs = 'Hay documentos soportes realizados en la DIAN pero el valor causado en SIIGO es menor'
                else:
                    obs = 'OK'
                
                sheet.append([
                    tercero,
                    tercero_data['tercero'],
                    tercero_data['siigo'],
                    tercero_data['dian'],
                    diferencia,
                    obs
                ])
                
                if diferencia!=0:
                    st = wb.create_sheet(tercero_data['tercero'][:31].strip())
                    st.append([
                        'SIIGO',
                    ])
                    st.append([
                        'CUENTA DESCRIPCION','CUENTA','DESCRIPCION','SALDO INICIAL','COMPROBANTE','FECHA','NIT','NOMBRE','DESCRIPCION','INV','BASE','CC SC','DEBITOS','CREDITOS','SALDO MOV'
                    ])
                    
                    siigo_font  = copy(sheet_siigo.cell(7,1).font )
                    siigo_fill = copy(sheet_siigo.cell(7,1).fill )
                    dian_font  = copy(sheet_ds.cell(1,1).font )
                    dian_fill = copy(sheet_ds.cell(1,1).fill )
                    
                    
                    i=2
                    for row in tercero_data['rows_siigo']:
                        rw_lst = []
                        for col in range(1,16):
                            cell= sheet_siigo.cell(row,col).value
                            rw_lst.append(cell)
                        st.append(rw_lst)
                        i += 1
                    
                    st.append([])
                    st.append([
                        'DIAN',
                    ])
                    st.append([
                        'CUFE/CUDE','Folio','Prefijo','Fecha Emisión','Fecha Recepción','NIT Emisor','Nombre Emisor','NIT Receptor','Nombre Receptor','IVA','ICA','IPC','Total','Estado','Grupo',
                    ])
                    
                    j = i+3
                    
                    for row in tercero_data['rows_dian']:
                        rw_lst = []
                        for col in range(2,16):
                            cell= sheet_ds.cell(row,col).value
                            rw_lst.append(cell)
                        st.append(rw_lst)
                    
                    
                    for col in range(1,16):
                        cell = st.cell(2,col)
                        cell.font = siigo_font
                        cell.fill = siigo_fill
                        
                        cell = st.cell(i+3,col)
                        cell.font = dian_font
                        cell.fill = dian_fill

                    st.column_dimensions['A'].width = 42
                    st.column_dimensions['B'].width = 8.5
                    st.column_dimensions['C'].width = 15
                    st.column_dimensions['D'].width = 14
                    st.column_dimensions['E'].width = 20
                    st.column_dimensions['F'].width = 16.5
                    st.column_dimensions['G'].width = 14
                    st.column_dimensions['H'].width = 23
                    st.column_dimensions['I'].width = 38
                    st.column_dimensions['M'].width = 9.5
                    st.column_dimensions['M'].number_format = 'Comma'
            
            sheet.column_dimensions['A'].width = 12
            sheet.column_dimensions['B'].width = 35
            sheet.column_dimensions['C'].width = 13
            sheet.column_dimensions['D'].width = 13
            sheet.column_dimensions['E'].width = 13
            sheet.column_dimensions['F'].width = 80
                
                
                
            wb.save('static_media/tmp/conciliacion_ds.xlsx')               
            
            
            msj = 'Puedes descargar el archivo <strong><a href="/media/tmp/conciliacion_ds.xlsx">aquí</a></strong>' 
                
            data = {
                'msj':msj
            }
                
            return JsonResponse(data)

def active_prop_by_project(request):
    if request.is_ajax():
        if request.method == 'GET':
            proyecto = request.GET.get('proyecto')
            
            obj_adj = Vista_Adjudicacion.objects.using(proyecto).filter(
                Estado = 'Aprobado'
            ).exclude(Origen = 'Canje').order_by('Nombre')
            
                
            data = {
                'data':JSONRender(obj_adj).render()
            }
            
            return JsonResponse(data)
@login_required
def info_clientes(request):
    if request.method == 'GET':
        if request.is_ajax():
            q = request.GET.get('client')
            lista_clientes = clientes.objects.filter(
                Q(idTercero__icontains=q)|
                Q(nombrecompleto__icontains = q)
            ).exclude(pk__icontains='alt')
            
            data = {
                'results':JSONRender(lista_clientes).render()
            }
            
            return JsonResponse(data)
            
def dlocalpayments(request):
   if request.method == 'GET':
       amount = request.GET.get('amount')
       contrato = request.GET.get('contrato')
       print(amount,contrato)
       api = Dlocal()
       
       url = api.createpayment(amount = amount, contrato = contrato)
       
       data = {
           'url':url
       }
              
       return JsonResponse(data)


def info_adj_otrosi(request):
    project = request.GET.get('project') 
    adj =  request.GET.get('adj') 
    
    obj_adj = Adjudicacion.objects.using(project).get(pk=adj)
    
    
    
    return JsonResponse(obj_adj.data_otrosi())
          

class JSONRender():
    
    def __init__(self,queryset,reverse=False,query_functions=[]):
        self.queryset = queryset
        self.reverse = reverse
        self.query_functions = query_functions
    
    def render(self):
        object_dict = list()
        if self.queryset.count() > 0:
            fields = [f for f in self.queryset[0]._meta._get_fields(reverse=self.reverse)]
        else: 
            fields = []
        for obj in self.queryset:
            item = {}
            for field in fields:
                field_value = eval("obj."+field.name)     
                if type(field) == ForeignKey:
                    field_value = self.ForeingKeyRender(field,field_value)
                elif type(field) == ManyToManyField:
                    field_value = 'ManytoManyField'
                elif type(field) == FileField:
                    field_value = str(field_value)
                item[field.name] = field_value
            for func in self.query_functions:
                item[func] = eval('obj.'+func+'()')
            object_dict.append(item)
        return object_dict

    def ForeingKeyRender(self,fk,queryset_item):
        query_dict = {}
        field_list = fk.related_model._meta._get_fields(reverse = self.reverse)
        for field in field_list:
            if queryset_item == None:
                field_value = None
            else:
                field_value = eval(f'queryset_item.{field.name}')
                if type(field) == ForeignKey:
                    field_value = self.ForeingKeyRender(field,field_value)
                elif type(field) == ManyToManyField:
                    field_value = 'ManytoManyField'
                elif type(field) == FileField or ImageField:
                    field_value = str(field_value)
            query_dict[field.name] = field_value
        return query_dict


@login_required
def ajax_empresas(request):
    """Retorna lista de empresas"""
    empresas_list = empresas.objects.all().order_by('nombre')
    data = []
    for emp in empresas_list:
        data.append({
            'Nit': emp.Nit,
            'nombre': emp.nombre
        })
    return JsonResponse({'data': data})

@login_required
def ajax_cuentas_bancarias(request):
    """Retorna cuentas bancarias filtradas por empresa"""
    empresa_nit = request.GET.get('empresa')

    if empresa_nit:
        cuentas = cuentas_pagos.objects.filter(nit_empresa_id=empresa_nit).order_by('empresa')
    else:
        cuentas = cuentas_pagos.objects.all().order_by('empresa')

    data = []
    for cuenta in cuentas:
        data.append({
            'idcuenta': cuenta.idcuenta,
            'cuentabanco': cuenta.cuentabanco,
            'empresa': cuenta.empresa
        })
    return JsonResponse({'data': data})

urls = [
    path('datainmuebles/<proyecto>',data_inmueble),
    path('dataotrosi',info_adj_otrosi),
    path('datatable_spanish',spanish_datatables),
    path('searchbuilder',spansh_search_conditions),
    path('sb_pinned',sb_pinned),
    path('propietarybase',base_propietarios),
    path('comisions',cargos_comisiones),
    path('salesmen',info_asesores),
    path('activeclients',active_prop_by_project),
    path('datacountries',ajax_info_pais),
    path('dataclients',info_clientes),
    path('dlocalpayment',dlocalpayments),
    path('linderos',ajax_estructura_linderos),
    path('revisionds',ajax_revision_ds),
    path('empresas',ajax_empresas),
    path('cuentas_bancarias',ajax_cuentas_bancarias)
]