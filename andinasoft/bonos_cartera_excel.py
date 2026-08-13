"""
Generacion dinamica del Excel de bonos / override de cartera.

Reemplaza la dependencia de resources/excel_formats/Bonos_cartera.xlsx
(plantilla con gestores y proyectos hardcodeados).

Logica equivalente a la hoja Override del template historico:
- Por (gestor, proyecto): suma presupuesto, rcdo pptado, rcdo no pptado sin cashout
- Cumplimiento = rcdo_pptado / presupuesto
- Total bonificable = rcdo_pptado + rcdo_no_pptado_sin_cashout
- Bono override = total_bonificable * tasa_ov (default 0.2%)

En el detalle: Recaudo presupuestado = MIN(total, ppto) y siempre entra al bono.
No esperado = 0 si Cashout=Si (ese excedente ya pago bono aparte); si no, total-ppto.
El Override suma presupuestado (col M) + no esperado (col N).

La hoja Escalas se incluye como referencia parametrizable; el Override
historico usaba tasa fija (no consultaba Escalas).
"""
from __future__ import annotations

from collections import defaultdict
from decimal import Decimal, ROUND_HALF_UP
from itertools import groupby
from types import SimpleNamespace

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

from andinasoft.cartera_gestor_service import GESTOR_JURIDICO

ENCABEZADOS_DETALLE = [
    'Adjudicación',
    'Cliente',
    'Estado',
    'Origen',
    'Venta del mes',
    'Tipo de cartera',
    'Edad',
    'Cuota del mes',
    'Recaudo del mes',
    'Cuotas vencidas',
    'Recaudo vencido',
    'Presupuesto total',
    'Recaudo presupuestado',
    'No esperado',
    'Recaudo total',
    'Gestor',
    'Cashout',
]

ENCABEZADOS_OVERRIDE = [
    'Gestor',
    'Proyecto',
    'Presupuesto',
    'Recaudo presupuestado',
    'Cumplimiento',
    'No esperado (sin cashout)',
    'Recaudo bonificable',
    'Tasa OV',
    'Bono OV',
    'Bono por escala (ref.)',
]

MESES = (
    '', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
    'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre',
)

# Paleta del reporte (lectura en pantalla e impresión).
COLOR_TITULO = '1F4E79'
COLOR_HEADER = '1F4E79'
COLOR_HEADER_FONT = 'FFFFFF'
COLOR_SUBTOTAL = 'D6EAF8'
COLOR_TOTAL = '1F4E79'
COLOR_ZEBRA = 'F7F9FC'
COLOR_BONO = 'E8F8F0'
COLOR_BORDE = 'BFBFBF'
COLOR_ESCALAS = 'FFF8E1'

FMT_MONEY = '"$"#,##0.00'
FMT_PCT = '0.00%'

THIN = Border(
    left=Side(style='thin', color=COLOR_BORDE),
    right=Side(style='thin', color=COLOR_BORDE),
    top=Side(style='thin', color=COLOR_BORDE),
    bottom=Side(style='thin', color=COLOR_BORDE),
)
ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal='right', vertical='center')

# Escala historica (hoja Escalas del template). Cumplimiento minimo -> % bono.
ESCALAS_DEFAULT = (
    (Decimal('0.40'), Decimal('0.0010')),
    (Decimal('0.60'), Decimal('0.0012')),
    (Decimal('0.90'), Decimal('0.0014')),
    (Decimal('1.00'), Decimal('0.0016')),
)

TASA_OV_DEFAULT = Decimal('0.002')  # 0.2% como en Override!H
ZERO = Decimal('0')


def _dec(value) -> Decimal:
    if value is None or value == '':
        return ZERO
    if isinstance(value, Decimal):
        return value
    try:
        return Decimal(str(value))
    except Exception:
        return ZERO


def _money(value: Decimal) -> Decimal:
    return value.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)


def _pct(value: Decimal) -> Decimal:
    return value.quantize(Decimal('0.0001'), rounding=ROUND_HALF_UP)


def normalizar_gestor(asesor) -> str:
    nombre = (asesor or '').strip()
    return nombre.upper() if nombre else 'SIN GESTOR'


def es_cashout(valor) -> bool:
    """True si la fila es cashout: el no esperado no entra al bono; el presupuestado sí."""
    if valor is None:
        return False
    text = str(valor).strip().lower()
    if not text or text in ('no', 'n', '0', 'false'):
        return False
    return text in ('si', 'sí', 'yes', 'y', '1', 'true', 'cashout')


def bono_por_escala(cumplimiento: Decimal, escalas=ESCALAS_DEFAULT) -> Decimal:
    """Devuelve el % de bono segun la escala (mayor umbral alcanzado)."""
    tasa = ZERO
    for minimo, bono in escalas:
        if cumplimiento >= minimo:
            tasa = bono
    return tasa


def agregar_fila_override(acc, proyecto: str, fila, *, excluir_juridico=True):
    """Acumula montos de una fila de informe en el dict (gestor, proyecto)."""
    gestor = normalizar_gestor(getattr(fila, 'asesor', None))
    if excluir_juridico and gestor == GESTOR_JURIDICO:
        return
    key = (gestor, proyecto)
    bucket = acc[key]
    bucket['presupuesto'] += _dec(getattr(fila, 'presupuesto', 0))
    bucket['rcdo_pptado'] += _dec(getattr(fila, 'recaudo_pptado', 0))
    no_ppt = _dec(getattr(fila, 'recaudo_nopptado', 0))
    cashout = getattr(fila, 'cashout', None)
    if not es_cashout(cashout):
        bucket['rcdo_nopptado_sin_cashout'] += no_ppt
    bucket['filas'] += 1


def calcular_override_row(gestor, proyecto, montos, *, tasa_ov=TASA_OV_DEFAULT, escalas=ESCALAS_DEFAULT):
    ppto = montos.get('presupuesto', ZERO)
    pptado = montos.get('rcdo_pptado', ZERO)
    noppt = montos.get('rcdo_nopptado_sin_cashout', ZERO)
    cumplimiento = (pptado / ppto) if ppto > 0 else ZERO
    total_bonificable = pptado + noppt
    bono_ov = total_bonificable * tasa_ov
    bono_escala = total_bonificable * bono_por_escala(cumplimiento, escalas)
    return SimpleNamespace(
        gestor=gestor,
        proyecto=proyecto,
        presupuesto=_money(ppto),
        rcdo_pptado=_money(pptado),
        cumplimiento=_pct(cumplimiento),
        rcdo_nopptado_sin_cashout=_money(noppt),
        total_bonificable=_money(total_bonificable),
        tasa_ov=tasa_ov,
        bono_override=_money(bono_ov),
        tasa_escala=bono_por_escala(cumplimiento, escalas),
        bono_escala=_money(bono_escala),
        filas=montos.get('filas', 0),
    )


def construir_override(por_gestor_proyecto, *, tasa_ov=TASA_OV_DEFAULT, escalas=ESCALAS_DEFAULT):
    """
    por_gestor_proyecto: dict[(gestor, proyecto)] -> montos
    Returns: lista de filas detalle + lista de totales por gestor.
    """
    filas = []
    for (gestor, proyecto) in sorted(por_gestor_proyecto.keys(), key=lambda k: (k[0], k[1])):
        filas.append(
            calcular_override_row(
                gestor,
                proyecto,
                por_gestor_proyecto[(gestor, proyecto)],
                tasa_ov=tasa_ov,
                escalas=escalas,
            )
        )
    totales = []
    by_gestor = defaultdict(lambda: {
        'presupuesto': ZERO,
        'rcdo_pptado': ZERO,
        'rcdo_nopptado_sin_cashout': ZERO,
        'total_bonificable': ZERO,
        'bono_override': ZERO,
        'bono_escala': ZERO,
    })
    for row in filas:
        g = by_gestor[row.gestor]
        g['presupuesto'] += row.presupuesto
        g['rcdo_pptado'] += row.rcdo_pptado
        g['rcdo_nopptado_sin_cashout'] += row.rcdo_nopptado_sin_cashout
        g['total_bonificable'] += row.total_bonificable
        g['bono_override'] += row.bono_override
        g['bono_escala'] += row.bono_escala
    for gestor in sorted(by_gestor.keys()):
        g = by_gestor[gestor]
        cumplimiento = (g['rcdo_pptado'] / g['presupuesto']) if g['presupuesto'] > 0 else ZERO
        totales.append(
            SimpleNamespace(
                gestor=gestor,
                presupuesto=_money(g['presupuesto']),
                rcdo_pptado=_money(g['rcdo_pptado']),
                cumplimiento=_pct(cumplimiento),
                rcdo_nopptado_sin_cashout=_money(g['rcdo_nopptado_sin_cashout']),
                total_bonificable=_money(g['total_bonificable']),
                bono_override=_money(g['bono_override']),
                bono_escala=_money(g['bono_escala']),
            )
        )
    return filas, totales


def etiqueta_periodo(periodo) -> str:
    if not periodo:
        return ''
    s = str(periodo).strip()
    if len(s) == 6 and s.isdigit():
        mes = int(s[4:6])
        if 1 <= mes <= 12:
            return f'{MESES[mes]} {s[:4]}'
    return s


def _fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)


def _font_header():
    return Font(name='Calibri', bold=True, color=COLOR_HEADER_FONT, size=11)


def _font_title():
    return Font(name='Calibri', bold=True, color=COLOR_TITULO, size=16)


def _font_subtitle():
    return Font(name='Calibri', italic=True, color='5D6D7E', size=11)


def _font_normal(bold=False, color='000000', size=10):
    return Font(name='Calibri', bold=bold, color=color, size=size)


def _apply_print_setup(ws, *, landscape=True, freeze='A2', title_rows='1:1'):
    ws.page_setup.orientation = 'landscape' if landscape else 'portrait'
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.paperSize = 1  # Letter
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.horizontalCentered = True
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.6, bottom=0.6, header=0.3, footer=0.3)
    ws.print_title_rows = title_rows
    ws.freeze_panes = freeze
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 100
    ws.page_setup.horizontalDpi = 300
    ws.oddFooter.left.text = '&A'
    ws.oddFooter.right.text = 'Página &P de &N'


def _style_header_row(ws, row, ncol, fill_color=COLOR_HEADER):
    fill = _fill(fill_color)
    font = _font_header()
    for c in range(1, ncol + 1):
        cell = ws.cell(row, c)
        cell.font = font
        cell.fill = fill
        cell.alignment = ALIGN_CENTER
        cell.border = THIN
    ws.row_dimensions[row].height = 32


def _style_data_cell(cell, *, money=False, pct=False, center=False, zebra=False, highlight=False, bold=False):
    cell.font = _font_normal(bold=bold)
    cell.border = THIN
    if money:
        cell.number_format = FMT_MONEY
        cell.alignment = ALIGN_RIGHT
    elif pct:
        cell.number_format = FMT_PCT
        cell.alignment = ALIGN_CENTER
    elif center:
        cell.alignment = ALIGN_CENTER
    else:
        cell.alignment = ALIGN_LEFT
    if highlight:
        cell.fill = _fill(COLOR_BONO)
    elif zebra:
        cell.fill = _fill(COLOR_ZEBRA)


def _set_widths(ws, widths):
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def _cashout_label(valor) -> str:
    if valor is None or str(valor).strip() == '':
        return 'No'
    return 'Sí' if es_cashout(valor) else 'No'


def _excel_sheet_ref(nombre: str) -> str:
    title = (nombre or '')[:31].replace("'", "''")
    return f"'{title}'"


def _formula_pptado(row: int) -> str:
    """Recaudo esperado: siempre entra al bono, aunque haya cashout."""
    return f'=MIN(O{row},L{row})'


def _formula_nopptado(row: int) -> str:
    """Cashout=Si: solo el no esperado queda en 0; el presupuestado se paga igual."""
    return (
        f'=IF(OR(UPPER(TRIM(Q{row}))="SI",UPPER(TRIM(Q{row}))="SÍ"),'
        f'0,MAX(0,O{row}-L{row}))'
    )


def _formula_sumif_col(proyecto: str, col_letter: str, override_row: int) -> str:
    sheet = _excel_sheet_ref(proyecto)
    return (
        f'=SUMIF({sheet}!P:P,A{override_row},{sheet}!{col_letter}:{col_letter})'
    )


def _formula_sum_rows(col_letter: str, rows) -> str:
    refs = ','.join(f'{col_letter}{r}' for r in rows)
    return f'=SUM({refs})' if refs else '=0'


def _write_override_money_row(ws, row_i, values, *, zebra=False, subtotal=False, grand=False, highlight_bono=True):
    """values: tuple aligned with ENCABEZADOS_OVERRIDE."""
    ncol = len(ENCABEZADOS_OVERRIDE)
    money_cols = {3, 4, 6, 7, 9, 10}
    pct_cols = {5, 8}
    if grand:
        fill = _fill(COLOR_TOTAL)
        font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    elif subtotal:
        fill = _fill(COLOR_SUBTOTAL)
        font = _font_normal(bold=True, size=10)
    else:
        fill = None
        font = _font_normal(bold=False, size=10)

    for c in range(1, ncol + 1):
        cell = ws.cell(row_i, c, values[c - 1])
        cell.font = font
        cell.border = THIN
        if c in money_cols:
            cell.number_format = FMT_MONEY
            cell.alignment = ALIGN_RIGHT
        elif c in pct_cols:
            cell.number_format = FMT_PCT
            cell.alignment = ALIGN_CENTER
        else:
            cell.alignment = ALIGN_LEFT if c == 1 else ALIGN_CENTER
        if fill is not None:
            cell.fill = fill
        elif highlight_bono and c == 9:
            cell.fill = _fill(COLOR_BONO)
        elif zebra:
            cell.fill = _fill(COLOR_ZEBRA)
    ws.row_dimensions[row_i].height = 18


def _write_detalle_sheet(wb, proyecto: str, filas):
    title = proyecto[:31]
    if title in wb.sheetnames:
        wb.remove(wb[title])
    ws = wb.create_sheet(title)
    ws.sheet_properties.tabColor = '5DADE2'

    ncol = len(ENCABEZADOS_DETALLE)
    for j, h in enumerate(ENCABEZADOS_DETALLE, start=1):
        ws.cell(1, j, h)
    _style_header_row(ws, 1, ncol)

    money_cols = {8, 9, 10, 11, 12, 13, 14, 15}
    for i, fila in enumerate(filas, start=2):
        zebra = (i % 2) == 0
        values = [
            getattr(fila, 'pk', '') or '',
            getattr(fila, 'cliente', '') or '',
            getattr(fila, 'estado', '') or '',
            getattr(fila, 'origen', '') or '',
            getattr(fila, 'venta_mes', '') or '',
            getattr(fila, 'tipocartera', '') or '',
            getattr(fila, 'edad', '') or '',
            float(_dec(getattr(fila, 'ppto_mes', 0))),
            float(_dec(getattr(fila, 'recaudo_mes', 0))),
            float(_dec(getattr(fila, 'ppto_vencido', 0))),
            float(_dec(getattr(fila, 'recaudo_vencido', 0))),
            float(_dec(getattr(fila, 'presupuesto', 0))),
            _formula_pptado(i),
            _formula_nopptado(i),
            float(_dec(getattr(fila, 'recaudo_total', 0))),
            normalizar_gestor(getattr(fila, 'asesor', None)),
            _cashout_label(getattr(fila, 'cashout', None)),
        ]
        for c, val in enumerate(values, start=1):
            cell = ws.cell(i, c, val)
            _style_data_cell(
                cell,
                money=c in money_cols,
                center=c in (1, 3, 4, 5, 6, 7, 17),
                zebra=zebra,
            )
        ws.row_dimensions[i].height = 16

    last_row = max(1, 1 + len(filas))
    ws.auto_filter.ref = f'A1:{get_column_letter(ncol)}{last_row}'
    _set_widths(ws, [16, 32, 14, 14, 14, 16, 12, 14, 16, 16, 16, 18, 20, 22, 16, 22, 12])
    _apply_print_setup(ws, freeze='A2', title_rows='1:1')
    return ws


def _write_escalas_sheet(wb, escalas=ESCALAS_DEFAULT):
    if 'Escalas' in wb.sheetnames:
        del wb['Escalas']
    ws = wb.create_sheet('Escalas', 1)
    ws.sheet_properties.tabColor = 'F4D03F'

    ws.merge_cells('A1:B1')
    title = ws.cell(1, 1, 'Escalas de bono por cumplimiento')
    title.font = _font_title()
    title.alignment = ALIGN_LEFT
    ws.row_dimensions[1].height = 24

    ws.merge_cells('A2:B2')
    sub = ws.cell(2, 1, 'Tabla de referencia según el porcentaje de cumplimiento')
    sub.font = _font_subtitle()
    ws.row_dimensions[2].height = 18

    headers = ('Cumplimiento mínimo', 'Porcentaje de bono')
    for j, h in enumerate(headers, start=1):
        ws.cell(4, j, h)
    _style_header_row(ws, 4, 2)

    for i, (minimo, bono) in enumerate(escalas, start=5):
        zebra = (i % 2) == 1
        c1 = ws.cell(i, 1, float(minimo))
        c2 = ws.cell(i, 2, float(bono))
        _style_data_cell(c1, pct=True, zebra=zebra)
        _style_data_cell(c2, pct=True, zebra=zebra, highlight=True)
        ws.row_dimensions[i].height = 18

    _set_widths(ws, [28, 24])
    _apply_print_setup(ws, landscape=False, freeze='A5', title_rows='4:4')
    return ws


def _write_override_sheet(wb, filas_ov, totales, *, tasa_ov=TASA_OV_DEFAULT, periodo=None):
    if 'Override' in wb.sheetnames:
        del wb['Override']
    ws = wb.create_sheet('Override', 0)
    ws.sheet_properties.tabColor = '1E8449'

    ncol = len(ENCABEZADOS_OVERRIDE)
    last_col = get_column_letter(ncol)

    ws.merge_cells(f'A1:{last_col}1')
    titulo = 'Reporte Override (OV) de cartera'
    etiqueta = etiqueta_periodo(periodo)
    if etiqueta:
        titulo = f'{titulo}  —  {etiqueta}'
    cell_t = ws.cell(1, 1, titulo)
    cell_t.font = _font_title()
    cell_t.alignment = ALIGN_LEFT
    ws.row_dimensions[1].height = 26

    ws.merge_cells(f'A2:{last_col}2')
    cell_s = ws.cell(
        2,
        1,
        f'Tasa OV: {float(tasa_ov):.2%}     ·     Cashout=Sí: se excluye solo el no esperado; el presupuestado sí entra al bono',
    )
    cell_s.font = _font_subtitle()
    ws.row_dimensions[2].height = 18

    header_row = 4
    for j, h in enumerate(ENCABEZADOS_OVERRIDE, start=1):
        ws.cell(header_row, j, h)
    _style_header_row(ws, header_row, ncol)

    row_i = header_row + 1
    zebra_on = False
    totales_map = {t.gestor: t for t in totales}

    if not filas_ov:
        ws.merge_cells(f'A{row_i}:{last_col}{row_i}')
        empty = ws.cell(row_i, 1, 'No hay datos de override para este periodo.')
        empty.font = _font_subtitle()
        empty.alignment = ALIGN_CENTER
        row_i += 1
    else:
        subtotal_rows = []
        for gestor, grupo in groupby(filas_ov, key=lambda r: r.gestor):
            data_rows = []
            for row in grupo:
                zebra_on = not zebra_on
                _write_override_money_row(
                    ws,
                    row_i,
                    (
                        row.gestor,
                        row.proyecto,
                        float(row.presupuesto),
                        _formula_sumif_col(row.proyecto, 'M', row_i),
                        f'=IF(C{row_i}=0,0,D{row_i}/C{row_i})',
                        _formula_sumif_col(row.proyecto, 'N', row_i),
                        f'=D{row_i}+F{row_i}',
                        float(tasa_ov),
                        f'=G{row_i}*H{row_i}',
                        float(row.bono_escala),
                    ),
                    zebra=zebra_on,
                )
                data_rows.append(row_i)
                row_i += 1
            tot = totales_map.get(gestor)
            if tot and data_rows:
                _write_override_money_row(
                    ws,
                    row_i,
                    (
                        f'TOTAL  {gestor}',
                        '',
                        _formula_sum_rows('C', data_rows),
                        _formula_sum_rows('D', data_rows),
                        f'=IF(C{row_i}=0,0,D{row_i}/C{row_i})',
                        _formula_sum_rows('F', data_rows),
                        f'=D{row_i}+F{row_i}',
                        float(tasa_ov),
                        f'=G{row_i}*H{row_i}',
                        _formula_sum_rows('J', data_rows),
                    ),
                    subtotal=True,
                )
                subtotal_rows.append(row_i)
                row_i += 1
            row_i += 1  # espacio entre gestores

        if subtotal_rows:
            _write_override_money_row(
                ws,
                row_i,
                (
                    'TOTAL GENERAL',
                    '',
                    _formula_sum_rows('C', subtotal_rows),
                    _formula_sum_rows('D', subtotal_rows),
                    f'=IF(C{row_i}=0,0,D{row_i}/C{row_i})',
                    _formula_sum_rows('F', subtotal_rows),
                    f'=D{row_i}+F{row_i}',
                    float(tasa_ov),
                    f'=G{row_i}*H{row_i}',
                    _formula_sum_rows('J', subtotal_rows),
                ),
                grand=True,
            )
            row_i += 1

    _set_widths(ws, [28, 22, 16, 22, 14, 24, 20, 12, 14, 16])
    _apply_print_setup(ws, freeze='A5', title_rows='4:4')
    return ws


def _write_errores_sheet(wb, errores):
    if 'Errores' in wb.sheetnames:
        del wb['Errores']
    ws = wb.create_sheet('Errores')
    ws.sheet_properties.tabColor = 'C0392B'
    ws.merge_cells('A1:B1')
    t = ws.cell(1, 1, 'Proyectos que no se pudieron incluir en el reporte')
    t.font = _font_title()
    ws.row_dimensions[1].height = 24
    for j, h in enumerate(('Proyecto', 'Detalle del error'), start=1):
        ws.cell(3, j, h)
    _style_header_row(ws, 3, 2)
    for i, (proyecto, msg) in enumerate(errores, start=4):
        c1 = ws.cell(i, 1, proyecto)
        c2 = ws.cell(i, 2, (msg or '')[:500])
        _style_data_cell(c1, zebra=(i % 2) == 0)
        _style_data_cell(c2, zebra=(i % 2) == 0)
    _set_widths(ws, [28, 80])
    _apply_print_setup(ws, landscape=False, freeze='A4', title_rows='3:3')
    return ws


def generar_libro_bonos(
    datos_por_proyecto: dict,
    *,
    tasa_ov=TASA_OV_DEFAULT,
    escalas=ESCALAS_DEFAULT,
    excluir_juridico=True,
    errores=None,
    periodo=None,
):
    """
    datos_por_proyecto: {nombre_proyecto: lista de filas informe_cartera_rows}
    errores: lista opcional [(proyecto, mensaje)]
    periodo: YYYYMM para el titulo del Override
    """
    wb = openpyxl.Workbook()
    default = wb.active
    wb.remove(default)

    acc = defaultdict(lambda: {
        'presupuesto': ZERO,
        'rcdo_pptado': ZERO,
        'rcdo_nopptado_sin_cashout': ZERO,
        'filas': 0,
    })

    for proyecto in sorted(datos_por_proyecto.keys()):
        filas = datos_por_proyecto[proyecto] or []
        _write_detalle_sheet(wb, proyecto, filas)
        for fila in filas:
            agregar_fila_override(acc, proyecto, fila, excluir_juridico=excluir_juridico)

    filas_ov, totales = construir_override(acc, tasa_ov=tasa_ov, escalas=escalas)
    _write_override_sheet(wb, filas_ov, totales, tasa_ov=tasa_ov, periodo=periodo)
    _write_escalas_sheet(wb, escalas)

    if errores:
        _write_errores_sheet(wb, errores)

    return wb

