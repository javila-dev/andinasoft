import decimal
import csv
import io
from django.shortcuts import render, redirect
from django.template import RequestContext
from django.http import HttpResponseRedirect, JsonResponse, HttpResponse, FileResponse, QueryDict
from django.contrib.admin.views.decorators import staff_member_required
from django.core.exceptions import PermissionDenied
from django.contrib.auth.decorators import permission_required, user_passes_test,login_required
from django.contrib.auth.models import User
from django.db.models import Avg, Max, Min, Sum, Q, OuterRef, Subquery, DecimalField, Exists, Value
from django.db import connections, transaction
from django.db.utils import InternalError as DBInternalError
from django.db.models.functions import Coalesce
from django.core import serializers
from django.forms import ValidationError
from django.conf import settings
from django.template.loader import get_template
from django.core.mail import EmailMultiAlternatives
from django.core.files.storage import default_storage
from django.views.decorators.csrf import csrf_protect, csrf_exempt
from django.views.decorators.http import require_http_methods
from django.urls import path
from andina.decorators import group_perm_required
from api_auth.decorators import api_token_auth
from andinasoft.ajax_request import JSONRender
from andinasoft.forms import form_buscar_reestr, form_nueva_reestr, registrar_asesor, detalle_adjudicacion as form_detalle_adj,documentos_contrato as form_docs_contratos
from andinasoft.forms import form_lista_proyectos as projects, form_nuevo_cliente_PJ
from andinasoft.forms import form_info_titular, form_nuevo_recibo, form_lista_recaudos, form_radicar_recibo
from andinasoft.forms import form_nueva_venta, form_modificar_recibo, form_seguimiento, form_reestructuracion
from andinasoft.forms import form_escala_comision, form_recaudo_noradicado, form_revision_op, form_retiro_asesor
from andinasoft.forms import form_procesabilidad, form_inv_admin, form_ver_ppto, form_detalle_comisiones, form_nuevo_cliente
from andinasoft.forms import form_radicar_factura, form_causar, form_pagar, form_nuevo_recaudoNR, form_int_comi_banco
from andinasoft.models import ( asesores, bk_bfchangeplan, bk_planpagos, bk_recaudodetallado,clientes,Facturas, reestructuraciones_otrosi,
                                timeline_radicados, Pagos, cuentas_pagos, Usuarios_Proyectos, Gtt, Detalle_gtt,
                                GastosInforme, CuentasAsociadas, ItemsInforme, CentrosCostos, proyectos, entidades_bancarias, 
                                Profiles, Avatars, Countries, States, Cities, sagrilaft_info, CIUU)
from andinasoft.shared_models import Adjudicacion, Vista_Adjudicacion, documentos_contratos, fractales_ventas,saldos_adj,titulares_por_adj, fractales_ventas
from andinasoft.shared_models import Recaudos, consecutivos, Recaudos_general, AsignacionComisiones, CargosFijos, InfoCartera, Cargos_comisiones
from andinasoft.shared_models import timeline,seguimientos, Inmuebles, ventas_nuevas, RecaudosNoradicados, Pagocomision
from andinasoft.shared_models import VerificacionOperaciones, DescuentosCondicionados, PlanPagos, DescuentosCondicionados, formas_pago
from andinasoft.shared_models import Promesas, PresupuestoCartera, Parametros_Operaciones, EntregaManzanas, Pqrs
from andinasoft.handlers_functions import upload_docs_asesores, upload_docs_contratos, upload_docs_radicados, upload_docs
from andinasoft.handlers_functions import aplicar_pago, respuesta_reestructuracion, envio_notificacion, envio_email_template
from andinasoft.handlers_functions import cargar_gastos_informe
from andinasoft.create_pdf import GenerarPDF
from andinasoft.utilities import Utilidades, pdf_gen, pdf_gen_weasy, calcular_tabla_amortizacion
from andinasoft.passes_test import perms_test
from buildingcontrol import models as building_model
from accounting import forms as forms_accounting
from finance.models import recibos_internos
from decimal import Decimal
from dateutil.relativedelta import relativedelta
from django.utils.dateparse import parse_date
from uuid import uuid4
from django.utils.html import strip_tags
import traceback
import openpyxl
import datetime
import json
import os
import math
import calendar
import random
import string
from urllib.parse import urlparse
from openpyxl.styles import Font, Color, Alignment, Border, Side, colors, PatternFill
from copy import copy


# Create your views here.
tasas = [
    {'name':'0%','value':'0.0000'},
    {'name':'0.8%','value':'0.0080'},
    {'name':'0.99%','value':'0.0099'},
    {'name':'1.5%','value':'0.0150'}
]


def _move_storage_prefix(prefix_src, prefix_dst):
    src = prefix_src.rstrip("/") + "/"
    dst = prefix_dst.rstrip("/") + "/"
    pending = [src]

    while pending:
        current_prefix = pending.pop()
        current_dst = current_prefix.replace(src, dst, 1)
        dirs, files = default_storage.listdir(current_prefix)

        for dirname in dirs:
            pending.append(f"{current_prefix}{dirname}/")

        for filename in files:
            src_path = f"{current_prefix}{filename}"
            dst_path = f"{current_dst}{filename}"
            with default_storage.open(src_path, "rb") as file_obj:
                if default_storage.exists(dst_path):
                    default_storage.delete(dst_path)
                default_storage.save(dst_path, file_obj)
            default_storage.delete(src_path)


def _to_storage_key(path):
    normalized = os.path.normpath(path)
    media_root = os.path.normpath(settings.MEDIA_ROOT)
    if normalized.startswith(media_root):
        return os.path.relpath(normalized, media_root).replace("\\", "/")
    return normalized.lstrip("/").replace("\\", "/")


def _normalize_soporte_key(soporte_path):
    if not soporte_path:
        return None

    media_url = getattr(settings, 'MEDIA_URL', '/media/') or '/media/'
    candidate = str(soporte_path).strip()
    if not candidate:
        return None

    if candidate.startswith('http://') or candidate.startswith('https://'):
        parsed = urlparse(candidate)
        candidate = parsed.path or candidate

    if media_url and candidate.startswith(media_url):
        candidate = candidate[len(media_url):]
    elif candidate.startswith('/media/'):
        candidate = candidate[len('/media/'):]

    candidate = candidate.lstrip('/').replace("\\", "/")

    if candidate.startswith('static_media/'):
        candidate = candidate[len('static_media/'):]

    return candidate or None


def _resolve_soporte_url(soporte_path):
    if not soporte_path:
        return None
    raw_value = str(soporte_path).strip()
    if not raw_value:
        return None
    if raw_value.startswith('http://') or raw_value.startswith('https://'):
        return raw_value

    key = _normalize_soporte_key(raw_value)
    if not key:
        return None
    try:
        return default_storage.url(key)
    except Exception:
        media_url = getattr(settings, 'MEDIA_URL', '/media/') or '/media/'
        return f"{media_url.rstrip('/')}/{key.lstrip('/')}"


def _save_workbook_with_dirs(book, path):
    directory = os.path.dirname(path)
    if directory:
        os.makedirs(directory, exist_ok=True)
    book.save(path)


INVENTARIO_IMPORT_COLUMNS = [
    'idinmueble',
    'etapa',
    'manzananumero',
    'lotenumero',
    'matricula',
    'vrmetrocuadrado',
    'estado',
    'finobra',
    'areaprivada',
    'areaconstruida',
    'area_lt',
    'area_mz',
    'norte',
    'lindero_norte',
    'colindante_norte',
    'sur',
    'lindero_sur',
    'colindante_sur',
    'este',
    'lindero_este',
    'colidante_este',
    'oeste',
    'lindero_oeste',
    'colindante_oeste',
    'fac_valor_via_principal',
    'fac_valor_area_social',
    'fac_valor_esquinero',
    'obsbloqueo',
    'meses',
]


def _normalize_header(value):
    return str(value or '').strip().lower()


def _clean_value(value):
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value if value != '' else None
    return value


def _parse_decimal(value):
    value = _clean_value(value)
    if value is None:
        return None
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))
    value = str(value).replace(' ', '')
    if ',' in value and '.' in value:
        if value.rfind(',') > value.rfind('.'):
            value = value.replace('.', '').replace(',', '.')
        else:
            value = value.replace(',', '')
    else:
        value = value.replace(',', '.')
    return Decimal(value)


def _parse_int(value):
    value = _clean_value(value)
    if value is None:
        return None
    if isinstance(value, int):
        return value
    return int(Decimal(str(value)))


def _parse_datetime_value(value):
    value = _clean_value(value)
    if value is None:
        return None
    if isinstance(value, datetime.datetime):
        return value
    if isinstance(value, datetime.date):
        return datetime.datetime.combine(value, datetime.time.min)
    value = str(value)
    parsed = parse_date(value)
    if parsed:
        return datetime.datetime.combine(parsed, datetime.time.min)
    try:
        return datetime.datetime.strptime(value, '%Y-%m-%d %H:%M:%S')
    except ValueError:
        try:
            return datetime.datetime.strptime(value, '%Y-%m-%d')
        except ValueError as exc:
            raise ValueError(f'Fecha inválida: {value}') from exc


def _parse_inventory_row(row_data):
    parsed = {}
    parsed['idinmueble'] = _clean_value(row_data.get('idinmueble'))
    if not parsed['idinmueble']:
        raise ValueError('idinmueble es obligatorio')

    parsed['etapa'] = _clean_value(row_data.get('etapa'))
    parsed['manzananumero'] = _clean_value(row_data.get('manzananumero'))
    parsed['lotenumero'] = _clean_value(row_data.get('lotenumero'))
    parsed['matricula'] = _clean_value(row_data.get('matricula'))
    parsed['estado'] = _clean_value(row_data.get('estado'))
    parsed['obsbloqueo'] = _clean_value(row_data.get('obsbloqueo'))

    decimal_fields = [
        'vrmetrocuadrado', 'areaprivada', 'areaconstruida', 'area_lt', 'area_mz',
        'norte', 'sur', 'este', 'oeste',
        'fac_valor_via_principal', 'fac_valor_area_social', 'fac_valor_esquinero'
    ]
    for field in decimal_fields:
        parsed[field] = _parse_decimal(row_data.get(field))

    parsed['meses'] = _parse_int(row_data.get('meses'))
    parsed['finobra'] = _parse_datetime_value(row_data.get('finobra'))

    text_fields = [
        'lindero_norte', 'colindante_norte', 'lindero_sur', 'colindante_sur',
        'lindero_este', 'colidante_este', 'lindero_oeste', 'colindante_oeste'
    ]
    for field in text_fields:
        parsed[field] = _clean_value(row_data.get(field))

    return parsed


def _iter_inventory_rows(uploaded_file):
    extension = os.path.splitext(uploaded_file.name or '')[1].lower()
    if extension == '.csv':
        content = uploaded_file.read().decode('utf-8-sig')
        reader = csv.DictReader(io.StringIO(content))
        headers = {_normalize_header(h): h for h in (reader.fieldnames or [])}
        missing = [col for col in INVENTARIO_IMPORT_COLUMNS if col not in headers]
        if missing:
            raise ValueError(f'Columnas faltantes en CSV: {", ".join(missing)}')
        for index, row in enumerate(reader, start=2):
            normalized = {_normalize_header(k): v for k, v in row.items()}
            if not any(_clean_value(v) is not None for v in normalized.values()):
                continue
            yield index, normalized
        return

    if extension == '.xlsx':
        wb = openpyxl.load_workbook(uploaded_file, data_only=False)
        ws = wb.active
        header_cells = [cell.value for cell in ws[1]]
        headers = [_normalize_header(cell) for cell in header_cells]
        missing = [col for col in INVENTARIO_IMPORT_COLUMNS if col not in headers]
        if missing:
            raise ValueError(f'Columnas faltantes en XLSX: {", ".join(missing)}')
        header_index = {name: idx for idx, name in enumerate(headers)}

        for row_idx in range(2, ws.max_row + 1):
            row_values = {}
            is_empty = True
            for field in INVENTARIO_IMPORT_COLUMNS:
                cell = ws.cell(row=row_idx, column=header_index[field] + 1)
                if cell.data_type == 'f':
                    raise ValueError(f'No se permiten fórmulas (fila {row_idx}, columna {field})')
                row_values[field] = cell.value
                if _clean_value(cell.value) is not None:
                    is_empty = False
            if is_empty:
                continue
            yield row_idx, row_values
        return

    raise ValueError('Formato no soportado. Usa .xlsx o .csv')


@login_required
def proyecto_popup(request,redireccion):
    redirecciones={
        'gtt':'comercial/gtt',
        'lista_asesores':'comercial/lista_asesores',
        'parametros':'operaciones/parametros',
        'informe_mes':'operaciones/informes',
        'promesas':'operaciones/promesas',
        'descuentos':'operaciones/descuentos_condicionados',
        'interfaces':'contabilidad/interfaces',
        'detalle_comisiones':'comercial/detalle_comisiones',
        'ver_ppto':'cartera/ver_presupuesto',
        'inv_admin':'operaciones/inventario_administrativo',
        'comisiones':'operaciones/comisiones',
        'desistidos':'desistidos',
        'recaudos_nr':'tesoreria/recaudos_nr',
        'nueva_adjudicacion':'operaciones/por_adjudicar',
        'lista_adj':'adjudicaciones',
        'ventas_sin_aprobar':'comercial/ventas_sin_aprobar',
        'inventario_ccial':'venta/inventario',
        'graph_ventas':'graphs/ventas',
        'graph_rcdo_mes':'graphs/recaudos',
        'graph_rcdo_year':'graphs/recaudos/comparativo',
        'graph_cartera_year':'graphs/cartera/year',
        'pqrs':'servicio_cliente/pqrs',
        'nuevo_recaudo':'tesoreria/adjudicaciones',
        'interfaces_banco':'tesoreria/interfaces_bancarias',
        'gestion_asesores':'comercial/lista_asesores_general',
        'edades_cartera':'cartera/edades_cartera',
    }
    logos_proyecto = {
        'Tesoro Escondido': 'img/logo-Tesoro-Escondido.png',
        'Vegas de Venecia': 'img/logo-camisas-vegas-de-venecia.png',
        'Carmelo Reservado': 'img/logo_carmelo_reservado.png',
        'Sandville Beach': 'img/sandville beach.png',
        'Perla del Mar': 'img/logo-perla.png',
        'Fractal': 'img/fractal-logo.jpg',
        'Casas de Verano': 'img/casas-de-verano450x.png',
        'Oasis': 'img/logo_oasis.png',
    }
    direccion=redirecciones[redireccion]
    lista_proyectos = []
    for proyecto_item in proyectos.objects.values('proyecto', 'activo').order_by('proyecto'):
        nombre = proyecto_item['proyecto']
        if nombre == 'default':
            continue
        if nombre not in connections.databases:
            continue
        lista_proyectos.append({
            'nombre': nombre,
            'logo': logos_proyecto.get(nombre),
            'activo': bool(proyecto_item.get('activo')),
        })
    context={
        'redireccion':direccion,
        'proyectos': lista_proyectos,
    }
    return render(request,'proyectos_popup.html',context)

@login_required
def welcome(request):
    
    if check_groups(request,groups=['ExternoFractal'],raise_exception=False):
        if not request.user.is_superuser:
            return HttpResponseRedirect('/fractal')
    
    avatar_actual=Profiles.objects.get(user=request.user).avatar.pk
    show_info=False
    if avatar_actual == 9999999:
        show_info = True
    
    birthdays = [] 
    active_users = Profiles.objects.filter(user__is_active=True)
    hoy=datetime.date.today()
    futuro = hoy+relativedelta(days=15)
    for user in active_users:
        dia = user.fecha_nacimiento.day
        mes= user.fecha_nacimiento.month
        año = datetime.date.today().year
        cumpleaños = datetime.date(año,mes,dia)
        if (cumpleaños>=hoy and cumpleaños<=futuro):
            birthdays.append(user)
    proyectos=(
        'Tesoro Escondido',
        'Vegas de Venecia',
        'Perla del Mar',
        'Sandville Beach',
        'Sotavento',
        'Oasis',
    )   
    mes=datetime.date.today().month
    año=datetime.date.today().year
    ultimo_dia=calendar.monthrange(año,mes)[1]
    inicio_mes=datetime.date(año,mes,1)
    fin_mes=datetime.date(año,mes,ultimo_dia)
    contratos_por_aprobar = 0
    ventas_mes=0
    lotes_disponibles = 0
    gtt_pendientes=0
    recibos=0
    recibos_pormi=0
    recibos_nr=0
    clientes_cobro_ccial=0
    clientes_cobro_admin=0
    pqrs_rad=0
    pqrs_abierta=0
    pqrs_cerr=0
    ventas_jefe=0
    ventas_pdtes_jefe=0
    lotes_libres_jefe=0
    ventas_por_adjudicar=0
    ventas_adjudicadas=0
    contratos_anulados=0
    for proyecto in proyectos:
        valor=ventas_nuevas.objects.using(proyecto).filter(estado='Pendiente').count()
        ventas_mes+=ventas_nuevas.objects.using(proyecto).filter(fecha_contrato__gte=inicio_mes,fecha_contrato__lte=fin_mes).exclude(estado='Anulado').count()
        contratos_por_aprobar+=valor
        #---------------------------
        lotes=Inmuebles.objects.using(proyecto).filter(estado='libre').count()
        lotes_disponibles+=lotes  
        #---------------------------
        gtt=Gtt.objects.filter(proyecto=proyecto,estado='pendiente').count()
        gtt_pendientes+=gtt
        #---------------------------
        recibos+=Recaudos_general.objects.using(proyecto).filter(fecha__gte=inicio_mes,fecha__lte=fin_mes).count()
        recibos_pormi+=Recaudos_general.objects.using(proyecto).filter(fecha__gte=inicio_mes,fecha__lte=fin_mes,usuario=request.user).count()
        #---------------------------
        recibos_nr+=RecaudosNoradicados.objects.using(proyecto).all().count()
        #---------------------------
        periodo = f'{año}{mes:02d}'
        clientes_cobro_ccial+=PresupuestoCartera.objects.using(proyecto).filter(periodo=periodo,tipocartera='Comercial').values('idadjudicacion').annotate(Sum('cuota')).count()
        clientes_cobro_admin+=PresupuestoCartera.objects.using(proyecto).filter(periodo=periodo,tipocartera ='Administrativa').values('idadjudicacion').annotate(Sum('cuota')).count()
        #---------------------------
        pqrs_rad+=Pqrs.objects.using(proyecto).filter(fecha_radicado__gte=inicio_mes,fecha_radicado__lte=fin_mes).count()
        pqrs_cerr+=Pqrs.objects.using(proyecto).filter(fecha_respuesta__gte=inicio_mes,fecha_respuesta__lte=fin_mes,estado='Cerrado').count()
        pqrs_abierta+=Pqrs.objects.using(proyecto).filter(estado='Abierta').count()
        #Jefe ventas----------------
        ventas_jefe+=ventas_nuevas.objects.using(proyecto).filter(usuario=request.user,fecha_contrato__gte=inicio_mes,fecha_contrato__lte=fin_mes).exclude(estado='Anulado').count()
        ventas_pdtes_jefe+=ventas_nuevas.objects.using(proyecto).filter(estado='Pendiente',usuario=request.user).count()
        if check_project(request,proyecto,raise_exception=False):
            lotes_libres_jefe+=Inmuebles.objects.using(proyecto).filter(estado='libre').count()
        #Asistente Operaciones
        if check_project(request,proyecto,raise_exception=False):
            ventas_por_adjudicar+=ventas_nuevas.objects.using(proyecto).filter(estado='Aprobado').count()
            ventas_adjudicadas+=Adjudicacion.objects.using(proyecto).filter(fecha__gte=inicio_mes,fecha__lte=fin_mes).count()
            contratos_anulados+=ventas_nuevas.objects.using(proyecto).filter(estado='Anulado',fecha_contrato__gte=inicio_mes,fecha_contrato__lte=fin_mes).count()
        #Proyectos
        ordenes_abiertas = building_model.contratos.objects.filter(estado='Pendiente').count()
        ordenes_delmes = building_model.contratos.objects.filter(fecha_creacion__gte=inicio_mes).count()
        mis_ordenes_abiertas = building_model.contratos.objects.filter(estado='Pendiente',usuario_crea=request.user).count()
        mis_ordenes_delmes = building_model.contratos.objects.filter(fecha_creacion__gte=inicio_mes,usuario_crea=request.user).count()
        actas_abiertas = building_model.actas_contratos.objects.filter(estado='Pendiente').count()
        actas_delmes = building_model.actas_contratos.objects.filter(fecha_acta__gte=inicio_mes).count()
        mis_actas_abiertas = building_model.actas_contratos.objects.filter(estado='pendiente',usuario_crea=request.user).count()
        mis_actas_delmes = building_model.actas_contratos.objects.filter(fecha_acta__gte=inicio_mes,usuario_crea=request.user).count()
    
    list_reminders={
        'Gerencia Comercial':{
            'name':'Gerencia Comercial',
            'reminders':[
                {
                    'title':'Ventas del mes',
                    'value':ventas_mes
                },
                {
                    'title':'Contratos por aprobar',
                    'value':contratos_por_aprobar
                },
                {
                    'title':'Lotes disponibles',
                    'value':lotes_disponibles,
                },
                {
                    'title':'GTT por aprobar',
                    'value': gtt_pendientes
                },
                {
                    'title':'Asesores activos',
                    'value':asesores.objects.filter(estado='Activo',tipo_asesor='Externo').count()
                },
                         ]
        },
        'Tesoreria':{
            'name':'Tesoreria',
            'reminders':[
                {
                'title':'Recibos hechos en el mes',
                'value':recibos
                },
                {
                'title':'Recibos hechos por mi',
                'value':recibos_pormi
                },
                {
                'title':'Recibos no radicados',
                'value':recibos_nr
                },
                {
                'title':'Pagos registrados en el mes',
                'value':Pagos.objects.filter(fechapago__gte=inicio_mes,fechapago__lte=fin_mes).count()
                },
                {
                'title':'Causaciones recibidas',
                'value':Facturas.objects.filter(fechacausa__gte=inicio_mes,fechacausa__lte=fin_mes).count()
                },
            ]
        },
        'Contabilidad':{
            'name':'Contabilidad',
            'reminders':[
                {
                'title':'Facturas recibidas mes',
                'value':Facturas.objects.filter(fecharadicado__gte=inicio_mes,fecharadicado__lte=fin_mes,).count()
                },
                {
                'title':'Causaciones registradas mes',
                'value':Facturas.objects.filter(fechacausa__gte=inicio_mes,fechacausa__lte=fin_mes).count()
                },
                {
                'title':'Facturas sin causar',
                'value':Facturas.objects.filter(fechacausa__isnull=True).count()
                },
            ]
        },
        'Supervisor Cartera':{
            'name':'Supervisor Cartera',
            'reminders':[
                {
                'title':'Clientes cartera ccial',
                'value':clientes_cobro_ccial
                },
                {
                'title':'Clientes cartera admin',
                'value':clientes_cobro_admin
                },
            ]
        },
        'Gestor Cartera':{
            'name':'Gestor Cartera',
            'reminders':[
                {
                'title':'Clientes cartera ccial',
                'value':clientes_cobro_ccial
                },
                {
                'title':'Clientes cartera admin',
                'value':clientes_cobro_admin
                },
            ]
        },
        'Servicio Cliente':{
            'name':'Servicio Cliente',
            'reminders':[
                {
                'title':'Pqrs radicadas en el mes',
                'value':pqrs_rad
                },
                {
                'title':'Pqrs abiertas',
                'value':pqrs_abierta,
                },
                {
                'title':'Pqrs cerradas en el mes',
                'value':pqrs_cerr
                },
            ]
        },
        'Jefe Ventas':{
            'name':'Jefe Ventas',
            'reminders':[
                {
                'title':'Mis ventas del mes',
                'value':ventas_jefe
                },
                {
                'title':'Ventas sin aprobar',
                'value':ventas_pdtes_jefe,
                },
                {
                'title':'Lotes disponibles',
                'value':lotes_libres_jefe 
                },
            ]
        },
        'Recepcion':{
            'name':'Recepcion',
            'reminders':[
                {
                'title':'Facturas radicadas en el mes',
                'value':Facturas.objects.filter(fecharadicado__gte=inicio_mes,fecharadicado__lte=fin_mes,).count()
                },
                {
                'title':'Facturas sin causar',
                'value':Facturas.objects.filter(fechacausa__isnull=True).count()
                },
            ]
        },
        'Asistente Operaciones':{
            'name':'Operaciones',
            'reminders':[
                {
                'title':'Ventas por adjudicar',
                'value':ventas_por_adjudicar
                },
                {
                'title':'Ventas adjudicadas en el mes',
                'value':ventas_adjudicadas
                },
                {
                'title':'Contratos anulados en el mes',
                'value':contratos_anulados
                },
            ]
        },
        'Gerente de Proyectos':{
            'name':'Gerente de proyectos',
            'reminders':[
                {
                'title':'Ordenes sin aprobar',
                'value':ordenes_abiertas
                },
                {
                'title':'Ordenes del mes',
                'value':ordenes_delmes
                },
                {
                'title':'Actas sin aprobar',
                'value':actas_abiertas
                },
                {
                'title':'Actas del mes',
                'value':actas_delmes
                },
            ]
        },
        'Asistente de Proyectos':{
            'name':'Asistente de proyectos',
            'reminders':[
                {
                'title':'Ordenes sin aprobar',
                'value':mis_ordenes_abiertas
                },
                {
                'title':'Ordenes del mes',
                'value':mis_ordenes_delmes
                },
                {
                'title':'Actas sin aprobar',
                'value':mis_actas_abiertas
                },
                {
                'title':'Actas del mes',
                'value':mis_actas_delmes
                },
            ]
        }
    }
    grupos_usuario=[]
    grupos = request.user.groups.all()
    for grupo in grupos:
        grupos_usuario.append(list_reminders.get(grupo.name))
    if request.user.is_superuser:
        grupos_usuario=list_reminders.values()
    context={
        'birthdays':birthdays,
        'recordatorios':grupos_usuario,
        'hoy':datetime.date.today(),
        'showinfo':show_info,
    }
    return render(request,'welcome.html',context)

def get_avatars(request):
    data = {'avatars': []}
    if request.is_ajax() and request.method == 'GET':
        avatars = Avatars.objects.exclude(name='noavatar')
        data['avatars'] = [
            {
                'name': avatar.name,
                'image': str(avatar.image),
                'image_url': avatar.image.url if avatar.image else '',
            }
            for avatar in avatars
        ]
    return JsonResponse(data)

def cambiar_avatar(request,avatar):
    user_profile = Profiles.objects.get(user=request.user.pk)
    new_avatar = Avatars.objects.get(name=avatar)
    user_profile.avatar = new_avatar
    user_profile.save()
    return HttpResponseRedirect('/welcome')

def registrar_cliente(request):
    data=open(settings.STATICFILES_DIRS[0]+'/JSON/colombia.json','r')
    data=data.read()
    context={
        'data_municipios':data,
        'form':form_nuevo_cliente,
        'form_pj':form_nuevo_cliente_PJ
    }
    if request.method == 'POST':
        form_datos_titular=request.POST
        form_datos_pj=form_nuevo_cliente_PJ(request.POST)
        
        if form_datos_titular and not form_datos_pj.is_valid():
            idtercero=form_datos_titular.get('idTercero')
            tipodoc = form_datos_titular.get('tipo_doc_id')
            lugar_exp_id = form_datos_titular.get('lugar_expedicion_id')
            fecha_exp_id = form_datos_titular.get('fecha_exp_id')
            nombres = form_datos_titular.get('nombres')
            if nombres.endswith(' '): nombres = nombres[:-1]
            apellidos = form_datos_titular.get('apellidos')
            if apellidos.endswith(' '): apellidos = apellidos[:-1]
            nombrecompleto=nombres+' '+apellidos
            celular1=form_datos_titular.get('celular1')
            telefono1=form_datos_titular.get('telefono1')
            telefono2=form_datos_titular.get('telefono2')
            domicilio=form_datos_titular.get('domicilio')
            ciudad=request.POST.get('ciudad')
            oficina=form_datos_titular.get('oficina')
            ciudad_ofic=request.POST.get('ciudad_ofic')
            email=form_datos_titular.get('email')
            fecha_nac=form_datos_titular.get('fecha_nac')
            ocupacion=form_datos_titular.get('ocupacion')
            hijos=form_datos_titular.get('hijos')
            nivel_educativo=form_datos_titular.get('nivel_educativo')
            estado_civil=form_datos_titular.get('estado_civil')
            nivel_ingresos=form_datos_titular.get('nivel_ingresos')
            vehiculo=form_datos_titular.get('vehiculo')
            vivienda=form_datos_titular.get('vivienda')
            fecha_act=datetime.date.today()
            pais = form_datos_titular.get('pais')
            pais = Countries.objects.get(pk=pais)
            estado = form_datos_titular.get('estado')
            estado = States.objects.get(pk=estado)
            ciudad = form_datos_titular.get('ciudad')
            ciudad = Cities.objects.get(pk=ciudad)
            
            lugar_nac = form_datos_titular.get('lugar_nacimiento')
            nacionalidad = form_datos_titular.get('nacionalidad')
            tipo_ocupacion = form_datos_titular.get('tipo_ocupacion')
            cargo_ocupacion = form_datos_titular.get('cargo_ocupacion')
            ingresos_provienen_de = form_datos_titular.get('ingresos_provienen_de')
            declara_renta = True if form_datos_titular.get('declara_renta') == 'True' else False
            tiene_rut = True if form_datos_titular.get('tiene_rut') == 'True' else False
            codigo_ciuu = form_datos_titular.get('ciuu')
            if codigo_ciuu == "": 
                codigo_ciuu = None
            else:
                codigo_ciuu = CIUU.objects.get(pk=codigo_ciuu)
            
            p1 = 'Si' if form_datos_titular.get("p_compl_1") == 'True' else 'No'
            p2 = 'Si' if form_datos_titular.get("p_compl_2") == 'True' else 'No'
            p3 = 'Si' if form_datos_titular.get("p_compl_3") == 'True' else 'No'
            p4 = 'Si' if form_datos_titular.get("p_compl_4") == 'True' else 'No'
            p5 = 'Si' if form_datos_titular.get("p_compl_5") == 'True' else 'No'
            p6 = 'Si' if form_datos_titular.get("p_compl_6") == 'True' else 'No'
            
            
            preguntas_complementarias = f'''{{
                "data":[{{
                    "pregunta":"¿Es sujeto de obligaciones tributarias en otro país o grupo de paises?",
                    "respuesta":"{p1}",
                    "complemento":"{form_datos_titular.get("compl_p_compl_1")}"
                }},{{
                    "pregunta":"¿Realiza transacciones en moneda extrajera?",
                    "respuesta":"{p2}",
                    "complemento":"{form_datos_titular.get("compl_p_compl_2")}"
                }},{{
                    "pregunta":"¿Posee inversiones en el exterior?",
                    "respuesta":"{p3}",
                    "complemento":"{form_datos_titular.get("compl_p_compl_3")}"
                }},{{
                    "pregunta":"¿Posee cuentas en moneda extranjera?",
                    "respuesta":"{p4}",
                    "complemento":"{form_datos_titular.get("compl_p_compl_4")}"
                }},{{
                    "pregunta":"¿Realiza transferencias al extranjero?",
                    "respuesta":"{p5}",
                    "complemento":"{form_datos_titular.get("compl_p_compl_5")}"
                }},{{
                    "pregunta":"¿Recibe transferencias del extranjero?",
                    "respuesta":"{p6}",
                    "complemento":"{form_datos_titular.get("compl_p_compl_6")}"
                }}]
            }}'''
            
            es_peps = True if form_datos_titular.get('es_peps') == 'True' else False
            peps_desde = None if form_datos_titular.get('peps_desde') == "" else form_datos_titular.get('peps_desde')
            
            peps_hasta = None if form_datos_titular.get('peps_hasta') == "" else form_datos_titular.get('peps_hasta')
            
            peps_entidad = form_datos_titular.get('entidad_peps')             
            peps_cargo = form_datos_titular.get('cargo_peps')       
            peps_familiar = form_datos_titular.get('es_familiar_peps')       
            peps_familiar_parentesco = form_datos_titular.get('parentesco_familiar_peps') 
            peps_familiar_entidad = form_datos_titular.get('entidad_familiar_peps') 
            peps_familiar_cargo = form_datos_titular.get('cargo_familiar_peps') 
            referencia_famliar = form_datos_titular.get('ref_familiar_nombre') 
            referencia_familiar_telefono = form_datos_titular.get('ref_familiar_telefono') 
            referencia_personal = form_datos_titular.get('ref_personal_nombre') 
            referencia_personal_telefono =form_datos_titular.get('ref_personal_telefono') 
            
            cliente = clientes.objects.using('default').create(
                idTercero=idtercero, tipo_doc = tipodoc, lugar_exp_id = lugar_exp_id,
                fecha_exp_id = fecha_exp_id, nombres=nombres.upper(), apellidos=apellidos.upper(),
                nombrecompleto=nombrecompleto.upper(), celular1=celular1, telefono1=telefono1,
                domicilio=domicilio, pais = pais, estado = estado, city = ciudad, 
                nacionalidad = nacionalidad, email=email, fecha_nac=fecha_nac,
                nivel_educativo=nivel_educativo, estado_civil=estado_civil,
                nivel_ingresos=nivel_ingresos, vivienda=vivienda,
                fecha_actualizacion=fecha_act, ocupacion = tipo_ocupacion,
                lugar_nac = lugar_nac
            )
            
            sagrilaft_info.objects.create(
                cliente = cliente, empresa_labora = ocupacion, cargo_actual = cargo_ocupacion,
                origen_ingresos = ingresos_provienen_de, declara_renta = declara_renta, 
                tiene_rut = tiene_rut, codigo_ciuu = codigo_ciuu , 
                preguntas_complementarias = preguntas_complementarias,
                es_peps = es_peps, peps_desde = peps_desde, peps_hasta = peps_hasta,
                peps_entidad = peps_entidad, peps_cargo = peps_cargo, peps_familiar = peps_familiar,
                peps_familiar_parentesco = peps_familiar_parentesco, peps_familiar_entidad = peps_familiar_entidad,
                peps_familiar_cargo = peps_familiar_cargo, referencia_familiar = referencia_famliar,
                referencia_familiar_telefono = referencia_familiar_telefono, 
                referencia_personal = referencia_personal, 
                referencia_personal_telefono = referencia_personal_telefono
                 
            )
                
            
            return redirect('/reg_cliente/success')
        elif form_datos_pj.is_valid():
            id_tercero = request.POST.get('idTercero')
            razon_social = request.POST.get('nombres')
            rep_legal = request.POST.get('apellidos')
            celular = request.POST.get('celular1')
            telefono1 = request.POST.get('telefono1')
            telefono2 = request.POST.get('telefono2')
            domicilio = request.POST.get('domicilio')
            ciudad = request.POST.get('ciudad')
            email = request.POST.get('email')
            fecha_creacion = request.POST.get('fecha_creac')
            ocupacion = request.POST.get('ocupacion')
            clientes.objects.create(
                idTercero=id_tercero,
                nombrecompleto=razon_social,
                nombres=razon_social,
                apellidos=rep_legal,
                celular1=celular,
                telefono1=telefono1,
                telefono2=telefono2,
                domicilio=domicilio,
                ciudad=ciudad,
                email=email,
                fecha_nac=fecha_creacion,
                ocupacion=ocupacion,
                fecha_actualizacion=datetime.date.today()                
            )
            return redirect('/reg_cliente/success')
        elif not form_datos_titular.is_valid():
            context['form']=form_nuevo_cliente(request.POST)
            context['PNnovalid']=True
            
    return render(request,'registrar_cliente.html',context)

def registro_asesor(request):
    if request.method == 'POST':
        form=registrar_asesor(request.POST,request.FILES)
        if form.is_valid():
            cedula=form.cleaned_data.get('Cedula')
            nombre=form.cleaned_data.get('Nombre')
            email=form.cleaned_data.get('Email')
            fecha_nac=form.cleaned_data.get('fecha_nacimiento')
            direccion=form.cleaned_data.get('Direccion')
            estado_civil=form.cleaned_data.get('Estado_civil')
            nivel_educativo=form.cleaned_data.get('Nivel_educativo')
            equipo=form.cleaned_data.get('Equipo')
            rut=form.cleaned_data.get('doc_rut')
            cc=form.cleaned_data.get('doc_cc')
            cert_bancaria=form.cleaned_data.get('cert_bancaria')
            banco=form.cleaned_data.get('Banco')
            num_cta=form.cleaned_data.get('num_cta')
            telefono=form.cleaned_data.get('Telefono')
            tipo_cta=form.cleaned_data.get('tipo_cta')
            hv = form.cleaned_data.get('hv')
            afiliaciones = form.cleaned_data.get('afiliaciones')
            
            if rut==None:
                check_rut=False
            else:
                check_rut=True
                upload_docs_asesores(request.FILES['doc_rut'],cedula,'RUT')
            if cc==None:
                check_cc=False
            else:
                check_cc=True
                upload_docs_asesores(request.FILES['doc_cc'],cedula,'CC')
            if cert_bancaria==None:
                check_cert=False
            else:
                check_cert=True
                upload_docs_asesores(request.FILES['cert_bancaria'],cedula,'Cert_bancaria')    
            check_hv = False
            if hv:
                check_hv = True
                upload_docs_asesores(request.FILES['hv'],cedula,'Hoja_de_vida')   
            check_afiliacion = False  
            if afiliaciones:
                check_afiliacion = True
                upload_docs_asesores(request.FILES['afiliaciones'],cedula,'Soporte_Afiliaciones')
                
            instance=asesores.objects.create(cedula=cedula,
                                            nombre=nombre.upper(),
                                            email=email,
                                            direccion=direccion,
                                            estado_civil=estado_civil,
                                            nivel_educativo=nivel_educativo,
                                            equipo=equipo,
                                            rut=check_rut,
                                            cc=check_cc,
                                            cert_banc=check_cert,
                                            fecha_nacimiento=fecha_nac,
                                            banco=banco,
                                            cuenta=num_cta,
                                            telefono=telefono,
                                            tipo_cuenta=tipo_cta,
                                            hv=check_hv,
                                            afiliaciones=check_afiliacion)
            """ msj_notif = f'Hola Paola!, Te informamos que se registró un nuevo asesor ({nombre}) en el proyecto {equipo}.'
            envio_notificacion(msj_notif,f'Se registró un nuevo Asesor',['gestionhumana@somosandina.co']) """
            return HttpResponseRedirect('reg_asesor/success')
    else:
        form=registrar_asesor(None)
    context={
        'form':form,
    }
    return render(request,'registro_asesor.html',context)

def registro_exitoso(request):
    return render(request,'registro_exitoso.html')

def pago_exitoso(request):
    return render(request,'pago_exitoso.html')

def pago_proceso(request):
    return render(request,'pago_proceso.html')

@group_perm_required(('andinasoft.view_asesores',),raise_exception=True)
def lista_asesores(request,proyecto):
    check_project(request,proyecto)
    form=form_retiro_asesor
    if request.method == 'POST':
        form = form_retiro_asesor(request.POST)
        if form.is_valid():
            cedula = form.cleaned_data.get('cedula')
            fecha_retiro = form.cleaned_data.get('fecha_retiro')
            obj_asesor = asesores.objects.get(pk=cedula)
            obj_asesor.fecha_baja=fecha_retiro
            obj_asesor.estado='Retirado'
            obj_asesor.save()
    listado_asesores=asesores.objects.filter(estado='Activo',equipo=proyecto)
    context={
        'list_asesores':listado_asesores,
        'form':form,
        'proyecto':proyecto,
    }
    return render(request,'listado_asesores.html',context)

@group_perm_required(('andinasoft.view_asesores','andinasoft.change_asesores',),raise_exception=True)
def lista_asesores_general(request,proyecto):
    check_project(request,proyecto)
    form=form_retiro_asesor
    if request.method == 'POST':
        form = form_retiro_asesor(request.POST)
        if form.is_valid():
            cedula = form.cleaned_data.get('cedula')
            fecha_retiro = form.cleaned_data.get('fecha_retiro')
            obj_asesor = asesores.objects.get(pk=cedula)
            obj_asesor.fecha_baja=fecha_retiro
            obj_asesor.estado='Retirado'
            obj_asesor.save()
    listado_asesores=asesores.objects.filter(equipo=proyecto)
    context={
        'list_asesores':listado_asesores,
        'form':form,
        'proyecto':proyecto,
    }
    return render(request,'listado_asesores.html',context)

@group_perm_required(('andinasoft.view_gtt',),raise_exception=True)
def lista_gtt(request,proyecto):
    if request.method == 'GET':
        if request.is_ajax():
            item_gtt = request.GET.get('id_gtt')
            obj_gtt = Detalle_gtt.objects.filter(gtt=item_gtt)
            data_gtt = []
            for gtt in obj_gtt:
                asesor = gtt.asesor.nombre
                valor = gtt.valor 
                data_gtt.append({'asesor':asesor,'valor':valor})
            return JsonResponse({'data':data_gtt})
    if request.method == 'POST':
        if request.POST.get('btnAprobar'):
            check_perms(request,('andinasoft.change_gtt',))
            item_gtt = request.POST.get('gtt_aprobar')
            obj_gtt=Gtt.objects.get(pk=item_gtt,proyecto=proyecto)
            obj_gtt.estado='Aprobado'
            obj_gtt.usuario_aprueba=str(request.user)
            obj_gtt.save()
        
    obj_gtt = Gtt.objects.filter(proyecto=proyecto).annotate(valor=Sum('detalle_gtt__valor'))
    context= {
        'gtts':obj_gtt,
        'proyecto':proyecto,
    }
    return render(request,'lista_gtt.html',context)

@group_perm_required(('andinasoft.add_gtt',),raise_exception=True)
def nuevo_gtt(request,proyecto):
    asesores_activos=asesores.objects.filter(equipo=proyecto,estado='Activo',tipo_asesor="Externo")
    valor_inicial=112500
    context = {
        'asesores':asesores_activos,
        'valor_inicial':valor_inicial,
    }
    if request.method == 'POST':
        desde = request.POST.get('dia_desde')
        hasta = request.POST.get('dia_hasta')
        list_asesores = request.POST.getlist('cedula')
        list_valores = request.POST.getlist('valor')
        
        Gtt.objects.create(proyecto=proyecto,estado='Pendiente',fecha_desde=desde,fecha_hasta=hasta,usuario_crea=str(request.user))
        obj_gtt = Gtt.objects.last()
        for i in range(0,len(list_asesores)):
            asesor=asesores.objects.get(pk=list_asesores[i])
            valor=int(list_valores[i].replace('.',''))
            if valor > 0:
                Detalle_gtt.objects.create(valor=valor,gtt=obj_gtt,asesor=asesor)
        
        return HttpResponseRedirect('/comercial/gtt/'+proyecto)            
        
    return render(request,'nuevo_gtt.html',context)

def printGtt(request,proyecto,gtt):
    obj_gtt = Gtt.objects.get(pk=gtt)
    detalle_gtt = Detalle_gtt.objects.filter(gtt=gtt)
    ruta=settings.DIR_EXPORT+f'GTT_{proyecto}_{gtt}.pdf'
    pdf = GenerarPDF()
    pdf.gtt(ruta=ruta,proyecto=proyecto,desde=obj_gtt.fecha_desde,hasta=obj_gtt.fecha_hasta,contenido=detalle_gtt,
            crea=obj_gtt.usuario_crea,aprueba=obj_gtt.usuario_aprueba)
    return FileResponse(open(ruta,'rb'),as_attachment=True,filename=f'GTT_{proyecto}_{gtt}.pdf')

def simulador(request):
    context = {
        
    }
    return render(request,'simulator.html',context)

@group_perm_required(('andinasoft.view_pagocomision',),raise_exception=True)
def printComisiones(request,proyecto,desde,hasta):
    stmt=f'CALL detalle_comisiones_fecha("{desde}","{hasta}")'
    comisiones=Pagocomision.objects.using(proyecto).raw(stmt)
    ruta=settings.DIR_EXPORT+f'Comisiones_{proyecto}_{desde}_{hasta}.pdf'
    pdf = GenerarPDF()
    pdf.comisiones(ruta=ruta,proyecto=proyecto,desde=desde,hasta=hasta,contenido=comisiones,
            crea=str(request.user))
    return FileResponse(open(ruta,'rb'),as_attachment=True,filename=f'Comisiones_{proyecto}_{desde}_{hasta}.pdf')

@group_perm_required(('andinasoft.view_vista_adjudicacion',),raise_exception=True)
def lista_adjudicaciones(request,proyecto):
    check_project(request,proyecto)
    if request.user.is_authenticated:
        lista_adj=Vista_Adjudicacion.objects.using(proyecto).all().exclude(Estado__istartswith="Desistido")
        context={
            'proyecto':proyecto,
            'form':projects,
            'adjudicaciones':lista_adj
            }
        return render(request,'listado_adj.html',context)
    else:
        return redirect('/accounts/login')

@group_perm_required(('andinasoft.view_vista_adjudicacion',),raise_exception=True)
def lista_desistidos(request,proyecto):
    check_project(request,proyecto)
    if request.user.is_authenticated:
        lista_adj=Adjudicacion.objects.using(proyecto).raw('CALL desistimientos()')
        context={
            'proyecto':proyecto,
            'form':projects,
            'adjudicaciones':lista_adj
            }
        return render(request,'listado_desistidos.html',context)
    else:
        return redirect('/accounts/login')

def _parse_bool(value):
    return True if str(value).lower() in ('true', 'on', '1', 'si') else False


def _validar_recaudo_context(request, proyecto, adj, titulares, saldos_cuotas, consecutivo, form_token, post_data):
    total_cap = 0
    total_intcte = 0
    total_intmora = 0
    count_cuota = 0

    abonocapital = _parse_bool(post_data.get('abonocapital'))
    fecha = post_data.get('fecha')
    if not fecha:
        fecha = datetime.date.today()
        if settings.LIVE:
            fecha = datetime.date.today() + relativedelta(days=1)

    fecha_pago = post_data.get('fecha_pago')
    # Intentar parsear diferentes formatos de fecha
    fecha_pago_str = str(fecha_pago).strip()
    if ' ' in fecha_pago_str:
        # Formato: YYYY-MM-DD HH:MM:SS
        fecha_pago = datetime.datetime.strptime(fecha_pago_str.split(' ')[0], '%Y-%m-%d')
    elif '-' in fecha_pago_str:
        # Formato: YYYY-MM-DD
        fecha_pago = datetime.datetime.strptime(fecha_pago_str, '%Y-%m-%d')
    elif len(fecha_pago_str) == 8 and fecha_pago_str.isdigit():
        # Formato: YYYYMMDD
        fecha_pago = datetime.datetime.strptime(fecha_pago_str, '%Y%m%d')
    else:
        # Intentar con formato por defecto
        fecha_pago = datetime.datetime.strptime(fecha_pago_str, '%Y-%m-%d')
    forma_pago = post_data.get('forma_pago')
    condonacion = post_data.get('condonacion_mora')
    condonacion = 'Si' if condonacion == '1' or condonacion == 'Si' else 'No'
    valor_pagado = post_data.get('valor', '')
    if isinstance(valor_pagado, (int, float)):
        valor_pagado = str(valor_pagado)
    valor_pagado = valor_pagado.replace(',', '')
    valor_pagado = decimal.Decimal(valor_pagado)
    valor_recibo = int(valor_pagado)
    concepto = post_data.get('concepto')
    procentaje_condonado = post_data.get('condonacion_porc')
    numsolic = post_data.get('numsolicitud')

    if procentaje_condonado == None or procentaje_condonado == '':
        cobro_mora = 1
    else:
        cobro_mora = decimal.Decimal(1 - (float(procentaje_condonado) / 100))
    recaudo = []
    for cuota in saldos_cuotas:
        if valor_pagado == 0:
            break
        elif cuota.pendiente().get('total') <= 0:
            pass
        else:
            if abonocapital and (cuota.fecha > datetime.date.today()):
                recaudo.append((fecha_pago, 'ABONO', valor_pagado, 0, 0, 0, valor_pagado))
                total_cap += valor_pagado
                valor_pagado = 0
            else:
                datos_mora = cuota.mora(dia_pago=fecha_pago)
                saldo_pendiente = cuota.pendiente()
                if datos_mora.get('dias') == 0:
                    count_cuota += 1
                capital = saldo_pendiente.get('capital')
                intcte = saldo_pendiente.get('interes')
                intmora = round(datos_mora.get('valor') * cobro_mora, 0)
                if valor_pagado >= intmora:
                    mora_pagada = intmora
                    valor_pagado -= intmora
                else:
                    mora_pagada = valor_pagado
                    valor_pagado -= valor_pagado
                if valor_pagado >= intcte:
                    intcte_pagado = intcte
                    valor_pagado -= intcte
                else:
                    intcte_pagado = valor_pagado
                    valor_pagado -= valor_pagado
                if valor_pagado >= capital:
                    capital_pagado = capital
                    valor_pagado -= capital
                else:
                    capital_pagado = valor_pagado
                    valor_pagado -= valor_pagado
                total_cap += capital_pagado
                total_intcte += intcte_pagado
                total_intmora += mora_pagada
                total_rcdo_cta = capital_pagado + intcte_pagado + mora_pagada
                recaudo.append((cuota.fecha, cuota.idcta, capital_pagado, intcte_pagado, datos_mora.get('dias'), mora_pagada, total_rcdo_cta))

    totales_recaudo = (total_cap, total_intcte, total_intmora, total_cap + total_intcte + total_intmora)
    verif_pago = True
    if valor_recibo > (total_cap + total_intcte + total_intmora):
        verif_pago = False

    context = {
        'form': form_nuevo_recibo(proyecto=proyecto, initial={
            'fecha': fecha,
            'fecha_pago': fecha_pago,
            'numsolicitud': numsolic
        }),
        'adj': adj,
        'titulares': titulares,
        'num_recibo': consecutivo,
        'saldo_cuotas': saldos_cuotas,
        'detalle_recaudo': recaudo,
        'fecha': fecha,
        'concepto': concepto,
        'forma_pago': forma_pago,
        'valor_recibo': f'{int(valor_recibo):,}',
        'abono_capital': abonocapital,
        'condonacion': condonacion,
        'procentaje_condonado': procentaje_condonado,
        'totales': totales_recaudo,
        'guardar': True,
        'verif_valor': verif_pago,
        'count_cuota': count_cuota,
        'form_token': form_token,
        'bloquear_formulario': False
    }

    alerts = []
    if not verif_pago:
        alerts.append({
            'code': 'AMOUNT_EXCEEDS_PLAN',
            'message': 'El valor a pagar no puede ser mayor al plan de pagos pendientes.',
            'severity': 'error'
        })
    if count_cuota > 2:
        alerts.append({
            'code': 'MANY_FUTURE_INSTALLMENTS',
            'message': 'Este pago se esta aplicando a muchas cuotas futuras, revisalo!',
            'severity': 'warning'
        })

    # Permitir abono a capital aunque haya CI pendiente (se aplicará solo a FN)

    return context, alerts


def _validar_integridad_post_abono(proyecto, adj):
    """
    Valida la integridad del plan de pagos DESPUÉS de hacer el abono a capital.
    Esta función se ejecuta dentro de la transacción, antes del commit.

    Retorna (es_valido, errores[])

    Validaciones:
    1. Capital del plan debe coincidir con saldos
    2. Capital recaudado debe coincidir con saldos
    3. Capital pendiente debe ser coherente
    4. Todas las cuotas deben estar cuadradas (capital = cuota - interes)
    """
    errores = []
    tolerancia = 1  # 1 peso de tolerancia por redondeos

    # 1. Validar coherencia: Capital total del plan
    cap_plan = PlanPagos.objects.using(proyecto).filter(
        adj=adj
    ).aggregate(Sum('capital'))['capital__sum'] or 0

    cap_saldos = saldos_adj.objects.using(proyecto).filter(
        adj=adj
    ).aggregate(Sum('capital'))['capital__sum'] or 0

    if abs(cap_plan - cap_saldos) > tolerancia:
        errores.append({
            'code': 'CAPITAL_PLAN_MISMATCH',
            'message': f'Capital del plan ({cap_plan:,.0f}) no coincide con saldos ({cap_saldos:,.0f}). Diferencia: {abs(cap_plan - cap_saldos):,.0f}',
            'severity': 'critical',
            'data': {'cap_plan': float(cap_plan), 'cap_saldos': float(cap_saldos)}
        })

    # 2. Validar coherencia: Capital recaudado
    cap_recaudado = Recaudos.objects.using(proyecto).filter(
        idadjudicacion=adj,
        estado='Aprobado'
    ).aggregate(Sum('capital'))['capital__sum'] or 0

    cap_rcdo_saldos = saldos_adj.objects.using(proyecto).filter(
        adj=adj
    ).aggregate(Sum('rcdocapital'))['rcdocapital__sum'] or 0

    if abs(cap_recaudado - cap_rcdo_saldos) > tolerancia:
        errores.append({
            'code': 'CAPITAL_RECAUDADO_MISMATCH',
            'message': f'Capital recaudado ({cap_recaudado:,.0f}) no coincide con saldos ({cap_rcdo_saldos:,.0f}). Diferencia: {abs(cap_recaudado - cap_rcdo_saldos):,.0f}',
            'severity': 'critical',
            'data': {'cap_recaudado': float(cap_recaudado), 'cap_rcdo_saldos': float(cap_rcdo_saldos)}
        })

    # 3. Validar coherencia: Capital pendiente
    cap_pdte_calculado = cap_plan - cap_recaudado
    cap_pdte_saldos = saldos_adj.objects.using(proyecto).filter(
        adj=adj
    ).aggregate(Sum('saldocapital'))['saldocapital__sum'] or 0

    if abs(cap_pdte_calculado - cap_pdte_saldos) > tolerancia:
        errores.append({
            'code': 'CAPITAL_PENDIENTE_MISMATCH',
            'message': f'Capital pendiente calculado ({cap_pdte_calculado:,.0f}) no coincide con saldos ({cap_pdte_saldos:,.0f}). Diferencia: {abs(cap_pdte_calculado - cap_pdte_saldos):,.0f}',
            'severity': 'critical',
            'data': {'cap_pdte_calc': float(cap_pdte_calculado), 'cap_pdte_saldos': float(cap_pdte_saldos)}
        })

    # 4. Validar que cada cuota esté cuadrada: capital + intcte = cuota
    cuotas_plan = PlanPagos.objects.using(proyecto).filter(adj=adj)
    cuotas_descuadradas = []

    for cuota in cuotas_plan:
        suma_componentes = cuota.capital + cuota.intcte
        if abs(suma_componentes - cuota.cuota) > tolerancia:
            cuotas_descuadradas.append({
                'idcta': cuota.idcta,
                'capital': float(cuota.capital),
                'intcte': float(cuota.intcte),
                'cuota': float(cuota.cuota),
                'diferencia': float(abs(suma_componentes - cuota.cuota))
            })

    if cuotas_descuadradas:
        errores.append({
            'code': 'CUOTAS_DESCUADRADAS',
            'message': f'{len(cuotas_descuadradas)} cuotas descuadradas (capital + interés ≠ cuota)',
            'severity': 'critical',
            'data': {'cuotas': cuotas_descuadradas[:5]}  # Primeras 5 cuotas con error
        })

    # 5. Validar que no haya capital pendiente negativo en ninguna cuota
    cuotas_negativas = saldos_adj.objects.using(proyecto).filter(
        adj=adj,
        saldocapital__lt=0
    )

    if cuotas_negativas.exists():
        lista_negativas = list(cuotas_negativas.values('idcta', 'saldocapital')[:5])
        errores.append({
            'code': 'SALDO_CAPITAL_NEGATIVO',
            'message': f'{cuotas_negativas.count()} cuotas con saldo de capital negativo',
            'severity': 'critical',
            'data': {'cuotas': lista_negativas}
        })

    es_valido = len(errores) == 0
    return es_valido, errores


def _guardar_recaudo(
    request,
    proyecto,
    adj,
    titulares,
    saldo_cuotas,
    saldos_cuotas,
    consecutivo,
    obj_adj,
    form_recibo,
    form_token
):
    context_updates = {}
    alerts = []
    processed_tokens = request.session.setdefault('processed_receipt_tokens', [])
    if form_token in processed_tokens:
        context_updates['alerta'] = True
        context_updates['titulo'] = 'Andinasoft dice:'
        context_updates['mensaje'] = 'Este formulario ya fue procesado. No puedes guardar el recibo nuevamente.'
        context_updates['link'] = False
        alerts.append({
            'code': 'FORM_ALREADY_PROCESSED',
            'message': context_updates['mensaje'],
            'severity': 'error'
        })
        return context_updates, alerts, True
    if not form_recibo.is_valid():
        context_updates['form_token'] = form_token
        context_updates['form_errors'] = form_recibo.errors
        return context_updates, alerts, False

    nro_recibo = f'{consecutivo.consecutivo}'
    abonocapital = form_recibo.cleaned_data.get('abonocapital')
    fecha = form_recibo.cleaned_data.get('fecha')
    fecha_pago = form_recibo.cleaned_data.get('fecha_pago')
    forma_pago = form_recibo.cleaned_data.get('forma_pago')
    condonacion = form_recibo.cleaned_data.get('condonacion_mora')
    valor_pagado = int(float(form_recibo.cleaned_data.get('valor', 0).replace(',', '')))
    valor_recibo = valor_pagado
    concepto = form_recibo.cleaned_data.get('concepto')
    procentaje_condonado = form_recibo.cleaned_data.get('condonacion_porc')
    numsolic = form_recibo.cleaned_data.get('numsolicitud')
    if numsolic:
        try:
            solicitud_existente = recibos_internos.objects.using('default').get(pk=numsolic)
            if solicitud_existente.recibo_asociado:
                context_updates['alerta'] = True
                context_updates['titulo'] = 'Andinasoft dice:'
                context_updates['mensaje'] = f'La solicitud #{numsolic} ya está asociada al recibo {solicitud_existente.recibo_asociado}.'
                context_updates['link'] = False
                alerts.append({
                    'code': 'REQUEST_ALREADY_LINKED',
                    'message': context_updates['mensaje'],
                    'severity': 'error'
                })
                return context_updates, alerts, True
        except recibos_internos.DoesNotExist:
            pass
    es_ci = False
    es_fn = False
    es_co = False
    es_ce = False
    if procentaje_condonado == None or procentaje_condonado == '':
        cobro_mora = 1
    else:
        cobro_mora = 1 - (procentaje_condonado / 100)

    # ABONO A CAPITAL: Transacción TODO O NADA
    # Si falla cualquier parte, se revierte TODO (backup + pagos + recálculo)
    if abonocapital:
        try:
            # Usar transacción atómica: si algo falla, se revierte TODO
            with transaction.atomic(using=proyecto):
                # Crear backup
                obj_proyecto = proyectos.objects.get(pk=proyecto)
                obj_bk = bk_bfchangeplan.objects.create(
                    proyecto=obj_proyecto,
                    usuario_bk=request.user, adj=adj
                )
                full_planpagos = PlanPagos.objects.using(proyecto).filter(adj=adj)
                for i in full_planpagos:
                    bk_planpagos.objects.create(
                        id_bk=obj_bk,
                        proyecto=obj_proyecto,
                        idcta=i.idcta, tipocta=i.tipocta, nrocta=i.nrocta,
                        adj=i.adj, capital=i.capital, intcte=i.intcte,
                        cuota=i.cuota, fecha=i.fecha
                    )
                full_recaudodetalle = Recaudos.objects.using(proyecto).filter(idadjudicacion=adj)
                for i in full_recaudodetalle:
                    bk_recaudodetallado.objects.create(
                        id_bk=obj_bk,
                        proyecto=obj_proyecto, recibo=i.recibo,
                        fecha=i.fecha, idcta=i.idcta, idadjudicacion=i.idadjudicacion,
                        capital=i.capital, interescte=i.interescte, interesmora=i.interesmora,
                        moralqd=i.moralqd, fechaoperacion=i.fechaoperacion,
                        usuario=i.usuario, estado=i.estado
                    )

                # PASO 1: Pagar cuotas vencidas primero (solo las que tienen fecha <= fecha_pago)
                saldo_disponible = valor_pagado
                cuotas_vencidas = saldo_cuotas.filter(fecha__lte=fecha_pago)

                for cuota in cuotas_vencidas:
                    if saldo_disponible == 0:
                        break

                    datos_mora = PlanPagos.objects.using(proyecto).get(idcta=cuota.idcta).mora(dia_pago=fecha_pago)

                    if cuota.idcta[:2] == 'CI':
                        es_ci = True
                    elif cuota.idcta[:2] == 'FN':
                        es_fn = True
                    elif cuota.idcta[:2] == 'CE':
                        es_ce = True
                    elif cuota.idcta[:2] == 'CO':
                        es_co = True

                    capital = cuota.saldocapital
                    intcte = cuota.saldointcte
                    intmora = round(datos_mora.get('valor') * cobro_mora, 0)

                    # Pagar mora
                    if saldo_disponible >= intmora:
                        mora_pagada = intmora
                        saldo_disponible -= intmora
                    else:
                        mora_pagada = saldo_disponible
                        saldo_disponible = 0

                    # Pagar interés
                    if saldo_disponible >= intcte:
                        intcte_pagado = intcte
                        saldo_disponible -= intcte
                    else:
                        intcte_pagado = saldo_disponible
                        saldo_disponible = 0

                    # Pagar capital
                    if saldo_disponible >= capital:
                        capital_pagado = capital
                        saldo_disponible -= capital
                    else:
                        capital_pagado = saldo_disponible
                        saldo_disponible = 0

                    # Crear recaudo solo si se pagó algo
                    if mora_pagada > 0 or intcte_pagado > 0 or capital_pagado > 0:
                        Recaudos.objects.using(proyecto).create(
                            recibo=nro_recibo,
                            fecha=fecha,
                            idcta=cuota.idcta,
                            idadjudicacion=adj,
                            capital=capital_pagado,
                            interescte=intcte_pagado,
                            interesmora=mora_pagada,
                            moralqd=cuota.saldomora,
                            fechaoperacion=datetime.datetime.today(),
                            usuario=request.user,
                            estado='Aprobado'
                        )

                # PASO 2: Si queda saldo, hacer abono a capital
                if saldo_disponible > 0:
                    tasa = obj_adj.tasafnc

                    # Obtener consecutivo de cuotas CO
                    obj_planpagos = PlanPagos.objects.using(proyecto).filter(adj=adj, tipocta='CO')
                    if obj_planpagos.exists():
                        ultimacta = obj_planpagos.aggregate(Max('nrocta'))['nrocta__max']
                    else:
                        ultimacta = 0
                    ultimacta += 1
                    
                    

                    # Crear cuota CO con el saldo disponible
                    PlanPagos.objects.using(proyecto).create(
                        adj=adj,
                        fecha=fecha,
                        tipocta='CO',
                        idcta=f'CO{ultimacta}{adj}',
                        capital=saldo_disponible,
                        cuota=saldo_disponible,
                        nrocta=ultimacta,
                        intcte=0
                    )

                    # Crear recaudo CO
                    Recaudos.objects.using(proyecto).create(
                        recibo=nro_recibo,
                        fecha=fecha,
                        idcta=f'CO{ultimacta}{adj}',
                        idadjudicacion=adj,
                        capital=saldo_disponible,
                        interescte=0,
                        interesmora=0,
                        moralqd=0,
                        fechaoperacion=datetime.datetime.today(),
                        usuario=request.user,
                        estado='Aprobado'
                    )

                    # Calcular capital pendiente de TODO el plan
                                        
                    cap_total = obj_adj.valor

                    cap_pagado = Recaudos.objects.using(proyecto).filter(
                        idadjudicacion=adj,
                        estado='Aprobado'
                    ).aggregate(Sum('capital'))['capital__sum'] or 0

                    # Este es el monto EXACTO que debe sumar el nuevo plan de pagos
                    saldo_por_distribuir = cap_total - cap_pagado

                    # Validación de seguridad: Si pagaron de más, el saldo es 0
                    if saldo_por_distribuir < 0:
                        saldo_por_distribuir = 0

                    # -------------------------------------------------------------
                    # 2. OBTENER CUOTAS DISPONIBLES (VIVAS)
                    # -------------------------------------------------------------
                    # Usamos la vista para filtrar solo las que no se han pagado totalmente
                    # Excluyendo la CO que acabamos de crear
                    ids_cuotas_vivas = saldos_adj.objects.using(proyecto).filter(
                        adj=adj,
                        saldocapital__gt=0
                    ).values_list('idcta', flat=True)
                    
                    # Traemos los objetos reales ordenados por fecha
                    cuotas_para_recalcular = PlanPagos.objects.using(proyecto).filter(
                        adj=adj,
                        idcta__in=ids_cuotas_vivas
                    ).order_by('fecha')

                    cuotas_a_borrar = []

                    # -------------------------------------------------------------
                    # 3. DISTRIBUCIÓN EXACTA (ALGORITMO DE LLENADO)
                    # -------------------------------------------------------------
                    for cta in cuotas_para_recalcular:
                        
                        # Si ya distribuimos todo el dinero, esta cuota sobra -> Se borra
                        if saldo_por_distribuir <= 0:
                            cuotas_a_borrar.append(cta.idcta)
                            continue

                        # --- A. Calcular Intereses de esta cuota sobre el saldo actual ---
                        if cta.intcte > 0:
                            intcte_nuevo = round(saldo_por_distribuir * tasa)
                        else:
                            intcte_nuevo = 0

                        # --- B. Calcular cuánto Capital abonaría esta cuota normalmente ---
                        # Fórmula: La cuota fija que paga el cliente MENOS los intereses generados
                        cap_calculado = cta.cuota - intcte_nuevo

                        # --- C. EL MOMENTO DE LA VERDAD (Ajuste Final) ---
                        # Verificamos si este capital calculado es mayor o igual a lo que falta.
                        # O si el capital calculado es negativo/cero (amortización negativa).
                        
                        if cap_calculado >= saldo_por_distribuir:
                            # ES LA ÚLTIMA CUOTA REAL
                            # Ajustamos el capital EXACTAMENTE a lo que falta para cerrar el crédito
                            cta.capital = saldo_por_distribuir
                            cta.intcte = intcte_nuevo
                            
                            # La cuota a pagar cambia en este último pago para cuadrar
                            cta.cuota = cta.capital + cta.intcte
                            
                            cta.save()
                            
                            # Ya no queda deuda
                            saldo_por_distribuir = 0
                            
                        else:
                            # ES UNA CUOTA NORMAL (INTERMEDIA)
                            cta.capital = cap_calculado
                            cta.intcte = intcte_nuevo
                            
                            # Mantenemos la cuota fija original
                            # cta.cuota se queda igual
                            
                            cta.save()
                            
                            # Restamos lo que acabamos de asignar del saldo global
                            saldo_por_distribuir -= cap_calculado

                    # 4. Limpieza de cuotas excedentes
                    if cuotas_a_borrar:
                        PlanPagos.objects.using(proyecto).filter(
                            adj=adj,
                            idcta__in=cuotas_a_borrar
                        ).delete()

                # VALIDACIÓN POST-PROCESO: Verificar que todo cuadre
                # Se ejecuta SIEMPRE al final (haya o no abono a capital)
                es_valido, errores_validacion = _validar_integridad_post_abono(proyecto, adj)

                if not es_valido:
                    # Preparar mensaje de error detallado
                    mensajes_error = []
                    for error in errores_validacion:
                        mensajes_error.append(f"• {error['message']}")

                    mensaje_completo = "ABONO A CAPITAL RECHAZADO - Validación falló:\n\n" + "\n".join(mensajes_error)
                    mensaje_completo += f"\n\nAdjudicación: {adj}"
                    mensaje_completo += f"\nRecibo: {nro_recibo}"
                    mensaje_completo += f"\nValor total: ${valor_recibo:,.0f}"
                    mensaje_completo += "\n\nTODOS LOS CAMBIOS HAN SIDO REVERTIDOS (incluyendo pagos de cuotas vencidas)."

                    # Lanzar excepción para forzar rollback
                    raise ValidationError(mensaje_completo)

                # Si llegamos aquí, la validación pasó → commit automático al salir del with

        except ValidationError as e:
            # El rollback ya ocurrió automáticamente - TODO fue revertido
            context_updates['alerta'] = True
            context_updates['titulo'] = 'Abono a capital rechazado'
            context_updates['mensaje'] = str(e)
            context_updates['link'] = False
            alerts.append({
                'code': 'ABONO_CAPITAL_VALIDATION_FAILED',
                'message': str(e),
                'severity': 'error',
                'errors': errores_validacion if 'errores_validacion' in locals() else []
            })
            return context_updates, alerts, True

        except Exception as e:
            # Cualquier otro error también hace rollback de TODO
            context_updates['alerta'] = True
            context_updates['titulo'] = 'Error en abono a capital'
            context_updates['mensaje'] = f'Ocurrió un error inesperado durante el abono a capital: {str(e)}. Todos los cambios fueron revertidos (el recibo NO fue creado).'
            context_updates['link'] = False
            alerts.append({
                'code': 'ABONO_CAPITAL_UNEXPECTED_ERROR',
                'message': str(e),
                'severity': 'error'
            })
            return context_updates, alerts, True

    # Procesamiento normal de cuotas (no abono a capital)
    else:
        for cuota in saldo_cuotas:
            if valor_pagado == 0:
                break
            else:
                if not abonocapital:
                    datos_mora = PlanPagos.objects.using(proyecto
                            ).get(idcta=cuota.idcta
                                ).mora(dia_pago=fecha_pago)
                    if cuota.idcta[:2] == 'CI':
                        es_ci = True
                    elif cuota.idcta[:2] == 'FN':
                        es_fn = True
                    elif cuota.idcta[:2] == 'CE':
                        es_ce = True
                    elif cuota.idcta[:2] == 'CO':
                        es_co = True
                    capital = cuota.saldocapital
                    intcte = cuota.saldointcte
                    intmora = round(datos_mora.get('valor') * cobro_mora, 0)
                    if valor_pagado >= intmora:
                        mora_pagada = intmora
                        valor_pagado -= intmora
                    else:
                        mora_pagada = valor_pagado
                        valor_pagado -= valor_pagado
                    if valor_pagado >= intcte:
                        intcte_pagado = intcte
                        valor_pagado -= intcte
                    else:
                        intcte_pagado = valor_pagado
                        valor_pagado -= valor_pagado
                    if valor_pagado >= capital:
                        capital_pagado = capital
                        valor_pagado -= capital
                    else:
                        capital_pagado = valor_pagado
                        valor_pagado -= valor_pagado
                    Recaudos.objects.using(proyecto).create(recibo=nro_recibo,
                                                            fecha=fecha,
                                                            idcta=cuota.idcta,
                                                            idadjudicacion=adj,
                                                            capital=capital_pagado,
                                                            interescte=intcte_pagado,
                                                            interesmora=mora_pagada,
                                                            moralqd=cuota.saldomora,
                                                            fechaoperacion=datetime.datetime.today(),
                                                            usuario=request.user,
                                                            estado='Aprobado')
        
    if es_ci:
        operacion = 'Cuota Inicial'
    elif es_fn:
        operacion = 'Financiacion'
    elif es_ce:
        operacion = 'Extraordinaria'
    elif es_co:
        operacion = 'Contado'
    else:
        operacion = 'Abono Capital'
    Recaudos_general.objects.using(proyecto).create(idadjudicacion=adj,
                                                    fecha=fecha,
                                                    fecha_pago=fecha_pago,
                                                    numrecibo=nro_recibo,
                                                    idtercero=titulares.IdTercero1,
                                                    operacion=operacion,
                                                    valor=valor_recibo,
                                                    formapago=str(forma_pago),
                                                    concepto=concepto,
                                                    usuario=request.user)
    consecutivo.consecutivo = consecutivo.consecutivo + 1
    consecutivo.save()
    obj_saldos = saldos_adj.objects.using(proyecto).filter(adj=adj)
    saldos = obj_saldos.aggregate(Sum('saldocuota'))
    saldos = saldos['saldocuota__sum']
    if saldos <= 0:
        obj_adj = Adjudicacion.objects.using(proyecto).get(idadjudicacion=adj)
        obj_adj.estado = 'Pagado'
        obj_adj.save()

    filename = f'Recibo_caja_{nro_recibo}_{proyecto}.pdf'
    ruta = settings.DIR_EXPORT + filename

    obj_recibo = Recaudos_general.objects.using(proyecto).get(numrecibo=nro_recibo)

    context_rc = {
        'recibo': obj_recibo
    }

    if proyecto == 'Oasis':
        pdf = pdf_gen_weasy(f'pdf/{proyecto}/recibo.html', context_rc, filename)
    else:
        pdf = pdf_gen(f'pdf/{proyecto}/recibo.html', context_rc, filename)

    file = pdf.get('root')

    if numsolic != '' and numsolic != None:
        obj_rec_int = recibos_internos.objects.using('default').get(pk=numsolic)
        obj_rec_int.recibo_asociado = nro_recibo
        obj_rec_int.usuario_confirma = request.user
        obj_rec_int.fecha_confirma = datetime.date.today()
        # Limpiar flag de revisión manual ya que el recibo fue procesado
        obj_rec_int.requiere_revision_manual = False
        obj_rec_int.motivo_revision = None
        obj_rec_int.save()

    # Marcar movimiento bancario como usado si se proporcionó
    movimiento_banco_id = form_recibo.cleaned_data.get('movimiento_banco_id')
    if movimiento_banco_id and movimiento_banco_id.strip():
        try:
            from accounting.models import egresos_banco
            from django.utils import timezone
            movimiento = egresos_banco.objects.using('default').get(pk=int(movimiento_banco_id))
            movimiento.usado_agente = True
            movimiento.fecha_uso_agente = timezone.now()
            movimiento.recibo_asociado_agente = nro_recibo
            movimiento.proyecto_asociado_agente = proyecto
            movimiento.save()
        except (egresos_banco.DoesNotExist, ValueError):
            pass  # Si no se encuentra el movimiento, continuar sin error

    context_updates['ruta_recibo'] = settings.DIR_DOWNLOADS + filename
    context_updates['nro_recibo'] = nro_recibo
    context_updates['grabado'] = True
    context_updates['alerta'] = True
    context_updates['titulo'] = '¡Listo!'
    context_updates['mensaje'] = 'Descarga el recibo aquí'
    context_updates['link'] = True
    context_updates['ruta_link'] = settings.DIR_DOWNLOADS + filename
    context_updates['redireccion'] = True
    context_updates['redirect'] = f'/tesoreria/adjudicaciones/{proyecto}'
    processed_tokens.append(form_token)
    if len(processed_tokens) > 50:
        processed_tokens = processed_tokens[-50:]
    request.session['processed_receipt_tokens'] = processed_tokens
    form_token = uuid4().hex
    context_updates['form_token'] = form_token
    return context_updates, alerts, False


def _resolve_proyecto_alias(proyecto_param):
    raw_alias = (proyecto_param or '').strip()
    alias = raw_alias.replace('_',' ').replace('-', ' ').strip()
    if not alias:
        raise proyectos.DoesNotExist('Debes enviar el proyecto.')
    if alias in connections.databases:
        return alias
    for key in connections.databases.keys():
        if key.lower() == alias.lower():
            return key
    try:
        proyectos.objects.get(pk=alias)
        return alias
    except proyectos.DoesNotExist:
        match = proyectos.objects.filter(proyecto__iexact=alias).values_list('proyecto', flat=True).first()
        if match:
            return match.strip()
        compact = ' '.join(alias.split())
        if compact != alias:
            match = proyectos.objects.filter(proyecto__iexact=compact).values_list('proyecto', flat=True).first()
            if match:
                return match.strip()
        raise


def _find_adj_record(proyecto_alias, adj_param):
    adj_clean = (adj_param or '').strip()
    if not adj_clean:
        raise titulares_por_adj.DoesNotExist('Debes enviar el adj.')

    variants = []
    variants.append(adj_clean)
    variants.append(adj_clean.upper())
    variants.append(adj_clean.lower())
    if adj_clean.upper().startswith('ADJ'):
        variants.append(adj_clean.upper()[3:])
        variants.append(adj_clean[3:])
    else:
        variants.append(f'ADJ{adj_clean}')
        variants.append(f'adj{adj_clean}')

    for candidate in variants:
        candidate = candidate.strip()
        if not candidate:
            continue
        obj = titulares_por_adj.objects.using(proyecto_alias).filter(adj__iexact=candidate).first()
        if obj:
            return obj
    raise titulares_por_adj.DoesNotExist(f'No encontramos la adjudicación {adj_param} en {proyecto_alias}.')


def _get_recaudo_setup(proyecto, adj):
    proyecto_alias = _resolve_proyecto_alias(proyecto)
    titulares = _find_adj_record(proyecto_alias, adj)
    adj_real = titulares.adj.strip()

    try:
        datos_cliente = clientes.objects.using('default').get(idTercero=titulares.IdTercero1)
    except clientes.DoesNotExist:
        error = clientes.DoesNotExist(f'Cliente con idTercero={titulares.IdTercero1} no encontrado')
        error.idTercero = titulares.IdTercero1
        raise error

    try:
        consecutivo = consecutivos.objects.using(proyecto_alias).get(documento='RC')
    except consecutivos.DoesNotExist:
        error = consecutivos.DoesNotExist(f'Consecutivo RC no encontrado en proyecto {proyecto_alias}')
        error.documento = 'RC'
        error.proyecto = proyecto_alias
        raise error

    saldo_cuotas = saldos_adj.objects.using(proyecto_alias).filter(adj=adj_real, saldocuota__gt=0)
    saldos_cuotas = PlanPagos.objects.using(proyecto_alias).filter(adj=adj_real).order_by('fecha')

    try:
        obj_adj = Adjudicacion.objects.using(proyecto_alias).get(idadjudicacion=adj_real)
    except Adjudicacion.DoesNotExist:
        error = Adjudicacion.DoesNotExist(f'Adjudicación {adj_real} no encontrada en {proyecto_alias}')
        error.adj = adj_real
        error.proyecto = proyecto_alias
        raise error

    return proyecto_alias, adj_real, titulares, datos_cliente, consecutivo, saldo_cuotas, saldos_cuotas, obj_adj


def _parse_api_payload(request):
    content_type = (request.content_type or '').lower()
    if 'json' in content_type:
        try:
            raw_data = request.body.decode('utf-8') or '{}'
            data = json.loads(raw_data)
        except (ValueError, UnicodeDecodeError):
            raise ValueError('El cuerpo de la solicitud no es un JSON válido.')
    else:
        data = request.POST.dict()
    return data


def _normalize_api_receipt_payload(payload: dict) -> dict:
    """Normalize API-specific flags so they match the form expectations."""
    normalized = dict(payload)

    abono_value = normalized.pop('abono_capital', None)
    if abono_value is not None:
        normalized['abonocapital'] = 'on' if str(abono_value).lower() in ('1', 'true', 'on') else ''
    elif 'abonocapital' in normalized:
        normalized['abonocapital'] = 'on' if str(normalized['abonocapital']).lower() in ('1', 'true', 'on') else ''

    cond_value = normalized.get('condonacion_mora')
    if cond_value is not None:
        cond_truthy = str(cond_value).lower() in ('1', 'true', 'si', 'on')
        normalized['condonacion_mora'] = 'Si' if cond_truthy else 'No'
        normalized['condonacion_porc'] = '100' if cond_truthy else '0'

    # Normalizar fechas a formato YYYY-MM-DD
    date_fields = ['fecha_pago', 'fecha']
    for field in date_fields:
        if field in normalized and normalized[field]:
            fecha_str = str(normalized[field]).strip()
            if ' ' in fecha_str:
                # Formato: YYYY-MM-DD HH:MM:SS - extraer solo la fecha
                normalized[field] = fecha_str.split(' ')[0]
            elif len(fecha_str) == 8 and fecha_str.isdigit():
                # Formato: YYYYMMDD - convertir a YYYY-MM-DD
                try:
                    fecha_dt = datetime.datetime.strptime(fecha_str, '%Y%m%d')
                    normalized[field] = fecha_dt.strftime('%Y-%m-%d')
                except ValueError:
                    pass  # Dejar el valor original si falla el parseo

    return normalized


def _serialize_recaudo_detalle(detalle):
    serialized = []
    if not detalle:
        return serialized
    for row in detalle:
        fecha = row[0]
        if isinstance(fecha, (datetime.datetime, datetime.date)):
            fecha = fecha.strftime('%Y-%m-%d')
        serialized.append({
            'fecha': fecha,
            'cuota_id': row[1],
            'capital': int(row[2]),
            'interes_corriente': int(row[3]),
            'dias_mora': row[4],
            'valor_mora': int(row[5]),
            'total_aplicado': int(row[6]),
        })
    return serialized


def _serialize_totales(totales):
    if not totales:
        return {}
    return {
        'capital': int(totales[0]),
        'interes_corriente': int(totales[1]),
        'mora': int(totales[2]),
        'total': int(totales[3]),
    }


@group_perm_required(('andinasoft.add_recaudos_general',),raise_exception=True)
def nuevo_recaudo(request,proyecto,adj,):
    if not request.user.is_authenticated:
        return redirect('/accounts/login')

    proyecto_alias, adj_real, titulares, datos_cliente, consecutivo, saldo_cuotas, saldos_cuotas, obj_adj = _get_recaudo_setup(proyecto, adj)
    proyecto = proyecto_alias
    adj = adj_real

    detalle_exists = Recaudos.objects.using(proyecto_alias).filter(
        idadjudicacion=adj_real,
        recibo=OuterRef('numrecibo')
    )
    tiene_recaudos_pendientes = Recaudos_general.objects.using(proyecto_alias).filter(
        idadjudicacion=adj_real
    ).annotate(
        aplicado=Exists(detalle_exists)
    ).filter(aplicado=False).exists()

    form_token = request.POST.get('form_token') or uuid4().hex
    numsolic_post = request.POST.get('numsolicitud') if request.method == 'POST' else None
    allow_pending_flow = False
    if request.method == 'POST':
        allow_pending_flow = bool(request.POST.get('validar-recaudo') or (numsolic_post and str(numsolic_post).strip() != ''))

    bloquear_formulario = tiene_recaudos_pendientes and not allow_pending_flow
    context={
        'form':form_nuevo_recibo(proyecto=proyecto_alias),
        'adj':adj_real,
        'titulares':titulares,
        'num_recibo':consecutivo,
        'saldo_cuotas':saldos_cuotas,
        'count_cuota':0,
        'proyecto':proyecto_alias,
        'tiene_recaudos_pendientes':tiene_recaudos_pendientes,
        'form_token':form_token,
        'bloquear_formulario':bloquear_formulario
    }
    if bloquear_formulario:
        context['mensaje_bloqueo']='Existen recaudos pendientes por validar. Debes aplicarlos antes de generar un nuevo recibo.'

    if request.method == 'POST':
        if tiene_recaudos_pendientes and not allow_pending_flow:
            return render(request,'nuevo_recaudo.html',context)

        form_recibo = form_nuevo_recibo(request.POST or None,proyecto=proyecto)
        validation_alerts = []
        creation_alerts = []

        if request.POST.get('validar-recaudo'):
            context, validation_alerts = _validar_recaudo_context(
                request=request,
                proyecto=proyecto,
                adj=adj,
                titulares=titulares,
                saldos_cuotas=saldos_cuotas,
                consecutivo=consecutivo,
                form_token=form_token,
                post_data=request.POST
            )

        if request.POST.get('guardar-recaudo'):
            updates, creation_alerts, should_return = _guardar_recaudo(
                request=request,
                proyecto=proyecto,
                adj=adj,
                titulares=titulares,
                saldo_cuotas=saldo_cuotas,
                saldos_cuotas=saldos_cuotas,
                consecutivo=consecutivo,
                obj_adj=obj_adj,
                form_recibo=form_recibo,
                form_token=form_token
            )
            context.update(updates)
            if should_return:
                return render(request,'nuevo_recaudo.html',context)

        if validation_alerts:
            context['validation_alerts'] = validation_alerts
        if creation_alerts:
            context['creation_alerts'] = creation_alerts

    return render(request,'nuevo_recaudo.html',context)

@csrf_exempt
@api_token_auth
@require_http_methods(["POST"])
def api_validate_recaudo(request):
    if not request.user.is_authenticated or request.user.is_anonymous:
        return JsonResponse({'detail': 'Token inválido o no autenticado'}, status=401)
    try:
        payload = _normalize_api_receipt_payload(_parse_api_payload(request))
    except ValueError as exc:
        return JsonResponse({'error': str(exc)}, status=400)

    proyecto = payload.get('proyecto')
    adj = payload.get('adj')
    if not proyecto or not adj:
        return JsonResponse({'error': 'Debes enviar los campos proyecto y adj.'}, status=400)
    print(proyecto, adj)
    try:
        proyecto_alias, adj_real, titulares, datos_cliente, consecutivo, saldo_cuotas, saldos_cuotas, obj_adj = _get_recaudo_setup(proyecto, adj)
    except proyectos.DoesNotExist:
        return JsonResponse({'error': f'No encontramos el proyecto "{proyecto}".'}, status=404)
    except titulares_por_adj.DoesNotExist as e:
        return JsonResponse({'error': f'No encontramos la adjudicación "{adj}" en el proyecto "{proyecto}". Detalle: {str(e)}'}, status=404)
    except clientes.DoesNotExist as e:
        id_tercero = getattr(e, 'idTercero', 'desconocido')
        return JsonResponse({'error': f'El cliente asociado a la adjudicación no existe en la base de datos. IdTercero: {id_tercero}'}, status=404)
    except consecutivos.DoesNotExist as e:
        documento = getattr(e, 'documento', 'RC')
        proyecto_err = getattr(e, 'proyecto', proyecto)
        return JsonResponse({'error': f'No existe el consecutivo "{documento}" configurado para el proyecto "{proyecto_err}".'}, status=404)
    except Adjudicacion.DoesNotExist as e:
        adj_err = getattr(e, 'adj', adj)
        proyecto_err = getattr(e, 'proyecto', proyecto)
        return JsonResponse({'error': f'No encontramos el registro de adjudicación "{adj_err}" en el proyecto "{proyecto_err}".'}, status=404)

    form_token = payload.get('form_token') or uuid4().hex
    context, alerts = _validar_recaudo_context(
        request=request,
        proyecto=proyecto_alias,
        adj=adj_real,
        titulares=titulares,
        saldos_cuotas=saldos_cuotas,
        consecutivo=consecutivo,
        form_token=form_token,
        post_data=payload
    )
    data = {
        'alerts': alerts,
        'form_token': form_token,
        'verif_valor': context.get('verif_valor'),
        'count_cuota': context.get('count_cuota'),
        'detalle_recaudo': _serialize_recaudo_detalle(context.get('detalle_recaudo')),
        'totales': _serialize_totales(context.get('totales')),
        'valor_recibo': context.get('valor_recibo'),
        'abono_capital': context.get('abono_capital'),
        'condonacion': context.get('condonacion'),
        'procentaje_condonado': context.get('procentaje_condonado'),
    }
    return JsonResponse(data, status=200)


@csrf_exempt
@api_token_auth
@require_http_methods(["POST"])
def api_crear_recaudo(request):
    if not request.user.is_authenticated or request.user.is_anonymous:
        return JsonResponse({'detail': 'Token inválido o no autenticado'}, status=401)
    try:
        payload = _normalize_api_receipt_payload(_parse_api_payload(request))
    except ValueError as exc:
        return JsonResponse({'error': str(exc)}, status=400)

    proyecto = payload.get('proyecto')
    adj = payload.get('adj')
    if not proyecto or not adj:
        return JsonResponse({'error': 'Debes enviar los campos proyecto y adj.'}, status=400)

    try:
        proyecto_alias, adj_real, titulares, datos_cliente, consecutivo, saldo_cuotas, saldos_cuotas, obj_adj = _get_recaudo_setup(proyecto, adj)
    except proyectos.DoesNotExist:
        return JsonResponse({'error': f'No encontramos el proyecto "{proyecto}".'}, status=404)
    except titulares_por_adj.DoesNotExist as e:
        return JsonResponse({'error': f'No encontramos la adjudicación "{adj}" en el proyecto "{proyecto}". Detalle: {str(e)}'}, status=404)
    except clientes.DoesNotExist as e:
        id_tercero = getattr(e, 'idTercero', 'desconocido')
        return JsonResponse({'error': f'El cliente asociado a la adjudicación no existe en la base de datos. IdTercero: {id_tercero}'}, status=404)
    except consecutivos.DoesNotExist as e:
        documento = getattr(e, 'documento', 'RC')
        proyecto_err = getattr(e, 'proyecto', proyecto)
        return JsonResponse({'error': f'No existe el consecutivo "{documento}" configurado para el proyecto "{proyecto_err}".'}, status=404)
    except Adjudicacion.DoesNotExist as e:
        adj_err = getattr(e, 'adj', adj)
        proyecto_err = getattr(e, 'proyecto', proyecto)
        return JsonResponse({'error': f'No encontramos el registro de adjudicación "{adj_err}" en el proyecto "{proyecto_err}".'}, status=404)

    form_token = payload.get('form_token') or uuid4().hex
    form_data = QueryDict('', mutable=True)

    # Campos obligatorios que deben ser llenados automáticamente
    form_data.update({
        'id_adj': adj_real,
        'numrecibo': str(consecutivo.consecutivo),
        'nombre_cliente': datos_cliente.nombrecompleto,
        'idtercero': datos_cliente.idTercero,
    })

    # Agregar los datos del payload
    for key, value in payload.items():
        if isinstance(value, list):
            for item in value:
                form_data.update({key: item})
        else:
            form_data.update({key: value})

    # Mapear forma_pago si viene en el payload
    if 'forma_pago' in payload:
        forma_pago_nombre = payload['forma_pago']
        try:
            # Buscar la forma de pago en el proyecto (case-insensitive)
            forma = formas_pago.objects.using(proyecto_alias).filter(
                descripcion__iexact=forma_pago_nombre
            ).first()

            if forma:
                form_data['forma_pago'] = forma.descripcion
            else:
                # Si no se encuentra, intentar buscar por coincidencia parcial
                forma = formas_pago.objects.using(proyecto_alias).filter(
                    descripcion__icontains=forma_pago_nombre
                ).first()

                if forma:
                    form_data['forma_pago'] = forma.descripcion
                else:
                    return JsonResponse({
                        'error': f'No se encontró la forma de pago "{forma_pago_nombre}" en el proyecto "{proyecto_alias}". Usa /api/formas-pago?proyecto={proyecto_alias} para ver las opciones disponibles.'
                    }, status=400)
        except Exception as e:
            return JsonResponse({
                'error': f'Error al buscar la forma de pago: {str(e)}'
            }, status=500)

    form_recibo = form_nuevo_recibo(form_data, proyecto=proyecto_alias)
    updates, alerts, should_return = _guardar_recaudo(
        request=request,
        proyecto=proyecto_alias,
        adj=adj_real,
        titulares=titulares,
        saldo_cuotas=saldo_cuotas,
        saldos_cuotas=saldos_cuotas,
        consecutivo=consecutivo,
        obj_adj=obj_adj,
        form_recibo=form_recibo,
        form_token=form_token
    )
    if updates.get('form_errors'):
        return JsonResponse({'alerts': alerts, 'errors': updates['form_errors']}, status=400)
    if should_return and alerts:
        return JsonResponse({'alerts': alerts}, status=400)
    response_data = {
        'alerts': alerts,
        'nro_recibo': updates.get('nro_recibo'),
        'pdf_url': updates.get('ruta_link'),
        'ruta_recibo': updates.get('ruta_recibo'),
        'mensaje': updates.get('mensaje'),
        'redirect': updates.get('redirect'),
        'form_token': updates.get('form_token', form_token)
    }
    status_code = 201 if updates.get('nro_recibo') else 202
    return JsonResponse(response_data, status=status_code)

@csrf_exempt
@api_token_auth
@require_http_methods(["GET"])
def api_formas_pago(request):
    """
    Endpoint para obtener las formas de pago disponibles de un proyecto.
    GET /api/formas-pago?proyecto=Tesoro Escondido
    """
    if not request.user.is_authenticated or request.user.is_anonymous:
        return JsonResponse({'detail': 'Token inválido o no autenticado'}, status=401)

    proyecto = request.GET.get('proyecto')
    if not proyecto:
        return JsonResponse({'error': 'Debes enviar el parámetro "proyecto".'}, status=400)

    try:
        proyecto_alias = _resolve_proyecto_alias(proyecto)
    except proyectos.DoesNotExist:
        return JsonResponse({'error': f'No encontramos el proyecto "{proyecto}".'}, status=404)

    try:
        formas = formas_pago.objects.using(proyecto_alias).all().order_by('descripcion')
        formas_list = []
        for forma in formas:
            formas_list.append({
                'id': forma.descripcion,
                'descripcion': forma.descripcion,
                'cuenta_banco': forma.cuenta_banco,
                'cuenta_contable': forma.cuenta_contable,
            })

        return JsonResponse({
            'proyecto': proyecto_alias,
            'total': len(formas_list),
            'formas_pago': formas_list
        }, status=200)
    except Exception as e:
        return JsonResponse({'error': f'Error al obtener las formas de pago: {str(e)}'}, status=500)


def _serialize_cliente(cliente_obj):
    """Serializa información básica de un cliente/tercero."""
    if not cliente_obj:
        return None
    return {
        'id': cliente_obj.idTercero,
        'tipo_documento': cliente_obj.tipo_doc,
        'nombre_completo': cliente_obj.nombrecompleto,
        'nombres': cliente_obj.nombres,
        'apellidos': cliente_obj.apellidos,
        'celular': cliente_obj.celular1,
        'celular2': cliente_obj.celular2,
        'telefono': cliente_obj.telefono1,
        'email': cliente_obj.email,
        'domicilio': cliente_obj.domicilio,
        'ciudad': cliente_obj.ciudad,
        'pais': cliente_obj.pais.country_name if cliente_obj.pais else None,
        'departamento': cliente_obj.estado.state_name if cliente_obj.estado else None,
        'ciudad_nombre': cliente_obj.city.city_name if cliente_obj.city else None,
        'fecha_nacimiento': cliente_obj.fecha_nac.isoformat() if cliente_obj.fecha_nac else None,
        'lugar_nacimiento': cliente_obj.lugar_nac,
        'nacionalidad': cliente_obj.nacionalidad,
        'ocupacion': cliente_obj.ocupacion,
        'estado_civil': cliente_obj.estado_civil,
    }


def _serialize_sagrilaft(cliente_id):
    """Serializa información sagrilaft de un cliente si está disponible."""
    try:
        sagrilaft = sagrilaft_info.objects.get(cliente_id=cliente_id)
        return {
            'empresa_labora': sagrilaft.empresa_labora,
            'cargo_actual': sagrilaft.cargo_actual,
            'origen_ingresos': sagrilaft.origen_ingresos,
            'declara_renta': sagrilaft.declara_renta,
            'tiene_rut': sagrilaft.tiene_rut,
            'codigo_ciuu': sagrilaft.codigo_ciuu.codigo if sagrilaft.codigo_ciuu else None,
            'es_peps': sagrilaft.es_peps,
            'peps_desde': sagrilaft.peps_desde.isoformat() if sagrilaft.peps_desde else None,
            'peps_hasta': sagrilaft.peps_hasta.isoformat() if sagrilaft.peps_hasta else None,
            'peps_entidad': sagrilaft.peps_entidad,
            'peps_cargo': sagrilaft.peps_cargo,
            'peps_familiar': sagrilaft.peps_familiar,
            'peps_familiar_parentesco': sagrilaft.peps_familiar_parentesco,
            'peps_familiar_entidad': sagrilaft.peps_familiar_entidad,
            'peps_familiar_cargo': sagrilaft.peps_familiar_cargo,
            'referencia_familiar': sagrilaft.referencia_familiar,
            'referencia_familiar_telefono': sagrilaft.referencia_familiar_telefono,
            'referencia_personal': sagrilaft.referencia_personal,
            'referencia_personal_telefono': sagrilaft.referencia_personal_telefono,
        }
    except sagrilaft_info.DoesNotExist:
        return None


def _serialize_inmueble(inmueble_obj):
    """Serializa información de un inmueble sin colindantes."""
    if not inmueble_obj:
        return None
    return {
        'id': inmueble_obj.idinmueble,
        'etapa': inmueble_obj.etapa,
        'manzana': inmueble_obj.manzananumero,
        'lote': inmueble_obj.lotenumero,
        'matricula': inmueble_obj.matricula,
        'estado': inmueble_obj.estado,
        'area_privada': float(inmueble_obj.areaprivada) if inmueble_obj.areaprivada else None,
        'area_construida': float(inmueble_obj.areaconstruida) if inmueble_obj.areaconstruida else None,
        'area_lote': float(inmueble_obj.area_lt) if inmueble_obj.area_lt else None,
        'area_manzana': float(inmueble_obj.area_mz) if inmueble_obj.area_mz else None,
    }


def _serialize_titular_completo(cliente_id):
    """Serializa información completa de un titular incluyendo sagrilaft."""
    try:
        cliente = clientes.objects.get(pk=cliente_id)
        data = _serialize_cliente(cliente)
        data['sagrilaft'] = _serialize_sagrilaft(cliente_id)
        return data
    except clientes.DoesNotExist:
        return None


def _serialize_adjudicacion_basica(adj_obj):
    """Serializa información básica de una adjudicación."""
    # Obtener titulares con id y nombre
    titulares = []
    for tercero_id in [adj_obj.idtercero1, adj_obj.idtercero2, adj_obj.idtercero3, adj_obj.idtercero4]:
        if tercero_id and tercero_id.strip():
            try:
                cliente = clientes.objects.get(pk=tercero_id)
                titulares.append({
                    'id': cliente.idTercero,
                    'nombre': cliente.nombrecompleto
                })
            except clientes.DoesNotExist:
                titulares.append({
                    'id': tercero_id,
                    'nombre': None
                })

    return {
        'id': adj_obj.idadjudicacion,
        'fecha': adj_obj.fecha.isoformat() if adj_obj.fecha else None,
        'tipo_contrato': adj_obj.tipocontrato,
        'contrato': adj_obj.contrato,
        'estado': adj_obj.estado,
        'inmueble_id': adj_obj.idinmueble,
        'titulares': titulares,
    }


def _serialize_adjudicacion_detalle(adj_obj, proyecto_alias):
    """Serializa información completa de una adjudicación con relaciones."""
    data = {
        'id': adj_obj.idadjudicacion,
        'fecha': adj_obj.fecha.isoformat() if adj_obj.fecha else None,
        'tipo_contrato': adj_obj.tipocontrato,
        'contrato': adj_obj.contrato,
        'estado': adj_obj.estado,
        'fecha_radicacion': adj_obj.fecharadicacion.isoformat() if adj_obj.fecharadicacion else None,
        'fecha_contrato': adj_obj.fechacontrato.isoformat() if adj_obj.fechacontrato else None,
        'fecha_desistimiento': adj_obj.fechadesistimiento.isoformat() if adj_obj.fechadesistimiento else None,
        'origen_venta': adj_obj.origenventa,
        'tipo_origen': adj_obj.tipoorigen,
        'oficina': adj_obj.oficina,
        'es_juridico': bool(adj_obj.es_juridico) if adj_obj.es_juridico is not None else None,
    }

    # Titulares con info completa y sagrilaft
    titulares = []
    for idx, tercero_id in enumerate([adj_obj.idtercero1, adj_obj.idtercero2, adj_obj.idtercero3, adj_obj.idtercero4], 1):
        if tercero_id and tercero_id.strip():
            titular_data = _serialize_titular_completo(tercero_id)
            if titular_data:
                titular_data['posicion'] = idx
                titulares.append(titular_data)
    data['titulares'] = titulares

    # Inmueble
    try:
        inmueble = Inmuebles.objects.using(proyecto_alias).get(idinmueble=adj_obj.idinmueble)
        data['inmueble'] = _serialize_inmueble(inmueble)
    except Inmuebles.DoesNotExist:
        data['inmueble'] = None

    return data


@csrf_exempt
@api_token_auth
@require_http_methods(["GET"])
def api_adjudicaciones(request):
    """
    Endpoint para obtener adjudicaciones de un proyecto.

    GET /api/adjudicaciones?proyecto=<proyecto>
        Retorna lista de adjudicaciones con información básica.

    GET /api/adjudicaciones?proyecto=<proyecto>&id=<ADJXX>
        Retorna detalle completo de una adjudicación específica con:
        - Información de la adjudicación
        - Titulares con info básica y sagrilaft
        - Inmueble (matrícula, áreas, sin colindantes)
    """
    if not request.user.is_authenticated or request.user.is_anonymous:
        return JsonResponse({'detail': 'Token inválido o no autenticado'}, status=401)

    proyecto = request.GET.get('proyecto')
    adj_id = request.GET.get('id')

    if not proyecto:
        return JsonResponse({'error': 'Debes enviar el parámetro "proyecto".'}, status=400)

    try:
        proyecto_alias = _resolve_proyecto_alias(proyecto)
    except proyectos.DoesNotExist:
        return JsonResponse({'error': f'No encontramos el proyecto "{proyecto}".'}, status=404)

    try:
        if adj_id:
            # Detalle de una adjudicación específica (solo Aprobado o Pagado)
            adj_clean = adj_id.strip().upper()
            if not adj_clean.startswith('ADJ'):
                adj_clean = f'ADJ{adj_clean}'

            try:
                adj = Adjudicacion.objects.using(proyecto_alias).filter(
                    estado__in=['Aprobado', 'Pagado']
                ).get(idadjudicacion__iexact=adj_clean)
            except Adjudicacion.DoesNotExist:
                # Intentar sin el prefijo ADJ
                adj_clean = adj_id.strip()
                try:
                    adj = Adjudicacion.objects.using(proyecto_alias).filter(
                        estado__in=['Aprobado', 'Pagado']
                    ).get(idadjudicacion__iexact=adj_clean)
                except Adjudicacion.DoesNotExist:
                    return JsonResponse({'error': f'No encontramos la adjudicación "{adj_id}" en {proyecto_alias} o no está en estado Aprobado/Pagado.'}, status=404)

            return JsonResponse({
                'proyecto': proyecto_alias,
                'adjudicacion': _serialize_adjudicacion_detalle(adj, proyecto_alias)
            }, status=200)
        else:
            # Lista de adjudicaciones con info básica (solo Aprobado o Pagado)

            # Parámetros de paginación
            page = int(request.GET.get('page', 1))
            page_size = int(request.GET.get('page_size', request.GET.get('limit', 100)))
            page_size = min(page_size, 500)  # Máximo 500 registros por página
            offset = int(request.GET.get('offset', (page - 1) * page_size))

            # Parámetros de filtro incremental
            updated_since = request.GET.get('updated_since')
            since_id = request.GET.get('since_id')

            # Parámetro de ordenamiento
            order = request.GET.get('order', '-fecha')
            valid_orders = ['fecha', '-fecha', 'idadjudicacion', '-idadjudicacion']
            if order not in valid_orders:
                order = '-fecha'

            # Query base
            adjudicaciones = Adjudicacion.objects.using(proyecto_alias).filter(
                estado__in=['Aprobado', 'Pagado']
            )

            # Aplicar filtro updated_since (fecha >= timestamp)
            if updated_since:
                try:
                    from django.utils.dateparse import parse_datetime
                    dt = parse_datetime(updated_since)
                    if dt:
                        adjudicaciones = adjudicaciones.filter(fecha__gte=dt)
                except (ValueError, TypeError):
                    pass

            # Aplicar filtro since_id (id > since_id para paginación por cursor)
            if since_id:
                adjudicaciones = adjudicaciones.filter(idadjudicacion__gt=since_id)

            # Contar total antes de paginar
            total_count = adjudicaciones.count()

            # Aplicar ordenamiento y paginación
            adjudicaciones = adjudicaciones.order_by(order)[offset:offset + page_size]

            adj_list = [_serialize_adjudicacion_basica(adj) for adj in adjudicaciones]

            # Calcular info de paginación
            total_pages = (total_count + page_size - 1) // page_size if page_size > 0 else 1

            return JsonResponse({
                'proyecto': proyecto_alias,
                'pagination': {
                    'page': page,
                    'page_size': page_size,
                    'total_pages': total_pages,
                    'total_records': total_count,
                },
                'filters': {
                    'updated_since': updated_since,
                    'since_id': since_id,
                    'order': order,
                },
                'total': len(adj_list),
                'adjudicaciones': adj_list
            }, status=200)
    except Exception as e:
        return JsonResponse({'error': f'Error al obtener adjudicaciones: {str(e)}'}, status=500)


@csrf_exempt
@api_token_auth
@require_http_methods(["GET"])
def api_terceros(request):
    """
    Endpoint para consultar terceros (clientes).

    GET /api/terceros
        Retorna listado paginado (max 15 por página) con info básica.

    GET /api/terceros?id=<idTercero>
        Retorna detalle completo del tercero con sagrilaft.
    """
    if not request.user.is_authenticated or request.user.is_anonymous:
        return JsonResponse({'detail': 'Token inválido o no autenticado'}, status=401)

    tercero_id = request.GET.get('id')

    try:
        if tercero_id:
            # Detalle de un tercero específico
            try:
                cliente = clientes.objects.get(pk=tercero_id.strip())
            except clientes.DoesNotExist:
                return JsonResponse({'error': f'No encontramos el tercero "{tercero_id}".'}, status=404)

            data = _serialize_cliente(cliente)
            data['sagrilaft'] = _serialize_sagrilaft(tercero_id.strip())

            return JsonResponse({
                'tercero': data
            }, status=200)
        else:
            # Listado paginado
            page = int(request.GET.get('page', 1))
            page_size = int(request.GET.get('page_size', request.GET.get('limit', 15)))
            page_size = min(page_size, 15)
            offset = int(request.GET.get('offset', (page - 1) * page_size))

            # Filtro por nombre
            search = request.GET.get('search')

            # Query base
            terceros_qs = clientes.objects.all()

            if search:
                terceros_qs = terceros_qs.filter(nombrecompleto__icontains=search.strip())

            # Ordenamiento
            order = request.GET.get('order', 'nombrecompleto')
            valid_orders = ['nombrecompleto', '-nombrecompleto', 'idTercero', '-idTercero', 'fecha_actualizacion', '-fecha_actualizacion']
            if order not in valid_orders:
                order = 'nombrecompleto'

            total_count = terceros_qs.count()
            terceros_qs = terceros_qs.order_by(order)[offset:offset + page_size]

            terceros_list = []
            for c in terceros_qs:
                terceros_list.append({
                    'id': c.idTercero,
                    'tipo_documento': c.tipo_doc,
                    'nombre_completo': c.nombrecompleto,
                    'celular': c.celular1,
                    'email': c.email,
                    'ciudad': c.ciudad,
                })

            total_pages = (total_count + page_size - 1) // page_size if page_size > 0 else 1

            return JsonResponse({
                'pagination': {
                    'page': page,
                    'page_size': page_size,
                    'total_pages': total_pages,
                    'total_records': total_count,
                },
                'total': len(terceros_list),
                'terceros': terceros_list
            }, status=200)
    except Exception as e:
        return JsonResponse({'error': f'Error al obtener terceros: {str(e)}'}, status=500)


@group_perm_required(perms=('andinasoft.view_recaudos_general',),raise_exception=True)
def lista_recaudos(request):
    if request.user.is_authenticated:
        alerta=False
        def _get_formas_pago_choices(project_alias):
            if not project_alias:
                return []
            alias = str(project_alias)
            try:
                qs = formas_pago.objects.using(alias).all().order_by('descripcion')
            except Exception:
                return []
            return [(fp.descripcion, fp.descripcion) for fp in qs]
        post_data = request.POST if request.method == 'POST' else None
        selected_project_value = request.POST.get('proyecto') or request.POST.get('Proyecto')
        form_modificaciones = form_modificar_recibo(post_data, formas_pago_choices=_get_formas_pago_choices(selected_project_value))
        context={"form":form_lista_recaudos(),
                 "form_modificaciones":form_modificaciones
                 }
        if request.method == 'POST':
            form=form_lista_recaudos(request.POST)
            if request.POST.get('boton-buscar'):
                if form.is_valid():
                    proyecto=form.cleaned_data.get('Proyecto')
                    fecha_desde=form.cleaned_data.get('Fecha_Desde')
                    fecha_hasta=form.cleaned_data.get('Fecha_Hasta')
                    recibos=Recaudos_general.objects.using(str(proyecto)).filter(fecha__gte=fecha_desde,fecha__lte=fecha_hasta)
                    fd=datetime.datetime.strftime(fecha_desde,"%Y-%m-%d")
                    fh=datetime.datetime.strftime(fecha_hasta,"%Y-%m-%d")
                    stmt=f'CALL verlistarecaudos("{fd}","{fh}");'
                    forma_raw=Recaudos_general.objects.using(str(proyecto)).raw(stmt)
                    recibos_list = list(forma_raw)
                    if recibos_list:
                        numeros = [rec.numrecibo for rec in recibos_list]
                        soportes_map = {}
                        qs_soportes = recibos_internos.objects.using('default').filter(
                            proyecto=proyecto,
                            recibo_asociado__in=numeros
                        )
                        for sol in qs_soportes:
                            if sol.soporte:
                                soportes_map[sol.recibo_asociado] = sol.soporte.url
                        for rec in recibos_list:
                            rec.soporte_finance = soportes_map.get(rec.numrecibo)
                    context['recibos']=recibos_list
                    context['proyecto']=proyecto
                    context['desde']=fecha_desde
                    context['formas_pago']=[label for value,label in _get_formas_pago_choices(proyecto)]
                    context['form']=form
                else: 
                    context['alerta']=True
                    context['mensaje_alerta']='La fecha de inicio no puede ser mayor a la inicial'
            context['form_modificaciones']=form_modificaciones
            if request.POST.get('reimprimir'):
                if form_modificaciones.is_valid():
                    proyecto=form_modificaciones.cleaned_data.get('proyecto')
                    nro_recibo=form_modificaciones.cleaned_data.get('nrorecibo')
                    datos_recaudo=Recaudos_general.objects.using(proyecto).get(numrecibo=nro_recibo)
                    idtitular1=datos_recaudo.idtercero
                    datos_cliente=clientes.objects.using('default').get(idTercero=idtitular1)
                    titular1=datos_cliente.nombrecompleto
                    fecha=datos_recaudo.fecha
                    concepto=datos_recaudo.concepto
                    usuario=datos_recaudo.usuario
                    if concepto==None: concepto='Abono a credito Lote'
                    valor_recibo=datos_recaudo.valor
                    forma_pago=datos_recaudo.formapago
                    filename=f'recibo_caja_{proyecto}_{nro_recibo}.pdf'
                    ruta=settings.DIR_EXPORT+f'recibo_caja_{proyecto}_{nro_recibo}.pdf'
                    
                    filename = f'Recibo_caja_{nro_recibo}_{proyecto}.pdf'
                    
                    obj_recibo = Recaudos_general.objects.using(proyecto).get(numrecibo = nro_recibo)
                                        
                    context = {
                        'recibo':obj_recibo
                    }
                                        
                    pdf = pdf_gen(f'pdf/{proyecto}/recibo.html', context,filename)
                    
                    file = pdf.get('root')
                                            
                    return FileResponse(open(file,'rb'),as_attachment=True,filename=filename)
                
            if request.POST.get('boton-modificar-fecha') or request.POST.get('boton-modificar-consec'):
                if not check_perms(request,('andinasoft.change_recaudos_general',),raise_exception=False):
                    raise PermissionDenied
                if form_modificaciones.is_valid():
                    nro_recibo=form_modificaciones.cleaned_data.get('nrorecibo')
                    proyecto=form_modificaciones.cleaned_data.get('proyecto')
                    fecha_modific=form_modificaciones.cleaned_data.get('fecha_modif')
                    consec_modific=form_modificaciones.cleaned_data.get('consecutivo_modif')
                    recibogral=Recaudos_general.objects.using(proyecto).get(numrecibo=nro_recibo)
                    recibodetalle=Recaudos.objects.using(proyecto).filter(recibo=nro_recibo)
                    if request.POST.get('boton-modificar-fecha'):
                        recibogral.fecha=fecha_modific
                        for linea in recibodetalle:
                            linea.fecha=fecha_modific
                            linea.save()
                        mensaje=f'La fecha del recibo {nro_recibo} fue cambiada a {fecha_modific}'
                    if request.POST.get('boton-modificar-consec'):
                        recibogral.numrecibo=consec_modific
                        for linea in recibodetalle:
                            linea.recibo=consec_modific
                            linea.save()
                        mensaje=f'El consecutivo del recibo {nro_recibo} fue cambiado a {consec_modific}'
                    recibogral.save()
                    context['alerta']=True
                    context['mensaje_alerta']=mensaje
                else: 
                    context['alerta']=True
                    context['mensaje_alerta']='El consecutivo que intentas asignar ya existe'
            if request.POST.get('eliminar-regresando') or request.POST.get('eliminar-sinregresar') or request.POST.get('eliminar-anulando'):
                if not check_perms(request,('andinasoft.change_recaudos_general',),raise_exception=False):
                    raise PermissionDenied
                if form_modificaciones.is_valid():
                    proyecto=form_modificaciones.cleaned_data.get('proyecto')
                    nro_recibo=form_modificaciones.cleaned_data.get('nrorecibo')
                    recibogral=Recaudos_general.objects.using(proyecto).get(numrecibo=nro_recibo)
                    recibodetalle=Recaudos.objects.using(proyecto).filter(recibo=nro_recibo)
                    for linea in recibodetalle:
                        linea.delete()
                    if request.POST.get('eliminar-anulando'):
                        recibogral.idadjudicacion='ADJ0'
                        recibogral.idtercero=''
                        recibogral.operacion='Anulado'
                        recibogral.valor=0
                        recibogral.formapago=''
                        recibogral.save()
                        mensaje=f'El Recibo {nro_recibo} fue eliminado y el consecutivo anulado'
                    else:
                        recibogral.delete()
                    if request.POST.get('eliminar-regresando'):
                        prefijo='M'
                        consecutivo=consecutivos.objects.using(proyecto).get(documento='RC',prefijo=prefijo)
                        consecutivo.consecutivo-=1
                        consecutivo.save()
                        mensaje=f'El Recibo {nro_recibo} fue eliminado, el consecutivo actual es {consecutivo.consecutivo}'
                    elif request.POST.get('eliminar-sinregresar'):
                        mensaje=f'El Recibo {nro_recibo} fue eliminado'
                    context['alerta']=True
                    context['mensaje_alerta']=mensaje
            if request.POST.get('boton-modificar-forma'):
                if not check_perms(request,('andinasoft.change_recaudos_general',),raise_exception=False):
                    raise PermissionDenied
                if form_modificaciones.is_valid():
                    proyecto=form_modificaciones.cleaned_data.get('proyecto')
                    nro_recibo=form_modificaciones.cleaned_data.get('nrorecibo')
                    nueva_forma=form_modificaciones.cleaned_data.get('forma_pago_modif')
                    if not proyecto or not nueva_forma:
                        context['alerta']=True
                        context['mensaje_alerta']='Debes seleccionar la forma de pago y el proyecto.'
                    else:
                        if not formas_pago.objects.using(proyecto).filter(descripcion=nueva_forma).exists():
                            context['alerta']=True
                            context['mensaje_alerta']='La forma de pago seleccionada no existe en el proyecto.'
                        else:
                            recibogral=Recaudos_general.objects.using(proyecto).get(numrecibo=nro_recibo)
                            recibogral.formapago=nueva_forma
                            recibogral.save(update_fields=['formapago'])
                            context['alerta']=True
                            context['mensaje_alerta']=f'La forma de pago del recibo {nro_recibo} fue cambiada a {nueva_forma}'
                else:
                    context['alerta']=True
                    context['mensaje_alerta']='Completa la información para cambiar la forma de pago.'
            if request.POST.get('boton-modificar-valor'):
                if not check_perms(request,('andinasoft.change_recaudos_general',),raise_exception=False):
                    raise PermissionDenied
                if form_modificaciones.is_valid():
                    proyecto=form_modificaciones.cleaned_data.get('proyecto')
                    nro_recibo=form_modificaciones.cleaned_data.get('nrorecibo')
                    nuevo_valor=form_modificaciones.cleaned_data.get('valor_modif')
                    if not proyecto or nuevo_valor is None:
                        context['alerta']=True
                        context['mensaje_alerta']='Debes ingresar el nuevo valor y el proyecto.'
                    else:
                        tiene_detalles = Recaudos.objects.using(proyecto).filter(recibo=nro_recibo).exists()
                        if tiene_detalles:
                            context['alerta']=True
                            context['mensaje_alerta']='No se puede modificar el valor porque el recaudo ya tiene detalles aplicados.'
                        else:
                            recibogral=Recaudos_general.objects.using(proyecto).get(numrecibo=nro_recibo)
                            recibogral.valor=nuevo_valor
                            recibogral.save(update_fields=['valor'])
                            context['alerta']=True
                            context['mensaje_alerta']=f'El valor del recibo {nro_recibo} fue cambiado a {nuevo_valor}'
                else:
                    context['alerta']=True
                    context['mensaje_alerta']='Completa la información para cambiar el valor.'
        else:
            context['form_modificaciones']=form_modificaciones
        if 'formas_pago' not in context and selected_project_value:
            context['formas_pago']=[label for value,label in _get_formas_pago_choices(selected_project_value)]
        
        return render(request,'listado_recaudos.html',context)
    else:
        return redirect('/accounts/login')

@group_perm_required(('andinasoft.add_recaudos_general',),raise_exception=True)
def lista_adj_recaudos(request,proyecto):
    context={
        'proyecto':proyecto,
        'adjudicaciones':Vista_Adjudicacion.objects.using(proyecto).all().exclude(Q(Estado__istartswith="Desistido")|Q(Estado__istartswith="Pagado"))
    }
    
    return render(request,'listado_adj_recaudos.html',context)

@group_perm_required(perms=('andinasoft.view_adjudicacion',),raise_exception=True)
def detalle_adjudicacion(request,proyecto,adj):
    check_project(request,proyecto)
    alerta=False
    titulo=None
    mensaje=None
    dir_link=None
    link=False
    obj_adj=Adjudicacion.objects.using(proyecto).get(idadjudicacion=adj)
    proyecto_obj = proyectos.objects.get(pk=proyecto)
    procesable='Si'
    if obj_adj.p_enmendadura==1: procesable='No'
    if obj_adj.p_doccompleta==1: procesable='No'
    if obj_adj.p_valincorrectos==1: procesable='No'
    if request.is_ajax():
        if request.method == 'POST':
            todo = request.POST.get('todo')
            
            if todo == 'restore_bk':
                # Obtener ID del backup a restaurar
                bk_id = request.POST.get('bk_id')

                if bk_id:
                    # Restaurar backup específico
                    try:
                        obj_bk_selected = bk_bfchangeplan.objects.get(pk=bk_id, proyecto=proyecto, adj=adj)
                        bk = obj_bk_selected.pk
                    except bk_bfchangeplan.DoesNotExist:
                        return JsonResponse({'status': 404, 'message': 'Backup no encontrado'}, status=404)
                else:
                    # Fallback: restaurar último backup
                    obj_bk = bk_bfchangeplan.objects.filter(proyecto=proyecto, adj=adj)
                    if obj_bk.exists():
                        bk = obj_bk.last().pk
                    else:
                        return JsonResponse({'status': 404, 'message': 'No hay backups disponibles'}, status=404)

                if bk:
                    bk_plan = bk_planpagos.objects.filter(id_bk = bk)
                    #Borramos el plan de pagos vigentee introducimos el bk
                    actual_plan = PlanPagos.objects.using(proyecto).filter(
                        adj = adj
                    )
                    for i in actual_plan:
                        i.delete()
                    
                    for i in bk_plan:
                        PlanPagos.objects.using(proyecto).create(
                            idcta = i.idcta, tipocta = i.tipocta, nrocta = i.nrocta,
                            adj = i.adj, capital = i.capital, intcte = i.intcte,
                            cuota = i.cuota, fecha = i.fecha
                        )
                    
                    bk_rec = bk_recaudodetallado.objects.filter(id_bk = bk)
                    actual_rec = Recaudos.objects.using(proyecto).filter(idadjudicacion = adj)
                    
                    for i in actual_rec:
                        i.delete()
                    
                    for i in bk_rec:
                        Recaudos.objects.using(proyecto).create(
                            recibo = i.recibo, fecha = i.fecha, idcta = i.idcta,
                            idadjudicacion = i.idadjudicacion, capital = i.capital,
                            interescte = i.interescte, interesmora = i.interesmora,
                            moralqd = i.moralqd, fechaoperacion = i.fechaoperacion,
                            usuario = i.usuario, estado = i.estado
                        )
                    obj_timeline=timeline.objects.using(proyecto)
                    obj_timeline.create(adj=adj,
                                    fecha=datetime.date.today(),
                                    usuario=request.user,
                                    accion='Restauró backup de plan de pagos')
                    data = {
                        'msj':'Backup restaurado'
                    }
                    
                    return JsonResponse(data)
            
            elif todo == 'desvincular_recaudos':
                check_perms(request,('andinasoft.delete_recaudos',),raise_exception=True)
                recibos = request.POST.getlist('recibos[]')
                if not recibos:
                    return JsonResponse({'status':400,'message':'Debes seleccionar al menos un recaudo.'},status=400)
                
                recibos_desc = list(
                    Recaudos_general.objects.using(proyecto).filter(
                        idadjudicacion=adj
                    ).order_by('-fecha','-idrecaudo').values_list('numrecibo', flat=True)
                )
                if len(recibos) > len(recibos_desc):
                    return JsonResponse({'status':400,'message':'Los recaudos seleccionados no son válidos.'},status=400)
                
                if recibos != recibos_desc[:len(recibos)]:
                    return JsonResponse({'status':400,'message':'Solo puedes desvincular los recaudos más recientes en orden descendente.'},status=400)
                
                Recaudos.objects.using(proyecto).filter(
                    idadjudicacion=adj,
                    recibo__in=recibos
                ).delete()
                
                timeline.objects.using(proyecto).create(
                    adj=adj,
                    fecha=datetime.date.today(),
                    usuario=request.user,
                    accion=f'Desvinculó los recaudos {", ".join(recibos)}'
                )
                
                saldo_pendiente = saldos_adj.objects.using(proyecto).filter(
                    adj=adj
                ).aggregate(total=Sum('saldocapital')).get('total') or 0
                
                data = {
                    'status':200,
                    'message':'Los recaudos seleccionados fueron desvinculados.',
                    'saldo_pendiente':float(saldo_pendiente)
                }
                
                return JsonResponse(data)
            
            elif todo == 'reaplicar_recaudos':
                check_perms(request,('andinasoft.add_recaudos_general',),raise_exception=True)
                recibos = request.POST.getlist('recibos[]')
                if not recibos:
                    return JsonResponse({'status':400,'message':'Selecciona al menos un recibo pendiente.'},status=400)
                
                pendientes_qs = list(Recaudos_general.objects.using(proyecto).filter(
                    idadjudicacion=adj
                ).annotate(
                    aplicado=Exists(
                        Recaudos.objects.using(proyecto).filter(
                            idadjudicacion=adj,
                            recibo=OuterRef('numrecibo')
                        )
                    )
                ).filter(aplicado=False).order_by('fecha','idrecaudo'))
                pending_nums = [rec.numrecibo for rec in pendientes_qs]
                if not pending_nums:
                    return JsonResponse({'status':400,'message':'No hay recaudos pendientes por aplicar.'},status=400)
                
                if len(recibos) > len(pending_nums):
                    return JsonResponse({'status':400,'message':'La selección no es válida.'},status=400)
                
                expected = pending_nums[:len(recibos)]
                if recibos != expected:
                    return JsonResponse({'status':400,'message':'Debes seleccionar los recibos en orden desde el más antiguo.'},status=400)
                
                titulares = titulares_por_adj.objects.using(proyecto).get(adj=adj)
                timeline_msgs = []
                pending_map = {rec.numrecibo: rec for rec in pendientes_qs}
                allow_abono = request.POST.get('allow_abono_capital') == 'true'
                for num in recibos:
                    rec_obj = pending_map.get(num)
                    if rec_obj is None:
                        return JsonResponse({'status':400,'message':f'El recibo {num} ya fue aplicado.'},status=400)
                    solicitud = recibos_internos.objects.using('default').filter(recibo_asociado=num).first()
                    if solicitud and solicitud.abono_capital:
                        if not allow_abono:
                            return JsonResponse({
                                'status':409,
                                'needs_confirmation':True,
                                'message':f'El recibo {num} es un abono a capital. Confirma que deseas reaplicarlo.'
                            },status=409)
                    porcentaje_cond = Decimal('100') if solicitud and solicitud.condonacion else Decimal('0')
                    saldo_cuotas = list(saldos_adj.objects.using(proyecto).filter(adj=adj,saldocuota__gt=0))
                    if len(saldo_cuotas) == 0:
                        return JsonResponse({'status':400,'message':'No hay saldos pendientes para aplicar el recaudo.'},status=400)
                    aplicar_pago(
                        request=request,
                        adj=adj,
                        fecha=rec_obj.fecha,
                        forma_pago=rec_obj.formapago or '',
                        valor_pagado=rec_obj.valor,
                        concepto=rec_obj.concepto or 'PAGO RECIBIDO DE CLIENTE',
                        valor_recibo=rec_obj.valor,
                        porcentaje_condonado=porcentaje_cond,
                        saldo_cuotas=saldo_cuotas,
                        consecutivo=rec_obj.numrecibo,
                        Recaudos=Recaudos,
                        Recaudos_general=Recaudos_general,
                        titulares=titulares,
                        proyecto=proyecto,
                        crear_recibo=False,
                        cobrar_mora=False
                    )
                    if solicitud:
                        solicitud.usuario_confirma = request.user
                        solicitud.fecha_confirma = datetime.date.today()
                        solicitud.save(update_fields=['usuario_confirma','fecha_confirma'])
                    timeline_msgs.append(num)
                
                if timeline_msgs:
                    timeline.objects.using(proyecto).create(
                        adj=adj,
                        fecha=datetime.date.today(),
                        usuario=request.user,
                        accion=f'Reaplicó los recaudos pendientes {", ".join(timeline_msgs)}'
                    )
                
                return JsonResponse({'status':200,'message':'Los recaudos seleccionados fueron reaplicados correctamente.'})
    else:
        if request.method == 'POST':
            form_doc=form_docs_contratos(request.POST,request.FILES)
            form_datos_titular=form_info_titular(request.POST)
            form_seguimientos=form_seguimiento(request.POST)
            form_procesable=form_procesabilidad(request.POST)
            if request.POST.get('btnProcesabilidad'):
                if form_procesable.is_valid():
                    enmendaduras=form_procesable.cleaned_data.get('Enmendaduras')
                    documentacion=form_procesable.cleaned_data.get('Documentacion_Incompleta')
                    valores_incorr=form_procesable.cleaned_data.get('Valores_Incorrectos')
                    obs=form_procesable.cleaned_data.get('Observaciones')
                    obj_adj.p_enmendadura=enmendaduras
                    obj_adj.p_doccompleta=documentacion
                    obj_adj.p_valincorrectos=valores_incorr
                    obj_adj.p_obs=obs
                    obj_adj.save()
                    obj_timeline=timeline.objects.using(proyecto)
                    obj_timeline.create(adj=adj,
                                    fecha=datetime.date.today(),
                                    usuario=request.user,
                                    accion='Cambió el estado de Procesabilidad')
                    obj_adj=Adjudicacion.objects.using(proyecto).get(idadjudicacion=adj)
                    procesable='Si'
                    if obj_adj.p_enmendadura==1: procesable='No'
                    if obj_adj.p_doccompleta==1: procesable='No'
                    if obj_adj.p_valincorrectos==1: procesable='No'
            if request.POST.get('btnDesistir'):
                if check_perms(request,('andinasoft.delete_adjudicacion',),raise_exception=False):
                    obj_adj=Adjudicacion.objects.using(proyecto).get(idadjudicacion=adj)
                    inmueble_liberar=obj_adj.idinmueble
                    obj_inmueble=Inmuebles.objects.using(proyecto).get(idinmueble=inmueble_liberar)
                    obj_adj.estado='Desistido'
                    obj_adj.fechadesistimiento=datetime.date.today()
                    obj_adj.save()
                    
                    nota_dev = request.POST.get('nota_desist')
                    capital_dev = int(request.POST.get('cap_devuelto').replace(',','')) * -1
                    interes_dev = int(request.POST.get('int_devueltos').replace(',','')) * -1
                    total_dev = capital_dev + interes_dev
                    nuevo_estado_lote = request.POST.get("nuevoestadolote")
                    
                    
                    Recaudos_general.objects.using(proyecto).create(
                        idadjudicacion = adj, fecha = datetime.date.today(),
                        numrecibo = nota_dev, idtercero = obj_adj.idtercero1,
                        operacion = 'Desistimiento', valor = total_dev,
                        formapago = 'Desistimiento', concepto = 'Desistimiento',
                        usuario = request.user.username
                    )
                    
                    Recaudos.objects.using(proyecto).create(
                        recibo = nota_dev, fecha = datetime.date.today(),
                        idcta = adj, idadjudicacion = adj,
                        capital=capital_dev,interescte=interes_dev,
                        interesmora = 0, moralqd = 0,
                        fechaoperacion = datetime.date.today(),
                        usuario = request.user.username, estado = 'Aprobado'
                    )
                    
                    obj_inmueble.estado=nuevo_estado_lote
                    obj_inmueble.save()
                    obj_timeline=timeline.objects.using(proyecto)
                    obj_timeline.create(adj=adj,
                                fecha=datetime.date.today(),
                                usuario=request.user,
                                accion='Desistió el Contrato')
                    año_actual=datetime.date.today().year
                    mes_actual=datetime.date.today().month
                    periodo_actual=f'{año_actual}{mes_actual:02d}'
                    obj_ppto=PresupuestoCartera.objects.using(proyecto).filter(idadjudicacion=adj,periodo=periodo_actual)
                    if obj_ppto.exists():
                        for cuota in obj_ppto:
                            cuota.delete()
                    alerta=True
                    titulo='¡Hecho!'
                    mensaje=f'El {adj} fue desistido con exito'
                else: raise PermissionDenied
            if request.POST.get('btnCesion'):
                if check_perms(request,('andinasoft.change_adjudicacion',),raise_exception=False):
                    titular_ceder=request.POST.get('Titular-Ceder')
                    nuevo_titular=request.POST.get('nuevo-titular')
                    if titular_ceder=='Titular 1':
                        obj_adj.idtercero1=nuevo_titular
                    elif titular_ceder=='Titular 2':
                        obj_adj.idtercero2=nuevo_titular
                    elif titular_ceder=='Titular 3':
                        obj_adj.idtercero3=nuevo_titular
                    elif titular_ceder=='Titular 4':
                        obj_adj.idtercero4=nuevo_titular
                    obj_adj.save()
                    obj_timeline=timeline.objects.using(proyecto)
                    obj_timeline.create(adj=adj,
                                fecha=datetime.date.today(),
                                usuario=request.user,
                                accion=f'Aplicó una cesion al {titular_ceder}')
                    alerta=True
                    titulo='¡Hecho!'
                    mensaje=f'El {titular_ceder} fue cambiado con exito'
                else: raise PermissionDenied
            if request.POST.get('actualizar_info_titular'):
                if form_datos_titular.is_valid():
                    idtercero=form_datos_titular.cleaned_data.get('idTercero')
                    nombres=form_datos_titular.cleaned_data.get('nombres')
                    apellidos=form_datos_titular.cleaned_data.get('apellidos')
                    nombrecompleto=nombres+' '+apellidos
                    celular1=form_datos_titular.cleaned_data.get('celular1')
                    telefono1=form_datos_titular.cleaned_data.get('telefono1')
                    telefono2=form_datos_titular.cleaned_data.get('telefono2')
                    domicilio=form_datos_titular.cleaned_data.get('domicilio')
                    ciudad=form_datos_titular.cleaned_data.get('ciudad')
                    oficina=form_datos_titular.cleaned_data.get('oficina')
                    email=form_datos_titular.cleaned_data.get('email')
                    fecha_nac=form_datos_titular.cleaned_data.get('fecha_nac')
                    ocupacion=form_datos_titular.cleaned_data.get('ocupacion')
                    hijos=form_datos_titular.cleaned_data.get('hijos')
                    nivel_educativo=form_datos_titular.cleaned_data.get('nivel_educativo')
                    estado_civil=form_datos_titular.cleaned_data.get('estado_civil')
                    nivel_ingresos=form_datos_titular.cleaned_data.get('nivel_ingresos')
                    vehiculo=form_datos_titular.cleaned_data.get('vehiculo')
                    vivienda=form_datos_titular.cleaned_data.get('vivienda')
                    pasatiempos=form_datos_titular.cleaned_data.get('pasatiempos')
                    id_conyuge=form_datos_titular.cleaned_data.get('id_conyuge')
                    nombre_cony=form_datos_titular.cleaned_data.get('nombre_cony')
                    apellido_cony=form_datos_titular.cleaned_data.get('apellido_cony')
                    celular_cony=form_datos_titular.cleaned_data.get('celular_cony')
                    email_cony=form_datos_titular.cleaned_data.get('email_cony')
                    fechanac_cony=form_datos_titular.cleaned_data.get('fechanac_cony')
                    if fechanac_cony=='': fechanac_cony=None
                    ocupacion_cony=form_datos_titular.cleaned_data.get('ocupacion_cony')
                    fecha_act=datetime.datetime.today()
                    cliente=clientes.objects.using('default').get(idTercero=idtercero)
                    cliente.nombres=nombres
                    cliente.apellidos=apellidos
                    cliente.nombrecompleto=nombrecompleto
                    cliente.celular1=celular1
                    cliente.telefono1=telefono1
                    cliente.telefono2=telefono2
                    cliente.domicilio=domicilio
                    cliente.ciudad=ciudad
                    cliente.oficina=oficina
                    cliente.email=email
                    cliente.fecha_nac=fecha_nac
                    cliente.ocupacion=ocupacion
                    cliente.hijos=hijos
                    cliente.nivel_educativo=nivel_educativo
                    cliente.estado_civil=estado_civil
                    cliente.nivel_ingresos=nivel_ingresos
                    cliente.vehiculo=vehiculo
                    cliente.vivienda=vivienda
                    cliente.pasatiempos=pasatiempos
                    cliente.id_conyuge=id_conyuge
                    cliente.nombre_cony=nombre_cony
                    cliente.apellido_cony=apellido_cony
                    cliente.celular_cony=celular_cony
                    cliente.fechanac_cony=fechanac_cony
                    cliente.ocupacion_cony=ocupacion_cony
                    cliente.fecha_actualizacion=fecha_act
                    cliente.email_cony=email_cony
                    cliente.ocupacion_cony=ocupacion_cony
                    cliente.save()            
            if request.POST.get('cargar_docs'):
                if form_doc.is_valid():
                    descrip_doc=f"{form_doc.cleaned_data.get('tipo_doc')}_{datetime.datetime.today()}"
                    fecha_carga=datetime.datetime.today()
                    usuario_carga=request.user
                    upload_docs_contratos(request.FILES['documento_cargar'],adj,proyecto,descrip_doc)
                    documentos_contratos.objects.using(proyecto).create(adj=adj,descripcion_doc=descrip_doc,
                                                            fecha_carga=fecha_carga,usuario_carga=usuario_carga)
            if request.POST.get('agregarSeguimiento'):
                if form_seguimientos.is_valid():
                    forma_contacto=form_seguimientos.cleaned_data.get('forma_contacto')
                    tipo_seguimiento=form_seguimientos.cleaned_data.get('tipo_seguimiento')
                    comentarios=form_seguimientos.cleaned_data.get('comentarios')
                    compromiso=form_seguimientos.cleaned_data.get('tiene_compromiso')
                    fecha_compromiso=form_seguimientos.cleaned_data.get('fecha_compromiso')
                    valor_compromiso=form_seguimientos.cleaned_data.get('valor_compromiso')
                    seguimientos.objects.using(proyecto).create(adj=adj,
                                                                fecha=datetime.date.today(),
                                                                tipo_seguimiento=tipo_seguimiento,
                                                                forma_contacto=forma_contacto,
                                                                respuesta_cliente=comentarios,
                                                                valor_compromiso=valor_compromiso,
                                                                fecha_compromiso=fecha_compromiso,
                                                                usuario=request.user)
            if request.POST.get('eliminar_docs'):
                check_perms(request,('andinasoft.delete_documentos_contratos',))
                doc_eliminar=request.POST.get('docAccion')
                obj_documento=documentos_contratos.objects.using(proyecto).get(adj=adj,descripcion_doc=doc_eliminar)
                obj_documento.delete()
            if request.POST.get('btnVerifPagado'):
                estado=obj_adj.estado
                if estado=='Desistido':
                    alerta=True
                    titulo='Error'
                    mensaje='Este ADJ esta desistido, no se puede clasificar como pagado'
                else:
                    obj_saldos=saldos_adj.objects.using(proyecto).filter(adj=adj).aggregate(Sum('saldocapital'))
                    sum_saldos=obj_saldos['saldocapital__sum']
                    if sum_saldos==0:
                        obj_adj.estado='Pagado'
                        obj_adj.save()
                        alerta=True
                        titulo='¡Listo!'
                        mensaje=f'El nuevo estado del {adj} es Pagado'
                    else:
                        alerta=True
                        titulo='Error'
                        mensaje=f'Este adj tiene saldos pendientes, por favor verificalo y vuelve a intentar'
            if request.POST.get('btnRadicarPQRS'):
                check_perms(request,('andinasoft.add_pqrs',))
                tipo=request.POST.get('tipoPQRS')
                req_rta=request.POST.get('reqRespuesta')
                archivo=request.FILES.get('peticion')
                fecha_rad = request.POST.get('fecha_recibido')
                fecha_ven = request.POST.get('fecha_vencimiento')
                nombre_archivo = f'{tipo}_{datetime.datetime.now()}'.replace(':','.')
                Pqrs.objects.using(proyecto).create(idadjudicacion=adj,estado='Abierta',tipo=tipo,
                                                    fecha_radicado=fecha_rad,fecha_vencimiento=fecha_ven,
                                                    doc_peticion=nombre_archivo+'.pdf',usuario_radica=str(request.user))
                ruta =f'{settings.DIR_DOCS}/doc_contratos/{proyecto}/{adj}/'
                upload_docs(archivo,ruta,nombre_archivo)
                documentos_contratos.objects.using(proyecto).create(adj=adj,descripcion_doc=nombre_archivo,
                                                                    fecha_carga=datetime.date.today(),usuario_carga=str(request.user))
                pqrs=Pqrs.objects.using(proyecto).last()
                obj_timeline=timeline.objects.using(proyecto)
                accion=f'Radicó una {tipo} del {fecha_rad} (Radicado #{pqrs.pk})'
                obj_timeline.create(adj=adj,
                            fecha=datetime.date.today(),
                            usuario=request.user,
                            accion=accion)
                alerta=True
                titulo='PQRS Radicada'
                mensaje=f'La {tipo} fue radicada con el id = {pqrs.pk}'
            if request.POST.get('btnCambiarLote'):
                check_perms(request,('andinasoft.change_adjudicacion',))
                nuevo_lote = request.POST.get('lotecambiar')
                nuevo_estado = request.POST.get('nuevoEstadoLote')
                lote_actual = obj_adj.idinmueble
                #Primero hacemos el cambio del estado del lote
                obj_nuevolote = Inmuebles.objects.using(proyecto).get(idinmueble=nuevo_lote)
                obj_nuevolote.estado = 'Adjudicado'
                obj_loteanterior = Inmuebles.objects.using(proyecto).get(idinmueble=lote_actual)
                obj_loteanterior.estado = nuevo_estado
                obj_nuevolote.save()
                obj_loteanterior.save()
                
                obj_adj.idinmueble=nuevo_lote
                obj_adj.save()
                obj_adj=Adjudicacion.objects.using(proyecto).get(idadjudicacion=adj)
                
                obj_timeline=timeline.objects.using(proyecto)
                obj_timeline.create(adj=adj,
                            fecha=datetime.date.today(),
                            usuario=request.user,
                            accion=f'Cambió el lote de {lote_actual} a {nuevo_lote}')
                alerta=True
                titulo='¡Hecho!'
                mensaje=f'El lote fue cambiado con exito'
            
            
        lista_clientes=clientes.objects.using('default').all()
        dir_redireccion=f'{settings.MEDIA_URL}docs_andinasoft/tmp/{proyecto}_estadocuenta_{adj}'
        timeline_adj=timeline.objects.using(proyecto).filter(adj=adj).order_by('fecha').reverse()
        seguimientos_adj=seguimientos.objects.using(proyecto).filter(adj=adj).order_by('fecha').reverse()
        saldos_por_cuotas=saldos_adj.objects.using(proyecto).filter(adj=adj).order_by('fecha')
        capital_subquery = Recaudos.objects.using(proyecto).filter(
            idadjudicacion=adj,
            recibo=OuterRef('numrecibo')
        ).values('recibo').annotate(total_cap=Sum('capital')).values('total_cap')
        detalle_exists = Recaudos.objects.using(proyecto).filter(
            idadjudicacion=adj,
            recibo=OuterRef('numrecibo')
        )
        recibos_qs=Recaudos_general.objects.using(proyecto).filter(
            idadjudicacion=adj
        ).annotate(
            capital_total=Coalesce(
                Subquery(capital_subquery, output_field=DecimalField(max_digits=20, decimal_places=2)),
                Value(0, output_field=DecimalField(max_digits=20, decimal_places=2))
            ),
            aplicado=Exists(detalle_exists)
        ).order_by('-fecha','-idrecaudo')
        recibos_detallados = list(recibos_qs)
        numeros_recibos = [rec.numrecibo for rec in recibos_detallados if rec.numrecibo]
        soportes_map = {}
        abono_map = {}
        condonacion_map = {}
        if numeros_recibos:
            qs_soportes = recibos_internos.objects.using('default').filter(
                proyecto=proyecto_obj,
                recibo_asociado__in=numeros_recibos
            ).values('recibo_asociado','soporte','abono_capital','condonacion')
            for row in qs_soportes:
                num = row.get('recibo_asociado')
                if not num:
                    continue
                if row.get('soporte'):
                    soporte_url = _resolve_soporte_url(row.get('soporte'))
                    if soporte_url:
                        soportes_map[num] = soporte_url
                abono_map[num] = row.get('abono_capital')
                condonacion_map[num] = row.get('condonacion')
        tiene_recaudos_pendientes = False
        for rec in recibos_detallados:
            rec.soporte_url = soportes_map.get(rec.numrecibo)
            rec.abono_capital = abono_map.get(rec.numrecibo, False)
            rec.condonacion_flag = condonacion_map.get(rec.numrecibo)
            if not rec.aplicado:
                tiene_recaudos_pendientes = True
        recaudos_pendientes = sorted(
            [r for r in recibos_detallados if not r.aplicado],
            key=lambda r: (r.fecha, r.idrecaudo)
        )
        detalle_recaudos=Recaudos.objects.using(proyecto).filter(idadjudicacion=adj)
        total_recaudo=Recaudos_general.objects.using(proyecto).filter(idadjudicacion=adj).aggregate(Sum('valor'))
        vista_adj=Vista_Adjudicacion.objects.using(proyecto).get(IdAdjudicacion=adj)
        lista_docs=documentos_contratos.objects.using(proyecto).filter(adj=adj)
        saldo_cuotas=saldos_adj.objects.using(proyecto).filter(adj=adj)
        titulares=titulares_por_adj.objects.using(proyecto).get(adj=adj)
        try:info_tit1=clientes.objects.using('default').get(idTercero=titulares.IdTercero1)
        except:info_tit1=[]
        try:info_tit2=clientes.objects.using('default').get(idTercero=titulares.IdTercero2)
        except:info_tit2=[]
        try:info_tit3=clientes.objects.using('default').get(idTercero=titulares.IdTercero3)
        except:info_tit3=[]
        try:info_tit4=clientes.objects.using('default').get(idTercero=titulares.IdTercero4)
        except:info_tit4=[]
        saldo=0
        pagado=0
        for cuota in saldo_cuotas:
            saldo+=cuota.saldocapital if cuota.saldocapital else 0
            pagado+=cuota.rcdocapital if cuota.rcdocapital else 0
        
        if request.method == 'POST':
            nombre=info_tit1.nombrecompleto
            direccion=info_tit1.domicilio
            inmueble=vista_adj.Inmueble
            tipodoc=obj_adj.tipocontrato
            nrocontrato=obj_adj.contrato
            formapago=obj_adj.formapago
            telefono=str(info_tit1.celular1)+'-'+str(info_tit1.telefono1)
            valor_inmueble=vista_adj.Valor
            valor_ci=saldos_por_cuotas.filter(tipocta='CI').aggregate(Sum('capital'))
            valor_ci=valor_ci['capital__sum']
            rcdo_ci=saldos_por_cuotas.filter(tipocta='CI').aggregate(Sum('rcdocapital'))
            rcdo_ci=rcdo_ci['rcdocapital__sum']
            saldo_ci=valor_ci-rcdo_ci
            cant_fn=saldos_por_cuotas.filter(tipocta='FN').count()
            vr_cta_fn=saldos_por_cuotas.filter(tipocta='FN',saldocuota__gt=0)
            if vr_cta_fn.exists():
                if len(vr_cta_fn)>1: vr_cta_fn=vr_cta_fn[1]
                else: vr_cta_fn=vr_cta_fn[0]
                if vr_cta_fn is None: vr_cta_fn=0
                else: vr_cta_fn=vr_cta_fn.cuota
            else:
                vr_cta_fn=0
            valor_fn=saldos_por_cuotas.filter(tipocta='FN').aggregate(Sum('capital'))
            valor_fn=valor_fn['capital__sum']
            if valor_fn is None: valor_fn=0
            rcdo_fn=saldos_por_cuotas.filter(tipocta='FN').aggregate(Sum('rcdocapital'))
            rcdo_fn=rcdo_fn['rcdocapital__sum']
            if rcdo_fn is None: rcdo_fn=0
            saldo_fn=valor_fn-rcdo_fn
            
            if saldos_por_cuotas.filter(tipocta='CE').exists():
                cant_ce=saldos_por_cuotas.filter(tipocta='CE').count()
                vr_cta_ce=saldos_por_cuotas.filter(tipocta='CE').last()
                vr_cta_ce=vr_cta_ce.cuota
                valor_ce=saldos_por_cuotas.filter(tipocta='CE').aggregate(Sum('capital'))
                valor_ce=valor_ce['capital__sum']
                rcdo_ce=saldos_por_cuotas.filter(tipocta='CE').aggregate(Sum('rcdocapital'))
                rcdo_ce=rcdo_ce['rcdocapital__sum']
                saldo_ce=valor_ce-rcdo_ce
            else:
                cant_ce=0
                vr_cta_ce=0
                valor_ce=0
                rcdo_ce=0
                saldo_ce=0
            if saldos_por_cuotas.filter(tipocta='CO').exists():
                valor_co=saldos_por_cuotas.filter(tipocta='CO').aggregate(Sum('capital'))
                valor_co=valor_co['capital__sum']
                rcdo_co=saldos_por_cuotas.filter(tipocta='CO').aggregate(Sum('rcdocapital'))
                rcdo_co=rcdo_co['rcdocapital__sum']
                saldo_co=valor_co-rcdo_co
            else:
                cant_co=0
                vr_cta_co=0
                valor_co=0
                rcdo_co=0
                saldo_co=0
            total_rcdo=rcdo_ci+rcdo_fn+rcdo_ce+rcdo_co
            saldo_cap_pendiente=saldos_por_cuotas.filter(saldocuota__gt=0,).aggregate(Sum('saldocapital'))
            saldo_cap_pendiente=saldo_cap_pendiente['saldocapital__sum']
            if saldo_cap_pendiente==None: saldo_cap_pendiente=0
            saldo_int_pendiente=saldos_por_cuotas.filter(fecha__lte=datetime.datetime.today()).aggregate(Sum('saldointcte'))
            saldo_int_pendiente=saldo_int_pendiente['saldointcte__sum']
            if saldo_int_pendiente==None: saldo_int_pendiente=0
            saldo_pendiente=saldo_cap_pendiente+saldo_int_pendiente
            interes_mora=saldos_por_cuotas.aggregate(Sum('saldomora'))
            interes_mora=interes_mora['saldomora__sum']
            cuotas_vencidas=saldos_por_cuotas.filter(saldocuota__gt=0,fecha__lte=datetime.datetime.today()).count()
            dias_mora=saldos_por_cuotas.filter(diasmora__lte=0).aggregate(Max('diasmora'))
            dias_mora=dias_mora['diasmora__max']
            saldo_vencido=saldos_por_cuotas.filter(saldocuota__gt=0,fecha__lte=datetime.datetime.today()).aggregate(Sum('saldocuota'))
            saldo_vencido=saldo_vencido['saldocuota__sum']
            detalle_ctas_pendientes=saldos_por_cuotas.filter(fecha__lte=datetime.datetime.today(),saldocuota__gt=0)
            resumen_pagos=detalle_recaudos.order_by('fecha','recibo')
            
            if request.POST.get('impEstadodeCuenta'):
                
                obj_adj = Adjudicacion.objects.using(proyecto).get(pk=adj)
                today = datetime.date.today()
                next_30_days = today+ datetime.timedelta(days=30)
                cuotas_a_la_fecha = PlanPagos.objects.using(proyecto).filter(adj=adj, fecha__lte=today).order_by('fecha')
                cuotas_futuras = PlanPagos.objects.using(proyecto).filter(
                    adj=adj, fecha__gt=today,fecha__lte = next_30_days).order_by('fecha')
                
                cuotas_vencidas = []
                
                total_cuotas_vencidas = {
                    'valor': 0,
                    'intereses_mora':0,
                    'total': 0
                }
                
                recaudador = {
                    'Tesoro Escondido': 'STATUS COMERCIALIZADORA S.A.S. NIT: 901018375-4',
                    'Vegas de Venecia': 'STATUS COMERCIALIZADORA S.A.S. NIT: 901018375-4',
                    'Perla del Mar': 'ANDINA CONCEPTOS INMOBILIARIOS S.A.S. NIT: 900993044-9',
                    'Sandville Beach': 'ANDINA CONCEPTOS INMOBILIARIOS S.A.S. NIT: 900993044-9',
                    'Carmelo Reservado': 'ANDINA CONCEPTOS INMOBILIARIOS S.A.S. NIT: 900993044-9', 
                }
                
                for q in cuotas_a_la_fecha:
                    pendiente = q.pendiente()
                    if pendiente.get('total', 0) > 0:
                        mora = q.mora()
                        cuotas_vencidas.append({
                            'fecha':q.fecha,
                            'idcta': q.pk.split('ADJ')[0],
                            'pendiente':pendiente,
                            'mora':mora,
                        })
                        total_cuotas_vencidas['valor'] += pendiente.get('total')
                        total_cuotas_vencidas['intereses_mora'] += mora.get('valor')
                        total_cuotas_vencidas['total'] += pendiente.get('total') + mora.get('valor')


                context = {
                    'adj':obj_adj,
                    'cuotas_a_la_fecha': cuotas_vencidas,
                    'cuotas_futuras':cuotas_futuras,
                    'user': request.user,
                    'now': datetime.datetime.now(),
                    'totals': total_cuotas_vencidas,
                    'recaudador': recaudador.get(proyecto)
                }
                
                
                filename = f'Estado_de_cuenta_{adj}_{proyecto}.pdf'
                        
                            
                pdf = pdf_gen(f'pdf/statement_of_account.html',context,filename)
                
                ruta = pdf.get('root')
                                        
                ruta_link=pdf.get('url')
                
                dir_link = ruta_link
                alerta=True
                titulo='¡Ya puedes descargar tu documento!'
                mensaje='Puedes descargarlo aqui'
                link=True
                
            if request.POST.get('impResumenPagos'):
                ruta=settings.DIR_EXPORT+f'{proyecto}_resumencredito_{adj}.pdf'
                GenerarPDF().resumen_pagos(ruta=ruta,
                            fecha=str(datetime.datetime.today()),
                            adj=adj,
                            nombre=nombre,
                            proyecto=proyecto,
                            direccion=direccion,
                            inmueble=inmueble,
                            telefono=telefono,
                            valor_contrato=valor_inmueble,
                            vr_ci=valor_ci,
                            rcdo_ci=rcdo_ci,
                            saldo_ci=saldo_ci,
                            vr_fn=valor_fn,
                            rcdo_fn=rcdo_fn,
                            saldo_fn=saldo_fn,
                            ctas_fn=cant_fn,
                            vr_cta_fn=vr_cta_fn,
                            tasa=1.5,
                            vr_ce=valor_ce,
                            ctas_ce=cant_ce,
                            rcdo_ce=rcdo_ce,
                            saldo_ce=saldo_ce,
                            vr_co=valor_co,
                            rcdo_co=rcdo_co,
                            saldo_co=saldo_co,
                            rcdo_total=total_rcdo,
                            saldo_total=saldo_pendiente,
                            saldo_mora=interes_mora,
                            ctas_vencidas=cuotas_vencidas,
                            saldo_vencido=saldo_vencido,
                            dias_mora=dias_mora,
                            cuotas_pendientes=detalle_ctas_pendientes,
                            cuotas_pagadas=resumen_pagos)
                dir_link=settings.DIR_DOWNLOADS+f'{proyecto}_resumencredito_{adj}.pdf'
                alerta=True
                titulo='¡Ya puedes descargar tu documento!'
                mensaje='Puedes descargarlo aqui'
                link=True
            if request.POST.get('impPortada'):
                check_perms(request,('andinasoft.view_pagocomision',))
                planpagos='Regular'
                if saldo_fn>0 and saldo_ce>0:
                    planpagos='Extraordinario'
                ruta=settings.DIR_EXPORT+f'{proyecto}_portada_{adj}.pdf'
                fecha=datetime.datetime.strftime(datetime.datetime.today(),'%d-%m-%Y')
                escala=AsignacionComisiones.objects.using(proyecto).raw(f'CALL portadacomisiones("{adj}")')
                if info_tit2==[]: nombret2=''
                else:nombret2=info_tit2.nombrecompleto
                if info_tit3==[]: nombret3=''
                else:nombret3=info_tit3.nombrecompleto
                if info_tit4==[]: nombret4=''
                else:nombret4=info_tit4.nombrecompleto
                GenerarPDF().portada_adj(ruta,proyecto,adj,fecha,inmueble,tipodoc,nrocontrato,
                                         info_tit1.nombrecompleto,nombret2,nombret3,nombret4,
                                         valor_inmueble,valor_ci,saldo_fn+saldo_ce+saldo_co,formapago,planpagos,
                                         vr_cta_fn,vr_cta_ce,escala)
                dir_link=settings.DIR_DOWNLOADS+f'{proyecto}_portada_{adj}.pdf'
                alerta=True
                titulo='¡Ya puedes descargar tu documento!'
                mensaje='Puedes descargarlo aqui'
                link=True
        
        
        context={
            'alerta':alerta,
            'mensaje':mensaje,
            'titulo_alerta':titulo,
            'ruta_link':dir_link,
            'link':link,
            'titulares':titulares,
            'proyecto':proyecto,
            'adj':adj,
            'info_adj':vista_adj,
            'datos_adj':obj_adj,
            'info_titular1':info_tit1,
            'info_titular2':info_tit2,
            'info_titular3':info_tit3,
            'info_titular4':info_tit4,
            'form':form_detalle_adj,
            'form_info_adj':form_detalle_adj,
            'form_info_titular':form_info_titular,
            'form_docs':form_docs_contratos,
            'form_procesabilidad':form_procesabilidad,
            'timeline':timeline_adj,
            'seguimientos':seguimientos_adj,
            'pagado':pagado,
            'recaudos':recibos_detallados,
            'recaudos_pendientes':recaudos_pendientes,
            'detalle_recaudos':detalle_recaudos,
            'saldos_cuotas':saldos_por_cuotas,
            'total_recaudo':total_recaudo,
            'saldo_capital':saldo,
            'tiene_recaudos_pendientes':tiene_recaudos_pendientes,
            'lista_documentos':lista_docs,
            'form_seguimiento':form_seguimiento,
            'clientes':lista_clientes,
            'procesable':procesable,
            'pqrs':Pqrs.objects.using(proyecto).filter(idadjudicacion=adj),
            'lotes_libres':Inmuebles.objects.using(proyecto).filter(estado='Libre'),
            'ultimobk': bk_bfchangeplan.objects.filter(adj=adj,proyecto=proyecto).order_by('-fecha_bk')
        }
        return render(request,'detalle_adj.html',context)
    

@login_required
def seleccionar_proyecto(request):
    return render(request,'seleccion_proyectos.html')

@group_perm_required(perms=('andinasoft.add_ventas_nuevas',),raise_exception=True)
def nueva_venta(request,proyecto,inmueble):
    
    check_project(request,proyecto)
    lista_clientes=clientes.objects.using('default').all()
    datos_inmueble=Inmuebles.objects.using(proyecto).get(idinmueble=inmueble)
    incrementos=datos_inmueble.fac_valor_via_principal*datos_inmueble.fac_valor_area_social*datos_inmueble.fac_valor_esquinero
    valor_lote=(datos_inmueble.areaprivada*datos_inmueble.vrmetrocuadrado*incrementos)
    valor_lote=Utilidades().redondear_numero(numero=valor_lote,multiplo=1000000,redondeo='>')
    default_ci=40
    default_saldo=60
    
    context={
        'form':form_nueva_venta,
        'clientes':lista_clientes,
        'inmueble':inmueble,
        'datos_inmueble':datos_inmueble,
        'valor_lote':valor_lote,
        'default_ci':default_ci,
        'default_saldo':default_saldo,
        'tasa':tasas,
        'proyecto':proyecto,
        'btnName':'Guardar',
    }
    if datos_inmueble.estado!='Libre':
        context['alerta']=True
        context['mensaje']='Este inmueble ya no esta disponible'
        context['redirect']=f"/venta/inventario/{proyecto}"
        context['titulo_alerta']='Error'
    if request.method == 'POST':
        form=form_nueva_venta(request.POST)
        if form.is_valid():
            idtercero1=strip_tags(form.cleaned_data.get('titular1') or '')
            idtercero2=strip_tags(form.cleaned_data.get('titular2') or '')
            idtercero3=strip_tags(form.cleaned_data.get('titular3') or '')
            idtercero4=strip_tags(form.cleaned_data.get('titular4') or '')
            inmueble=form.cleaned_data.get('inmueble')
            valor_lote=form.cleaned_data.get('valor')
            valor_letras=form.cleaned_data.get('valor_letras')
            porc_ci=form.cleaned_data.get('porc_cta_ini')
            valor_ci=form.cleaned_data.get('vr_ci')
            porc_saldo=form.cleaned_data.get('porc_saldo')
            valor_saldo=form.cleaned_data.get('vr_saldo')
            forma_pago=form.cleaned_data.get('forma_pago')
            plan_pago=form.cleaned_data.get('forma_plan_pagos')
            cant_ci1=form.cleaned_data.get('cant_ci_1')
            cant_ci2=form.cleaned_data.get('cant_ci_2')
            cant_ci3=form.cleaned_data.get('cant_ci_3')
            cant_ci4=form.cleaned_data.get('cant_ci_4')
            cant_ci5=form.cleaned_data.get('cant_ci_5')
            cant_ci6=form.cleaned_data.get('cant_ci_6')
            cant_ci7=form.cleaned_data.get('cant_ci_7')
            fecha_ci1=form.cleaned_data.get('fecha_ini_ci1')
            fecha_ci2=form.cleaned_data.get('fecha_ini_ci2')
            fecha_ci3=form.cleaned_data.get('fecha_ini_ci3')
            fecha_ci4=form.cleaned_data.get('fecha_ini_ci4')
            fecha_ci5=form.cleaned_data.get('fecha_ini_ci5')
            fecha_ci6=form.cleaned_data.get('fecha_ini_ci6')
            fecha_ci7=form.cleaned_data.get('fecha_ini_ci7')
            valor_ci1=form.cleaned_data.get('valor_ci1')
            valor_ci2=form.cleaned_data.get('valor_ci2')
            valor_ci3=form.cleaned_data.get('valor_ci3')
            valor_ci4=form.cleaned_data.get('valor_ci4')
            valor_ci5=form.cleaned_data.get('valor_ci5')
            valor_ci6=form.cleaned_data.get('valor_ci6')
            valor_ci7=form.cleaned_data.get('valor_ci7')
            cant_fn=form.cleaned_data.get('cant_fn')
            fecha_fn=form.cleaned_data.get('fecha_ini_fn')
            valor_fn=form.cleaned_data.get('valor_fn')
            cant_ce=form.cleaned_data.get('cant_ce')
            fecha_ce=form.cleaned_data.get('fecha_ini_ce')
            valor_ce=form.cleaned_data.get('valor_ce')
            periodo_ce=form.cleaned_data.get('periodo_ce')
            observaciones=form.cleaned_data.get('observaciones')
            tasa = request.POST.get('tasafn')
            nueva_venta=ventas_nuevas.objects.using(proyecto)
            nueva_venta.create(id_t1=idtercero1,id_t2=idtercero2,id_t3=idtercero3,id_t4=idtercero4,
                            inmueble=inmueble,valor_venta=valor_lote,forma_pago=forma_pago,cuota_inicial=valor_ci,
                            cant_ci1=cant_ci1,cant_ci2=cant_ci2,cant_ci3=cant_ci3,cant_ci4=cant_ci4,
                            cant_ci5=cant_ci5,cant_ci6=cant_ci6,cant_ci7=cant_ci7,
                            fecha_ci1=fecha_ci1,fecha_ci2=fecha_ci2,fecha_ci3=fecha_ci3,
                            fecha_ci4=fecha_ci4,fecha_ci5=fecha_ci5,fecha_ci6=fecha_ci6,fecha_ci7=fecha_ci7,
                            valor_ci1=valor_ci1,valor_ci2=valor_ci2,valor_ci3=valor_ci3,valor_ci4=valor_ci4,
                            valor_ci5=valor_ci5,valor_ci6=valor_ci6,valor_ci7=valor_ci7,
                            saldo=valor_saldo,forma_saldo=plan_pago,valor_ctas_fn=valor_fn,inicio_fn=fecha_fn,nro_cuotas_fn=cant_fn,
                            valor_ctas_ce=valor_ce,inicio_ce=fecha_ce,nro_cuotas_ce=cant_ce,period_ce=periodo_ce,
                            observaciones=observaciones,fecha_contrato=datetime.datetime.today(),usuario=request.user,estado='Pendiente',
                            tasa=Decimal(tasa))
            lote=Inmuebles.objects.using(proyecto).get(idinmueble=inmueble)
            lote.estado='Reservado'
            lote.save()
            msj_notif=f'Hola!, te informamos que {request.user.first_name} ha creado un nuevo contrato en {proyecto} sobre el inmueble {inmueble} por valor de ${valor_lote:,} y está disponible para tu revisión. '
            """ if proyecto=='Vegas de Venecia' or proyecto=='Tesoro Escondido':
                envio_notificacion(msj_notif,f'Se creó un contrato sobre el inmueble {inmueble} en {proyecto}',['jorgeavila@somosandina.co','sb@somosandina.co'])
            else:
                    envio_notificacion(msj_notif,f'Se creó un contrato sobre el inmueble {inmueble} en {proyecto}',['jorgeavila@somosandina.co']) """
            context['alerta']=True
            context['mensaje']='El Contrato fué creado con exito'
            context['redirect']=f"/comercial/ventas_sin_aprobar/{proyecto}"
            context['titulo_alerta']='¡Todo salió Perfecto!'
    
    
    if proyecto.lower() == 'fractal':
        return render(request, 'venta_fractal.html',context)
    
    return render(request,'nueva_venta.html',context)

@group_perm_required(perms=('andinasoft.add_ventas_nuevas',),raise_exception=True)
def ventas_sin_aprobar(request,proyecto):
    check_project(request,proyecto)
    
    if proyecto == 'Fractal':
        ventas=ventas_nuevas.objects.using(proyecto).filter(estado='Pendiente')
        context={
            'proyecto':proyecto,
            'ventas':ventas,
            'formaspago': formas_pago.objects.using(proyecto).all().order_by('descripcion'),
            'asesores': asesores.objects.filter(estado__icontains='Activo').order_by('nombre')
            }
        return render(request,'venta_fractal.html',context)
    
    ventas=ventas_nuevas.objects.using(proyecto).filter(estado='Pendiente').values()
    
    for venta in ventas:
        nombre_cliente=clientes.objects.using('default').get(idTercero=venta['id_t1'])
        venta['nombre_cliente']=nombre_cliente.nombrecompleto
    context={
        'proyecto':proyecto,
        'ventas':ventas,
        'btnName':'Modificar',
    }
    
    
    return render(request,'ventas_sin_aprobar.html',context)


def ventas_aprobadas(request,proyecto):
    check_project(request,proyecto)
    
    if proyecto == 'Fractal':
        check_perms(request,('andinasoft.view_adjudicacion',))
        ventas=ventas_nuevas.objects.using(proyecto).filter(estado='Aprobado')
        context={
            'proyecto':proyecto,
            'ventas':ventas,
        }
        return render(request,'aprobadas_fractal.html',context)
    
    elif proyecto == 'Fractalbasic':
        proyecto='Fractal'
        check_project(request,proyecto)
        check_perms(request,('andinasoft.add_adjudicacion',))
    
    check_perms(request,('andinasoft.add_adjudicacion',))
    ventas=ventas_nuevas.objects.using(proyecto).filter(estado='Aprobado').values()
    for venta in ventas:
        nombre_cliente=clientes.objects.using('default').get(idTercero=venta['id_t1'])
        venta['nombre_cliente']=nombre_cliente.nombrecompleto
    context={
        'proyecto':proyecto,
        'ventas':ventas,
        'btnName':'Modificar',
    }
    return render(request,'ventas_aprobadas.html',context)

@group_perm_required(perms=('andinasoft.add_ventas_nuevas',),raise_exception=True)
def acciones_venta(request,proyecto,contrato):
    check_project(request,proyecto)
    if request.user.is_authenticated:
        context={}
        alerta=False
        titulo=None
        mensaje=None
        redireccion=False
        dir_redirect=None
        link=False
        ruta_link=None
        data_ini_venta=ventas_nuevas.objects.using(proyecto).get(id_venta=contrato)
        obj_parametro=Parametros_Operaciones.objects.using(proyecto)
        if data_ini_venta.estado!='Pendiente':
            titulo='Error'
            alerta=True
            mensaje='Este contrato ya no esta disponible'
            redireccion=True
            dir_redirect=f'/comercial/ventas_sin_aprobar/{proyecto}'
        if request.method == 'POST':
            form=form_nueva_venta(request.POST)
            if form.is_valid():
                venta_modificar=ventas_nuevas.objects.using(proyecto).get(id_venta=contrato)
                if request.POST.get('btnAnular'):
                    if not check_perms(request,('andinasoft.delete_ventas_nuevas',),raise_exception=False):
                        alerta=True
                        titulo='Error'
                        mensaje='No tienes permisos suficientes para anular una venta'
                    else:
                        lote=venta_modificar.inmueble
                        obj_inmueble=Inmuebles.objects.using(proyecto).get(idinmueble=lote)
                        obj_inmueble.estado='Libre'
                        obj_inmueble.save()
                        venta_modificar.estado='Anulado'
                        venta_modificar.save()
                        alerta=True
                        mensaje=f'El contrato {contrato} fue anulado'
                        titulo='¡Listo!'
                        redireccion=True
                        dir_redirect=f"/comercial/ventas_sin_aprobar/{proyecto}"
                if request.POST.get('btnAprobar'):
                    if not check_perms(request,('andinasoft.delete_ventas_nuevas',),raise_exception=False):
                        alerta=True
                        titulo='Error'
                        mensaje='No tienes permisos suficientes para aprobar una venta'
                    else:
                        venta_modificar.estado='Aprobado'
                        venta_modificar.usuarioaprueba=str(request.user)
                        venta_modificar.fecha_aprueba=datetime.date.today()
                        venta_modificar.save()
                        subject=f'Se ha aprobado el contrato Nº{contrato} {proyecto}'
                        mensaje=f'Hola! el contrato {contrato} de {proyecto} fue aprobado en el modulo comercial y ya esta disponible para su adjudicación'
                        template='emails/notificacion_con_lista.html'
                        variables={
                            'mensaje':mensaje,
                        }
                        destinatarios={
                            'Sandville Beach':['jorgeavila@somosandina.co','tatianamontes@somosandina.co'],
                            'Perla del Mar':['jorgeavila@somosandina.co','tatianamontes@somosandina.co'],
                            'Sandville del Sol':['jorgeavila@somosandina.co'],
                            'Tesoro Escondido':['jorgeavila@somosandina.co','recepcion@somosandina.co'],
                            'Vegas de Venecia':['jorgeavila@somosandina.co','recepcion@somosandina.co'],
                            'Sotavento':['jorgeavila@somosandina.co','tatianamontes@somosandina.co'],
                        }
                        """ envio_email_template(subject,'autoreports@andinasoft.com.co',destinatarios[proyecto],template,variables) """
                        
                        alerta=True
                        mensaje=f'El contrato {contrato} fue aprobado exitosamente'
                        titulo='¡Todo salio Perfecto!'
                        redireccion=True
                        dir_redirect=f"/comercial/ventas_sin_aprobar/{proyecto}"
                if request.POST.get('btnformcontrato'):
                    venta_modificar.id_t1=form.cleaned_data.get('titular1')
                    venta_modificar.id_t2=form.cleaned_data.get('titular2')
                    venta_modificar.id_t3=form.cleaned_data.get('titular3')
                    venta_modificar.id_t4=form.cleaned_data.get('titular4')
                    venta_modificar.inmueble=form.cleaned_data.get('inmueble')
                    venta_modificar.valor_venta=form.cleaned_data.get('valor')
                    venta_modificar.forma_pago=form.cleaned_data.get('forma_pago')
                    venta_modificar.cuota_inicial=form.cleaned_data.get('vr_ci')
                    venta_modificar.cant_ci1=form.cleaned_data.get('cant_ci_1')
                    venta_modificar.cant_ci2=form.cleaned_data.get('cant_ci_2')
                    venta_modificar.cant_ci3=form.cleaned_data.get('cant_ci_3')
                    venta_modificar.cant_ci4=form.cleaned_data.get('cant_ci_4')
                    venta_modificar.cant_ci5=form.cleaned_data.get('cant_ci_5')
                    venta_modificar.cant_ci6=form.cleaned_data.get('cant_ci_6')
                    venta_modificar.cant_ci7=form.cleaned_data.get('cant_ci_7')
                    venta_modificar.fecha_ci1=form.cleaned_data.get('fecha_ini_ci1')
                    venta_modificar.fecha_ci2=form.cleaned_data.get('fecha_ini_ci2')
                    venta_modificar.fecha_ci3=form.cleaned_data.get('fecha_ini_ci3')
                    venta_modificar.fecha_ci4=form.cleaned_data.get('fecha_ini_ci4')
                    venta_modificar.fecha_ci5=form.cleaned_data.get('fecha_ini_ci5')
                    venta_modificar.fecha_ci6=form.cleaned_data.get('fecha_ini_ci6')
                    venta_modificar.fecha_ci7=form.cleaned_data.get('fecha_ini_ci7')
                    venta_modificar.valor_ci1=form.cleaned_data.get('valor_ci1')
                    venta_modificar.valor_ci2=form.cleaned_data.get('valor_ci2')
                    venta_modificar.valor_ci3=form.cleaned_data.get('valor_ci3')
                    venta_modificar.valor_ci4=form.cleaned_data.get('valor_ci4')
                    venta_modificar.valor_ci5=form.cleaned_data.get('valor_ci5')
                    venta_modificar.valor_ci6=form.cleaned_data.get('valor_ci6')
                    venta_modificar.valor_ci7=form.cleaned_data.get('valor_ci7')
                    venta_modificar.saldo=form.cleaned_data.get('vr_saldo')
                    venta_modificar.forma_saldo=form.cleaned_data.get('forma_plan_pagos')
                    venta_modificar.valor_ctas_fn=form.cleaned_data.get('valor_fn')
                    venta_modificar.inicio_fn=form.cleaned_data.get('fecha_ini_fn')
                    venta_modificar.nro_cuotas_fn=form.cleaned_data.get('cant_fn')
                    venta_modificar.valor_ctas_ce=form.cleaned_data.get('valor_ce')
                    venta_modificar.inicio_ce=form.cleaned_data.get('fecha_ini_ce')
                    venta_modificar.nro_cuotas_ce=form.cleaned_data.get('cant_ce')
                    venta_modificar.period_ce=form.cleaned_data.get('periodo_ce')
                    venta_modificar.observaciones=form.cleaned_data.get('observaciones')
                    venta_modificar.usuario=str(request.user)
                    venta_modificar.tasa = Decimal(request.POST.get('tasafn'))
                    venta_modificar.save()
                    titulo='¡Todo salio perfecto!'
                    mensaje=f'El contrato {contrato} fue modificado exitosamente'
                    alerta=True
                
            form_escala=form_escala_comision(request.POST)
            if form_escala.is_valid():
                cargos_fijos=CargosFijos.objects.using(proyecto).all()
                cc_generador=form_escala.cleaned_data.get('cc_generador')
                porc_generador=form_escala.cleaned_data.get('porc_generador')
                cc_linea=form_escala.cleaned_data.get('cc_linea')
                porc_linea=form_escala.cleaned_data.get('porc_linea')
                cc_cerrador=form_escala.cleaned_data.get('cc_cerrador')
                porc_cerrador=form_escala.cleaned_data.get('porc_cerrador')
                cc_gerventas=form_escala.cleaned_data.get('cc_gerventas')
                porc_gerventas=form_escala.cleaned_data.get('porc_gerventas')
                cc_jefep=form_escala.cleaned_data.get('cc_jefep')
                porc_jefep=form_escala.cleaned_data.get('porc_jefep')
                cc_jefev=form_escala.cleaned_data.get('cc_jefev')
                porc_jefev=form_escala.cleaned_data.get('porc_jefev')
                values=[
                    (cc_generador,porc_generador,8),
                    (cc_linea,porc_linea,9),
                    (cc_cerrador,porc_cerrador,99),
                    (cc_gerventas,porc_gerventas,14),
                    (cc_jefep,porc_jefep,15),
                    (cc_jefev,porc_jefev,13)
                    ]
                if request.POST.get('grabar-escala'):
                    try:
                        for valor in values:
                            if valor[0]!='':
                                AsignacionComisiones.objects.using(proyecto).create(id_comision=f'{valor[2]}-{contrato}',
                                                                                    idadjudicacion=contrato,
                                                                                    fecha=datetime.datetime.today(),
                                                                                    idgestor=valor[0],
                                                                                    idcargo=valor[2],
                                                                                    comision=valor[1],
                                                                                    usuario='Activo',
                                                                                    )
                        for fijo in cargos_fijos:
                            AsignacionComisiones.objects.using(proyecto).create(id_comision=f'{fijo.idcargo}-{contrato}',
                                                                                    idadjudicacion=contrato,
                                                                                    fecha=datetime.datetime.today(),
                                                                                    idgestor=fijo.cc_fija,
                                                                                    idcargo=fijo.idcargo,
                                                                                    comision=fijo.porc_fijo,
                                                                                    usuario='Activo',
                                                                                    )
                        alerta=True
                        mensaje='La Escala fue asignada de forma correcta'
                        titulo='¡Todo salio a la perfeccion!'
                    except:
                        alerta=True
                        mensaje='Ya hay una escala asignada a este contrato, si deseas cambiarla usa el boton Modificar'
                        titulo='Ha ocurrido un Error'
                if request.POST.get('modificar-escala'):
                    for valor in values:
                        cargo_existe=AsignacionComisiones.objects.using(proyecto).filter(idadjudicacion=contrato,idcargo=valor[2]).exists()
                        if cargo_existe:
                            cargo=AsignacionComisiones.objects.using(proyecto).get(idadjudicacion=contrato,idcargo=valor[2])
                        if cargo_existe and valor[0]=='':
                            cargo.delete()
                        elif cargo_existe and valor[0]!='':
                            cargo.idgestor=valor[0]
                            cargo.comision=valor[1]
                            cargo.save()
                        elif valor!=0 and not cargo_existe:
                            AsignacionComisiones.objects.using(proyecto).create(id_comision=f'{valor[2]}-{contrato}',
                                                                                idadjudicacion=contrato,
                                                                                fecha=datetime.datetime.today(),
                                                                                idgestor=valor[0],
                                                                                idcargo=valor[2],
                                                                                comision=valor[1],
                                                                                usuario=request.user,
                                                                                )
                    alerta=True
                    mensaje='La Escala fue modificada de forma correcta'
                    titulo='¡Todo salio a la perfeccion!'

            form_docs=form_docs_contratos(request.POST,request.FILES)
            if form_docs.is_valid():
                descrip_doc=f"{form_docs.cleaned_data.get('tipo_doc')}_{datetime.date.today()}"
                fecha_carga=datetime.datetime.today()
                usuario_carga=request.user
                upload_docs_contratos(request.FILES['documento_cargar'],contrato,proyecto,descrip_doc)
                documentos_contratos.objects.using(proyecto).create(adj=contrato,descripcion_doc=descrip_doc,
                                                        fecha_carga=fecha_carga,usuario_carga=usuario_carga)
                
        lista_clientes=clientes.objects.using('default').all()
        datos_venta=ventas_nuevas.objects.using(proyecto).get(id_venta=contrato)
        datos_inmueble=Inmuebles.objects.using(proyecto).get(idinmueble=datos_venta.inmueble)
        lista_asesores=asesores.objects.using('default').filter(estado='Activo')
        escala=AsignacionComisiones.objects.using(proyecto)
        recibos_generados=RecaudosNoradicados.objects.using(proyecto).filter(contrato=contrato)
        try:
            generador=escala.filter(idadjudicacion=contrato,idcargo=8).values()
            generador=generador[0]
            nombre_asesor=asesores.objects.using('default').get(cedula=generador['idgestor'])
            generador['nombre']=nombre_asesor.nombre
        except: generador=[]
        try: 
            linea=escala.filter(idadjudicacion=contrato,idcargo=9).values()
            linea=linea[0]
            nombre_asesor=asesores.objects.using('default').get(cedula=linea['idgestor'])
            linea['nombre']=nombre_asesor.nombre
        except: linea=[]
        try: 
            cerrador=escala.filter(idadjudicacion=contrato,idcargo=99).values()
            cerrador=cerrador[0]
            nombre_asesor=asesores.objects.using('default').get(cedula=cerrador['idgestor'])
            cerrador['nombre']=nombre_asesor.nombre
        except: cerrador=[]
        try: 
            gerenteventas=escala.filter(idadjudicacion=contrato,idcargo=14).values()
            gerenteventas=gerenteventas[0]
            nombre_asesor=asesores.objects.using('default').get(cedula=gerenteventas['idgestor'])
            gerenteventas['nombre']=nombre_asesor.nombre
        except: gerenteventas=[]
        try: 
            jefep=escala.filter(idadjudicacion=contrato,idcargo=15).values()
            jefep=jefep[0]
            nombre_asesor=asesores.objects.using('default').get(cedula=jefep['idgestor'])
            jefep['nombre']=nombre_asesor.nombre
        except : jefep=[]
        try: 
            jefev=escala.filter(idadjudicacion=contrato,idcargo=13).values()
            jefev=jefev[0]
            nombre_asesor=asesores.objects.using('default').get(cedula=jefev['idgestor'])
            jefev['nombre']=nombre_asesor.nombre
        except: jefev=[]
        try: datos_t1=clientes.objects.using('default').get(idTercero=datos_venta.id_t1)
        except: datos_t1=[]
        try: datos_t2=clientes.objects.using('default').get(idTercero=datos_venta.id_t2)
        except: datos_t2=[]
        try: datos_t3=clientes.objects.using('default').get(idTercero=datos_venta.id_t3)
        except: datos_t3=[]
        try: datos_t4=clientes.objects.using('default').get(idTercero=datos_venta.id_t4)
        except: datos_t4=[]
        tasa=1.5/100
        nro_ci=0
        obj_docs=documentos_contratos.objects.using(proyecto).filter(adj=contrato)

        fin_pagos_reg = datos_venta.inicio_fn + relativedelta(months=datos_venta.nro_cuotas_fn + 1)
        if datos_venta.inicio_ce:
            fin_pagos_ce = datos_venta.inicio_ce + relativedelta(months=datos_venta.nro_cuotas_ce + 1)
        else: fin_pagos_ce = fin_pagos_reg

        fin_pagos = max(fin_pagos_reg,fin_pagos_ce)

        meses_entrega_inmueble = int(datos_inmueble.meses or 0)
        min_meses_entrega = max(36, meses_entrega_inmueble)
        f_entrega = datos_venta.fecha_contrato + relativedelta(months=min_meses_entrega)

        if fin_pagos > f_entrega: 
            f_escritura = fin_pagos
        else:
            f_escritura = f_entrega + relativedelta(months = 1)


        context={
            'recibos_generados':recibos_generados,
            'form_recaudo':form_recaudo_noradicado,
            'form_docs':form_docs_contratos,
            'lista_documentos':obj_docs,
            'form':form_nueva_venta,
            'asesores':lista_asesores,
            'clientes':lista_clientes,
            'proyecto':proyecto,
            'contrato':contrato,
            'datos_venta':datos_venta,
            'datos_t1':datos_t1,
            'datos_t2':datos_t2,
            'datos_t3':datos_t3,
            'datos_t4':datos_t4,
            'tasa':tasas,
            'tasa_pactada':datos_venta.tasa,
            'datos_inmueble':datos_inmueble,
            'btnName':'Modificar',
            'form_escala':form_escala_comision,
            'generador':generador,
            'linea':linea,
            'cerrador':cerrador,
            'gerenteventas':gerenteventas,
            'jefep':jefep,
            'jefev':jefev,
            'tipo_form':'acciones',
            'cargos_comisiones':Cargos_comisiones.objects.using(proyecto).all(),
            'f_entrega':f_entrega,
            'f_escritura':f_escritura,
            'min_meses_entrega':min_meses_entrega,
        }
        
        pdf=GenerarPDF()
        if request.method == 'GET':
            if request.is_ajax():
                tipo=request.GET.get('tipo')
                if tipo=='recibo':
                    recibo=request.GET.get('recibo')
                    datos_recaudo=RecaudosNoradicados.objects.using(proyecto).get(recibo=recibo)
                    filename=f'{proyecto}_reciboNR_{recibo}.pdf'
                    if proyecto == 'Oasis':
                        result = pdf_gen_weasy(
                            f'pdf/{proyecto}/recibo_nr.html',
                            {'recibo': datos_recaudo},
                            filename,
                        )
                        dir_download = result['url']
                    else:
                        ruta=settings.DIR_EXPORT+filename
                        nombre_t1=datos_t1.nombrecompleto
                        resid_t1=str(datos_t1.domicilio)
                        cdresid_t1=str(datos_t1.ciudad)
                        cel_t1=str(datos_t1.celular1)
                        pdf.Recibo_caja(proyecto=proyecto,
                                        ruta=ruta,
                                        nroRecibo=recibo,
                                        titular1=nombre_t1,
                                        fecha=str(datos_recaudo.fecha),
                                        concepto=datos_recaudo.concepto,
                                        valor=datos_recaudo.valor,
                                        direccion=resid_t1,
                                        ciudad=cdresid_t1,
                                        telefono=cel_t1,
                                        formapag=datos_recaudo.formapago,
                                        user=request.user)
                        dir_download=settings.DIR_DOWNLOADS+filename
                    return JsonResponse({'instance':dir_download},status=200)
                elif tipo=='actualizar_fecha':
                    check_perms(request,('andinasoft.delete_ventas_nuevas',))
                    datos_venta.fecha_contrato=datetime.date.today()
                    datos_venta.save()
                    return JsonResponse({'instance':True},status=200)
        if request.method == 'POST':
            nro_contrato=contrato
            nombre_t1=datos_t1.nombrecompleto
            cc_t1=datos_t1.idTercero
            tel_t1=str(datos_t1.telefono1)
            cel_t1=str(datos_t1.celular1)
            ofic_t1=str(datos_t1.oficina)
            cdof_t1=''
            telof_t1=str(datos_t1.telefono2)
            resid_t1=str(datos_t1.domicilio)
            cdresid_t1=str(datos_t1.ciudad)
            telresid_t1=str(datos_t1.telefono1)
            email_t1=str(datos_t1.email)
            if datos_t2!=[]:
                nombre_t2=str(datos_t2.nombrecompleto)
                cc_t2=str(datos_t2.idTercero)
                tel_t2=str(datos_t2.telefono1)
                cel_t2=str(datos_t2.celular1)
                ofic_t2=str(datos_t2.oficina)
                cdof_t2=''
                telof_t2=str(datos_t2.telefono2)
                resid_t2=str(datos_t2.domicilio)
                cdresid_t2=str(datos_t2.ciudad)
                telresid_t2=str(datos_t2.telefono1)
                email_t2=str(datos_t2.email)
            else:
                nombre_t2=""
                cc_t2=""
                tel_t2=""
                cel_t2=""
                ofic_t2=""
                cdof_t2=''
                telof_t2=""
                resid_t2=""
                cdresid_t2=""
                telresid_t2=""
                email_t2=""
            if datos_t3!=[]:
                nombre_t3=str(datos_t3.nombrecompleto)
                cc_t3=str(datos_t3.idTercero)
                tel_t3=str(datos_t3.telefono1)
                cel_t3=str(datos_t3.celular1)
                ofic_t3=str(datos_t3.oficina)
                cdof_t3=''
                telof_t3=str(datos_t3.telefono2)
                resid_t3=str(datos_t3.domicilio)
                cdresid_t3=str(datos_t3.ciudad)
                telresid_t3=str(datos_t3.telefono1)
                email_t3=str(datos_t3.email)
            else:
                nombre_t3=""
                cc_t3=""
                tel_t3=""
                cel_t3=""
                ofic_t3=""
                cdof_t3=''
                telof_t3=""
                resid_t3=""
                cdresid_t3=""
                telresid_t3=""
                email_t3=""
            if datos_t4!=[]:
                nombre_t4=str(datos_t4.nombrecompleto)
                cc_t4=str(datos_t4.idTercero)
                tel_t4=str(datos_t4.telefono1)
                cel_t4=str(datos_t4.celular1)
                ofic_t4=str(datos_t4.oficina)
                cdof_t4=''
                telof_t4=str(datos_t4.telefono2)
                resid_t4=str(datos_t4.domicilio)
                cdresid_t4=str(datos_t4.ciudad)
                telresid_t4=str(datos_t4.telefono1)
                email_t4=str(datos_t4.email)
            else:
                nombre_t4=""
                cc_t4=""
                tel_t4=""
                cel_t4=""
                ofic_t4=""
                cdof_t4=''
                telof_t4=""
                resid_t4=""
                cdresid_t4=""
                telresid_t4=""
                email_t4=""
            lote=datos_inmueble.lotenumero
            manzana=datos_inmueble.manzananumero
            area=str(datos_inmueble.areaprivada)
            mtsnorte=str(datos_inmueble.norte)
            colnorte=datos_inmueble.colindante_norte
            mtseste=str(datos_inmueble.este)
            coleste=datos_inmueble.colidante_este
            mtssur=str(datos_inmueble.sur)
            colsur=datos_inmueble.colindante_sur
            mtsoeste=str(datos_inmueble.oeste)
            coloeste=datos_inmueble.colindante_oeste
            tiempo_entrega=str(datos_inmueble.meses)
            valor=datos_venta.valor_venta
            valor_letras=Utilidades().numeros_letras(valor)
            ci=datos_venta.cuota_inicial
            saldo=datos_venta.saldo
            if datos_venta.forma_pago=="Contado":
                fp=('x','','')
            elif datos_venta.forma_pago=="Credicontado":
                fp=('','x','')
            else:
                fp=('','','x')
            contado_x=fp[0]
            credic_x=fp[1]
            amort_x=fp[2]
            fci=''
            if datos_venta.cant_ci1!=None:
                if datos_venta.cant_ci1>1:
                    fci+=f'{datos_venta.cant_ci1} cuotas mensuales de ${datos_venta.valor_ci1:,} a partir del {datos_venta.fecha_ci1}'
                else:
                    fci+=f'1 cuota de ${datos_venta.valor_ci1:,} el {datos_venta.fecha_ci1}'
            if datos_venta.cant_ci2!=None:
                if datos_venta.cant_ci2>1:
                    fci+=f'; {datos_venta.cant_ci2} cuotas mensuales de ${datos_venta.valor_ci2:,} a partir del {datos_venta.fecha_ci2}'
                else:
                    fci+=f'; 1 cuota de ${datos_venta.valor_ci2:,} el {datos_venta.fecha_ci2}'
            if datos_venta.cant_ci3!=None:
                if datos_venta.cant_ci3>1:
                    fci+=f'; {datos_venta.cant_ci3} cuotas mensuales de ${datos_venta.valor_ci3:,} a partir del {datos_venta.fecha_ci3}'
                else:
                    fci+=f'; 1 cuota de ${datos_venta.valor_ci3:,} el {datos_venta.fecha_ci3}'
            if datos_venta.cant_ci4!=None:
                if datos_venta.cant_ci4>1:
                    fci+=f'; {datos_venta.cant_ci4} cuotas mensuales de ${datos_venta.valor_ci4:,} a partir del {datos_venta.fecha_ci4}'
                else:
                    fci+=f'; 1 cuota de ${datos_venta.valor_ci4:,} el {datos_venta.fecha_ci4}'
            if datos_venta.cant_ci5!=None:
                if datos_venta.cant_ci5>1:
                    fci+=f'; {datos_venta.cant_ci5} cuotas mensuales de ${datos_venta.valor_ci5:,} a partir del {datos_venta.fecha_ci5}'
                else:
                    fci+=f'; 1 cuota de ${datos_venta.valor_ci5:,} el {datos_venta.fecha_ci5}'
            if datos_venta.cant_ci6!=None:
                if datos_venta.cant_ci6>1:
                    fci+=f'; {datos_venta.cant_ci6} cuotas mensuales de ${datos_venta.valor_ci6:,} a partir del {datos_venta.fecha_ci6}'
                else:
                    fci+=f'; 1 cuota de ${datos_venta.valor_ci6:,} el {datos_venta.fecha_ci6}'
            if datos_venta.cant_ci7!=None:
                if datos_venta.cant_ci7>1:
                    fci+=f'; {datos_venta.cant_ci7} cuotas mensuales de ${datos_venta.valor_ci7:,} a partir del {datos_venta.fecha_ci7}'
                else:
                    fci+=f'; 1 cuota de ${datos_venta.valor_ci7:,} el {datos_venta.fecha_ci7}'
            formaCI=fci
            if datos_venta.nro_cuotas_fn>1:
                ffn=f'{datos_venta.nro_cuotas_fn} cuotas mensuales de ${datos_venta.valor_ctas_fn:,} a partir de {datos_venta.inicio_fn}'
            elif datos_venta.nro_cuotas_fn==0:
                ffn=''
            else:
                ffn=f'{datos_venta.nro_cuotas_fn} cuota de ${datos_venta.valor_ctas_fn:,} el {datos_venta.inicio_fn}'
            if datos_venta.nro_cuotas_ce!=None and datos_venta.nro_cuotas_ce!='':
                if datos_venta.nro_cuotas_ce>1:
                    ffn+=f' y {datos_venta.nro_cuotas_ce} cuotas {datos_venta.period_ce} por valor de ${datos_venta.valor_ctas_ce:,} a partir de {datos_venta.inicio_ce}'
                else:
                    ffn+=f' y {datos_venta.nro_cuotas_ce} cuota por valor de ${datos_venta.valor_ctas_ce:,} el {datos_venta.inicio_ce}'
            formaFN=ffn
            obs=datos_venta.observaciones
            dia_contrato=str(datos_venta.fecha_contrato.day)
            mes_contrato=datos_venta.fecha_contrato.month
            año_contrato=str(datos_venta.fecha_contrato.year)
            fecha_entrega=datos_inmueble.finobra
            today=datetime.date.today()
            obj_ctr = ventas_nuevas.objects.using(proyecto).get(pk=contrato)
            if (not (check_perms(request,('andinasoft.delete_ventas_nuevas',),raise_exception=False) 
                     or check_perms(request,('andinasoft.add_recaudos_general',),raise_exception=False))
                and today>datos_venta.fecha_contrato 
                and (request.POST.get('impContrato') 
                     or request.POST.get('impPagare') 
                     or request.POST.get('impVerificacion') 
                     or request.POST.get('impRecaudo'))):
                alerta=True
                mensaje='''Las opciones de impresion se pueden usar solo el mismo dia de la venta, si necesita reimpresiones solicitelas a Gerencia Comercial o al departamento de Operaciones'''
                titulo='Error'
                link=False
            else:
                if request.POST.get('impPromesa'):
                    fecha_entrega=request.POST.get('fechaentrega')
                    fecha_entrega=datetime.datetime.strptime(fecha_entrega,'%Y-%m-%d')
                    meses_entrega = (fecha_entrega - datetime.datetime.today()).days
                    meses_entrega = math.ceil(meses_entrega/30)
                    fecha_escritura=request.POST.get('fechaescritura')
                    fecha_escritura=datetime.datetime.strptime(fecha_escritura,'%Y-%m-%d')
                    oficina=request.POST.get('oficinaopcion')
                    ruta=settings.DIR_EXPORT+f'{proyecto}_contrato_{contrato}.pdf'

                    fecha_minima_entrega = datos_venta.fecha_contrato + relativedelta(months=min_meses_entrega)
                    if fecha_entrega.date() < fecha_minima_entrega:
                        alerta = True
                        titulo = 'Error'
                        link = False
                        mensaje = (
                            f'La fecha de entrega no puede ser menor a {min_meses_entrega} meses '
                            f'desde la fecha del contrato ({datos_venta.fecha_contrato}). '
                            f'Fecha mínima permitida: {fecha_minima_entrega}.'
                        )
                        context['alerta'] = True
                        context['titulo'] = titulo
                        context['mensaje'] = mensaje
                        context['link'] = False
                        return render(request,'acciones_venta.html',context)
                    if fecha_escritura.date() < fecha_entrega.date():
                        alerta = True
                        titulo = 'Error'
                        link = False
                        mensaje = (
                            f'La fecha de escritura ({fecha_escritura.date()}) no puede ser anterior '
                            f'a la fecha de entrega ({fecha_entrega.date()}).'
                        )
                        context['alerta'] = True
                        context['titulo'] = titulo
                        context['mensaje'] = mensaje
                        context['link'] = False
                        return render(request,'acciones_venta.html',context)
                    
                    parametro=obj_parametro.get(descripcion='formasOpcionManual')
                    if parametro.estado:
                        formaCI = request.POST.get('formaci')
                        formaFN = request.POST.get('formasaldo')
                    if proyecto=='Sandville Beach':
                        pdf.ExportPromesaSandvilleBeach(nro_contrato=nro_contrato,
                                                    nombre_t1=nombre_t1,
                                                    cc_t1=cc_t1,
                                                    tel_t1=tel_t1,
                                                    cel_t1=cel_t1,
                                                    ofic_t1=ofic_t1,
                                                    cdof_t1=cdof_t1,
                                                    telof_t1=telof_t1,
                                                    resid_t1=resid_t1,
                                                    cdresid_t1=cdresid_t1,
                                                    telresid_t1=telresid_t1,
                                                    email_t1=email_t1,
                                                    nombre_t2=nombre_t2,
                                                    cc_t2=cc_t2,
                                                    tel_t2=tel_t2,
                                                    cel_t2=cel_t2,
                                                    ofic_t2=ofic_t2,
                                                    cdof_t2=cdof_t2,
                                                    telof_t2=telof_t2,
                                                    resid_t2=resid_t2,
                                                    cdresid_t2=cdresid_t2,
                                                    telresid_t2=telresid_t2,
                                                    email_t2=email_t2,
                                                    nombre_t3=nombre_t3,
                                                    cc_t3=cc_t3,
                                                    tel_t3=tel_t3,
                                                    cel_t3=cel_t3,
                                                    ofic_t3=ofic_t3,
                                                    cdof_t3=cdof_t3,
                                                    telof_t3=telof_t3,
                                                    resid_t3=resid_t3,
                                                    cdresid_t3=cdresid_t3,
                                                    telresid_t3=telresid_t3,
                                                    email_t3=email_t3,
                                                    nombre_t4=nombre_t4,
                                                    cc_t4=cc_t4,
                                                    tel_t4=tel_t4,
                                                    cel_t4=cel_t4,
                                                    ofic_t4=ofic_t4,
                                                    cdof_t4=cdof_t4,
                                                    telof_t4=telof_t4,
                                                    resid_t4=resid_t4,
                                                    cdresid_t4=cdresid_t4,
                                                    telresid_t4=telresid_t4,
                                                    email_t4=email_t4,
                                                    lote=lote,
                                                    manzana=manzana,
                                                    area=area,
                                                    mtsnorte=mtsnorte,
                                                    colnorte=colnorte,
                                                    mtseste=mtseste,
                                                    coleste=coleste,
                                                    mtssur=mtssur,
                                                    colsur=colsur,
                                                    mtsoeste=mtsoeste,
                                                    coloeste=coloeste,
                                                    valor=valor,
                                                    valor_letras=valor_letras,
                                                    ci=ci,
                                                    saldo=saldo,
                                                    contado_x=contado_x,
                                                    credic_x=credic_x,
                                                    amort_x=amort_x,
                                                    formaCI=formaCI,
                                                    formaFN=formaFN,
                                                    obs=obs,
                                                    dia_contrato=dia_contrato,
                                                    mes_contrato=mes_contrato,
                                                    año_contrato=año_contrato,
                                                    fecha_escritura=fecha_escritura,
                                                    fecha_entrega=fecha_entrega,
                                                    ciudad_entrega=oficina,
                                                    ruta=ruta)
                        mensaje='Descarga el contrato aqui'
                        titulo='¡Listo!'
                        link=True
                    elif proyecto=='Perla del Mar':
                                                
                        context = {
                            'proyecto':proyecto,
                            'ctr':obj_ctr,
                            'fecha_escritura':fecha_escritura,
                            'meses_entrega':meses_entrega,
                            'oficina':oficina
                        }
                        
                        filename = f'Contrato_bien_futuro_{contrato}_{proyecto}.pdf'
                        
                            
                        pdf = pdf_gen(f'pdf/{proyecto}/contrato.html', context, filename)
                        
                        file = pdf.get('root')
                                                
                        return FileResponse(open(file,'rb'),as_attachment=True,filename=filename)
                    elif proyecto=='Tesoro Escondido':
                        porcDerecho=f'{(datos_inmueble.areaprivada*100/datos_inmueble.area_mz):.2f}'
                        pdf.ExportPromesaBugambilias(nro_contrato=nro_contrato,
                                                    nombre_t1=nombre_t1,
                                                    cc_t1=cc_t1,
                                                    tel_t1=tel_t1,
                                                    cel_t1=cel_t1,
                                                    ofic_t1=ofic_t1,
                                                    cdof_t1=cdof_t1,
                                                    telof_t1=telof_t1,
                                                    resid_t1=resid_t1,
                                                    cdresid_t1=cdresid_t1,
                                                    telresid_t1=telresid_t1,
                                                    email_t1=email_t1,
                                                    nombre_t2=nombre_t2,
                                                    cc_t2=cc_t2,
                                                    tel_t2=tel_t2,
                                                    cel_t2=cel_t2,
                                                    ofic_t2=ofic_t2,
                                                    cdof_t2=cdof_t2,
                                                    telof_t2=telof_t2,
                                                    resid_t2=resid_t2,
                                                    cdresid_t2=cdresid_t2,
                                                    telresid_t2=telresid_t2,
                                                    email_t2=email_t2,
                                                    nombre_t3=nombre_t3,
                                                    cc_t3=cc_t3,
                                                    tel_t3=tel_t3,
                                                    cel_t3=cel_t3,
                                                    ofic_t3=ofic_t3,
                                                    cdof_t3=cdof_t3,
                                                    telof_t3=telof_t3,
                                                    resid_t3=resid_t3,
                                                    cdresid_t3=cdresid_t3,
                                                    telresid_t3=telresid_t3,
                                                    email_t3=email_t3,
                                                    nombre_t4=nombre_t4,
                                                    cc_t4=cc_t4,
                                                    tel_t4=tel_t4,
                                                    cel_t4=cel_t4,
                                                    ofic_t4=ofic_t4,
                                                    cdof_t4=cdof_t4,
                                                    telof_t4=telof_t4,
                                                    resid_t4=resid_t4,
                                                    cdresid_t4=cdresid_t4,
                                                    telresid_t4=telresid_t4,
                                                    email_t4=email_t4,
                                                    lote=lote,
                                                    manzana=manzana,
                                                    area=area,
                                                    mtsnorte=mtsnorte,
                                                    colnorte=colnorte,
                                                    mtseste=mtseste,
                                                    coleste=coleste,
                                                    mtssur=mtssur,
                                                    colsur=colsur,
                                                    mtsoeste=mtsoeste,
                                                    coloeste=coloeste,
                                                    valor=valor,
                                                    valor_letras=valor_letras,
                                                    ci=ci,
                                                    saldo=saldo,
                                                    contado_x=contado_x,
                                                    credic_x=credic_x,
                                                    amort_x=amort_x,
                                                    formaCI=formaCI,
                                                    formaFN=formaFN,
                                                    obs=obs,
                                                    dia_contrato=dia_contrato,
                                                    mes_contrato=mes_contrato,
                                                    año_contrato=año_contrato,
                                                    fecha_escritura=fecha_escritura,
                                                    fecha_entrega=fecha_entrega,
                                                    ciudad_entrega=oficina,
                                                    ruta=ruta,
                                                    porcderecho=porcDerecho,
                                                    area_parcela=str(datos_inmueble.area_mz))
                        mensaje='Descarga el contrato aqui'
                        titulo='¡Listo!'
                        link=True
                    elif proyecto=='Sandville del Sol':
                        porcDerecho=f'{(datos_inmueble.areaprivada*100/datos_inmueble.area_mz):.2f}'
                        pdf.ExportOpcionAraza(nro_contrato=nro_contrato,
                                                    nombre_t1=nombre_t1,
                                                    cc_t1=cc_t1,
                                                    tel_t1=tel_t1,
                                                    cel_t1=cel_t1,
                                                    ofic_t1=ofic_t1,
                                                    cdof_t1=cdof_t1,
                                                    telof_t1=telof_t1,
                                                    resid_t1=resid_t1,
                                                    cdresid_t1=cdresid_t1,
                                                    telresid_t1=telresid_t1,
                                                    email_t1=email_t1,
                                                    nombre_t2=nombre_t2,
                                                    cc_t2=cc_t2,
                                                    tel_t2=tel_t2,
                                                    cel_t2=cel_t2,
                                                    ofic_t2=ofic_t2,
                                                    cdof_t2=cdof_t2,
                                                    telof_t2=telof_t2,
                                                    resid_t2=resid_t2,
                                                    cdresid_t2=cdresid_t2,
                                                    telresid_t2=telresid_t2,
                                                    email_t2=email_t2,
                                                    nombre_t3=nombre_t3,
                                                    cc_t3=cc_t3,
                                                    tel_t3=tel_t3,
                                                    cel_t3=cel_t3,
                                                    ofic_t3=ofic_t3,
                                                    cdof_t3=cdof_t3,
                                                    telof_t3=telof_t3,
                                                    resid_t3=resid_t3,
                                                    cdresid_t3=cdresid_t3,
                                                    telresid_t3=telresid_t3,
                                                    email_t3=email_t3,
                                                    nombre_t4=nombre_t4,
                                                    cc_t4=cc_t4,
                                                    tel_t4=tel_t4,
                                                    cel_t4=cel_t4,
                                                    ofic_t4=ofic_t4,
                                                    cdof_t4=cdof_t4,
                                                    telof_t4=telof_t4,
                                                    resid_t4=resid_t4,
                                                    cdresid_t4=cdresid_t4,
                                                    telresid_t4=telresid_t4,
                                                    email_t4=email_t4,
                                                    lote=lote,
                                                    manzana=manzana,
                                                    area=area,
                                                    porcderecho=porcDerecho,
                                                    mtsnorte=mtsnorte,
                                                    colnorte=colnorte,
                                                    mtseste=mtseste,
                                                    coleste=coleste,
                                                    mtssur=mtssur,
                                                    colsur=colsur,
                                                    mtsoeste=mtsoeste,
                                                    coloeste=coloeste,
                                                    valor=valor,
                                                    valor_letras=valor_letras,
                                                    ci=ci,
                                                    saldo=saldo,
                                                    contado_x=contado_x,
                                                    credic_x=credic_x,
                                                    amort_x=amort_x,
                                                    formaCI=formaCI,
                                                    formaFN=formaFN,
                                                    obs=obs,
                                                    dia_contrato=dia_contrato,
                                                    mes_contrato=mes_contrato,
                                                    año_contrato=año_contrato,
                                                    fecha_escritura=fecha_escritura,
                                                    fecha_entrega=fecha_entrega,
                                                    ciudad_entrega=oficina,
                                                    ruta=ruta)
                        mensaje='Descarga el contrato aqui'
                        titulo='¡Listo!'
                        link=True
                    elif proyecto == 'Vegas de Venecia':
                        pdf.ExportCBFVegasVenecia(nro_contrato=nro_contrato,
                                                    nombre_t1=nombre_t1,
                                                    cc_t1=cc_t1,
                                                    tel_t1=tel_t1,
                                                    cel_t1=cel_t1,
                                                    ofic_t1=ofic_t1,
                                                    cdof_t1=cdof_t1,
                                                    telof_t1=telof_t1,
                                                    resid_t1=resid_t1,
                                                    cdresid_t1=cdresid_t1,
                                                    telresid_t1=telresid_t1,
                                                    email_t1=email_t1,
                                                    nombre_t2=nombre_t2,
                                                    cc_t2=cc_t2,
                                                    tel_t2=tel_t2,
                                                    cel_t2=cel_t2,
                                                    ofic_t2=ofic_t2,
                                                    cdof_t2=cdof_t2,
                                                    telof_t2=telof_t2,
                                                    resid_t2=resid_t2,
                                                    cdresid_t2=cdresid_t2,
                                                    telresid_t2=telresid_t2,
                                                    email_t2=email_t2,
                                                    nombre_t3=nombre_t3,
                                                    cc_t3=cc_t3,
                                                    tel_t3=tel_t3,
                                                    cel_t3=cel_t3,
                                                    ofic_t3=ofic_t3,
                                                    cdof_t3=cdof_t3,
                                                    telof_t3=telof_t3,
                                                    resid_t3=resid_t3,
                                                    cdresid_t3=cdresid_t3,
                                                    telresid_t3=telresid_t3,
                                                    email_t3=email_t3,
                                                    nombre_t4=nombre_t4,
                                                    cc_t4=cc_t4,
                                                    tel_t4=tel_t4,
                                                    cel_t4=cel_t4,
                                                    ofic_t4=ofic_t4,
                                                    cdof_t4=cdof_t4,
                                                    telof_t4=telof_t4,
                                                    resid_t4=resid_t4,
                                                    cdresid_t4=cdresid_t4,
                                                    telresid_t4=telresid_t4,
                                                    email_t4=email_t4,
                                                    lote=lote,
                                                    manzana=manzana,
                                                    area=area,
                                                    mtsnorte=mtsnorte,
                                                    colnorte=colnorte,
                                                    mtseste=mtseste,
                                                    coleste=coleste,
                                                    mtssur=mtssur,
                                                    colsur=colsur,
                                                    mtsoeste=mtsoeste,
                                                    coloeste=coloeste,
                                                    valor=valor,
                                                    valor_letras=valor_letras,
                                                    ci=ci,
                                                    saldo=saldo,
                                                    contado_x=contado_x,
                                                    credic_x=credic_x,
                                                    amort_x=amort_x,
                                                    formaCI=formaCI,
                                                    formaFN=formaFN,
                                                    obs=obs,
                                                    dia_contrato=dia_contrato,
                                                    mes_contrato=mes_contrato,
                                                    año_contrato=año_contrato,
                                                    fecha_escritura=fecha_escritura,
                                                    fecha_entrega=fecha_entrega,
                                                    ciudad_entrega=oficina,
                                                    meses_entrega=tiempo_entrega,
                                                    ruta=ruta)
                        mensaje='Descarga el contrato aqui'
                        titulo='¡Listo!'
                        link=True
                    elif proyecto == 'Carmelo Reservado':
                        context = {
                            'proyecto':proyecto,
                            'ctr':obj_ctr,
                            'fecha_escritura':fecha_escritura,
                            'meses_entrega':meses_entrega,
                            'oficina':oficina
                        }
                        
                        filename = f'Contrato_bien_futuro_{contrato}_{proyecto}.pdf'
                        
                            
                        pdf = pdf_gen(f'pdf/{proyecto}/contrato.html', context, filename)
                        
                        file = pdf.get('root')
                                                
                        return FileResponse(open(file,'rb'),as_attachment=True,filename=filename)
                    elif proyecto == 'Casas de Verano' or proyecto == 'Oasis':
                        context = {
                            'proyecto':proyecto,
                            'ctr':obj_ctr,
                            'fecha_escritura':fecha_escritura,
                            'meses_entrega':meses_entrega,
                            'oficina':oficina,
                            'plan_pagos':calcular_tabla_amortizacion(obj_ctr),
                        }
                        
                        filename = f'Contrato_bien_futuro_{contrato}_{proyecto}.pdf'
                        
                            
                        if proyecto == 'Oasis':
                            pdf = pdf_gen_weasy(f'pdf/{proyecto}/contrato.html', context, filename)
                        else:
                            pdf = pdf_gen(f'pdf/{proyecto}/contrato.html', context, filename)
                        
                        file = pdf.get('root')
                                                
                        return FileResponse(open(file,'rb'),as_attachment=True,filename=filename)
                    else:
                        mensaje='El proyecto seleccionado no tiene formato de promesa asignado'
                        titulo='Error'
                    alerta=True
                    parameter=obj_parametro.get(descripcion='formasOpcionManual')
                    parameter.estado=False
                    parameter.save()
                    ruta_link=settings.DIR_DOWNLOADS+f'{proyecto}_contrato_{contrato}.pdf'
                    
                if request.POST.get('impContrato'):
                    ruta=settings.DIR_EXPORT+f'{proyecto}_contrato_{contrato}.pdf'
                    parametro=obj_parametro.get(descripcion='formasOpcionManual')
                    if parametro.estado:
                        formaCI = request.POST.get('formaci')
                        formaFN = request.POST.get('formasaldo')
                    if proyecto=='Vegas de Venecia':
                        pdf.ExportOpcionContratoVenecia(nro_contrato=nro_contrato,
                                                    nombre_t1=nombre_t1,
                                                    cc_t1=cc_t1,
                                                    tel_t1=tel_t1,
                                                    cel_t1=cel_t1,
                                                    ofic_t1=ofic_t1,
                                                    cdof_t1=cdof_t1,
                                                    telof_t1=telof_t1,
                                                    resid_t1=resid_t1,
                                                    cdresid_t1=cdresid_t1,
                                                    telresid_t1=telresid_t1,
                                                    email_t1=email_t1,
                                                    nombre_t2=nombre_t2,
                                                    cc_t2=cc_t2,
                                                    tel_t2=tel_t2,
                                                    cel_t2=cel_t2,
                                                    ofic_t2=ofic_t2,
                                                    cdof_t2=cdof_t2,
                                                    telof_t2=telof_t2,
                                                    resid_t2=resid_t2,
                                                    cdresid_t2=cdresid_t2,
                                                    telresid_t2=telresid_t2,
                                                    email_t2=email_t2,
                                                    nombre_t3=nombre_t3,
                                                    cc_t3=cc_t3,
                                                    tel_t3=tel_t3,
                                                    cel_t3=cel_t3,
                                                    ofic_t3=ofic_t3,
                                                    cdof_t3=cdof_t3,
                                                    telof_t3=telof_t3,
                                                    resid_t3=resid_t3,
                                                    cdresid_t3=cdresid_t3,
                                                    telresid_t3=telresid_t3,
                                                    email_t3=email_t3,
                                                    nombre_t4=nombre_t4,
                                                    cc_t4=cc_t4,
                                                    tel_t4=tel_t4,
                                                    cel_t4=cel_t4,
                                                    ofic_t4=ofic_t4,
                                                    cdof_t4=cdof_t4,
                                                    telof_t4=telof_t4,
                                                    resid_t4=resid_t4,
                                                    cdresid_t4=cdresid_t4,
                                                    telresid_t4=telresid_t4,
                                                    email_t4=email_t4,
                                                    lote=lote,
                                                    manzana=manzana,
                                                    area=area,
                                                    mtsnorte=mtsnorte,
                                                    colnorte=colnorte,
                                                    mtseste=mtseste,
                                                    coleste=coleste,
                                                    mtssur=mtssur,
                                                    colsur=colsur,
                                                    mtsoeste=mtsoeste,
                                                    coloeste=coloeste,
                                                    valor=valor,
                                                    valor_letras=valor_letras,
                                                    ci=ci,
                                                    saldo=saldo,
                                                    contado_x=contado_x,
                                                    credic_x=credic_x,
                                                    amort_x=amort_x,
                                                    formaCI=formaCI,
                                                    formaFN=formaFN,
                                                    obs=obs,
                                                    dia_contrato=dia_contrato,
                                                    mes_contrato=mes_contrato,
                                                    año_contrato=año_contrato,
                                                    ruta=ruta)
                        mensaje='Descarga el contrato aqui'
                        titulo='¡Listo!'
                        link=True
                    elif proyecto=='Tesoro Escondido':
                        porcDerecho=f'{(datos_inmueble.areaprivada*100/datos_inmueble.area_mz):.2f}'
                        pdf.ExportOpcionTesoro(nro_contrato=nro_contrato,
                                                    nombre_t1=nombre_t1,
                                                    cc_t1=cc_t1,
                                                    tel_t1=tel_t1,
                                                    cel_t1=cel_t1,
                                                    ofic_t1=ofic_t1,
                                                    cdof_t1=cdof_t1,
                                                    telof_t1=telof_t1,
                                                    resid_t1=resid_t1,
                                                    cdresid_t1=cdresid_t1,
                                                    telresid_t1=telresid_t1,
                                                    email_t1=email_t1,
                                                    nombre_t2=nombre_t2,
                                                    cc_t2=cc_t2,
                                                    tel_t2=tel_t2,
                                                    cel_t2=cel_t2,
                                                    ofic_t2=ofic_t2,
                                                    cdof_t2=cdof_t2,
                                                    telof_t2=telof_t2,
                                                    resid_t2=resid_t2,
                                                    cdresid_t2=cdresid_t2,
                                                    telresid_t2=telresid_t2,
                                                    email_t2=email_t2,
                                                    nombre_t3=nombre_t3,
                                                    cc_t3=cc_t3,
                                                    tel_t3=tel_t3,
                                                    cel_t3=cel_t3,
                                                    ofic_t3=ofic_t3,
                                                    cdof_t3=cdof_t3,
                                                    telof_t3=telof_t3,
                                                    resid_t3=resid_t3,
                                                    cdresid_t3=cdresid_t3,
                                                    telresid_t3=telresid_t3,
                                                    email_t3=email_t3,
                                                    nombre_t4=nombre_t4,
                                                    cc_t4=cc_t4,
                                                    tel_t4=tel_t4,
                                                    cel_t4=cel_t4,
                                                    ofic_t4=ofic_t4,
                                                    cdof_t4=cdof_t4,
                                                    telof_t4=telof_t4,
                                                    resid_t4=resid_t4,
                                                    cdresid_t4=cdresid_t4,
                                                    telresid_t4=telresid_t4,
                                                    email_t4=email_t4,
                                                    lote=lote[:-1],
                                                    manzana=manzana,
                                                    area=area,
                                                    mtsnorte=mtsnorte,
                                                    colnorte=colnorte,
                                                    mtseste=mtseste,
                                                    coleste=coleste,
                                                    mtssur=mtssur,
                                                    colsur=colsur,
                                                    mtsoeste=mtsoeste,
                                                    coloeste=coloeste,
                                                    valor=valor,
                                                    valor_letras=valor_letras,
                                                    ci=ci,
                                                    saldo=saldo,
                                                    contado_x=contado_x,
                                                    credic_x=credic_x,
                                                    amort_x=amort_x,
                                                    formaCI=formaCI,
                                                    formaFN=formaFN,
                                                    obs=obs,
                                                    dia_contrato=dia_contrato,
                                                    mes_contrato=mes_contrato,
                                                    año_contrato=año_contrato,
                                                    ruta=ruta,
                                                    parcelacion='1',
                                                    porcDerecho=str(porcDerecho),
                                                    fraccion=lote[-1],
                                                    meses_entrega=tiempo_entrega)
                        mensaje='Descarga el contrato aqui'
                        titulo='¡Listo!'
                        link=True
                    elif proyecto=='Sotavento':
                        pdf.ExportOpcionSotavento(nro_contrato=nro_contrato,
                                                    nombre_t1=nombre_t1,
                                                    cc_t1=cc_t1,
                                                    tel_t1=tel_t1,
                                                    cel_t1=cel_t1,
                                                    ofic_t1=ofic_t1,
                                                    cdof_t1=cdof_t1,
                                                    telof_t1=telof_t1,
                                                    resid_t1=resid_t1,
                                                    cdresid_t1=cdresid_t1,
                                                    telresid_t1=telresid_t1,
                                                    email_t1=email_t1,
                                                    nombre_t2=nombre_t2,
                                                    cc_t2=cc_t2,
                                                    tel_t2=tel_t2,
                                                    cel_t2=cel_t2,
                                                    ofic_t2=ofic_t2,
                                                    cdof_t2=cdof_t2,
                                                    telof_t2=telof_t2,
                                                    resid_t2=resid_t2,
                                                    cdresid_t2=cdresid_t2,
                                                    telresid_t2=telresid_t2,
                                                    email_t2=email_t2,
                                                    nombre_t3=nombre_t3,
                                                    cc_t3=cc_t3,
                                                    tel_t3=tel_t3,
                                                    cel_t3=cel_t3,
                                                    ofic_t3=ofic_t3,
                                                    cdof_t3=cdof_t3,
                                                    telof_t3=telof_t3,
                                                    resid_t3=resid_t3,
                                                    cdresid_t3=cdresid_t3,
                                                    telresid_t3=telresid_t3,
                                                    email_t3=email_t3,
                                                    nombre_t4=nombre_t4,
                                                    cc_t4=cc_t4,
                                                    tel_t4=tel_t4,
                                                    cel_t4=cel_t4,
                                                    ofic_t4=ofic_t4,
                                                    cdof_t4=cdof_t4,
                                                    telof_t4=telof_t4,
                                                    resid_t4=resid_t4,
                                                    cdresid_t4=cdresid_t4,
                                                    telresid_t4=telresid_t4,
                                                    email_t4=email_t4,
                                                    lote=lote,
                                                    manzana=manzana,
                                                    area=area,
                                                    mtsnorte=mtsnorte,
                                                    colnorte=colnorte,
                                                    mtseste=mtseste,
                                                    coleste=coleste,
                                                    mtssur=mtssur,
                                                    colsur=colsur,
                                                    mtsoeste=mtsoeste,
                                                    coloeste=coloeste,
                                                    valor=valor,
                                                    valor_letras=valor_letras,
                                                    ci=ci,
                                                    saldo=saldo,
                                                    contado_x=contado_x,
                                                    credic_x=credic_x,
                                                    amort_x=amort_x,
                                                    formaCI=formaCI,
                                                    formaFN=formaFN,
                                                    obs=obs,
                                                    meses_entrega=tiempo_entrega,
                                                    dia_contrato=dia_contrato,
                                                    mes_contrato=mes_contrato,
                                                    año_contrato=año_contrato,
                                                    ruta=ruta)
                        mensaje='Descarga el contrato aqui'
                        titulo='¡Listo!'
                        link=True
                    elif proyecto=='Sandville del Sol':
                        pdf.ExportOpcionTamarindos(nro_contrato="010",
                                                    nombre_t1=nombre_t1,
                                                    cc_t1=cc_t1,
                                                    tel_t1=tel_t1,
                                                    cel_t1=cel_t1,
                                                    ofic_t1=ofic_t1,
                                                    cdof_t1=cdof_t1,
                                                    telof_t1=telof_t1,
                                                    resid_t1=resid_t1,
                                                    cdresid_t1=cdresid_t1,
                                                    telresid_t1=telresid_t1,
                                                    email_t1=email_t1,
                                                    nombre_t2=nombre_t2,
                                                    cc_t2=cc_t2,
                                                    tel_t2=tel_t2,
                                                    cel_t2=cel_t2,
                                                    ofic_t2=ofic_t2,
                                                    cdof_t2=cdof_t2,
                                                    telof_t2=telof_t2,
                                                    resid_t2=resid_t2,
                                                    cdresid_t2=cdresid_t2,
                                                    telresid_t2=telresid_t2,
                                                    email_t2=email_t2,
                                                    nombre_t3=nombre_t3,
                                                    cc_t3=cc_t3,
                                                    tel_t3=tel_t3,
                                                    cel_t3=cel_t3,
                                                    ofic_t3=ofic_t3,
                                                    cdof_t3=cdof_t3,
                                                    telof_t3=telof_t3,
                                                    resid_t3=resid_t3,
                                                    cdresid_t3=cdresid_t3,
                                                    telresid_t3=telresid_t3,
                                                    email_t3=email_t3,
                                                    nombre_t4=nombre_t4,
                                                    cc_t4=cc_t4,
                                                    tel_t4=tel_t4,
                                                    cel_t4=cel_t4,
                                                    ofic_t4=ofic_t4,
                                                    cdof_t4=cdof_t4,
                                                    telof_t4=telof_t4,
                                                    resid_t4=resid_t4,
                                                    cdresid_t4=cdresid_t4,
                                                    telresid_t4=telresid_t4,
                                                    email_t4=email_t4,
                                                    lote=lote,
                                                    manzana=manzana,
                                                    area=area,
                                                    mtsnorte=mtsnorte,
                                                    colnorte=colnorte,
                                                    mtseste=mtseste,
                                                    coleste=coleste,
                                                    mtssur=mtssur,
                                                    colsur=colsur,
                                                    mtsoeste=mtsoeste,
                                                    coloeste=coloeste,
                                                    valor=valor,
                                                    valor_letras=valor_letras,
                                                    ci=ci,
                                                    saldo=saldo,
                                                    contado_x=contado_x,
                                                    credic_x=credic_x,
                                                    amort_x=amort_x,
                                                    formaCI=formaCI,
                                                    formaFN=formaFN,
                                                    obs=obs,
                                                    dia_contrato=dia_contrato,
                                                    mes_contrato=mes_contrato,
                                                    año_contrato=año_contrato,
                                                    ruta=ruta,
                                                    parcelacion='3',
                                                    porcDerecho='100',
                                                    fraccion='-')
                    else:
                        mensaje='El proyecto seleccionado no tiene formato de opcion de promesa asignado'
                        titulo='Error'
                    alerta=True
                    ruta_link=settings.DIR_DOWNLOADS+f'{proyecto}_contrato_{contrato}.pdf'
                    parameter=obj_parametro.get(descripcion='formasOpcionManual')
                    parameter.estado=False
                    parameter.save()
                if request.POST.get('impPagare'):
                    filename = f'{proyecto}_pagare_{contrato}.pdf'
                    ruta=settings.DIR_EXPORT+filename
                    if proyecto=='Oasis':
                        context_pagare = {
                            'proyecto': proyecto,
                            'ctr': obj_ctr,
                        }
                        pdf_file = pdf_gen_weasy(f'pdf/{proyecto}/pagare.html', context_pagare, filename)
                        ruta_link = pdf_file.get('url')
                    elif proyecto=='Vegas de Venecia':
                        pdf.PagareQuadrata(nroPagare=contrato,
                                        nombreT1=nombre_t1,ccT1=cc_t1,nombreT2=nombre_t2,ccT2=cc_t2,
                                        nombreT3=nombre_t3,ccT3=cc_t3,nombreT4=nombre_t4,ccT4=cc_t4,
                                        diaPagare=dia_contrato,mesPagare=str(mes_contrato),añoPagare=año_contrato,ciudad='Medellin',ruta=ruta)
                    elif proyecto=='Tesoro Escondido':
                        pdf.PagareTesoro(nroPagare=contrato,
                                        nombreT1=nombre_t1,ccT1=cc_t1,nombreT2=nombre_t2,ccT2=cc_t2,
                                        nombreT3=nombre_t3,ccT3=cc_t3,nombreT4=nombre_t4,ccT4=cc_t4,
                                        diaPagare=dia_contrato,mesPagare=str(mes_contrato),añoPagare=año_contrato,ciudad='Medellin',ruta=ruta)
                    
                    elif proyecto=='Carmelo Reservado':
                        pdf.PagareCarmelo(nroPagare=contrato,
                                        nombreT1=nombre_t1,ccT1=cc_t1,nombreT2=nombre_t2,ccT2=cc_t2,
                                        nombreT3=nombre_t3,ccT3=cc_t3,nombreT4=nombre_t4,ccT4=cc_t4,
                                        diaPagare=dia_contrato,mesPagare=str(mes_contrato),añoPagare=año_contrato,ciudad='Medellin',ruta=ruta)
                    elif proyecto=='Casas de Verano':
                        pdf.PagareCasasdeVerano(nroPagare=contrato,
                                        nombreT1=nombre_t1,ccT1=cc_t1,nombreT2=nombre_t2,ccT2=cc_t2,
                                        nombreT3=nombre_t3,ccT3=cc_t3,nombreT4=nombre_t4,ccT4=cc_t4,
                                        diaPagare=dia_contrato,mesPagare=str(mes_contrato),añoPagare=año_contrato,ciudad='Medellin',ruta=ruta)
                    
                    elif proyecto=='Sandville Beach':
                        pdf.PagareSandvilleBeach(nroPagare=contrato,
                                        nombreT1=nombre_t1,ccT1=cc_t1,nombreT2=nombre_t2,ccT2=cc_t2,
                                        nombreT3=nombre_t3,ccT3=cc_t3,nombreT4=nombre_t4,ccT4=cc_t4,
                                        diaPagare=dia_contrato,mesPagare=str(mes_contrato),añoPagare=año_contrato,ciudad='Medellin',ruta=ruta)
                    elif proyecto=='Perla del Mar' or proyecto=='Sandville del Sol':
                        pdf.PagareSandvilleMar(nroPagare=contrato,
                                        nombreT1=nombre_t1,ccT1=cc_t1,nombreT2=nombre_t2,ccT2=cc_t2,
                                        nombreT3=nombre_t3,ccT3=cc_t3,nombreT4=nombre_t4,ccT4=cc_t4,
                                        diaPagare=dia_contrato,mesPagare=str(mes_contrato),añoPagare=año_contrato,ciudad='Medellin',ruta=ruta)
                    elif proyecto=='Sotavento':
                        pdf.PagareSotavento(nroPagare=contrato,
                                        nombreT1=nombre_t1,ccT1=cc_t1,nombreT2=nombre_t2,ccT2=cc_t2,
                                        nombreT3=nombre_t3,ccT3=cc_t3,nombreT4=nombre_t4,ccT4=cc_t4,
                                        diaPagare=dia_contrato,mesPagare=str(mes_contrato),añoPagare=año_contrato,ciudad='Monteria',ruta=ruta)
                    alerta=True
                    mensaje='Descarga el Pagaré aqui'
                    titulo='¡Listo!'
                    link=True
                    if proyecto!='Oasis':
                        ruta_link=settings.DIR_DOWNLOADS+filename
                if request.POST.get('impVerificacion'):
                    ruta=settings.DIR_EXPORT+f'{proyecto}_verificacion_{contrato}.pdf'
                    parametro=obj_parametro.get(descripcion='formasVerificacionManual')
                    if parametro.estado:
                        formaCI = request.POST.get('formaci')
                        formaFN = request.POST.get('formasaldo')
                    if proyecto=='Vegas de Venecia':
                        pdf.VerificacionQuadrata(ruta=ruta,nombreT1=nombre_t1,nombreT2=nombre_t2,nombreT3=nombre_t3,nombreT4=nombre_t4,
                                                ccTitular1=cc_t1,ccTitular2=cc_t2,ccTitular3=cc_t3,ccTitular4=cc_t4,lote=lote,manzana=manzana,
                                                area=area,valor=str(valor),ci=str(ci),formaci=formaCI,saldo=str(saldo),formasaldo=formaFN,fechaEntrega=str(fecha_entrega.date()),
                                                fechaEscritura='',diactr=dia_contrato,mesctr=str(mes_contrato),anioctr=año_contrato,ciudad='Medellin',meses_entrega=tiempo_entrega
                                                )
                        alerta=True
                        mensaje='Descarga la verificacion aqui'
                        titulo='¡Listo!'
                        link=True
                        ruta_link=settings.DIR_DOWNLOADS+f'{proyecto}_verificacion_{contrato}.pdf'
                    elif proyecto=='Tesoro Escondido':
                        pdf.VerificacionTerranova(ruta=ruta,nombreT1=nombre_t1,nombreT2=nombre_t2,nombreT3=nombre_t3,nombreT4=nombre_t4,
                                                ccTitular1=cc_t1,ccTitular2=cc_t2,ccTitular3=cc_t3,ccTitular4=cc_t4,lote=lote,manzana=manzana,
                                                area=area,valor=str(valor),ci=str(ci),formaci=formaCI,saldo=str(saldo),formasaldo=formaFN,fechaEntrega=f'{tiempo_entrega} meses despues de la firma de la opcion/promesa',
                                                fechaEscritura=''
                                                )
                        alerta=True
                        mensaje='Descarga la verificacion aqui'
                        titulo='¡Listo!'
                        link=True
                        ruta_link=settings.DIR_DOWNLOADS+f'{proyecto}_verificacion_{contrato}.pdf'
                    elif proyecto=='Sotavento':
                        pdf.VerificacionSotavento(ruta=ruta,nombreT1=nombre_t1,nombreT2=nombre_t2,nombreT3=nombre_t3,nombreT4=nombre_t4,
                                                ccTitular1=cc_t1,ccTitular2=cc_t2,ccTitular3=cc_t3,ccTitular4=cc_t4,lote=lote,manzana=manzana,
                                                area=area,valor=str(valor),ci=str(ci),formaci=formaCI,saldo=str(saldo),formasaldo=formaFN,fechaEntrega=f'{tiempo_entrega} meses despues de la firma de la opcion/promesa',
                                                fechaEscritura=''
                                                )
                        alerta=True
                        mensaje='Descarga la verificacion aqui'
                        titulo='¡Listo!'
                        link=True
                        ruta_link=settings.DIR_DOWNLOADS+f'{proyecto}_verificacion_{contrato}.pdf'
                    elif proyecto=='Perla del Mar':
                        context = {
                            'proyecto':proyecto,
                            'ctr':obj_ctr,
                            'meses_entrega':24,
                            'oficina':'Medellín'
                        }
                        
                        filename = f'Verificacion_contrato_{contrato}_{proyecto}.pdf'
                        
                            
                        pdf = pdf_gen(f'pdf/{proyecto}/verificacion.html',context,filename)
                        
                        file = pdf.get('root')
                                                
                        return FileResponse(open(file,'rb'),as_attachment=True,filename=filename)
                    
                    elif proyecto=='Carmelo Reservado':
                        
                        context = {
                            'proyecto':proyecto,
                            'ctr':obj_ctr,
                            'meses_entrega':36,
                            'oficina':'Medellín'
                        }
                        
                        filename = f'Verificacion_contrato_{contrato}_{proyecto}.pdf'
                        
                            
                        pdf = pdf_gen(f'pdf/{proyecto}/verificacion.html',context,filename)
                        
                        file = pdf.get('root')
                                                
                        return FileResponse(open(file,'rb'),as_attachment=True,filename=filename)
                    elif proyecto=='Casas de Verano':
                        
                        context = {
                            'proyecto':proyecto,
                            'ctr':obj_ctr,
                            'meses_entrega':36,
                            'oficina':'Medellín'
                        }
                        
                        filename = f'Verificacion_contrato_{contrato}_{proyecto}.pdf'
                        
                            
                        pdf = pdf_gen(f'pdf/{proyecto}/verificacion.html',context,filename)
                        
                        file = pdf.get('root')
                                                
                        return FileResponse(open(file,'rb'),as_attachment=True,filename=filename)
                    
                    else:
                        alerta=True
                        mensaje='Este proyecto no tiene formato de verificacion asociado'
                        titulo='¡Error!'
                    parameter=obj_parametro.get(descripcion='formasVerificacionManual')
                    parameter.estado=False
                    parameter.save()
                    parameter=obj_parametro.get(descripcion='entregaVerificacionManual')
                    parameter.estado=False
                    parameter.save()
                if request.POST.get('impRecaudo'):
                    form_rec=form_recaudo_noradicado(request.POST,request.FILES)
                    if form_rec.is_valid():
                        today=datetime.date.today()
                        obj_consecutivo=consecutivos.objects.using(proyecto).get(documento='RC')
                        concepto=form_rec.cleaned_data.get('Concepto')
                        valor=form_rec.cleaned_data.get('Valor')
                        formapago=form_rec.cleaned_data.get('formapago')
                        soporte = form_rec.cleaned_data.get('soporte')
                        name_doc = f'soporte_rc_{obj_consecutivo.consecutivo}_{proyecto}_CTR{contrato}'
                        typedoc = soporte.name.split('.')[-1].lower()
                        file_dir = f'{settings.MEDIA_ROOT}/soportes_recibos/ventas_nuevas/{proyecto}/'
                        file_key = _to_storage_key(f'{file_dir}{name_doc}.{typedoc}')
                        upload_docs(soporte,file_dir,name_doc,typedoc)
                        RecaudosNoradicados.objects.using(proyecto).create(recibo=obj_consecutivo.consecutivo,
                                                                        contrato=contrato,
                                                                        fecha=today,
                                                                        concepto=concepto,
                                                                        valor=valor,
                                                                        formapago=formapago,
                                                                        usuario=request.user.username,
                                                                        soportepago=file_key)
                        nrorecibo = obj_consecutivo.consecutivo
                        obj_consecutivo.consecutivo+=1
                        obj_consecutivo.save()
                        filename=f'{proyecto}_reciboNR_{nrorecibo}.pdf'
                        if proyecto == 'Oasis':
                            datos_recaudo = RecaudosNoradicados.objects.using(proyecto).get(recibo=nrorecibo)
                            result = pdf_gen_weasy(
                                f'pdf/{proyecto}/recibo_nr.html',
                                {'recibo': datos_recaudo},
                                filename,
                            )
                            ruta_link = result['url']
                        else:
                            ruta=settings.DIR_EXPORT+filename
                            pdf.Recibo_caja(proyecto=proyecto,
                                            ruta=ruta,
                                            nroRecibo=nrorecibo,
                                            titular1=nombre_t1,
                                            fecha=str(datetime.date.today()),
                                            concepto=concepto,
                                            valor=valor,
                                            direccion=resid_t1,
                                            ciudad=cdresid_t1,
                                            telefono=cel_t1,
                                            formapag=formapago,
                                            user=request.user)
                            ruta_link=settings.DIR_DOWNLOADS+filename
                        alerta=True
                        mensaje='Descarga el Recibo aqui'
                        titulo='¡Listo!'
                        link=True     
                if request.POST.get('impTerminosAlttum'):
                    beneficiarios = request.POST.get('beneficiarios')
                    ruta=settings.DIR_EXPORT+f'{proyecto}_Terminos y condiciones alttum_contrato_{contrato}.pdf'
                    pdf.terminosAlttum(
                        proyecto=proyecto,
                        beneficiarios=beneficiarios,
                        cliente=nombre_t1,
                        cc_cliente=cc_t1,
                        email=email_t1,
                        ocupacion=datos_t1.ocupacion,
                        telefono=datos_t1.telefono1,
                        direccion=resid_t1,
                        cel=cel_t1,
                        fecha_contrato=datos_venta.fecha_contrato,
                        ruta=ruta
                    )
                    alerta=True
                    mensaje='Descarga los terminos y condiciones aqui'
                    titulo='¡Listo!'
                    link=True
                    ruta_link=settings.DIR_DOWNLOADS+f'{proyecto}_Terminos y condiciones alttum_contrato_{contrato}.pdf'
        obj_parametro=Parametros_Operaciones.objects.using(proyecto)
        jsonparametros=serializers.serialize('json',obj_parametro.all())
        
        context['parameters']=jsonparametros
        context['adj']=contrato
        context['alerta']=alerta
        context['mensaje']=mensaje
        context['titulo_alerta']=titulo
        context['redireccion']=redireccion
        context['redirect']=dir_redirect
        context['link']=link
        context['ruta_link']=ruta_link
        
        context['adj']=contrato
        
        return render(request,'acciones_venta.html',context)
    else:
        return redirect('/accounts/login')

@group_perm_required(perms=('andinasoft.add_adjudicacion',),raise_exception=True)
def adjudicar_venta(request,proyecto,contrato):
    check_project(request,proyecto)
    alerta=False
    titulo=None
    mensaje=None
    redireccion=False
    dir_redirect=None
    lista_clientes=clientes.objects.using('default').all()
    datos_venta=ventas_nuevas.objects.using(proyecto).get(id_venta=contrato)
    datos_inmueble=Inmuebles.objects.using(proyecto).get(idinmueble=datos_venta.inmueble)
    lista_asesores=asesores.objects.using('default').filter(estado='Activo')
    escala=AsignacionComisiones.objects.using(proyecto)
    recibos_generados=RecaudosNoradicados.objects.using(proyecto).filter(contrato=contrato)
    try:
        generador=escala.filter(idadjudicacion=contrato,idcargo=8).values()
        generador=generador[0]
        nombre_asesor=asesores.objects.using('default').get(cedula=generador['idgestor'])
        generador['nombre']=nombre_asesor.nombre
    except: generador=[]
    try: 
        linea=escala.filter(idadjudicacion=contrato,idcargo=9).values()
        linea=linea[0]
        nombre_asesor=asesores.objects.using('default').get(cedula=linea['idgestor'])
        linea['nombre']=nombre_asesor.nombre
    except: linea=[]
    try: 
        cerrador=escala.filter(idadjudicacion=contrato,idcargo=99).values()
        cerrador=cerrador[0]
        nombre_asesor=asesores.objects.using('default').get(cedula=cerrador['idgestor'])
        cerrador['nombre']=nombre_asesor.nombre
    except: cerrador=[]
    try: 
        gerenteventas=escala.filter(idadjudicacion=contrato,idcargo=14).values()
        gerenteventas=gerenteventas[0]
        nombre_asesor=asesores.objects.using('default').get(cedula=gerenteventas['idgestor'])
        gerenteventas['nombre']=nombre_asesor.nombre
    except: gerenteventas=[]
    try: 
        jefep=escala.filter(idadjudicacion=contrato,idcargo=15).values()
        jefep=jefep[0]
        nombre_asesor=asesores.objects.using('default').get(cedula=jefep['idgestor'])
        jefep['nombre']=nombre_asesor.nombre
    except : jefep=[]
    try: 
        jefev=escala.filter(idadjudicacion=contrato,idcargo=13).values()
        jefev=jefev[0]
        nombre_asesor=asesores.objects.using('default').get(cedula=jefev['idgestor'])
        jefev['nombre']=nombre_asesor.nombre
    except: jefev=[]
    try: datos_t1=clientes.objects.using('default').get(idTercero=datos_venta.id_t1)
    except: datos_t1=[]
    try: datos_t2=clientes.objects.using('default').get(idTercero=datos_venta.id_t2)
    except: datos_t2=[]
    try: datos_t3=clientes.objects.using('default').get(idTercero=datos_venta.id_t3)
    except: datos_t3=[]
    try: datos_t4=clientes.objects.using('default').get(idTercero=datos_venta.id_t4)
    except: datos_t4=[]
    tasa=1.5/100
    nro_ci=0
    if datos_venta.estado=='Adjudicado':
        alerta=True
        titulo='Error'
        mensaje='Este contrato ya fue adjudicado'
        redireccion=True
        dir_redirect=f'/operaciones/por_adjudicar/{proyecto}'
    obj_docs=documentos_contratos.objects.using(proyecto).filter(adj=contrato)
    context={
        'recibos_generados':recibos_generados,
        'form_docs':form_docs_contratos,
        'form_recaudo':form_recaudo_noradicado,
        'form_op':form_revision_op,
        'form':form_nueva_venta,
        'lista_documentos':obj_docs,
        'asesores':lista_asesores,
        'clientes':lista_clientes,
        'proyecto':proyecto,
        'contrato':contrato,
        'datos_venta':datos_venta,
        'datos_t1':datos_t1,
        'datos_t2':datos_t2,
        'datos_t3':datos_t3,
        'datos_t4':datos_t4,
        'tasa':tasas,
        'datos_inmueble':datos_inmueble,
        'btnName':'Modificar',
        'form_escala':form_escala_comision,
        'generador':generador,
        'linea':linea,
        'cerrador':cerrador,
        'tlmk':gerenteventas,
        'jefep':jefep,
        'jefev':jefev,
        'tipo_form':'adjudicacion',
        'adj':contrato
    }
    if request.method == 'POST':
        if request.POST.get('btnDesaprobar'):
            obj_venta=ventas_nuevas.objects.using(proyecto).get(pk=contrato)
            obj_venta.estado='Pendiente'
            obj_venta.save()
            alerta=True
            titulo='Listo'
            mensaje=f'El contrato {contrato} fue desaprobado'
            redireccion=True
            dir_redirect=f'/comercial/ventas_sin_aprobar/{proyecto}'
        form=form_revision_op(request.POST)
        if form.is_valid():
            ob_adj=Adjudicacion.objects.using(proyecto)
            obj_consecutivo=consecutivos.objects.using(proyecto).get(documento='ADJ')
            idadj=obj_consecutivo.consecutivo
            tipocontrato=form.cleaned_data.get('Tipo_Contrato')
            origen=form.cleaned_data.get('Origen_Venta')
            descuento=form.cleaned_data.get('valor_dcto')
            oficina=form.cleaned_data.get('Oficina')
            p_enmendadura=form.cleaned_data.get('Enmendaduras')
            p_doccompleta=form.cleaned_data.get('Documentacion_Incompleta')
            p_valincorrectos=form.cleaned_data.get('Valores_Incorrectos')
            p_obs=form.cleaned_data.get('Observaciones')
            fecha_condicion=form.cleaned_data.get('fecha_dcto')
            valor_condicion=form.cleaned_data.get('valor_pagodcto')
            nuevo_ci=form.cleaned_data.get('nuevo_ci')
            nuevo_saldo=form.cleaned_data.get('nuevo_saldo')
            cartera_asignar=form.cleaned_data.get('Cartera_Asignar')
            
            if descuento==None: descuento=0
            base_comision=datos_venta.valor_venta-descuento
            
            ADJ = ob_adj.create(fecha=datetime.date.today(),
                          idadjudicacion=f'ADJ{idadj}',
                          tipocontrato=tipocontrato,
                          contrato=contrato,
                          idinmueble=datos_venta.inmueble,
                          idtercero1=datos_venta.id_t1,
                          idtercero2=datos_venta.id_t2,
                          idtercero3=datos_venta.id_t3,
                          idtercero4=datos_venta.id_t4,
                          valor=datos_venta.valor_venta,
                          formapago=datos_venta.forma_pago,
                          cuotainicial=datos_venta.cuota_inicial,
                          financiacion=datos_venta.saldo,
                          plazofnc=datos_venta.nro_cuotas_fn,
                          cuotafnc=datos_venta.valor_ctas_fn,
                          iniciofnc=datos_venta.inicio_fn,
                          plazoextra=datos_venta.nro_cuotas_ce,
                          inicioextra=datos_venta.inicio_ce,
                          cuotaextra=datos_venta.valor_ctas_ce,
                          fechacontrato=datos_venta.fecha_contrato,
                          estado='Aprobado',
                          origenventa=origen,
                          basecomision=base_comision,
                          oficina=oficina,
                          usuario=request.user,
                          p_enmendadura=p_enmendadura,
                          p_doccompleta=p_doccompleta,
                          p_valincorrectos=p_valincorrectos,
                          p_obs=p_obs,
                          tasafnc = datos_venta.tasa
                          )
            obj_consecutivo.consecutivo=idadj+1
            obj_consecutivo.save()
            datos_venta.estado='Adjudicado'
            datos_venta.adj = ADJ
            datos_venta.save()
            obj_inmueble=Inmuebles.objects.using(proyecto).get(idinmueble=datos_venta.inmueble)
            if proyecto != 'Fractal':
                obj_inmueble.estado='Adjudicado'
                obj_inmueble.save() 
            if fecha_condicion!='' and fecha_condicion!=None:
                obj_descuentos=DescuentosCondicionados.objects.using(proyecto)
                obj_descuentos.create(idadjudicacion=f'ADJ{idadj}',
                                      fecha_condicion=fecha_condicion,
                                      valor_condicion=valor_condicion,
                                      valor_dcto=descuento,
                                      nueva_ci=nuevo_ci,
                                      nuevo_saldo=nuevo_saldo,
                                      estado='Pendiente')
          #cambiar nombre de carpeta si existe
            obj_docs=documentos_contratos.objects.using(proyecto).filter(adj=contrato)
            for doc in obj_docs:
                doc.adj=f'ADJ{idadj}'
                doc.save()
            try:
                src_prefix = _to_storage_key(f'{settings.DIR_DOCS}/doc_contratos/{proyecto}/{contrato}')
                dst_prefix = _to_storage_key(f'{settings.DIR_DOCS}/doc_contratos/{proyecto}/ADJ{idadj}')
                _move_storage_prefix(src_prefix, dst_prefix)
            except Exception:
                traceback.print_exc()
            #creacion plan de pagos
           #ci
            obj_planpagos=PlanPagos.objects.using(proyecto)
            
            count_ci=1
            cantidad=datos_venta.cant_ci1
            fecha=datos_venta.fecha_ci1
            valor=datos_venta.valor_ci1
            if cantidad is not None:
                for j in range(0,cantidad):
                    fecha_ci=fecha+relativedelta(months=j)
                    idcta=f'CI{count_ci}ADJ{idadj}'
                    obj_planpagos.create(idcta=idcta,
                                        tipocta='CI',
                                        nrocta=count_ci,
                                        adj=f'ADJ{idadj}',
                                        capital=valor,
                                        intcte=0,
                                        cuota=valor,
                                        fecha=fecha_ci)
                    count_ci+=1
            cantidad=datos_venta.cant_ci2
            fecha=datos_venta.fecha_ci2
            valor=datos_venta.valor_ci2
            if cantidad is not None:
                for j in range(0,cantidad):
                    fecha_ci=fecha+relativedelta(months=j)
                    idcta=f'CI{count_ci}ADJ{idadj}'
                    obj_planpagos.create(idcta=idcta,
                                        tipocta='CI',
                                        nrocta=count_ci,
                                        adj=f'ADJ{idadj}',
                                        capital=valor,
                                        intcte=0,
                                        cuota=valor,
                                        fecha=fecha_ci)
                    count_ci+=1
            cantidad=datos_venta.cant_ci3
            fecha=datos_venta.fecha_ci3
            valor=datos_venta.valor_ci3
            if cantidad is not None:
                for j in range(0,cantidad):
                    fecha_ci=fecha+relativedelta(months=j)
                    idcta=f'CI{count_ci}ADJ{idadj}'
                    obj_planpagos.create(idcta=idcta,
                                        tipocta='CI',
                                        nrocta=count_ci,
                                        adj=f'ADJ{idadj}',
                                        capital=valor,
                                        intcte=0,
                                        cuota=valor,
                                        fecha=fecha_ci)
                    count_ci+=1
            cantidad=datos_venta.cant_ci4
            fecha=datos_venta.fecha_ci4
            valor=datos_venta.valor_ci4
            if cantidad is not None:
                for j in range(0,cantidad):
                    fecha_ci=fecha+relativedelta(months=j)
                    idcta=f'CI{count_ci}ADJ{idadj}'
                    obj_planpagos.create(idcta=idcta,
                                        tipocta='CI',
                                        nrocta=count_ci,
                                        adj=f'ADJ{idadj}',
                                        capital=valor,
                                        intcte=0,
                                        cuota=valor,
                                        fecha=fecha_ci)
                    count_ci+=1
            cantidad=datos_venta.cant_ci5
            fecha=datos_venta.fecha_ci5
            valor=datos_venta.valor_ci5
            if cantidad is not None:
                for j in range(0,cantidad):
                    fecha_ci=fecha+relativedelta(months=j)
                    idcta=f'CI{count_ci}ADJ{idadj}'
                    obj_planpagos.create(idcta=idcta,
                                        tipocta='CI',
                                        nrocta=count_ci,
                                        adj=f'ADJ{idadj}',
                                        capital=valor,
                                        intcte=0,
                                        cuota=valor,
                                        fecha=fecha_ci)
                    count_ci+=1
            cantidad=datos_venta.cant_ci6
            fecha=datos_venta.fecha_ci6
            valor=datos_venta.valor_ci6
            if cantidad is not None:
                for j in range(0,cantidad):
                    fecha_ci=fecha+relativedelta(months=j)
                    idcta=f'CI{count_ci}ADJ{idadj}'
                    obj_planpagos.create(idcta=idcta,
                                        tipocta='CI',
                                        nrocta=count_ci,
                                        adj=f'ADJ{idadj}',
                                        capital=valor,
                                        intcte=0,
                                        cuota=valor,
                                        fecha=fecha_ci)
                    count_ci+=1
            cantidad=datos_venta.cant_ci7
            fecha=datos_venta.fecha_ci7
            valor=datos_venta.valor_ci7
            if cantidad is not None:
                for j in range(0,cantidad):
                    fecha_ci=fecha+relativedelta(months=j)
                    idcta=f'CI{count_ci}ADJ{idadj}'
                    obj_planpagos.create(idcta=idcta,
                                        tipocta='CI',
                                        nrocta=count_ci,
                                        adj=f'ADJ{idadj}',
                                        capital=valor,
                                        intcte=0,
                                        cuota=valor,
                                        fecha=fecha_ci)
                    count_ci+=1
           #saldo
            cantidad=datos_venta.nro_cuotas_fn
            fecha=datos_venta.inicio_fn
            valor=datos_venta.valor_ctas_fn
            tasa = datos_venta.tasa
            if datos_venta.forma_saldo=='CONTADO':
                tipocta='CO'
            else: tipocta='FN'
            if cantidad is not None:
                if datos_venta.forma_saldo=='Regular':
                    valor_presente=datos_venta.saldo
                else:
                    valor_presente=Utilidades().CalcularVP(valor,tasa,cantidad)
                capital_fn=valor_presente
                for i in range (0,cantidad):
                    fecha_fn=fecha+relativedelta(months=i)
                    idcta=f'{tipocta}{i+1}ADJ{idadj}'
                    interes=round(tasa*valor_presente,0)
                    capital=valor-interes
                    if i==cantidad-1:
                        capital=valor_presente
                        valor=capital+interes
                    valor_presente-=capital
                    
                    obj_planpagos.create(idcta=idcta,
                                        tipocta=tipocta,
                                        nrocta=i+1,
                                        adj=f'ADJ{idadj}',
                                        capital=capital,
                                        intcte=interes,
                                        cuota=valor,
                                        fecha=fecha_fn)
            cantidad=datos_venta.nro_cuotas_ce
            fecha=datos_venta.inicio_ce
            valor=datos_venta.valor_ctas_ce
            if cantidad is not None:
                periodo=datos_venta.period_ce
                periodos={
                        'Mensual':1,
                        'Trimestral':3,
                        'Semestral':6,
                        'Anual':12
                    }
                valor_presente=datos_venta.saldo - capital_fn
                for i in range (0,cantidad):
                    try: fecha_ce=fecha+relativedelta(months=i*periodos[periodo])
                    except: fecha_ce=fecha+relativedelta(months=i*int(periodo))
                    idcta=f'CE{i+1}ADJ{idadj}'
                    try: interes=round(tasa*periodos[periodo]*valor_presente,0)
                    except: interes=round(tasa*int(periodo)*valor_presente,0)
                    capital=valor-interes
                    if i==cantidad-1:
                        capital=valor_presente
                        valor=capital+interes
                    valor_presente-=capital
                    obj_planpagos.create(idcta=idcta,
                                        tipocta='CE',
                                        nrocta=i+1,
                                        adj=f'ADJ{idadj}',
                                        capital=capital,
                                        intcte=interes,
                                        cuota=valor,
                                        fecha=fecha_ce)
           #creacion timeline
            obj_timeline=timeline.objects.using(proyecto)
            obj_timeline.create(adj=f'ADJ{idadj}',
                               fecha=datetime.date.today(),
                               usuario=request.user,
                               accion='Adjudicó el Contrato')
            obj_timeline.create(adj=f'ADJ{idadj}',
                               fecha=datos_venta.fecha_contrato,
                               usuario=datos_venta.usuario,
                               accion='Creó el Contrato')
            obj_timeline.create(adj=f'ADJ{idadj}',
                               fecha=datos_venta.fecha_aprueba,
                               usuario=datos_venta.usuarioaprueba,
                               accion='Aprobó el Contrato')
           #modificacion escala a ADJ
            obj_comisiones=AsignacionComisiones.objects.using(proyecto)
            cargos=obj_comisiones.filter(idadjudicacion=datos_venta.id_venta)
            
            for cargo in cargos:
                cargo.id_comision=f'{cargo.idcargo}ADJ{idadj}'
                cargo.idadjudicacion=f'ADJ{idadj}'
                cargo.save()
           #Radicacion de recibo 
            adj=f'ADJ{idadj}'
            obj_recNR=RecaudosNoradicados.objects.using(proyecto).filter(contrato=datos_venta.id_venta)
            titulares=titulares_por_adj.objects.using(proyecto).get(adj=adj)
            proyecto_obj = proyectos.objects.get(pk=proyecto)
            for recibo in obj_recNR:
                saldo_cuotas=saldos_adj.objects.using(proyecto).filter(adj=adj,saldocuota__gt=0)
                aplicar_pago(request=request,adj=adj,fecha=recibo.fecha,forma_pago=recibo.formapago,
                             valor_pagado=recibo.valor,concepto=recibo.concepto,valor_recibo=recibo.valor,
                             porcentaje_condonado=100,saldo_cuotas=saldo_cuotas,consecutivo=recibo.recibo,
                             Recaudos=Recaudos,Recaudos_general=Recaudos_general,titulares=titulares,
                             proyecto=proyecto)
                soporte_rel = _normalize_soporte_key(recibo.soportepago)
                if soporte_rel:
                    fecha_pago_nr = recibo.fecha
                    if isinstance(fecha_pago_nr, datetime.date):
                        fecha_pago_dt = fecha_pago_nr
                    else:
                        fecha_pago_dt = parse_date(str(fecha_pago_nr))
                        if fecha_pago_dt is None:
                            try:
                                fecha_pago_dt = datetime.datetime.strptime(str(fecha_pago_nr),'%Y-%m-%d').date()
                            except ValueError:
                                fecha_pago_dt = datetime.date.today()
                    solicitud = recibos_internos.objects.using('default').filter(
                        recibo_asociado=recibo.recibo,
                        proyecto=proyecto_obj
                    ).first()
                    if not solicitud:
                        solicitud = recibos_internos.objects.using('default').create(
                            proyecto=proyecto_obj,
                            fecha_pago=fecha_pago_dt,
                            valor=recibo.valor,
                            soporte=soporte_rel,
                            cliente=adj,
                            usuario_solicita=request.user,
                            condonacion=0,
                            abono_capital=False
                        )
                    else:
                        if not solicitud.soporte:
                            solicitud.soporte = soporte_rel
                    solicitud.recibo_asociado = recibo.recibo
                    solicitud.usuario_confirma = request.user
                    solicitud.fecha_confirma = datetime.date.today()
                    solicitud.save()
                recibo.delete()
            # Registra el la info de cartera
            
            InfoCartera.objects.using(proyecto).create(idadjudicacion=adj,
                                                       gestorasignado=cartera_asignar)
            alerta=True
            titulo='¡Todo fue un exito!'
            mensaje=f'El contrato {contrato} fue adjudicado exitosamente con el {adj}'
            redireccion=True
            dir_redirect=f'/adjudicaciones/{proyecto}/{adj}'
        
        form_docs=form_docs_contratos(request.POST,request.FILES)
        if form_docs.is_valid():
            descrip_doc=f"{form_docs.cleaned_data.get('tipo_doc')}_{datetime.date.today()}"
            fecha_carga=datetime.datetime.today()
            usuario_carga=request.user
            upload_docs_contratos(request.FILES['documento_cargar'],contrato,proyecto,descrip_doc)
            documentos_contratos.objects.using(proyecto).create(adj=contrato,descripcion_doc=descrip_doc,
                                                        fecha_carga=fecha_carga,usuario_carga=usuario_carga)
            obj_docs=documentos_contratos.objects.using(proyecto).filter(adj=contrato)
            context['lista_documentos']=obj_docs
    context['alerta']=alerta
    context['titulo_alerta']=titulo       
    context['mensaje']=mensaje
    context['redireccion']=redireccion
    context['redirect']=dir_redirect
    
    return render(request,'crear_adjudicacion.html',context)

@login_required
@group_perm_required(perms=('andinasoft.add_ventas_nuevas',),raise_exception=True)
def inventario_comercial(request,proyecto):
    check_project(request,proyecto)
    inventario=Inmuebles.objects.using(proyecto).filter(estado='Libre')
    if proyecto.lower() == 'fractal':
        context={
            'proyecto':proyecto,
            'lotes':inventario
        }
        return render(request, 'inventario_fractal.html',context)
    
    lotes=[]
    for lote in inventario:
        if lote.fac_valor_esquinero>1:
            esquinero='Si'
        else: esquinero='No'
        casa=0
            
        incrementos=lote.fac_valor_via_principal*lote.fac_valor_area_social*lote.fac_valor_esquinero
        valor_lote=(lote.areaprivada*lote.vrmetrocuadrado*incrementos)+casa
        valor_lote=Utilidades().redondear_numero(numero=valor_lote,multiplo=1000000,redondeo='>')
        lotes.append((lote.idinmueble,lote.manzananumero,lote.lotenumero,lote.areaprivada,
                    esquinero,valor_lote))
    context={
        'proyecto':proyecto,
        'lotes':lotes
    }
    
    
    
    return render(request,'inventario_comercial.html',context)

@group_perm_required(perms=('andinasoft.view_recaudosnoradicados',),raise_exception=True)
def recaudos_noradicados(request,proyecto):
    if request.method == 'POST':
        formapago=request.POST.get('forma_pago')
        nro_recibo=request.POST.get('nro-recibo')
        if request.POST.get('radicarRecibo'):
            obj_recaudo=RecaudosNoradicados.objects.using(proyecto).get(recibo=nro_recibo)
            form=form_radicar_recibo(request.POST,proyecto=proyecto)
            if form.is_valid():
                check_perms(request,perms=('andinasoft.change_recaudosnoradicados',))
                obj_recaudo.formapago=str(formapago)
                obj_recaudo.save()
        if request.POST.get('AnularRecibo'):
            obj_recaudo=RecaudosNoradicados.objects.using(proyecto).get(recibo=nro_recibo)
            check_perms(request,perms=('andinasoft.delete_recaudosnoradicados',))
            Recaudos_general.objects.using(proyecto).create(fecha=obj_recaudo.fecha,idadjudicacion='ADJ0',formapago='',valor=0,
            usuario=request.user,numrecibo=nro_recibo,idtercero='',operacion='Anulado',concepto='Anulado')
            obj_recaudo.delete()
        if request.POST.get('crearRecibo'):
            form_nuevoRC=form_nuevo_recaudoNR(request.POST,proyecto=proyecto)
            if form_nuevoRC.is_valid():
                check_perms(request,perms=('andinasoft.add_recaudosnoradicados',))
                nro_recibo=form_nuevoRC.cleaned_data.get('nro_recibo')
                contrato=form_nuevoRC.cleaned_data.get('contrato')
                fecha_recibo=form_nuevoRC.cleaned_data.get('fecha_recibo')
                valor=form_nuevoRC.cleaned_data.get('valor')
                formapago=form_nuevoRC.cleaned_data.get('formapago')
                concepto=form_nuevoRC.cleaned_data.get('concepto')
                RecaudosNoradicados.objects.using(proyecto).create(fecha=fecha_recibo,contrato=contrato,formapago=formapago,
                                                valor=valor,concepto=concepto,recibo=nro_recibo)
        """ form_nuevoRC=form_nuevo_recaudoNR(request.POST,proyecto=proyecto)
        if form_nuevoRC.is_valid():
            if request.POST.get('crearRecibo'):
                check_perms(request,perms=('andinasoft.add_recaudosnoradicados',))
                nro_recibo=form_nuevoRC.cleaned_data.get('nro_recibo')
                contrato=form_nuevoRC.cleaned_data.get('contrato')
                fecha_recibo=form_nuevoRC.cleaned_data.get('fecha_recibo')
                valor=form_nuevoRC.cleaned_data.get('valor')
                formapago=form_nuevoRC.cleaned_data.get('formapago')
                RecaudosNoradicados.objects.create(fecha=fecha_recibo,contrato=contrato,formapago=formapago,
                                                   valor=valor,concepto=concepto,recibo=nro_recibo) """
                
                
    obj_rec_nr=RecaudosNoradicados.objects.using(proyecto).raw('CALL recaudos_nr()')
    context={
        'form':form_radicar_recibo(proyecto=proyecto),
        'form_nuevoRnr':form_nuevo_recaudoNR(proyecto=proyecto),
        'recaudos':obj_rec_nr,
    }
    return render(request,'recaudos_noradicados.html',context)


@group_perm_required(perms=('andinasoft.change_adjudicacion',),raise_exception=True)
def reestructuraciones(request,proyecto,adj):
    #datosclientes
    obj_terceros=titulares_por_adj.objects.using(proyecto).get(adj=adj)
    email_titulares=''
    email_titular1=clientes.objects.using('default').get(idTercero=obj_terceros.IdTercero1)
    email_titular1=email_titular1.email
    email_titulares+=email_titular1.lower()
    if obj_terceros.IdTercero2!='' and obj_terceros.IdTercero2!=None:
        email_titular2=clientes.objects.using('default').get(idTercero=obj_terceros.IdTercero2)
        email_titular2=email_titular2.email
        email_titulares+=','+email_titular2.lower()
    if obj_terceros.IdTercero3!='' and obj_terceros.IdTercero3!=None:
        email_titular3=clientes.objects.using('default').get(idTercero=obj_terceros.IdTercero3)
        email_titular3=email_titular3.email
        email_titulares+=','+email_titular3.lower()
    if obj_terceros.IdTercero4!='' and obj_terceros.IdTercero4!=None and obj_terceros.IdTercero4!='0':
        email_titular4=clientes.objects.using('default').get(idTercero=obj_terceros.IdTercero4)
        email_titular4=email_titular4.email
        email_titulares+=','+email_titular4.lower()
    datos_venta=Adjudicacion.objects.using(proyecto).get(idadjudicacion=adj)
    saldos=saldos_adj.objects.using(proyecto).filter(adj=adj)
    pendiente_ci=saldos.filter(tipocta='CI').aggregate(Sum('saldocapital'))
    pendiente_ci=pendiente_ci['saldocapital__sum']
    if pendiente_ci==None: pendiente_ci=0
    pendiente_saldo=saldos.exclude(tipocta='CI').aggregate(Sum('saldocapital'))
    pendiente_saldo=pendiente_saldo['saldocapital__sum']
    if pendiente_saldo==None: pendiente_saldo=0
    pendiente_capital=pendiente_ci+pendiente_saldo
    context={
        'adj':adj,
        'proyecto':proyecto,
        'form':form_reestructuracion,
        'datos_venta':datos_venta,
        'email_clientes':email_titulares,
        'saldos_adj':saldos,
        'pendiente_ci':pendiente_ci,
        'pendiente_saldo':pendiente_saldo,
        'pendiente_capital':pendiente_capital,
        'tasa':tasas,
    }
    if request.method == 'POST':
        if request.is_ajax():
            todo = request.POST.get('todo')
            if todo == 'reestruc_pagado':
                valor_ajuste = request.POST.get('vr_ajuste')
                obj_planpagos=PlanPagos.objects.using(proyecto).filter(adj=adj)
                tipos_saldo=('CI','FN','CE','CO')
                for tipo in tipos_saldo:
                    ctas_vigentes = saldos.filter(tipocta=tipo,saldocuota__gt=0).order_by('nrocta')
                    if ctas_vigentes.exists():
                        cta_modificar = ctas_vigentes.first()
                        nrocta_ajustar = cta_modificar.nrocta
                        if cta_modificar.saldocuota != cta_modificar.cuota:
                            capcta=cta_modificar.rcdocapital
                            intcta=cta_modificar.rcdointcte
                            cta_ajustar=PlanPagos.objects.using(proyecto).get(idcta = cta_modificar.idcta)
                                
                            cta_ajustar.capital=capcta
                            cta_ajustar.intcte=intcta
                            cta_ajustar.cuota=capcta+intcta
                            cta_ajustar.save()
                            
                        for cuota in obj_planpagos.filter(tipocta=tipo):
                            if int(cuota.nrocta) > nrocta_ajustar:
                                cuota.delete()
                                
                datos_venta.valor+=Decimal(valor_ajuste)
                datos_venta.estado = 'Pagado'
                datos_venta.save()
                
                obj_timeline=timeline.objects.using(proyecto)
                mensaje = f'Aplicó una reestructuración, con ajuste en precio de venta en {valor_ajuste}'
                obj_timeline.create(adj=adj,
                               fecha=datetime.date.today(),
                               usuario=request.user,
                               accion=mensaje)
                data = {
                    'status':'ok'    
                }
                
                return JsonResponse(data)
                
                
        else:
            form=form_reestructuracion(request.POST)
            if form.is_valid():
                #hacemos un backup del plan de pagos y del detalle de reacudos
                obj_proyecto = proyectos.objects.get(pk=proyecto)
                obj_bk = bk_bfchangeplan.objects.create(
                    proyecto = obj_proyecto,
                    usuario_bk = request.user, adj = adj
                )
                full_planpagos = PlanPagos.objects.using(proyecto).filter(adj=adj)
                for i in full_planpagos:
                    bk_planpagos.objects.create(
                        id_bk = obj_bk,
                        proyecto = obj_proyecto,
                        idcta = i.idcta, tipocta= i.tipocta, nrocta = i.nrocta,
                        adj = i.adj, capital = i.capital, intcte = i.intcte,
                        cuota = i.cuota, fecha = i.fecha
                    )
                full_recaudodetalle = Recaudos.objects.using(proyecto).filter(idadjudicacion=adj)
                for i in full_recaudodetalle:
                    bk_recaudodetallado.objects.create(
                        id_bk = obj_bk,
                        proyecto = obj_proyecto, recibo = i.recibo,
                        fecha = i.fecha, idcta = i.idcta, idadjudicacion = i.idadjudicacion,
                        capital = i.capital, interescte = i.interescte, interesmora = i.interesmora,
                        moralqd = i.moralqd, fechaoperacion = i.fechaoperacion, 
                        usuario = i.usuario, estado = i.estado
                    )
                    
                is_ci=False
                cant_ci1=form.cleaned_data.get('cant_ci_1')
                fecha_ci1=form.cleaned_data.get('fecha_ini_ci1')
                valor_ci1=form.cleaned_data.get('valor_ci1')
                if cant_ci1!=None: is_ci=True
                cant_ci2=form.cleaned_data.get('cant_ci_2')
                fecha_ci2=form.cleaned_data.get('fecha_ini_ci2')
                valor_ci2=form.cleaned_data.get('valor_ci2')
                cant_ci3=form.cleaned_data.get('cant_ci_3')
                fecha_ci3=form.cleaned_data.get('fecha_ini_ci3')
                valor_ci3=form.cleaned_data.get('valor_ci3')
                cant_ci4=form.cleaned_data.get('cant_ci_4')
                fecha_ci4=form.cleaned_data.get('fecha_ini_ci4')
                valor_ci4=form.cleaned_data.get('valor_ci4')
                plan_pagos=form.cleaned_data.get('forma_plan_pagos')
                cant_fn=form.cleaned_data.get('cant_fn')
                fecha_fn=form.cleaned_data.get('fecha_ini_fn')
                valor_fn=form.cleaned_data.get('valor_fn')
                cant_ce=form.cleaned_data.get('cant_ce')
                fecha_ce=form.cleaned_data.get('fecha_ini_ce')
                valor_ce=form.cleaned_data.get('valor_ce')
                formapago=form.cleaned_data.get('forma_pago')
                planpagos=form.cleaned_data.get('forma_plan_pagos')
                periodo_ce=form.cleaned_data.get('periodo_ce')
                saldo=form.cleaned_data.get('pendiente_saldo')
                pendiente_ci=form.cleaned_data.get('pendiente_ci')
                ajuste=form.cleaned_data.get('ajuste')
                correos=form.cleaned_data.get('correos_notificacion')
                notificar=form.cleaned_data.get('notificar')
                tasa = request.POST.get('tasafn')
                
                #consecutivos
                obj_planpagos=PlanPagos.objects.using(proyecto).filter(adj=adj)
                obj_saldos=saldos_adj.objects.using(proyecto).filter(adj=adj)
                obj_adjudicacion=Adjudicacion.objects.using(proyecto).get(idadjudicacion=adj)
                #ci
                if is_ci:
                    ultimacta_ci=obj_saldos.filter(tipocta='CI').exclude(saldocuota=0).aggregate(Min('nrocta'))
                    ultimacta_ci=ultimacta_ci['nrocta__min']
                    nrocta_ajustar=ultimacta_ci
                    valores_ajustar=obj_saldos.filter(tipocta='CI',nrocta=nrocta_ajustar)
                    if valores_ajustar.exists():
                        if valores_ajustar[0].saldocuota!= valores_ajustar[0].cuota:
                            capcta=valores_ajustar[0].rcdocapital
                            intcta=valores_ajustar[0].rcdointcte
                            #ajustamos la ultima cuota segun lo que haya pagado
                            cta_ajustar=PlanPagos.objects.using(proyecto).get(adj=adj,tipocta='CI',nrocta=nrocta_ajustar)
                            cta_ajustar.capital=capcta
                            cta_ajustar.intcte=intcta
                            cta_ajustar.cuota=capcta+intcta
                            cta_ajustar.save()
                            nrocta_ajustar+=1
                        else: nrocta_ajustar=ultimacta_ci
                    #borramos el plan que exista en adelante
                    if nrocta_ajustar!=None:
                        for cuota in obj_planpagos.filter(tipocta='CI'):
                            if int(cuota.nrocta)>=nrocta_ajustar:
                                cuota.delete()
                    #insertamos el nuevo plan de pagos
                    forma_ci=''
                    obj_planpagos=PlanPagos.objects.using(proyecto)
                    count_ci=nrocta_ajustar
                    cantidad=cant_ci1
                    fecha=fecha_ci1
                    valor=valor_ci1
                    if cantidad is not None:
                        for j in range(0,cantidad):
                            fecha_ci=fecha+relativedelta(months=j)
                            idcta=f'CI{count_ci}{adj}'
                            obj_planpagos.create(idcta=idcta,
                                                tipocta='CI',
                                                nrocta=count_ci,
                                                adj=adj,
                                                capital=valor,
                                                intcte=0,
                                                cuota=valor,
                                                fecha=fecha_ci)
                            count_ci+=1
                        forma_ci+=f'${pendiente_ci:,} pendientes por pagar, serán cancelados en {cantidad} cuota(s) fija(s) mensual(es) de ${valor:,} a partir del {fecha}'
                    cantidad=cant_ci2
                    fecha=fecha_ci2
                    valor=valor_ci2
                    if cantidad is not None:
                        for j in range(0,cantidad):
                            fecha_ci=fecha+relativedelta(months=j)
                            idcta=f'CI{count_ci}{adj}'
                            obj_planpagos.create(idcta=idcta,
                                                tipocta='CI',
                                                nrocta=count_ci,
                                                adj=adj,
                                                capital=valor,
                                                intcte=0,
                                                cuota=valor,
                                                fecha=fecha_ci)
                            count_ci+=1
                        forma_ci+=f', {cantidad} cuota(s) fija(s) mensual(es) de ${valor} a partir del {fecha}'
                    cantidad=cant_ci3
                    fecha=fecha_ci3
                    valor=valor_ci3
                    if cantidad is not None:
                        for j in range(0,cantidad):
                            fecha_ci=fecha+relativedelta(months=j)
                            idcta=f'CI{count_ci}{adj}'
                            obj_planpagos.create(idcta=idcta,
                                                tipocta='CI',
                                                nrocta=count_ci,
                                                adj=adj,
                                                capital=valor,
                                                intcte=0,
                                                cuota=valor,
                                                fecha=fecha_ci)
                            count_ci+=1
                        forma_ci+=f', {cantidad} cuota(s) fija(s) mensual(es) de ${valor} a partir del {fecha}'
                    cantidad=cant_ci4
                    fecha=fecha_ci4
                    valor=valor_ci4
                    if cant_ci4!=None:
                            for j in range(0,cantidad):
                                fecha_ci=fecha+relativedelta(months=j)
                                idcta=f'CI{count_ci}{adj}'
                                obj_planpagos.create(idcta=idcta,
                                                    tipocta='CI',
                                                    nrocta=count_ci,
                                                    adj=adj,
                                                    capital=valor,
                                                    intcte=0,
                                                    cuota=valor,
                                                    fecha=fecha_ci)
                                count_ci+=1
                            forma_ci+=f', {cantidad} cuota(s) fija(s) mensual(es) de ${valor} a partir del {fecha}'
                #saldo
                forma_saldo=''
                if formapago=='Amortizacion': fp='FN'
                elif formapago=='Contado': fp='CO' 
                elif formapago=='Credicontado': fp='CE'
                tipos_saldo=('FN','CE','CO')
                #FN
                obj_planpagos=PlanPagos.objects.using(proyecto).filter(adj=adj)
                ultimacta_fn=obj_planpagos.filter(tipocta='FN').aggregate(Max('nrocta'))
                ultimacta_fn=ultimacta_fn['nrocta__max']
                if ultimacta_fn==None: ultimacta_fn=1
                ultimacta_ce=obj_planpagos.filter(tipocta='CE').aggregate(Max('nrocta'))
                ultimacta_ce=ultimacta_ce['nrocta__max']
                if ultimacta_ce==None: ultimacta_ce=1
                ultimacta_co=obj_planpagos.filter(tipocta='CO').aggregate(Max('nrocta'))
                ultimacta_co=ultimacta_co['nrocta__max']
                if ultimacta_co==None: ultimacta_co=1
                for tipo in tipos_saldo:
                    ultimacta=obj_saldos.filter(tipocta=tipo).exclude(saldocuota__lte=0).aggregate(Min('nrocta'))
                    ultimacta=ultimacta['nrocta__min']
                    nrocta_ajustar=ultimacta
                    
                    if nrocta_ajustar == None:
                        pass
                    else:
                        valores_ajustar = obj_saldos.filter(tipocta=tipo,nrocta=nrocta_ajustar)
                        if valores_ajustar.exists():
                            if valores_ajustar[0].saldocuota!= valores_ajustar[0].cuota:
                                capcta=valores_ajustar[0].rcdocapital
                                intcta=valores_ajustar[0].rcdointcte
                                #ajustamos la ultima cuota segun lo que haya pagado
                                cta_ajustar=PlanPagos.objects.using(proyecto).get(adj=adj,tipocta=tipo,nrocta=nrocta_ajustar)
                                cta_ajustar.capital=capcta
                                cta_ajustar.intcte=intcta
                                cta_ajustar.cuota=capcta+intcta
                                cta_ajustar.save()
                                nrocta_ajustar += 1
                            else: nrocta_ajustar=ultimacta
                        #borramos el plan que exista en adelante
                        for cuota in obj_planpagos.filter(tipocta=tipo):
                            if int(cuota.nrocta) >= nrocta_ajustar:
                                cuota.delete()
                                
                        if tipo=='FN': ultimacta_fn = nrocta_ajustar
                        elif tipo=='CE': ultimacta_ce = nrocta_ajustar
                        elif tipo=='CO': ultimacta_co = nrocta_ajustar
                
                if pendiente_ci == 0:
                    ctas_vigentes_ci = obj_saldos.filter(tipocta='CI',saldocuota__gt=0).order_by('nrocta')
                    if ctas_vigentes_ci.exists():
                        cta_modificar = ctas_vigentes_ci.first()
                        nrocta_ci_ajustar = cta_modificar.nrocta
                        if cta_modificar.saldocuota != cta_modificar.cuota:
                            capcta=cta_modificar.rcdocapital
                            intcta=cta_modificar.rcdointcte
                            cta_ajustar=PlanPagos.objects.using(proyecto).get(adj=adj,tipocta='CI',nrocta=nrocta_ci_ajustar)
                                
                            cta_ajustar.capital=capcta
                            cta_ajustar.intcte=intcta
                            cta_ajustar.cuota=capcta+intcta
                            cta_ajustar.save()
                            nrocta_ajustar +=1
                            
                        for cuota in obj_planpagos.filter(tipocta='CI'):
                            if int(cuota.nrocta) >= nrocta_ajustar:
                                cuota.delete()
                #saldo
                cantidad=cant_fn
                fecha=fecha_fn
                valor=valor_fn
                forma_saldo=f'${saldo:,} pendientes por pagar, serán cancelados en {cantidad} cuotas mensuales de ${valor:,} a partir del {fecha}'
                
                if formapago=='Amortizacion':
                    tasa=Decimal(tasa)
                else:
                    tasa=Decimal(0)
                if formapago=='CONTADO':
                    tipocta='CO'
                    ultimacuota=ultimacta_co
                else: 
                    tipocta='FN'
                    ultimacuota=ultimacta_fn
                    
                if cantidad is not None:
                    if planpagos=='Regular':
                        valor_presente=saldo
                    else:
                        valor_presente=Utilidades().CalcularVP(valor,tasa,cantidad)
                    capital_fn=valor_presente
                    counter=0
                    
                    for i in range (ultimacuota,ultimacuota+cantidad):
                        fecha_fn = fecha + relativedelta(months=counter)
                        idcta=f'{tipocta}{i}{adj}'
                        cta = PlanPagos.objects.using(proyecto).filter(idcta=idcta).exists()
                        
                        while cta:
                            i += 1
                            idcta=f'{tipocta}{i}{adj}'
                            cta = PlanPagos.objects.using(proyecto).filter(idcta=idcta).exists()
                            
                        interes=round(tasa*valor_presente,0)
                        capital=valor-interes
                        if i==ultimacuota+cantidad-1:
                            capital=valor_presente
                            valor=capital+interes
                        valor_presente-=capital
                        obj_planpagos.create(idcta=idcta,
                                            tipocta=tipocta,
                                            nrocta=i,
                                            adj=adj,
                                            capital=capital,
                                            intcte=interes,
                                            cuota=valor,
                                            fecha=fecha_fn)
                        counter+=1
                        
                cantidad=cant_ce
                fecha=fecha_ce
                valor=valor_ce
                ultimacuota=ultimacta_ce
                tasa=Decimal(tasa)
                if cantidad is not None:
                    periodo=periodo_ce
                    periodos={
                            'Mensual':1,
                            'Trimestral':3,
                            'Semestral':6,
                            'Anual':12
                        }
                    valor_presente=saldo-capital_fn
                    counter=0
                    forma_saldo+=f' y {cantidad} cuota(s) extraordinarias {periodo}es de ${valor:,} a partir del {fecha}'
                    for i in range (ultimacta_ce,ultimacta_ce+cantidad):
                        try: fecha_ce=fecha+relativedelta(months=counter*periodos[periodo])
                        except: fecha_ce=fecha+relativedelta(months=counter*int(periodo))
                        idcta=f'CE{i}{adj}'
                        try: interes=round(tasa*periodos[periodo]*valor_presente,0)
                        except: interes=round(tasa*int(periodo)*valor_presente,0)
                        capital=valor-interes
                        if i==ultimacuota+cantidad-1:
                            capital=valor_presente
                            valor=capital+interes
                        valor_presente-=capital
                        cta = PlanPagos.objects.using(proyecto).filter(idcta=idcta).exists()
                        
                        while cta:
                            i += 1
                            idcta = f'{tipocta}{i}{adj}'
                            cta = PlanPagos.objects.using(proyecto).filter(idcta=idcta).exists()
                            
                        obj_planpagos.create(idcta=idcta,
                                            tipocta='CE',
                                            nrocta=i,
                                            adj=adj,
                                            capital=capital,
                                            intcte=interes,
                                            cuota=valor,
                                            fecha=fecha_ce)
                        counter+=1
                mensaje = 'Aplicó una reestructuración'

                if ajuste!=0:
                    obj_adjudicacion.valor+=Decimal(ajuste)
                    obj_adjudicacion.save()
                    mensaje += f', con ajuste en precio de venta en {ajuste:,}'
                
                tasa_actual = f'{datos_venta.tasafnc*100:.2f}'
                tasa_nueva = f'{tasa*100:.2f}'
                if tasa_actual != tasa_nueva:
                    mensaje += f', con cambio en tasa de financiacion de {datos_venta.tasafnc*100:.2f}% a {tasa*100:.2f}%'
                    obj_adjudicacion.tasafnc = Decimal(tasa)
                    obj_adjudicacion.save()
                
            #creacion timeline
                obj_timeline=timeline.objects.using(proyecto)
                obj_timeline.create(adj=adj,
                                fecha=datetime.date.today(),
                                usuario=request.user,
                                accion=mensaje)
                if notificar=='Si':
                    try:
                        inmueble=obj_adjudicacion.idinmueble
                        datos_inmueble=Inmuebles.objects.using(proyecto).get(idinmueble=inmueble)
                        lote=datos_inmueble.lotenumero
                        manzana=datos_inmueble.manzananumero
                        if forma_ci=='': forma_ci='Sin cambios'
                        destinatarios=correos.split(',')
                        destinatarios.append('no-reply@andinasoft.com.co')
                        respuesta_reestructuracion(proyecto,lote,manzana,forma_ci,forma_saldo,destinatarios)
                    except: pass
                context['alerta']=True
                context['titulo_alerta']='¡Todo fue un exito!'
                context['mensaje']='La reestructuracion fue aplicada de forma correcta'
                context['redireccion']=True
                context['redirect']=f'/adjudicaciones/{proyecto}/{adj}'
            
    return render(request,'reestructuracion.html',context)

@group_perm_required(perms=('andinasoft.change_adjudicacion',),raise_exception=True)
def mover_fechas(request,proyecto,adj):
    cuotas_pendientes=saldos_adj.objects.using(proyecto).filter(adj=adj,saldocuota__gt=0)
    context={
        'proyecto':proyecto,
        'adj':adj,
        'cuotas_pendientes':cuotas_pendientes,
    }
    if request.method == 'POST':
        fecha=request.POST.get('fechacuota')
        meses=request.POST.get('meses')
        fecha_inicio=datetime.datetime.strptime(fecha,'%d/%m/%Y')
        cuotas_modify=PlanPagos.objects.using(proyecto).filter(adj=adj,fecha__gte=fecha_inicio)        
        for cuota in cuotas_modify:
            cuota.fecha+=relativedelta(cuota.fecha,months=int(meses))
            cuota.save()
        obj_timeline=timeline.objects.using(proyecto)
        obj_timeline.create(adj=adj,
                        fecha=datetime.date.today(),
                        usuario=request.user,
                        accion='Aplicó un aplazamiento en fechas')
    return render(request,'mover_fechas.html',context)

@group_perm_required(perms=('andinasoft.add_pagocomision',),raise_exception=True)
def comisiones(request,proyecto):
    obj_comisiones=AsignacionComisiones.objects.using(proyecto).raw('CALL comisiones_porcentuales("adj","%%")')
    context={
        'adjudicaciones':obj_comisiones,
        'proyecto':proyecto,
    }
    if request.method == 'GET':
        if request.is_ajax():
            form=request.GET.get('form')
            if form=='form-liquidar':
                adj=request.GET.get('adj_buscar')
                escala=AsignacionComisiones.objects.using(proyecto).raw(f'CALL comisiones_porcentuales("id","{adj}")')
                comisiones_por_cargo=[]
                for cargo in escala:
                    if cargo.estado == 'Activo':
                        valores={}
                        valores['idcargo']=cargo.idcargo
                        valores['porc_comis']=cargo.porc_comis
                        valores['gestor']=cargo.nombregestor
                        valores['por_pagar']=cargo.por_pagar
                        if cargo.por_pagar==None: 
                            cargo.por_pagar=0
                            valores['por_pagar']=cargo.por_pagar
                        provision=round(cargo.por_pagar*Decimal(0.1),0)
                        if cargo.idcargo=='40': provision=0
                        valores['provision']=provision
                        neto=cargo.por_pagar-provision
                        valores['neto']=neto
                        comisiones_por_cargo.append(valores)
                return JsonResponse({'instance':comisiones_por_cargo},status=200)
            elif form=='form-liquidarrm':
                adj=request.GET.get('adj_buscar')
                escala=AsignacionComisiones.objects.using(proyecto).raw(f'CALL comisiones_porcentuales("id","{adj}")')
                comisiones_por_cargo=[]
                for cargo in escala:
                    if cargo.estado == 'Activo':
                        valores={}
                        valores['idcargo']=cargo.idcargo
                        valores['porc_comis']=cargo.porc_comis
                        valores['gestor']=cargo.nombregestor
                        if cargo.idcargo in ('8','9','99'):
                            val_pagar = round(cargo.porc_comis * cargo.basecomision / 100)
                        else:
                            val_pagar = cargo.por_pagar
                        valores['por_pagar'] = val_pagar
                        if cargo.por_pagar==None: 
                            cargo.por_pagar=0
                            valores['por_pagar']=cargo.por_pagar
                        provision=round(val_pagar*Decimal(0.1),0)
                        if cargo.idcargo=='40': provision=0
                        valores['provision']=provision
                        neto=val_pagar-provision
                        valores['neto']=neto
                        comisiones_por_cargo.append(valores)
                return JsonResponse({'instance':comisiones_por_cargo},status=200)
            elif form=='form-asignacion':
                adj=request.GET.get('adj_buscar')
                obj_asignacion = AsignacionComisiones.objects.using(proyecto).filter(idadjudicacion=adj)
                asignaciones = []
                for asignacion in obj_asignacion:
                    cargo = {
                        'idcargo':asignacion.id_comision,
                        'idgestor':asignacion.idgestor,
                        'gestor':asesores.objects.get(pk=asignacion.idgestor).nombre.upper(),
                        'comision':asignacion.comision,
                        'estado':asignacion.usuario
                    }
                    asignaciones.append(cargo)
                return JsonResponse({'data':asignaciones})
            
                
                
    if request.method == 'POST':
        if request.POST.get('btn-Liquidar'):
            adj=request.POST.get('adj-liquidacion')
            escala=AsignacionComisiones.objects.using(proyecto).raw(f'CALL comisiones_porcentuales("id","{adj}")')
            obj_pago=Pagocomision.objects.using(proyecto)
            for cargo in escala:
                if cargo.estado == 'Activo':
                    Id=cargo.id
                    fecha=datetime.date.today()
                    idadjudicacion=cargo.idadjudicacion
                    idgestor=cargo.idgestor
                    idcargo=cargo.idcargo
                    tasacomision=cargo.porc_comis
                    comision=cargo.por_pagar
                    if cargo.por_pagar==None: cargo.por_pagar=0
                    provision=round(cargo.por_pagar*Decimal(0.1),0)
                    if cargo.idcargo=='40': provision=0
                    retefuente=provision
                    dctoanticipo=0
                    pagoneto=cargo.por_pagar-provision
                    obj_pago.create(id_comision=Id,fecha=fecha,idadjudicacion=idadjudicacion,idgestor=idgestor,
                                    idcargo=idcargo,tasacomision=tasacomision,comision=comision,retefuente=retefuente,
                                    dctoanticipo=dctoanticipo,pagoneto=pagoneto,transaccion=1,veces=1,veriact=1)
        if request.POST.get('btn-Liquidar-rm'):
            adj=request.POST.get('adj-liquidacionrm')
            escala=AsignacionComisiones.objects.using(proyecto).raw(f'CALL comisiones_porcentuales("id","{adj}")')
            obj_pago=Pagocomision.objects.using(proyecto)
            for cargo in escala:
                if cargo.estado == 'Activo':
                    print('entra')
                    Id=cargo.id
                    fecha=datetime.date.today()
                    idadjudicacion=cargo.idadjudicacion
                    idgestor=cargo.idgestor
                    idcargo=cargo.idcargo
                    tasacomision=cargo.porc_comis
                    if cargo.idcargo in ('8','9','99'):
                        comision = round(cargo.porc_comis * cargo.basecomision / 100)
                        retirar = True
                    else:
                        comision = cargo.por_pagar
                        retirar = False
                    if comision==None: comision=0
                    provision=round(comision*Decimal(0.1),0)
                    if cargo.idcargo=='40': provision=0
                    retefuente=provision
                    dctoanticipo=0
                    pagoneto=comision-provision
                    obj_pago.create(id_comision=Id,fecha=fecha,idadjudicacion=idadjudicacion,idgestor=idgestor,
                                    idcargo=idcargo,tasacomision=tasacomision,comision=comision,retefuente=retefuente,
                                    dctoanticipo=dctoanticipo,pagoneto=pagoneto,transaccion=1,veces=1,veriact=1)
                    
                    if retirar:
                        cargo_retirar = AsignacionComisiones.objects.using(proyecto).get(pk = cargo.pk)
                        cargo_retirar.usuario = 'Retirado'
                        cargo_retirar.save()
           
                    
                    
        if request.POST.get('btn-asignarescala'):
            adj = request.POST.getlist('adj-asignacion')
            cargos = request.POST.getlist('idcargo')
            gestores = request.POST.getlist('idgestor')
            comision = request.POST.getlist('comision')
            estado = request.POST.getlist('estado')
            
            for i in range(0,len(cargos)):
                asignacion=AsignacionComisiones.objects.using(proyecto).get(id_comision=cargos[i])
                asignacion.idgestor = gestores[i]
                asignacion.comision = Decimal(comision[i])
                asignacion.usuario = estado[i]
                asignacion.save()
                
    
    return render(request,'comisiones.html',context)

@group_perm_required(perms=('andinasoft.view_inmuebles',),raise_exception=True)
def inventario_admin(request,proyecto):
    
    if request.method == 'GET':
        if request.GET.get('download_inventario_template') == '1':
            book = openpyxl.Workbook()
            sheet = book.active
            sheet.title = 'inventario'
            sheet.append(INVENTARIO_IMPORT_COLUMNS)
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename=plantilla_inventario.xlsx'
            book.save(response)
            return response
        if request.is_ajax():
            inmueble=request.GET.get('inmueble')
            data_inmueble=Inmuebles.objects.using(proyecto).filter(idinmueble=inmueble)
            ser_inmuebles=serializers.serialize('json',data_inmueble)
            return JsonResponse({'instance':ser_inmuebles},status=200)
    if request.method == 'POST':
        if request.POST.get('btn-cargar-inventario') == '1':
            check_perms(request, perms=('andinasoft.change_inmuebles',))
            upload = request.FILES.get('inventario_file')
            if not upload:
                return JsonResponse({'ok': False, 'message': 'Debes cargar un archivo .xlsx o .csv'}, status=400)

            created_count = 0
            updated_count = 0
            errors = []

            try:
                for row_number, row_data in _iter_inventory_rows(upload):
                    try:
                        parsed = _parse_inventory_row(row_data)
                        inmueble_id = parsed.pop('idinmueble')
                        _, created = Inmuebles.objects.using(proyecto).update_or_create(
                            idinmueble=inmueble_id,
                            defaults=parsed
                        )
                        if created:
                            created_count += 1
                        else:
                            updated_count += 1
                    except Exception as exc:
                        errors.append(f'Fila {row_number}: {exc}')
            except Exception as exc:
                return JsonResponse({'ok': False, 'message': str(exc)}, status=400)

            if errors:
                return JsonResponse({
                    'ok': False,
                    'message': f'Se procesaron con errores. Creados: {created_count}, actualizados: {updated_count}',
                    'errors': errors[:20]
                }, status=400)

            return JsonResponse({
                'ok': True,
                'message': f'Inventario cargado. Creados: {created_count}, actualizados: {updated_count}'
            })

        check_perms(request,perms=('andinasoft.change_inmuebles',))
        form_lote=form_inv_admin(request.POST)
        if form_lote.is_valid():
            inmueble=form_lote.cleaned_data.get('lote')
            estado=form_lote.cleaned_data.get('Estado_Lote')
            obsbloqueo=form_lote.cleaned_data.get('Observaciones_Bloqueo')
            vrmetro=form_lote.cleaned_data.get('nuevo_valor_lote')
            meses=form_lote.cleaned_data.get('Meses_Entrega')
            if estado=='Bloqueado':
                usuario=str(request.user)
                fecha=datetime.datetime.today()
            else:
                usuario=''
                fecha=None
            obj_inmueble=Inmuebles.objects.using(proyecto).get(idinmueble=inmueble)
            obj_inmueble.estado=estado
            obj_inmueble.usuariobloquea=usuario
            obj_inmueble.obsbloqueo=obsbloqueo
            obj_inmueble.fechadesbloque=fecha
            obj_inmueble.vrmetrocuadrado=vrmetro
            obj_inmueble.meses=meses
            obj_inmueble.save()
    inmuebles=Inmuebles.objects.using(proyecto).all()
    context={
        'proyecto':proyecto,
        'inmuebles':inmuebles,
        'form_lote':form_inv_admin
    }
    return render(request,'inv_admin.html',context)

@group_perm_required(('andinasoft.view_presupuestocartera',),raise_exception=True)
def presupuesto_cartera(request,proyecto,periodo):
    check_project(request,proyecto)
    context={}
    if request.method=='POST':
        adj=request.POST.get('adj')
        obj_ppto=PresupuestoCartera.objects.using(proyecto).filter(idadjudicacion=adj,periodo=periodo)
        obj_infocartera=InfoCartera.objects.using(proyecto).filter(idadjudicacion=adj)
        if request.POST.get('btnGestor'):
            check_perms(request,('andinasoft.change_presupuestocartera',))
            gestor=request.POST.get('nuevoGestor')
            for cuota in obj_ppto:
                cuota.asesor=gestor
                cuota.save()
            if obj_infocartera.exists():
                adj_cartera=InfoCartera.objects.using(proyecto).get(idadjudicacion=adj)
                adj_cartera.gestorasignado=gestor
                adj_cartera.save()
            else:
                InfoCartera.objects.using(proyecto).create(idadjudicacion=adj,
                                                            gestorasignado=gestor)
        if request.POST.get('btnAccion'):
            check_perms(request,('andinasoft.delete_presupuestocartera',))
            accion=request.POST.get('acciones')
            if accion=='eliminar':
                for cuota in obj_ppto:
                    cuota.delete()
            elif accion=='recalcular':
                for cuota in obj_ppto:
                    cuota.delete()
                año=periodo[:4]
                mes=periodo[-2:]
                dia=calendar.monthrange(int(año),int(mes))[1]
                fecha_hasta=datetime.datetime.strptime(f"{año}-{mes}-{dia}","%Y-%m-%d")
                stmt=f'CALL ver_presupuesto("{fecha_hasta}","{adj}")'
                obj_verpresupuesto=saldos_adj.objects.using(proyecto).raw(stmt)
                obj_nuevoppto=PresupuestoCartera.objects.using(proyecto)
                for cuota in obj_verpresupuesto:
                    obj_nuevoppto.create(id_ppto=cuota.id,periodo=periodo,idadjudicacion=cuota.adj,cliente=cuota.cliente,
                                        tipocta=cuota.tipocta,ncta=cuota.nrocta,idcta=cuota.idcta,tipocartera=cuota.tipocartera,
                                        fecha=cuota.fechacta,capital=cuota.saldocapital,interes=cuota.saldointcte,cuota=cuota.saldocuota,
                                        diasmora=cuota.diasmora,mora=cuota.saldomora,asesor=cuota.asesor,usuario=request.user,
                                        fechaoperacion=datetime.date.today(),edad=cuota.edad)
        
        if request.POST.get('btnExportar'):
            usuario_administrador=check_perms(request,('andinasoft.change_presupuestocartera',))
            if usuario_administrador:
                contenido_ppto=Adjudicacion.objects.using(proyecto).raw(f'CALL informe_cartera("{periodo}",NULL)')
            else:
                gestor=f'{request.user.first_name} {request.user.last_name}'.upper()
                contenido_ppto=Adjudicacion.objects.using(proyecto).raw(f'CALL informe_cartera("{periodo}",NULL)')
            book=openpyxl.Workbook()
            sheet=book.active
            encabezados=['Adjudicacion','Cliente','Estado','Origen','Venta Mes','Tipo Cartera','Edad',
                            'Cuota Mes','Recaudo Mes','Cuotas Vencidas','Recaudo Vencido','Presupuesto Total','Recaudo Presupuestado',
                            'Recaudo No Pptado','Recaudo Total','Asesor']
            sheet.append(encabezados)
            i=2
            for fila in contenido_ppto:
                sheet.cell(i,1,fila.pk)
                sheet.cell(i,2,fila.cliente)
                sheet.cell(i,3,fila.estado)
                sheet.cell(i,4,fila.origen)
                sheet.cell(i,5,fila.venta_mes)
                sheet.cell(i,6,fila.tipocartera)
                sheet.cell(i,7,fila.edad)
                sheet.cell(i,8,fila.ppto_mes)
                sheet.cell(i,9,fila.recaudo_mes)
                sheet.cell(i,10,fila.ppto_vencido)
                sheet.cell(i,11,fila.recaudo_vencido)
                sheet.cell(i,12,fila.presupuesto)
                sheet.cell(i,13,fila.recaudo_pptado)
                sheet.cell(i,14,fila.recaudo_nopptado)
                sheet.cell(i,15,fila.recaudo_total)
                sheet.cell(i,16,fila.asesor)
                i+=1
            filename = f'Presupuesto_{proyecto}_periodo_{periodo}.xlsx'
            ruta=settings.DIR_EXPORT+filename
            _save_workbook_with_dirs(book, ruta)
            return FileResponse(open(ruta,'rb'),as_attachment=True,filename=filename)
    
    usuario_administrador=check_perms(request,('andinasoft.change_presupuestocartera',),raise_exception=False)
    if usuario_administrador:
        obj_informe=Adjudicacion.objects.using(proyecto).raw(f'CALL informe_cartera("{periodo}",NULL)')
    else:
        gestor=f'{request.user.first_name} {request.user.last_name}'.upper()
        obj_informe=Adjudicacion.objects.using(proyecto).raw(f'CALL informe_cartera("{periodo}","{gestor}")')
    context['proyecto']=proyecto
    context['periodo']=periodo
    context['informe']=obj_informe
    context['periodo']=f'{periodo}'
            
    return render(request,'ver_ppto.html',context)

@group_perm_required(('andinasoft.view_presupuestocartera',),raise_exception=True)
def ver_presupuesto(request,proyecto):
    context={
        'proyecto':proyecto
    }
    if request.method=='POST':
        if request.POST.get('btnCargar'):
            check_perms(request,('andinasoft.add_presupuestocartera',))
            año=request.POST.get('periodoaño')
            mes=request.POST.get('periodomes')
            dia=calendar.monthrange(int(año),int(mes))[1]
            periodo=f'{año}{mes}'
            fecha_hasta=datetime.datetime.strptime(f"{año}-{mes}-{dia}","%Y-%m-%d")
            stmt=f'CALL ver_presupuesto("{fecha_hasta}","")'
            obj_verpresupuesto=saldos_adj.objects.using(proyecto).raw(stmt)
            obj_nuevoppto=PresupuestoCartera.objects.using(proyecto)
            for cuota in obj_verpresupuesto:
                obj_nuevoppto.create(id_ppto=cuota.id,periodo=periodo,idadjudicacion=cuota.adj,cliente=cuota.cliente,
                                     tipocta=cuota.tipocta,ncta=cuota.nrocta,idcta=cuota.idcta,tipocartera=cuota.tipocartera,
                                     fecha=cuota.fechacta,capital=cuota.saldocapital,interes=cuota.saldointcte,cuota=cuota.saldocuota,
                                     diasmora=cuota.diasmora,mora=cuota.saldomora,asesor=cuota.asesor,usuario=request.user,
                                     fechaoperacion=datetime.date.today(),edad=cuota.edad)
            context['alerta']=True
            context['mensaje']=f'El presupuesto del proyecto {proyecto} para el periodo {periodo} ha sido cargado'
            context['titulo_alerta']='¡Listo!'
            context['redireccion']=True
            context['redirect']=f'/cartera/ver_presupuesto/{proyecto}/{año}{mes}'
        if request.POST.get('btnEliminar'):
            check_perms(request,('andinasoft.delete_presupuestocartera',))
            año=request.POST.get('periodoaño')
            mes=request.POST.get('periodomes')
            periodo=f'{año}{mes}'
            obj_ppto=PresupuestoCartera.objects.using(proyecto).filter(periodo=periodo)
            for cuota in obj_ppto:
                cuota.delete()
            context['alerta']=True
            context['mensaje']=f'El presupuesto del proyecto {proyecto} para el periodo {periodo} ha sido eliminado'
            context['titulo_alerta']='¡Listo!'
        
    return render(request,'periodo_ppto.html',context)

@login_required
@group_perm_required(perms=('andinasoft.view_adjudicacion',),raise_exception=True)
def edades_cartera(request,proyecto):
    check_project(request,proyecto)
    
    if proyecto == 'Fractal':
        context = {
        }
        
        return render(request,'adjudicados_fractal.html',context)
    
    check_groups(request,('Supervisor Cartera',))
    obj_adj = Adjudicacion.objects.using(proyecto).filter(estado='Aprobado'
                        ).exclude(origenventa='Canje')
    adj_list = []
    for adj in obj_adj:
        adj_list.append( {
            'info':adj,
            'presupuesto':adj.presupuesto()
        })
    context = {
        'proyecto':proyecto,
        'adjudicaciones':adj_list,
    }
    
    return render(request,'edades_cartera.html',context)
    

@group_perm_required(perms=('andinasoft.add_ventas_nuevas',),raise_exception=True)
def detalle_comisiones(request,proyecto):
    check_project(request,proyecto)
    context={
        'proyecto':proyecto,
        'form':form_detalle_comisiones,
    }
    id_restringidos=('1037588511','79044027','1067913728','1037624224','1067868382')
    cargos_restringidos = ('8','9','99')
    if request.method == 'POST':
        form=form_detalle_comisiones(request.POST)
        if form.is_valid():
            if request.POST.get('buscar-porasesor'):
                cedula_asesor=form.cleaned_data.get('valor_buscado')
                if cedula_asesor in id_restringidos:
                    check_groups(request,('Gestion Humana','Coordinador operaciones'))
                fecha_desde=form.cleaned_data.get('fecha_desde')
                fecha_hasta=form.cleaned_data.get('fecha_hasta')
                
                try: 
                    asesor_obj = asesores.objects.using('default').get(cedula=cedula_asesor)
                    asesor=asesor_obj.nombre
                    
                except: asesor=''
                
                if asesor_obj.tipo_asesor == 'Externo':
                    if check_groups(request,('Gestion Humana','Coordinador operaciones'),raise_exception=False):
                        
                        stmt=f'CALL detalle_comisiones_asesor("{fecha_desde}","{fecha_hasta}","{cedula_asesor}")'
                        comisiones=Pagocomision.objects.using(proyecto).raw(stmt)
                        
                        context['respuesta']='asesor'
                        context['comisiones_asesor']=comisiones
                        context['asesor']=asesor
                        context['fecha_desde']=f'{fecha_desde}'
                        context['fecha_hasta']=f'{fecha_hasta}'
                        
                        form=form_detalle_comisiones(initial={
                            'valor_buscado':cedula_asesor,
                            'fecha_desde':fecha_desde,
                            'fecha_hasta':fecha_hasta,
                        })
                        context['form']=form
                    
                    else:
                        context['alerta']=True
                        context['mensaje']=f'Los asesores externos deben solicitar el detalle de comisiones a la coordinacion de operaciones.'
                        context['titulo_alerta']='No permitido'
                        
                else:
                    stmt=f'CALL detalle_comisiones_asesor("{fecha_desde}","{fecha_hasta}","{cedula_asesor}")'
                    comisiones=Pagocomision.objects.using(proyecto).raw(stmt)
                    
                    context['respuesta']='asesor'
                    context['comisiones_asesor']=comisiones
                    context['asesor']=asesor
                    context['fecha_desde']=f'{fecha_desde}'
                    context['fecha_hasta']=f'{fecha_hasta}'
                    
                    form=form_detalle_comisiones(initial={
                        'valor_buscado':cedula_asesor,
                        'fecha_desde':fecha_desde,
                        'fecha_hasta':fecha_hasta,
                    })
                    context['form']=form
            
            if request.POST.get('buscar-poradj'):
                if check_perms(request,('andinasoft.add_pagocomision',),raise_exception=False):
                    adj=form.cleaned_data.get('valor_buscado')
                    fecha_desde=form.cleaned_data.get('fecha_desde')
                    fecha_hasta=form.cleaned_data.get('fecha_hasta')
                    try: cliente=titulares_por_adj.objects.using(proyecto).get(adj=adj).titular1
                    except: cliente=''
                    stmt=f'CALL detalle_comisiones_cliente("{fecha_desde}","{fecha_hasta}","{adj}")'
                    comisiones=Pagocomision.objects.using(proyecto).raw(stmt)
                    context['respuesta']='adj'
                    context['comisiones_adj']=comisiones
                    context['cliente']=cliente
                    context['fecha_desde']=f'{fecha_desde}'
                    context['fecha_hasta']=f'{fecha_hasta}'
                    form=form_detalle_comisiones(initial={
                        'valor_buscado':adj,
                        'fecha_desde':fecha_desde,
                        'fecha_hasta':fecha_hasta,
                    })
                    context['form']=form
                else:
                    raise PermissionDenied
            
            if request.POST.get('buscar-porfecha'):
                if check_perms(request,('andinasoft.view_pagocomision',),raise_exception=False):
                    fecha_desde=form.cleaned_data.get('fecha_desde')
                    fecha_hasta=form.cleaned_data.get('fecha_hasta')
                    stmt=f'CALL detalle_comisiones_fecha("{fecha_desde}","{fecha_hasta}")'
                    comisiones=Pagocomision.objects.using(proyecto).raw(stmt)
                    context['respuesta']='fecha'
                    context['comisiones_fecha']=comisiones
                    context['fecha_desde']=f'{fecha_desde}'
                    context['fecha_hasta']=f'{fecha_hasta}'
                    form=form_detalle_comisiones(initial={
                        'fecha_desde':fecha_desde,
                        'fecha_hasta':fecha_hasta,
                    })
                    context['form']=form
                else:
                    raise PermissionDenied
            
    return render(request,'detalle_comisiones.html',context)

@group_perm_required(('andinasoft.add_facturas',),raise_exception=True)
def radicar_factura(request):
    context={
        'form': form_radicar_factura,
    }
    if request.method == 'POST':
        form=form_radicar_factura(request.POST,request.FILES)
        if form.is_valid():
            empresa=form.cleaned_data.get('empresa_radicado')
            num_factura=form.cleaned_data.get('num_factura')
            valor_factura=form.cleaned_data.get('valor')
            fecha_vencimiento=form.cleaned_data.get('fecha_vencimiento')
            fecha_factura=form.cleaned_data.get('fecha_factura')
            idtercero=form.cleaned_data.get('idtercero')
            nombretercero=form.cleaned_data.get('nombretercero')
            factura=form.cleaned_data.get('documento_cargar')
            
            existe_factura=Facturas.objects.filter(idtercero=idtercero,nrofactura=num_factura)
            if existe_factura.exists():
                context['alerta']=True
                context['mensaje']=f'La factura {num_factura} del tercero {nombretercero} ya existe'
                context['titulo_alerta']='¡Listo!'
            else:
                Facturas.objects.create(nrofactura=num_factura,
                                        fecharadicado=datetime.date.today(),
                                        idtercero=idtercero,
                                        nombretercero=nombretercero,
                                        empresa=empresa,
                                        valor=valor_factura,
                                        fechavenc=fecha_vencimiento,
                                        fechafactura=fecha_factura,
                                        estado='Recepcion')
                nro_radicado=Facturas.objects.aggregate(Max('nroradicado'))['nroradicado__max']
                name_doc=f'Radicado_{nro_radicado}_Factura_{num_factura}'
                if factura is not None:
                    documento=request.FILES['documento_cargar']
                    upload_docs_radicados(documento,'Facturas',name_doc)
                obj_timeline=timeline_radicados.objects.create(nroradicado=nro_radicado,
                                                               usuario=request.user,
                                                               accion='Radicó la factura',
                                                               fecha=datetime.date.today(),
                                                               tiporadicado='Factura')
                context['alerta']=True
                context['mensaje']=f'La factura fue radicada con numero de radicado {nro_radicado}'
                context['titulo_alerta']='¡Listo!'
            
    return render(request,'radicar_facturas.html',context)

@group_perm_required(perms=('andinasoft.view_facturas',),raise_exception=True)
def lista_facturas(request):
    obj_facturas=Facturas.objects.filter(origen='Radicado')
    obj_timeline=timeline_radicados.objects.all()
    context={
        'radicados':obj_facturas,
        'timeline':obj_timeline,
    }
    if request.is_ajax():
        if request.method == 'GET':
            radicado=request.GET.get('radicado')
            data_radicado=timeline_radicados.objects.filter(nroradicado=radicado,tiporadicado='Factura')
            ser_radicado=serializers.serialize('json',data_radicado)
            return JsonResponse({'instance':ser_radicado},status=200)
    if request.method == 'POST':
        if request.POST.get('btnRecibir'):
            factura=request.POST.get('numero-factura')
            radicado=request.POST.get('numero-radicado')
            obj_radicado=Facturas.objects.get(nroradicado=radicado)
            estado_actual=obj_radicado.estado
            recibe='Ninguno'
            if estado_actual=='Recepcion': estado_siguiente='Contabilidad'
            if estado_actual=='Contabilidad': estado_siguiente='Tesoreria'
            grupos=request.user.groups.all()
            for grupo in grupos:
                if 'Contabilidad' in grupo.name:
                    recibe='Contabilidad'
                elif 'Tesoreria' in grupo.name:
                    recibe='Tesoreria'
            check=False
            if request.user.is_superuser: check=True
            elif estado_actual=='Recepcion' and recibe=='Contabilidad': check=True
            elif estado_actual=='Contabilidad' and recibe=='Tesoreria': check=True 
            if check:
                if obj_radicado.nrocausa==None and estado_siguiente=='Tesoreria':
                    context['alerta']=True
                    context['mensaje']=f'No se puede recibir un documento en tesoreria sin numero de causacion'
                    context['titulo_alerta']='Error'
                else:
                    obj_radicado.estado=estado_siguiente
                    obj_radicado.save()
                    obj_timeline=timeline_radicados.objects.create(nroradicado=radicado,
                                                                usuario=request.user,
                                                                accion=f'Recibió la factura en {estado_siguiente}',
                                                                fecha=datetime.date.today(),
                                                               tiporadicado='Factura')
                    context['alerta']=True
                    context['mensaje']=f'Recibiste la factura {factura} del radicado {radicado}'
                    context['titulo_alerta']='¡Listo!'
            else:
                context['alerta']=True
                context['mensaje']=f'El flujo de un radicado es Recepcion-Contabilidad-Tesoreria, revisa tus permisos o el estado actual del radicado'
                context['titulo_alerta']='Error'
                
    return render(request,'lista_facturas.html',context)

@group_perm_required(('andinasoft.view_facturas','andinasoft.change_facturas'),raise_exception=True)
def causar_facturas(request):
    obj_radicados=Facturas.objects.filter(estado='Contabilidad',origen='Radicado')
    context={
        'radicados':obj_radicados,
        'form':form_causar
    }
    if request.method == 'POST':
        form=form_causar(request.POST)
        if form.is_valid():
            radicado=request.POST.get('numeroRadicado')
            nro_causa=form.cleaned_data.get('nrocausacion')
            pagoneto=form.cleaned_data.get('pagoneto')
            fechacausa=form.cleaned_data.get('fecha_causacion')
            obj_radicado=Facturas.objects.get(nroradicado=radicado)
            obj_radicado.nrocausa=nro_causa
            obj_radicado.pago_neto=pagoneto
            obj_radicado.fechacausa=datetime.date.today()
            obj_radicado.save()
            obj_timeline=timeline_radicados.objects.create(nroradicado=radicado,
                                                                usuario=request.user,
                                                                accion='Registró el numero de Causacion',
                                                                fecha=datetime.date.today(),
                                                               tiporadicado='Factura')
            context['alerta']=True
            context['mensaje']=f'Se registró el numero de causacion para el radicado {radicado}'
            context['titulo_alerta']='¡Listo!'
    
    return render(request,'causar_facturas.html',context)

@group_perm_required(('andinasoft.add_pagos',),raise_exception=True)
def pagar_facturas(request):
    obj_porpagar=Facturas.objects.raw('CALL facturas_porpagar()')
    context={
        'radicados':obj_porpagar,
        'form':form_pagar,
    }
    if request.method == 'GET':
        if request.is_ajax():
            solicitud=request.GET.get('solicitud')
            if solicitud=='formapago':
                empresa=request.GET.get('empresa')
                formaspago=cuentas_pagos.objects.filter(empresa=empresa)
                ser_formapspago=serializers.serialize('json',formaspago)
                return JsonResponse({'instance':ser_formapspago},status=200)
    if request.method == 'POST':
        form=form_pagar(request.POST)
        if form.is_valid():
            radicado=request.POST.get('numeroRadicado')
            fecha=form.cleaned_data.get('fecha_causacion')
            formapago=request.POST.get('formapago')
            valorpago=form.cleaned_data.get('valorpago')
            empresapago=form.cleaned_data.get('empresapago')
            obj_pagos=Pagos.objects.create(nroradicado=radicado,
                                           valor=valorpago,
                                           fechapago=fecha,
                                           formapago=formapago,
                                           empresapago=empresapago)
            obj_timeline=timeline_radicados.objects.create(nroradicado=radicado,
                                                               usuario=request.user,
                                                               accion=f'Registró un pago por ${valorpago:,}',
                                                               fecha=datetime.date.today(),
                                                               tiporadicado='Factura')
            obj_porpagar=Facturas.objects.raw('CALL facturas_porpagar()')
            context['radicados']=obj_porpagar
            context['alerta']=True
            context['mensaje']=f'El pago fue aplicado al radicado {radicado}'
            context['titulo_alerta']='¡Listo!'
    return render(request,'pagar_facturas.html',context)

@group_perm_required(('andinasoft.view_pagos',),raise_exception=True)
def lista_pagos(request):
    obj_pagos=Pagos.objects.raw('CALL pagosefectuados()')
    context={
        'pagos':obj_pagos,
    }
    if request.method == 'POST':
        file=request.FILES['cargar-soporte']
        nropago=request.POST.get('numero-pago')
        radicado=request.POST.get('numero-radicado')
        namedoc=f'Soporte_Pago_{nropago}_Radicado_{radicado}'
        upload_docs_radicados(file,'Soportes_Pago',namedoc)
        
    return render(request,'lista_pagos.html',context)

@group_perm_required(('andinasoft.view_pagos',),raise_exception=True)
def interfaces_bancarias(request):
    context= {
        'form-int-comis':form_int_comi_banco,
        'form-int-egr':forms_accounting.form_interfaz_egreso,
        'proyectos':proyectos.objects.all().exclude(proyecto__icontains='Alttum')
    }
    if request.method == 'GET':
        if request.is_ajax():
            data = request.GET.get('data_request')
            if data == 'cuentas_por_empresa':
                empresa = request.GET.get('empresa')
                cuentas = cuentas_pagos.objects.filter(empresa=empresa).exclude(Q(cuentabanco='Efectivo')|Q(cuentabanco='Tarjeta de Credito Corp'))
                json_cuentas = serializers.serialize('json',cuentas)
                return JsonResponse({'data':json_cuentas})
    if request.method == 'POST':
        desde = request.POST.get('fecha_desde')
        hasta = request.POST.get('fecha_hasta')
        empresa = request.POST.get('empresa')
        cuenta = request.POST.get('cuenta')
        proyecto = request.POST.get('proyecto')
        nits = {
            'Promotora Sandville':900993044,
            'Promotora Westville':901132949,
            'Status Comercializadora':901018375,
            'Quadrata Constructores':901004733,
            'Terranova Desarrolladora Turistica':900712229
        }
        nit_empresa = nits.get(empresa)
        cuenta_debitar = int(cuenta.replace('-',""))
        book=openpyxl.load_workbook("resources/excel_formats/InterfazPAB.xlsx")
        sheet=book.active
        sheet.cell(2,1,nit_empresa)
        sheet.cell(2,2,220)
        letra=random.choice(string.ascii_uppercase)
        numero=random.randint(1,9)
        secuencia=f'{letra}{numero}'
        sheet.cell(2,3,"I")
        sheet.cell(2,4,secuencia)
        sheet.cell(2,5,cuenta_debitar)
        nom_proy=proyecto.replace(' ',"").upper()[:4]
        sheet.cell(2,6,'D')
        sheet.cell(2,7,'COMISIONES')
        stmt=f'CALL detalle_comisiones_fecha("{desde}","{hasta}")'
        comisiones=Pagocomision.objects.using(proyecto).raw(stmt)
        
        row=4
        for line in comisiones:
            datos_asesor = asesores.objects.get(pk=line.idgestor)
            sheet.cell(row,1,1)
            sheet.cell(row,2,int(line.idgestor))
            sheet.cell(row,3,line.nombre.upper()[:31])
            if datos_asesor.tipo_cuenta == 'Ahorros': tipo_cuenta=37
            elif datos_asesor.tipo_cuenta == 'Corriente': tipo_cuenta=40
            else: tipo_cuenta=""
            sheet.cell(row,4,tipo_cuenta)
            try: banco=entidades_bancarias.objects.get(banco=datos_asesor.banco).codigo
            except: banco=""
            sheet.cell(row,5,banco)
            try: sheet.cell(row,6,int(datos_asesor.cuenta))
            except:pass
            referencia = proyecto.replace(" ","").upper()
            ref=f'COMIS{referencia}'[:20]
            sheet.cell(row,9,ref)
            sheet.cell(row,11,line.pagoneto)
            row+=1
        
        nombre_doc=f'Interfaz_PAB_Comisiones_{proyecto}_{desde}_{hasta}.xlsx'
        ruta=settings.DIR_EXPORT+nombre_doc
        _save_workbook_with_dirs(book, ruta)
        return FileResponse(open(ruta,'rb'),as_attachment=True,filename=nombre_doc)    
                
                
    return render(request,'interfaz_comisiones_banco.html',context)

@group_perm_required(('andinasoft.change_consecutivos',),raise_exception=True)
def interfaces_contabilidad(request,proyecto):
    obj_cuentas=consecutivos.objects.using(proyecto).get(documento='RC')
    try:
        obj_cuentas_comis = consecutivos.objects.using(proyecto).get(documento='COMISIONES')
    except:
        obj_cuentas_comis = ""
    context={
        'cuentas':obj_cuentas,
        'cuentas_comis':obj_cuentas_comis,
        'proyecto':proyecto,
        'form-int-egr':forms_accounting.form_interfaz_egreso,
        'proyectos':proyectos.objects.exclude(Q(pk__istartswith='alttum')|Q(pk__icontains='sol'))
    }
    
    if request.method == 'POST':
        if request.POST.get('btnIntRecibos'):
            desde=request.POST.get('Fecha_Desde')
            hasta=request.POST.get('Fecha_Hasta')
            book=openpyxl.load_workbook("resources/excel_formats/InterfazSIIGO.xlsx")
            sheet=book.active
            stmt=f'CALL interfaz_recibos("{desde}","{hasta}")'
            obj_recaudos=Recaudos_general.objects.using(proyecto).filter(fecha__range=(desde,hasta)).order_by('fecha')
            sheet.cell(1,1,f'{proyecto.upper()}_RECAUDOS_DESDE_{desde}_HASTA_{hasta}')
            row=6
            for r in obj_recaudos:
                
                if r.valor > 0:
                    for line in r.interfaz_contabilidad():
                        sheet.cell(row,1,line.get('tipocomprobante'))
                        sheet.cell(row,2,line.get('codigocomprobante'))
                        sheet.cell(row, 3,line.get('numerocomprobante'))
                        sheet.cell(row,4,line.get('cuenta'))
                        sheet.cell(row,5,line.get('naturaleza'))
                        sheet.cell(row,6,line.get('valor'))
                        sheet.cell(row,7,line.get('año'))
                        sheet.cell(row,8,line.get('mes'))
                        sheet.cell(row,9,line.get('dia'))
                        sheet.cell(row,16,line.get('nit'))
                        sheet.cell(row,18,line.get('descripcion'))
                        row+=1
            filename = f'Interfaz_Recibos_{proyecto}_{desde}_{hasta}.xlsx'
            ruta=settings.DIR_EXPORT+filename
            _save_workbook_with_dirs(book, ruta)            
            file = open(ruta,'rb')
            return FileResponse(file,as_attachment=True,filename=filename)
        
        if request.POST.get('btnIntComisiones'):
            fecha_desde=request.POST.get('Fecha_Desde')
            fecha_hasta=request.POST.get('Fecha_Hasta')
            consecutivo=request.POST.get('consecutivo')
            obj_centrocosto = CentrosCostos.objects.filter(proyecto=proyecto,empresa='Status Comercializadora')
            if obj_centrocosto.exists():
                cc=int(obj_centrocosto[0].centro)
                subcc=int(obj_centrocosto[0].subcentro)
            else:
                cc=""
                subcc=""       
            stmt=f'CALL detalle_comisiones_fecha("{fecha_desde}","{fecha_hasta}")'
            comisiones=Pagocomision.objects.using(proyecto).raw(stmt)
            documento = consecutivos.objects.using(proyecto).get(documento='COMISIONES')
            num_comprobante = documento.comprobante_contable
            cuenta_gasto = documento.cuenta_capital
            cuenta_provision = documento.cuenta_intcte
            cuenta_porpagar = documento.cuenta_inmora
            cuenta_anticipo = documento.cuenta_aux1
            book=openpyxl.load_workbook("resources/excel_formats/InterfazSIIGO.xlsx")
            sheet=book.active
            sheet.cell(1,1,f'{proyecto.upper()}_COMISIONES_DESDE_{fecha_desde}_HASTA_{fecha_hasta}')
            row=6
            cruce=1
            for line in comisiones:
                cuenta_de_gasto = cuenta_gasto
                provision = True
                gasto = line.comision
                asesor = asesores.objects.get(pk=line.idgestor)
                if asesor.estado=='Activo' and asesor.tipo_asesor=='Interno':
                    provision = False
                    cuenta_de_gasto = cuenta_anticipo
                    gasto=line.pagoneto
                if line.comision!=0:
                    descripcion = f'COMISIONES {line.nombre} {proyecto} {line.fecha}'
                    sheet.cell(row,1,"L")
                    sheet.cell(row,2,num_comprobante)
                    sheet.cell(row, 3,consecutivo)
                    sheet.cell(row,4,cuenta_de_gasto)
                    sheet.cell(row,5,'D')
                    sheet.cell(row,6,gasto)
                    sheet.cell(row,7,line.fecha.year)
                    sheet.cell(row,8,line.fecha.month)
                    sheet.cell(row,9,line.fecha.day)
                    sheet.cell(row,14,cc)
                    sheet.cell(row,15,subcc)
                    sheet.cell(row,16,line.idgestor)
                    sheet.cell(row,18,descripcion.upper())
                    if provision:
                        row+=1
                        descripcion = f'PROVISION COMISIONES {line.nombre} {proyecto} {line.fecha}'
                        sheet.cell(row,1,"L")
                        sheet.cell(row,2,num_comprobante)
                        sheet.cell(row, 3,consecutivo)
                        sheet.cell(row,4,cuenta_provision)
                        sheet.cell(row,5,'C')
                        sheet.cell(row,6,line.provision)
                        sheet.cell(row,7,line.fecha.year)
                        sheet.cell(row,8,line.fecha.month)
                        sheet.cell(row,9,line.fecha.day)
                        sheet.cell(row,14,cc)
                        sheet.cell(row,15,subcc)
                        sheet.cell(row,16,line.idgestor)
                        sheet.cell(row,18,descripcion.upper())
                    row+=1
                    descripcion = f'CXP COMISIONES {line.nombre} {proyecto} {line.fecha}'
                    sheet.cell(row,1,"L")
                    sheet.cell(row,2,num_comprobante)
                    sheet.cell(row,3,consecutivo)
                    sheet.cell(row,4,cuenta_porpagar)
                    sheet.cell(row,5,'C')
                    sheet.cell(row,6,line.pagoneto)
                    sheet.cell(row,7,line.fecha.year)
                    sheet.cell(row,8,line.fecha.month)
                    sheet.cell(row,9,line.fecha.day)
                    sheet.cell(row,14,cc)
                    sheet.cell(row,15,subcc)
                    sheet.cell(row,16,line.idgestor)
                    sheet.cell(row,18,descripcion.upper())
                    sheet.cell(row,58,f'L-{int(num_comprobante):03d}')
                    sheet.cell(row,59,consecutivo)
                    sheet.cell(row,60,cruce)
                    sheet.cell(row,61,line.fecha.year)
                    sheet.cell(row,62,line.fecha.month)
                    sheet.cell(row,63,line.fecha.day)
                    row+=1
                    cruce+=1
            name=f'Comisiones_{proyecto}_desde_{fecha_desde}_hasta_{fecha_hasta}.xlsx'
            ruta=settings.DIR_EXPORT+f'Comisiones_{proyecto}_desde_{fecha_desde}_hasta_{fecha_hasta}.xlsx'
            _save_workbook_with_dirs(book, ruta)
            return FileResponse(open(ruta,'rb'),as_attachment=True,filename=name)
            """ context['alerta']=True
            context['titulo_alerta']='¡Listo!'
            context['mensaje']='Descarga el archivo Aqui'
            context['link']=True
            context['ruta_link']=f'{settings.MEDIA_URL}tmp/Comisiones_{proyecto}_desde_{fecha_desde}_hasta_{fecha_hasta}.xlsx' """
        if request.POST.get('btnParamRecibos'):
            nrocomprobante=request.POST.get('nrocomprobante')
            ctacapital=request.POST.get('ctacapital')
            ctaintcte=request.POST.get('ctaintcte')
            ctaintmora=request.POST.get('ctaintmora')
            obj_cuentas=consecutivos.objects.using(proyecto).get(documento='RC')
            obj_cuentas.comprobante_contable=nrocomprobante
            obj_cuentas.cuenta_capital=ctacapital
            obj_cuentas.cuenta_intcte=ctaintcte
            obj_cuentas.cuenta_inmora=ctaintmora
            obj_cuentas.save()
            context['cuentas']=obj_cuentas
        if request.POST.get('btnParamComis'):
            nrocomprobante=request.POST.get('nrocomprobante')
            ctagasto=request.POST.get('ctagasto')
            ctaprovision=request.POST.get('ctaprovision')
            ctaporpagar=request.POST.get('ctaporpagar')
            ctaanticipo=request.POST.get('cta_anticipo')
            obj_cuentas=consecutivos.objects.using(proyecto).filter(documento='COMISIONES')
            if obj_cuentas.exists():
                obj_cuentas=obj_cuentas[0]
                obj_cuentas.comprobante_contable=nrocomprobante
                obj_cuentas.cuenta_capital=ctagasto
                obj_cuentas.cuenta_intcte=ctaprovision
                obj_cuentas.cuenta_inmora=ctaporpagar
                obj_cuentas.cuenta_aux1=ctaanticipo
                obj_cuentas.save()
            else:
                consecutivos.objects.using(proyecto).create(id_consec=4,documento='COMISIONES',consecutivo=0,
                                                            comprobante_contable=nrocomprobante,cuenta_capital=ctagasto,
                                                            cuenta_intcte=ctaprovision,cuenta_inmora=ctaporpagar,cuenta_aux1=ctaanticipo)
                obj_cuentas = consecutivos.objects.using(proyecto).last()
            context['cuentas_comis']=obj_cuentas
    if request.method == 'GET':
        tipo = request.GET.get('tipo')
        fecha = request.GET.get('fecha')
        if tipo == 'distribucionrecaudos':
            book=openpyxl.Workbook()
            sheet = book.active
            encabezados=['Cedula_T1','Nombre_T1','Cedula_T2','Nombre_T2','Cedula_T3','Nombre_T3',
                         'Cedula_T4','Nombre_T4','Adjudicacion','Estado','Origen','Capital','Int Cte',
                         'Int Mora','Total']
            sheet.append(encabezados)
            i=2
            obj_recaudos = Adjudicacion.objects.using(proyecto).all()
            for item in obj_recaudos:
                titulares = item.titulares()
                recaudos = item.recaudo_detallado_interfaz(proyecto=proyecto,date_until=fecha)
                sheet.cell(i,1,titulares.get('titular_1').pk)
                sheet.cell(i,2,titulares.get('titular_1').nombrecompleto)
                sheet.cell(i,3,titulares.get('titular_2').pk)
                sheet.cell(i,4,titulares.get('titular_2').nombrecompleto)
                sheet.cell(i,5,titulares.get('titular_3').pk)
                sheet.cell(i,6,titulares.get('titular_3').nombrecompleto)
                sheet.cell(i,7,titulares.get('titular_4').pk)
                sheet.cell(i,8,titulares.get('titular_4').nombrecompleto)
                sheet.cell(i,9,item.pk)
                sheet.cell(i,10,item.estado)
                sheet.cell(i,11,item.origenventa)
                sheet.cell(i,12,recaudos.get('capital'))
                sheet.cell(i,13,recaudos.get('interescte'))
                sheet.cell(i,14,recaudos.get('interesmora'))
                sheet.cell(i,15,recaudos.get('total'))
                i+=1
            filename=f'Recaudo_detallado_{proyecto}.xlsx'
            ruta=settings.MEDIA_ROOT+'/tmp/'+filename
            _save_workbook_with_dirs(book, ruta)
            return JsonResponse({'url': settings.MEDIA_URL + 'tmp/' + filename})
         
         
    return render(request,'interfaces_contabilidad.html',context)

def reporte_cartera(request,proyecto,año):
    book=openpyxl.load_workbook("resources/excel_formats/Reporte de cartera Anual.xlsx")
    sheet=book.get_sheet_by_name('Resumen x Periodo')
    agrupador={}
    for mes in range(1,13):
        contenido_ppto=Adjudicacion.objects.using(proyecto).raw(f'CALL informe_cartera("{año}{mes:02d}",NULL)')
        for fila in contenido_ppto:
            grupo=str(mes)+'-'+fila.tipocartera+'-'+fila.edad
            if grupo not in agrupador:
                agrupador.update({
                    grupo:{
                        'año':int(año),
                        'mes':mes,
                        'edad':fila.edad,
                        'tipocartera':fila.tipocartera,
                        'nroclientes':1,
                        'pptomes':fila.ppto_mes,
                        'rcdomes':fila.recaudo_mes,
                        'pptovenc':fila.ppto_vencido,
                        'rcdovenc':fila.recaudo_vencido,
                        'totalppto':fila.presupuesto,
                        'rcdopptado':fila.recaudo_pptado,
                        'rcdonopptado':fila.recaudo_nopptado,
                        'rcdototal':fila.recaudo_total,
                        'cumplimiento':fila.recaudo_pptado/fila.presupuesto
                    }
                })
            else:
                agrupador[grupo]['pptomes']+=fila.ppto_mes
                agrupador[grupo]['rcdomes']+=fila.recaudo_mes
                agrupador[grupo]['pptovenc']+=fila.ppto_vencido
                agrupador[grupo]['rcdovenc']+=fila.recaudo_vencido
                agrupador[grupo]['totalppto']+=fila.presupuesto
                agrupador[grupo]['rcdopptado']+=fila.recaudo_pptado
                agrupador[grupo]['rcdonopptado']+=fila.recaudo_nopptado
                agrupador[grupo]['rcdototal']+=fila.recaudo_total
                agrupador[grupo]['nroclientes']+=1
    row=2
    for line in agrupador.values():
        sheet.cell(row,1,line['año'])
        sheet.cell(row,2,line['mes'])
        sheet.cell(row,3,line['edad'])
        sheet.cell(row,4,line['tipocartera'])
        sheet.cell(row,5,line['pptomes'])
        sheet.cell(row,6,line['rcdomes'])
        sheet.cell(row,7,line['pptovenc'])
        sheet.cell(row,8,line['rcdovenc'])
        sheet.cell(row,9,line['totalppto'])
        sheet.cell(row,10,line['rcdopptado'])
        sheet.cell(row,11,line['rcdonopptado'])
        sheet.cell(row,12,line['rcdototal'])
        row+=1
        
    ruta=settings.DIR_EXPORT+f'prueba.xlsx'
    _save_workbook_with_dirs(book, ruta)
    file=open(ruta,'rb')
    response=FileResponse(file,as_attachment=True,filename='prueba.xlsx')
    return response

@group_perm_required(('andinasoft.change_descuentoscondicionados',),raise_exception=True)
def descuentos_condicionados(request,proyecto):
    stmt=f'CALL verificar_descuento()'
    obj_descuentos=DescuentosCondicionados.objects.using(proyecto).raw(stmt)
    context={
        'proyecto':proyecto,
        'descuentos':obj_descuentos,
    }
    return render(request,'descuentos_condicionados.html',context)

@group_perm_required(('andinasoft.view_promesas',),raise_exception=True)
def promesas(request,proyecto):
    alerta=False
    titulo_alerta=None
    mensaje_alerta=None
    link=False
    ruta_link=None
    pdf=GenerarPDF()
    impresion=False
    if request.method == 'POST':
        if not request.is_ajax():
            check_perms(request,('andinasoft.add_promesas',),raise_exception=True)
            if request.POST.get('btnOpciones'):
                adj=request.POST.get('adjopcion')
                tipo_fecha=request.POST.get('tipofechas')
                ciudad=request.POST.get('oficinaopcion')
                fecha_promesa = request.POST.get('fechapromesa')
                fecha_entrega=request.POST.get('fechaentrega')
                fecha_escritura=request.POST.get('fechaescritura')
                formaci = request.POST.get('formaci')
                formasaldo = request.POST.get('formasaldo')
                obs = request.POST.get('observaciones')
                formapago = request.POST.get('formapago')
                dias_prorroga = request.POST.get('prorroga')
                diacontrato=datetime.date.today().day
                mescontrato=datetime.date.today().month
                añocontrato=datetime.date.today().year
                impresion=True
            elif request.POST.get('EstadoPromesa'):
                adj=request.POST.get('adjOpcionPromesa')
                escriturado = request.POST.get('escriturado')
                if escriturado=='on': escriturado=True
                else: escriturado=False
                entregado = request.POST.get('entregado')
                if entregado=='on': entregado=True
                else: entregado=False
                prorrogado = request.POST.get('prorrogado')
                if prorrogado=='on': prorrogado=True
                else: prorrogado=False
                obj_promesa=Promesas.objects.using(proyecto).get(idadjudicacion=adj)
                obj_promesa.escriturado=escriturado
                obj_promesa.entregado=entregado
                obj_promesa.prorroga=prorrogado
                obj_promesa.save()
            elif request.POST.get('btnAprobar'):
                if check_perms(request,('andinasoft.delete_promesas',),raise_exception=False):
                    adj=request.POST.get('adjpromesa')
                    promesa=Promesas.objects.using(proyecto).get(idadjudicacion=adj)
                    obj_adj=Adjudicacion.objects.using(proyecto).get(idadjudicacion=adj)
                    obj_adj.tipocontrato='Promesa'
                    obj_adj.contrato=promesa.nropromesa
                    obj_adj.save()
                    promesa.estado='Aprobado'
                    promesa.usuarioaprueba=str(request.user)
                    promesa.fechaaprueba=datetime.date.today()
                    promesa.save()
                    obj_timeline=timeline.objects.using(proyecto).create(adj=adj,fecha=datetime.date.today(),usuario=request.user,accion='Aprobó promesa de compraventa')
                    alerta=True
                    titulo_alerta='¡Todo salio Perfecto!'
                    mensaje_alerta=f'La promesa del {adj} fue aprobada y ya se encuentra en modulo de promesas vigentes'
                else: raise PermissionDenied
            else:
                adj=request.POST.get('adjpromesa')
                tipo_fecha='Reimpresion'
                data_promesa=Promesas.objects.using(proyecto).get(idadjudicacion=adj)
                fecha_entrega=data_promesa.fechaentrega
                fecha_escritura=data_promesa.fechaescritura
                formaci=data_promesa.formaci
                formasaldo=data_promesa.formasaldo
                ciudad=data_promesa.ciudad
                fecha_prom=data_promesa.fechapromesa
                diacontrato=fecha_prom.day
                mescontrato=fecha_prom.month
                añocontrato=fecha_prom.year
                obs=data_promesa.observaciones
                if obs is None: obs=''
                formapago=data_promesa.formapago
                impresion=True
            if impresion:
                #objects
                obj_consecprom=consecutivos.objects.using(proyecto).get(documento='PROMESA')
                obj_adj=Adjudicacion.objects.using(proyecto).get(idadjudicacion=adj)
                obj_saldos=saldos_adj.objects.using(proyecto).filter(adj=adj)
                obj_promesas=Promesas.objects.using(proyecto).filter(idadjudicacion=adj)
                obj_planpagos=PlanPagos.objects.using(proyecto).filter(adj=adj)
                obj_inmueble=Inmuebles.objects.using(proyecto).get(idinmueble=obj_adj.idinmueble)
                if tipo_fecha=='Automatico':
                    fecha_entrega=datetime.date(obj_inmueble.finobra.year,obj_inmueble.finobra.month,obj_inmueble.finobra.day)
                    fin_pagos=obj_planpagos.aggregate(Max('fecha'))['fecha__max']
                    if fin_pagos<=fecha_entrega:
                        fecha_escritura=fecha_entrega+relativedelta(fecha_entrega,months=1)
                    elif fin_pagos>fecha_entrega:
                        fecha_escritura=fin_pagos+relativedelta(fin_pagos,months=1)
                elif tipo_fecha=='Manual':
                    fecha_escritura=datetime.datetime.strptime(fecha_escritura,"%Y-%m-%d")
                    fecha_entrega=datetime.datetime.strptime(fecha_entrega,"%Y-%m-%d")
                    
                #titulares
                titular1=clientes.objects.filter(idTercero=obj_adj.idtercero1)
                if titular1.exists(): titular1=titular1[0]
                else: titular1=clientes.objects.get(idTercero="")
                titular2=clientes.objects.filter(idTercero=obj_adj.idtercero2)
                if titular2.exists(): titular2=titular2[0]
                else: titular2=clientes.objects.get(idTercero="")
                titular3=clientes.objects.filter(idTercero=obj_adj.idtercero3)
                if titular3.exists(): titular3=titular3[0]
                else: titular3=clientes.objects.get(idTercero="")
                titular4=clientes.objects.filter(idTercero=obj_adj.idtercero4)
                if titular4.exists(): titular4=titular4[0]
                else: titular4=clientes.objects.get(idTercero="")
                #variables
                valor_letras=Utilidades().numeros_letras(obj_adj.valor)
                cuota_inicial=obj_planpagos.filter(tipocta='CI').aggregate(Sum('capital'))['capital__sum']
                if cuota_inicial==None: cuota_inicial=0
                saldo=obj_planpagos.exclude(tipocta='CI').aggregate(Sum('capital'))['capital__sum']
                if saldo==None: saldo=0
                if formapago=='Contado':
                    forma=('x','','')
                elif formapago=='Credicontado':
                    forma=('','x','')
                elif formapago=='Amortizacion':
                    forma=('','','x')
                
                
                if request.POST.get('btnReimpPromesa') or request.POST.get('btnOpciones'):
                    ruta=settings.DIR_EXPORT+f'Promesa_{proyecto}_{adj}.pdf'
                    alerta=True
                    titulo_alerta='¡Todo salio Perfecto!'
                    mensaje_alerta='Puedes previsualizar la promesa aquí'
                    link=True
                    ruta_link=settings.DIR_DOWNLOADS+f'Promesa_{proyecto}_{adj}.pdf'
                    if request.POST.get('btnOpciones'):
                        if obj_promesas.exists():
                            promesa=Promesas.objects.using(proyecto).get(idadjudicacion=adj)
                            promesa.ciudad=ciudad
                            promesa.fechapromesa = fecha_promesa
                            promesa.fechaentrega=fecha_entrega
                            promesa.fechaescritura=fecha_escritura
                            promesa.formaci=formaci
                            promesa.formasaldo=formasaldo
                            promesa.observaciones=obs
                            promesa.formapago=formapago
                            promesa.prorroga=False
                            promesa.dias_prorroga=dias_prorroga
                            try: 
                                consec_promesa=int(promesa.nropromesa)
                            except ValueError:
                                promesa.nropromesa = obj_consecprom.consecutivo
                                consec_promesa=obj_consecprom.consecutivo
                                obj_consecprom.consecutivo+=1
                                obj_consecprom.save()
                            promesa.save()
                            
                        else:
                            consec_promesa=obj_consecprom.consecutivo
                            promesa = Promesas.objects.using(proyecto).create(idadjudicacion=adj,formapago=formapago,estado='Pendiente',fechaentrega=fecha_entrega,
                                                    fechaescritura=fecha_escritura,fechapromesa=datetime.date.today(),nropromesa=consec_promesa,
                                                    formaci=formaci,formasaldo=formasaldo,usuariocrea=request.user,observaciones=obs,ciudad=ciudad,dias_prorroga=dias_prorroga)
                            obj_consecprom.consecutivo+=1
                            obj_consecprom.save()
                            obj_timeline=timeline.objects.using(proyecto).create(adj=adj,fecha=datetime.date.today(),usuario=request.user,accion='Creó promesa de compraventa')
                    else:
                        promesa=Promesas.objects.using(proyecto).get(idadjudicacion=adj)
                        dias_prorroga=promesa.dias_prorroga
                        consec_promesa=promesa.nropromesa
                        
                    if proyecto=='Sandville Beach':
                        GenerarPDF().ExportPromesaSandvilleBeach(nro_contrato=consec_promesa,
                        nombre_t1=titular1.nombrecompleto,cc_t1=titular1.idTercero,tel_t1=titular1.telefono1,cel_t1=titular1.celular1,ofic_t1=titular1.oficina,cdof_t1=titular1.ciudad,
                        telof_t1=titular1.telefono2,resid_t1=titular1.domicilio,cdresid_t1=titular1.ciudad,telresid_t1=titular1.telefono1,email_t1=titular1.email,
                        nombre_t2=titular2.nombrecompleto,cc_t2=titular2.idTercero,tel_t2=titular2.telefono1,cel_t2=titular2.celular1,ofic_t2=titular2.oficina,cdof_t2=titular2.ciudad,
                        telof_t2=titular2.telefono2,resid_t2=titular2.domicilio,cdresid_t2=titular2.ciudad,telresid_t2=titular2.telefono1,email_t2=titular2.email,
                        nombre_t3=titular3.nombrecompleto,cc_t3=titular3.idTercero,tel_t3=titular3.telefono1,cel_t3=titular3.celular1,ofic_t3=titular3.oficina,cdof_t3=titular3.ciudad,
                        telof_t3=titular3.telefono2,resid_t3=titular3.domicilio,cdresid_t3=titular3.ciudad,telresid_t3=titular3.telefono1,email_t3=titular3.email,
                        nombre_t4=titular4.nombrecompleto,cc_t4=titular4.idTercero,tel_t4=titular4.telefono1,cel_t4=titular4.celular1,ofic_t4=titular4.oficina,cdof_t4=titular4.ciudad,
                        telof_t4=titular4.telefono2,resid_t4=titular4.domicilio,cdresid_t4=titular4.ciudad,telresid_t4=titular4.telefono1,email_t4=titular4.email,
                        lote=obj_inmueble.lotenumero,manzana=obj_inmueble.manzananumero,area=str(obj_inmueble.areaprivada),
                        mtsnorte=str(obj_inmueble.norte),colnorte=obj_inmueble.colindante_norte,mtseste=str(obj_inmueble.este),coleste=obj_inmueble.colidante_este,
                        mtssur=str(obj_inmueble.sur),colsur=obj_inmueble.colindante_sur,mtsoeste=str(obj_inmueble.oeste),coloeste=obj_inmueble.colindante_oeste,
                        valor=int(obj_adj.valor),valor_letras=valor_letras,ci=int(cuota_inicial),saldo=int(saldo),contado_x=forma[0],credic_x=forma[1],amort_x=forma[2],
                        formaCI=formaci,formaFN=formasaldo,obs=obs,dia_contrato=str(diacontrato),mes_contrato=mescontrato,año_contrato=str(añocontrato),
                        fecha_entrega=fecha_entrega,fecha_escritura=fecha_escritura,ciudad_entrega=ciudad,ruta=ruta)
                        
                        filename = f'Promesa_{proyecto}_{adj}.pdf'
                        
                    elif proyecto=='Perla del Mar':
                        print('entra')
                        condominios = {
                            '1':'Alameda',
                            '2':'Caracola',
                        }
                        nro_condominio = obj_inmueble.etapa
                        nombre_condominio = condominios[nro_condominio]
                        
                        try:
                            meses_entrega = math.ceil((promesa.fechaentrega - datetime.date.today()).days / 30)
                        except:
                            meses_entrega = math.ceil((promesa.fechaentrega - datetime.datetime.today()).days / 30)
                        
                        context = {
                            'es_promesa':True,
                            'proyecto':proyecto,
                            'ctr':promesa,
                            'fecha_escritura':fecha_escritura,
                            'meses_entrega':meses_entrega,
                            'oficina':ciudad
                        }
                        
                        filename = f'Contrato_bien_futuro_{adj}_{proyecto}.pdf'
                        
                            
                        pdf = pdf_gen(f'pdf/{proyecto}/contrato.html', context, filename)
                        
                        ruta = pdf.get('root')
                                                
                        ruta_link=pdf.get('url')
                        
                        
                        """ GenerarPDF().ExportPromesaSandvilleMar(nro_contrato=consec_promesa,
                        nombre_t1=titular1.nombrecompleto,cc_t1=titular1.idTercero,tel_t1=titular1.telefono1,cel_t1=titular1.celular1,ofic_t1=titular1.oficina,cdof_t1=titular1.ciudad,
                        telof_t1=titular1.telefono2,resid_t1=titular1.domicilio,cdresid_t1=titular1.ciudad,telresid_t1=titular1.telefono1,email_t1=titular1.email,
                        nombre_t2=titular2.nombrecompleto,cc_t2=titular2.idTercero,tel_t2=titular2.telefono1,cel_t2=titular2.celular1,ofic_t2=titular2.oficina,cdof_t2=titular2.ciudad,
                        telof_t2=titular2.telefono2,resid_t2=titular2.domicilio,cdresid_t2=titular2.ciudad,telresid_t2=titular2.telefono1,email_t2=titular2.email,
                        nombre_t3=titular3.nombrecompleto,cc_t3=titular3.idTercero,tel_t3=titular3.telefono1,cel_t3=titular3.celular1,ofic_t3=titular3.oficina,cdof_t3=titular3.ciudad,
                        telof_t3=titular3.telefono2,resid_t3=titular3.domicilio,cdresid_t3=titular3.ciudad,telresid_t3=titular3.telefono1,email_t3=titular3.email,
                        nombre_t4=titular4.nombrecompleto,cc_t4=titular4.idTercero,tel_t4=titular4.telefono1,cel_t4=titular4.celular1,ofic_t4=titular4.oficina,cdof_t4=titular4.ciudad,
                        telof_t4=titular4.telefono2,resid_t4=titular4.domicilio,cdresid_t4=titular4.ciudad,telresid_t4=titular4.telefono1,email_t4=titular4.email,
                        lote=obj_inmueble.lotenumero,manzana=obj_inmueble.manzananumero,area=str(obj_inmueble.areaprivada),
                        mtsnorte=str(obj_inmueble.norte),colnorte=obj_inmueble.colindante_norte,mtseste=str(obj_inmueble.este),coleste=obj_inmueble.colidante_este,
                        mtssur=str(obj_inmueble.sur),colsur=obj_inmueble.colindante_sur,mtsoeste=str(obj_inmueble.oeste),coloeste=obj_inmueble.colindante_oeste,
                        valor=int(obj_adj.valor),valor_letras=valor_letras,ci=int(cuota_inicial),saldo=int(saldo),contado_x=forma[0],credic_x=forma[1],amort_x=forma[2],
                        formaCI=formaci,formaFN=formasaldo,obs=obs,dia_contrato=str(diacontrato),mes_contrato=mescontrato,año_contrato=str(añocontrato),
                        fecha_entrega=fecha_entrega,fecha_escritura=fecha_escritura,ciudad_entrega=ciudad,ruta=ruta,nro_condominio=nro_condominio,nombre_con=nombre_condominio,prorroga=dias_prorroga) """
                    
                    elif proyecto=='Tesoro Escondido':
                        porcDerecho=f'{(obj_inmueble.areaprivada*100/obj_inmueble.area_mz):.2f}'
                        GenerarPDF().ExportPromesaBugambilias(nro_contrato=consec_promesa,
                        nombre_t1=titular1.nombrecompleto,cc_t1=titular1.idTercero,tel_t1=titular1.telefono1,cel_t1=titular1.celular1,ofic_t1=titular1.oficina,cdof_t1=titular1.ciudad,
                        telof_t1=titular1.telefono2,resid_t1=titular1.domicilio,cdresid_t1=titular1.ciudad,telresid_t1=titular1.telefono1,email_t1=titular1.email,
                        nombre_t2=titular2.nombrecompleto,cc_t2=titular2.idTercero,tel_t2=titular2.telefono1,cel_t2=titular2.celular1,ofic_t2=titular2.oficina,cdof_t2=titular2.ciudad,
                        telof_t2=titular2.telefono2,resid_t2=titular2.domicilio,cdresid_t2=titular2.ciudad,telresid_t2=titular2.telefono1,email_t2=titular2.email,
                        nombre_t3=titular3.nombrecompleto,cc_t3=titular3.idTercero,tel_t3=titular3.telefono1,cel_t3=titular3.celular1,ofic_t3=titular3.oficina,cdof_t3=titular3.ciudad,
                        telof_t3=titular3.telefono2,resid_t3=titular3.domicilio,cdresid_t3=titular3.ciudad,telresid_t3=titular3.telefono1,email_t3=titular3.email,
                        nombre_t4=titular4.nombrecompleto,cc_t4=titular4.idTercero,tel_t4=titular4.telefono1,cel_t4=titular4.celular1,ofic_t4=titular4.oficina,cdof_t4=titular4.ciudad,
                        telof_t4=titular4.telefono2,resid_t4=titular4.domicilio,cdresid_t4=titular4.ciudad,telresid_t4=titular4.telefono1,email_t4=titular4.email,
                        lote=obj_inmueble.lotenumero,manzana=obj_inmueble.manzananumero,area=str(obj_inmueble.areaprivada),
                        mtsnorte=str(obj_inmueble.norte),colnorte=obj_inmueble.colindante_norte,mtseste=str(obj_inmueble.este),coleste=obj_inmueble.colidante_este,
                        mtssur=str(obj_inmueble.sur),colsur=obj_inmueble.colindante_sur,mtsoeste=str(obj_inmueble.oeste),coloeste=obj_inmueble.colindante_oeste,
                        valor=int(obj_adj.valor),valor_letras=valor_letras,ci=int(cuota_inicial),saldo=int(saldo),contado_x=forma[0],credic_x=forma[1],amort_x=forma[2],
                        formaCI=formaci,formaFN=formasaldo,obs=obs,dia_contrato=str(diacontrato),mes_contrato=mescontrato,año_contrato=str(añocontrato),
                        fecha_entrega=fecha_entrega,fecha_escritura=fecha_escritura,ciudad_entrega=ciudad,ruta=ruta,porcderecho=porcDerecho,area_parcela=str(obj_inmueble.area_mz))
                        ruta_link=ruta
                        
                        filename = f'Promesa_{proyecto}_{adj}.pdf'
                        
                    elif proyecto=='Vegas de Venecia':
                        try:
                            meses_entrega = math.ceil((promesa.fechaentrega - datetime.date.today()).days / 30)
                        except:
                            meses_entrega = math.ceil((promesa.fechaentrega - datetime.datetime.today()).days / 30)
                        GenerarPDF().ExportCBFVegasVenecia(nro_contrato=consec_promesa,
                        nombre_t1=titular1.nombrecompleto,cc_t1=titular1.idTercero,tel_t1=titular1.telefono1,cel_t1=titular1.celular1,ofic_t1=titular1.oficina,cdof_t1=titular1.ciudad,
                        telof_t1=titular1.telefono2,resid_t1=titular1.domicilio,cdresid_t1=titular1.ciudad,telresid_t1=titular1.telefono1,email_t1=titular1.email,
                        nombre_t2=titular2.nombrecompleto,cc_t2=titular2.idTercero,tel_t2=titular2.telefono1,cel_t2=titular2.celular1,ofic_t2=titular2.oficina,cdof_t2=titular2.ciudad,
                        telof_t2=titular2.telefono2,resid_t2=titular2.domicilio,cdresid_t2=titular2.ciudad,telresid_t2=titular2.telefono1,email_t2=titular2.email,
                        nombre_t3=titular3.nombrecompleto,cc_t3=titular3.idTercero,tel_t3=titular3.telefono1,cel_t3=titular3.celular1,ofic_t3=titular3.oficina,cdof_t3=titular3.ciudad,
                        telof_t3=titular3.telefono2,resid_t3=titular3.domicilio,cdresid_t3=titular3.ciudad,telresid_t3=titular3.telefono1,email_t3=titular3.email,
                        nombre_t4=titular4.nombrecompleto,cc_t4=titular4.idTercero,tel_t4=titular4.telefono1,cel_t4=titular4.celular1,ofic_t4=titular4.oficina,cdof_t4=titular4.ciudad,
                        telof_t4=titular4.telefono2,resid_t4=titular4.domicilio,cdresid_t4=titular4.ciudad,telresid_t4=titular4.telefono1,email_t4=titular4.email,
                        lote=obj_inmueble.lotenumero,manzana=obj_inmueble.manzananumero,area=str(obj_inmueble.areaprivada),
                        mtsnorte=str(obj_inmueble.norte),colnorte=obj_inmueble.colindante_norte,mtseste=str(obj_inmueble.este),coleste=obj_inmueble.colidante_este,
                        mtssur=str(obj_inmueble.sur),colsur=obj_inmueble.colindante_sur,mtsoeste=str(obj_inmueble.oeste),coloeste=obj_inmueble.colindante_oeste,
                        valor=int(obj_adj.valor),valor_letras=valor_letras,ci=int(cuota_inicial),saldo=int(saldo),contado_x=forma[0],credic_x=forma[1],amort_x=forma[2],
                        formaCI=formaci,formaFN=formasaldo,obs=obs,dia_contrato=str(diacontrato),mes_contrato=mescontrato,año_contrato=str(añocontrato),
                        fecha_entrega=fecha_entrega,fecha_escritura=fecha_escritura,ciudad_entrega=ciudad,ruta=ruta,meses_entrega=str(meses_entrega))
                        ruta_link=ruta
                        
                        filename = f'Promesa_{proyecto}_{adj}.pdf'
                        
                    elif proyecto=='Sandville del Sol':
                        GenerarPDF().ExportPromesaSandvilleMar(nro_contrato=consec_promesa,
                        nombre_t1=titular1.nombrecompleto,cc_t1=titular1.idTercero,tel_t1=titular1.telefono1,cel_t1=titular1.celular1,ofic_t1=titular1.oficina,cdof_t1=titular1.ciudad,
                        telof_t1=titular1.telefono2,resid_t1=titular1.domicilio,cdresid_t1=titular1.ciudad,telresid_t1=titular1.telefono1,email_t1=titular1.email,
                        nombre_t2=titular2.nombrecompleto,cc_t2=titular2.idTercero,tel_t2=titular2.telefono1,cel_t2=titular2.celular1,ofic_t2=titular2.oficina,cdof_t2=titular2.ciudad,
                        telof_t2=titular2.telefono2,resid_t2=titular2.domicilio,cdresid_t2=titular2.ciudad,telresid_t2=titular2.telefono1,email_t2=titular2.email,
                        nombre_t3=titular3.nombrecompleto,cc_t3=titular3.idTercero,tel_t3=titular3.telefono1,cel_t3=titular3.celular1,ofic_t3=titular3.oficina,cdof_t3=titular3.ciudad,
                        telof_t3=titular3.telefono2,resid_t3=titular3.domicilio,cdresid_t3=titular3.ciudad,telresid_t3=titular3.telefono1,email_t3=titular3.email,
                        nombre_t4=titular4.nombrecompleto,cc_t4=titular4.idTercero,tel_t4=titular4.telefono1,cel_t4=titular4.celular1,ofic_t4=titular4.oficina,cdof_t4=titular4.ciudad,
                        telof_t4=titular4.telefono2,resid_t4=titular4.domicilio,cdresid_t4=titular4.ciudad,telresid_t4=titular4.telefono1,email_t4=titular4.email,
                        lote=obj_inmueble.lotenumero,manzana=obj_inmueble.manzananumero,area=str(obj_inmueble.areaprivada),
                        mtsnorte=str(obj_inmueble.norte),colnorte=obj_inmueble.colindante_norte,mtseste=str(obj_inmueble.este),coleste=obj_inmueble.colidante_este,
                        mtssur=str(obj_inmueble.sur),colsur=obj_inmueble.colindante_sur,mtsoeste=str(obj_inmueble.oeste),coloeste=obj_inmueble.colindante_oeste,
                        valor=int(obj_adj.valor),valor_letras=valor_letras,ci=int(cuota_inicial),saldo=int(saldo),contado_x=forma[0],credic_x=forma[1],amort_x=forma[2],
                        formaCI=formaci,formaFN=formasaldo,obs=obs,dia_contrato=str(diacontrato),mes_contrato=mescontrato,año_contrato=str(añocontrato),
                        fecha_entrega=fecha_entrega,fecha_escritura=fecha_escritura,ciudad_entrega=ciudad,ruta=ruta)

                        filename = f'Promesa_{proyecto}_{adj}.pdf'
                        
                    elif proyecto == 'Sotavento' or proyecto == 'Oasis':
                        try:
                            meses_entrega = math.ceil((promesa.fechaentrega - datetime.date.today()).days / 30)
                        except:
                            meses_entrega = math.ceil((promesa.fechaentrega - datetime.datetime.today()).days / 30)
                        
                        context = {
                            'es_promesa':True,
                            'proyecto':proyecto,
                            'ctr':promesa,
                            'fecha_escritura':fecha_escritura,
                            'meses_entrega':meses_entrega,
                            'oficina':ciudad
                        }
                        
                        filename = f'Contrato_bien_futuro_{adj}_{proyecto}.pdf'
                        
                            
                        if proyecto == 'Oasis':
                            pdf = pdf_gen_weasy(f'pdf/{proyecto}/contrato.html', context, filename)
                        else:
                            pdf = pdf_gen(f'pdf/{proyecto}/contrato.html', context, filename)
                        
                        file = pdf.get('root')
                                                
                        ruta_link=pdf.get('url')
                    
                    return FileResponse(open(ruta,'rb'),as_attachment=True,filename=filename)
                    
                elif request.POST.get('btnReimpPagare'):
                    contrato=obj_promesas[0].nropromesa
                    nombre_t1=titular1.nombrecompleto
                    cc_t1=titular1.idTercero
                    nombre_t2=titular2.nombrecompleto
                    cc_t2=titular2.idTercero
                    nombre_t3=titular3.nombrecompleto
                    cc_t3=titular3.idTercero
                    nombre_t4=titular4.nombrecompleto
                    cc_t4=titular4.idTercero
                    ruta=settings.DIR_EXPORT+f'{proyecto}_pagare_{adj}.pdf'
                    if proyecto=='Perla del Mar':
                        pdf.PagareSandvilleMar(nroPagare=contrato,
                                                nombreT1=nombre_t1,ccT1=cc_t1,nombreT2=nombre_t2,ccT2=cc_t2,
                                                nombreT3=nombre_t3,ccT3=cc_t3,nombreT4=nombre_t4,ccT4=cc_t4,
                                                diaPagare=str(diacontrato),mesPagare=str(mescontrato),añoPagare=str(añocontrato),ciudad=ciudad,ruta=ruta)
                    elif proyecto=='Sandville Beach':
                        pdf.PagareSandvilleBeach(nroPagare=contrato,
                                                nombreT1=nombre_t1,ccT1=cc_t1,nombreT2=nombre_t2,ccT2=cc_t2,
                                                nombreT3=nombre_t3,ccT3=cc_t3,nombreT4=nombre_t4,ccT4=cc_t4,
                                                diaPagare=str(diacontrato),mesPagare=str(mescontrato),añoPagare=str(añocontrato),ciudad=ciudad,ruta=ruta)
                    elif proyecto=='Sandville del Sol':
                        pdf.PagareSandvilleBeach(nroPagare=contrato,
                                                nombreT1=nombre_t1,ccT1=cc_t1,nombreT2=nombre_t2,ccT2=cc_t2,
                                                nombreT3=nombre_t3,ccT3=cc_t3,nombreT4=nombre_t4,ccT4=cc_t4,
                                                diaPagare=str(diacontrato),mesPagare=str(mescontrato),añoPagare=str(añocontrato),ciudad=ciudad,ruta=ruta)
                    elif proyecto=='Tesoro Escondido':
                        pdf.PagareTesoro(nroPagare=contrato,
                                                nombreT1=nombre_t1,ccT1=cc_t1,nombreT2=nombre_t2,ccT2=cc_t2,
                                                nombreT3=nombre_t3,ccT3=cc_t3,nombreT4=nombre_t4,ccT4=cc_t4,
                                                diaPagare=str(diacontrato),mesPagare=str(mescontrato),añoPagare=str(añocontrato),ciudad=ciudad,ruta=ruta)
                    
                    alerta=True
                    mensaje_alerta='Descarga el Pagaré aqui'
                    titulo_alerta='¡Listo!'
                    link=True
                    ruta_link=settings.DIR_DOWNLOADS+f'{proyecto}_pagare_{adj}.pdf'

    if request.is_ajax():
        if request.method == 'GET':
            adj=request.GET.get('adj')
            obj_promesa=Promesas.objects.using(proyecto).filter(idadjudicacion=adj)
            escriturado = obj_promesa[0].escriturado
            entregado = obj_promesa[0].entregado
            prorrogado = obj_promesa[0].prorroga
            data={
                'escriturado':escriturado,
                'entregado':entregado,
                'prorrogado':prorrogado,
                'datos_promesa':serializers.serialize('json',obj_promesa)
            }
            return JsonResponse(data)
        
        if request.method == 'POST':
            tipo = request.POST.get('tipo')
            check_perms(request,('andinasoft.change_promesas',),raise_exception=True)
            adj = request.POST.get('adj')
            obj_promesa = Promesas.objects.using(proyecto).get(idadjudicacion=adj)
            
            if tipo == 'solofecha':
                obj_promesa.fechapromesa = request.POST.get('fecha_promesa')
                obj_promesa.fechaentrega = request.POST.get('fecha_entrega')
                obj_promesa.fechaescritura = request.POST.get('fecha_escritura')
                accion=f'Cambio las fechas de la promesa de compraventa'
                
            else:
                obj_promesa.estado = 'Pendiente'
                accion=f'Desaprobó la promesa de compraventa'
                
            obj_promesa.save()
            obj_timeline=timeline.objects.using(proyecto)
            obj_timeline.create(adj=adj,
                            fecha=datetime.date.today(),
                            usuario=request.user,
                            accion=accion)
            data = {
                'success':'si'
            }
            return JsonResponse(data)
        
        
    obj_opciones_vigentes=Adjudicacion.objects.using(proyecto).raw('CALL opciones_vigentes()')
    obj_opciones_vencidas=Adjudicacion.objects.using(proyecto).raw('CALL opciones_vencidas()')
    obj_opciones_porvencer=Adjudicacion.objects.using(proyecto).raw('CALL opciones_por_vencer()')
    obj_promesas_vigentes=Promesas.objects.using(proyecto).raw('CALL promesas_vigentes()')
    obj_entregas_vencidas=Promesas.objects.using(proyecto).raw('CALL entregas_vencidas()')
    obj_escrituras_vencidas=Promesas.objects.using(proyecto).raw('CALL escrituras_vencidas()')
    obj_entregas_porvencer=Promesas.objects.using(proyecto).raw('CALL entregas_por_vencer()')
    obj_escrituras_porvencer=Promesas.objects.using(proyecto).raw('CALL escrituras_por_vencer()')
    obj_promesas_pendientes=Promesas.objects.using(proyecto).raw('CALL promesas_sinaprobar()')
    context={
        'proyecto':proyecto,
        'opciones_vigentes':obj_opciones_vigentes,
        'opciones_vencidas':obj_opciones_vencidas,
        'promesas_vigentes':obj_promesas_vigentes,
        'entregas_vencidas':obj_entregas_vencidas,
        'escrituras_vencidas':obj_escrituras_vencidas,
        'opciones_porvencer':obj_opciones_porvencer,
        'entregas_porvencer':obj_entregas_porvencer,
        'escrituras_porvencer':obj_escrituras_porvencer,
        'promesas_pendientes':obj_promesas_pendientes,
        'alerta':alerta,
        'titulo_alerta':titulo_alerta,
        'mensaje':mensaje_alerta,
        'link':link,
        'ruta_link':ruta_link,
    }
    
    return render(request,'estados_promesas.html',context)

@group_perm_required(('andinasoft.view_pqrs',),raise_exception=True)
def lista_pqrs(request,proyecto):
    check_project(request,proyecto)
    context = {
        'pqrs':Pqrs.objects.using(proyecto).all(),
        'proyecto':proyecto,
        'today':datetime.date.today(),
    }
    if request.method == 'POST':
        check_perms(request,('andinasoft.change_pqrs',))
        radicado = request.POST.get('idPQRS')
        adj = request.POST.get('adjPQRS')
        fecha_rta = request.POST.get('fecha_respuesta')
        medio_rta = request.POST.get('forma_respuesta')
        soporte_rta = request.FILES.get('respuesta')
        soporte_envio = request.FILES.get('envio')
        obj_radicado = Pqrs.objects.using(proyecto).get(pk=radicado)
        archivo_rta = f'Respuesta_{obj_radicado.tipo}_{datetime.datetime.now()}'.replace(':','.')
        archivo_envio = f'Constancia_Envio_Respuesta_{obj_radicado.tipo}_{datetime.datetime.now()}'.replace(':','.')
        
        obj_radicado.fecha_respuesta=fecha_rta
        obj_radicado.doc_respuesta=archivo_rta+'.pdf'
        obj_radicado.tipo_respuesta=medio_rta
        obj_radicado.doc_envio=archivo_envio+'.pdf'
        obj_radicado.estado='Cerrado'
        obj_radicado.usuario_cierra=str(request.user)
        obj_radicado.save()
        
        ruta =f'{settings.DIR_DOCS}/doc_contratos/{proyecto}/{adj}/'
        upload_docs(soporte_rta,ruta,archivo_rta)
        documentos_contratos.objects.using(proyecto).create(adj=adj,descripcion_doc=archivo_rta,
                                                            fecha_carga=datetime.date.today(),usuario_carga=str(request.user))
        upload_docs(soporte_envio,ruta,archivo_envio)
        documentos_contratos.objects.using(proyecto).create(adj=adj,descripcion_doc=archivo_envio,
                                                            fecha_carga=datetime.date.today(),usuario_carga=str(request.user))
        obj_timeline=timeline.objects.using(proyecto)
        tipo=obj_radicado.tipo
        accion=f'Cerró la {tipo} #{radicado}'
        obj_timeline.create(adj=adj,
                    fecha=datetime.date.today(),
                    usuario=request.user,
                    accion=accion)
        alerta=True
        titulo='PQRS Cerrada'
        mensaje=f'La {tipo} #{radicado} fué cerrada correctamente'
        
    return render(request,'lista_pqrs.html',context)

@group_perm_required(('andinasoft.view_infocartera',),raise_exception=True)
def informe_mes(request,proyecto):
    context={
    }
    if request.method == 'POST':
        año=request.POST.get('año')
        mes=request.POST.get('mes')
        desde=f'{año}-{mes}-01'
        last_day=calendar.monthrange(int(año),int(mes))[1]
        hasta=f'{año}-{mes}-{last_day}'
        periodo=f'{año}{int(mes):02d}'
        if mes=='12':
            año_sig=int(año)+1
            mes_sig=1
        else:
            año_sig=int(año)
            mes_sig=int(mes)+1
        periodo_sig=f'{año_sig}{mes_sig:02d}'
        if request.POST.get('resultadosmes'):
            obj_ventas=Adjudicacion.objects.using(proyecto).raw(f'CALL informe_ventas("{desde}","{hasta}")')
            obj_detalleventas=Adjudicacion.objects.using(proyecto).raw(f'CALL detalle_ventas("{desde}","{hasta}")')
            obj_recaudos=Adjudicacion.objects.using(proyecto).raw(f'CALL recaudo_mes("{desde}","{hasta}")')
            obj_cartera=Adjudicacion.objects.using(proyecto).raw(f'CALL detalle_cartera("{periodo}")')
            obj_gastosMes=GastosInforme.objects.filter(proyecto=proyecto,fecha__gte=desde,fecha__lte=hasta)
            obj_gastosAcumulados=GastosInforme.objects.filter(proyecto=proyecto,fecha__lt=desde)
            obj_pptosig=Adjudicacion.objects.using(proyecto).raw(f'CALL detalle_cartera("{periodo_sig}")')
            obj_pagosobra = building_model.pagos_obras.objects.filter(
                proyecto_asume__proyecto=proyecto,
                pago__fechapago__gte=desde,
                pago__fechapago__lte=hasta,
            )
            
            book=openpyxl.load_workbook("resources/excel_formats/Formato Informe Mensual.xlsx")
            #pestaña ventas
            sheet_ini=book.get_sheet_by_name('RESULTADOS MES')
            sheet_ini.cell(1,1,f'INFORME DE RESULTADOS MES {proyecto.upper()} {año}-{mes}')
            sheet_ventas=book.get_sheet_by_name('Ventas')
            row=2
            for line in obj_ventas:
                sheet_ventas.cell(row,1,line.item)
                sheet_ventas.cell(row,2,line.cantidad_mes)
                sheet_ventas.cell(row,3,line.valor_mes)
                sheet_ventas.cell(row,4,line.cantidad_total)
                sheet_ventas.cell(row,5,line.valor_total)
                row+=1
            sheet_detalleventas=book.get_sheet_by_name('Detalle Ventas')
            row=2
            for line in obj_detalleventas:
                sheet_detalleventas.cell(row,1,line.fecha)
                sheet_detalleventas.cell(row,2,line.pk)
                sheet_detalleventas.cell(row,3,line.titular)
                sheet_detalleventas.cell(row,4,line.inmueble)
                sheet_detalleventas.cell(row,5,line.valor)
                sheet_detalleventas.cell(row,6,line.basecomision)
                sheet_detalleventas.cell(row,7,line.comisComercial)
                try:
                    sheet_detalleventas.cell(row,8,line.comisComercial*line.basecomision/100)
                except: pass
                sheet_detalleventas.cell(row,9,3*line.basecomision/100)
                sheet_detalleventas.cell(row,10,line.platamesa)
                sheet_detalleventas.cell(row,11,line.platames)
                row+=1
            sheet_detallercdos=book.get_sheet_by_name('Recaudos')
            row=2
            for line in obj_recaudos:
                sheet_detallercdos.cell(row,1,line.fecha)
                sheet_detallercdos.cell(row,2,line.pk)
                sheet_detallercdos.cell(row,3,line.titular)
                sheet_detallercdos.cell(row,5,line.valor)
                sheet_detallercdos.cell(row,6,line.operacion)
                sheet_detallercdos.cell(row,7,line.formapago)
                row+=1
            sheet_resultcartera=book.get_sheet_by_name('Cartera')
            row=2
            for line in obj_cartera:
                if line.asesor=='JURIDICO':
                    estado='Juridico'
                else: estado=line.estado
                sheet_resultcartera.cell(row,1,line.pk)
                sheet_resultcartera.cell(row,2,line.cliente)
                sheet_resultcartera.cell(row,3,estado)
                sheet_resultcartera.cell(row,4,line.origen)
                sheet_resultcartera.cell(row,5,line.tipocartera)
                sheet_resultcartera.cell(row,6,line.edad)
                sheet_resultcartera.cell(row,7,line.ppto_mes)
                sheet_resultcartera.cell(row,8,line.recaudo_mes)
                sheet_resultcartera.cell(row,9,line.ppto_vencido)
                sheet_resultcartera.cell(row,10,line.recaudo_vencido)
                sheet_resultcartera.cell(row,11,line.presupuesto)
                sheet_resultcartera.cell(row,12,line.recaudo_pptado)
                sheet_resultcartera.cell(row,13,line.recaudo_nopptado)
                sheet_resultcartera.cell(row,14,line.recaudo_total)
                row+=1
            sheet_resultcarterahist=book.get_sheet_by_name('CarteraHistorico')
            row=2
            for año_ppto in range(2015,int(año)+1):
                for mes_ppto in range(1,13):
                    if año_ppto==int(año) and mes_ppto==int(mes)+1:
                        break
                    periodo = f'{año_ppto}{mes_ppto:02d}'
                    obj_carterahist =  Adjudicacion.objects.using(proyecto).raw(f'CALL detalle_cartera("{periodo}")')
                    for line in obj_carterahist:
                        if line.asesor=='JURIDICO':
                            estado='Juridico'
                        else: estado=line.estado
                        sheet_resultcarterahist.cell(row,1,line.pk)
                        sheet_resultcarterahist.cell(row,2,line.cliente)
                        sheet_resultcarterahist.cell(row,3,estado)
                        sheet_resultcarterahist.cell(row,4,line.origen)
                        sheet_resultcarterahist.cell(row,5,line.tipocartera)
                        sheet_resultcarterahist.cell(row,6,line.edad)
                        sheet_resultcarterahist.cell(row,7,line.ppto_mes)
                        sheet_resultcarterahist.cell(row,8,line.recaudo_mes)
                        sheet_resultcarterahist.cell(row,9,line.ppto_vencido)
                        sheet_resultcarterahist.cell(row,10,line.recaudo_vencido)
                        sheet_resultcarterahist.cell(row,11,line.presupuesto)
                        sheet_resultcarterahist.cell(row,12,line.recaudo_pptado)
                        sheet_resultcarterahist.cell(row,13,line.recaudo_nopptado)
                        sheet_resultcarterahist.cell(row,14,line.recaudo_total)
                        sheet_resultcarterahist.cell(row,15,periodo)
                        row+=1
            row=2
            sheet_gastosmes=book.get_sheet_by_name('Gastos Mes')
            for line in obj_gastosMes:
                sheet_gastosmes.cell(row,1,line.fecha)
                sheet_gastosmes.cell(row,2,line.comprobante)
                sheet_gastosmes.cell(row,3,line.cuenta)
                sheet_gastosmes.cell(row,4,line.descrip_cuenta)
                sheet_gastosmes.cell(row,5,line.tercero)
                sheet_gastosmes.cell(row,6,line.descrip_gasto)
                sheet_gastosmes.cell(row,7,line.item_asociado)
                sheet_gastosmes.cell(row,8,line.valor)
                sheet_gastosmes.cell(row,9,line.empresa)
                row+=1
            row=2
            sheet_gastoshist=book.get_sheet_by_name('GastosHistorico')
            for line in obj_gastosAcumulados:
                sheet_gastoshist.cell(row,1,line.fecha)
                sheet_gastoshist.cell(row,2,line.comprobante)
                sheet_gastoshist.cell(row,3,line.cuenta)
                sheet_gastoshist.cell(row,4,line.descrip_cuenta)
                sheet_gastoshist.cell(row,5,line.tercero)
                sheet_gastoshist.cell(row,6,line.descrip_gasto)
                sheet_gastoshist.cell(row,7,line.item_asociado)
                sheet_gastoshist.cell(row,8,line.valor)
                sheet_gastoshist.cell(row,9,line.empresa)
                row+=1
            sheet_pptosig=book.get_sheet_by_name('PptoSiguiente')
            row=2
            for line in obj_pptosig:
                if line.asesor=='JURIDICO':
                    estado='Juridico'
                else: estado=line.estado
                sheet_pptosig.cell(row,1,line.pk)
                sheet_pptosig.cell(row,2,line.cliente)
                sheet_pptosig.cell(row,3,estado)
                sheet_pptosig.cell(row,4,line.origen)
                sheet_pptosig.cell(row,5,line.tipocartera)
                sheet_pptosig.cell(row,6,line.edad)
                sheet_pptosig.cell(row,7,line.ppto_mes)
                sheet_pptosig.cell(row,8,line.recaudo_mes)
                sheet_pptosig.cell(row,9,line.ppto_vencido)
                sheet_pptosig.cell(row,10,line.recaudo_vencido)
                sheet_pptosig.cell(row,11,line.presupuesto)
                sheet_pptosig.cell(row,12,line.recaudo_pptado)
                sheet_pptosig.cell(row,13,line.recaudo_nopptado)
                sheet_pptosig.cell(row,14,line.recaudo_total)
                row+=1
            sheet_pagosobra=book.get_sheet_by_name('PagosObra')
            row=2
            for line in obj_pagosobra:
                sheet_pagosobra.cell(row,1,line.pk)
                sheet_pagosobra.cell(row,2,line.pago.fechapago)
                sheet_pagosobra.cell(row,3,line.contrato_asociado.proveedor.nombre)
                sheet_pagosobra.cell(row,4,line.contrato_asociado.pk)
                sheet_pagosobra.cell(row,5,line.contrato_asociado.descripcion)
                sheet_pagosobra.cell(row,6,line.pago.valor)
                sheet_pagosobra.cell(row,7,line.item_informe.pk)
                sheet_pagosobra.cell(row,8,line.contrato_asociado.proyecto.proyecto)
                row+=1  
            filename=f'Informe_{proyecto}_periodo_{periodo}.xlsx'
            ruta=settings.DIR_EXPORT+filename
            _save_workbook_with_dirs(book, ruta)
            return FileResponse(open(ruta,'rb'),as_attachment=True,filename=filename)
        
        if request.POST.get('btnAlttum'):
            alttum = request.POST.get('alttum')
            book=openpyxl.load_workbook("resources/excel_formats/Formato Informe Mensual_Alttum.xlsx")
            #pestaña ventas
            sheet_ini=book.get_sheet_by_name('RESULTADOS MES')
            sheet_ini.cell(1,1,f'INFORME DE RESULTADOS MES {alttum.upper()} {año}-{mes}')
            obj_gastosMes=GastosInforme.objects.filter(proyecto=alttum,fecha__gte=desde,fecha__lte=hasta)
            obj_gastosAcumulados=GastosInforme.objects.filter(proyecto=alttum,fecha__lt=desde)
            row=2
            sheet_gastosmes=book.get_sheet_by_name('Gastos Mes')
            for line in obj_gastosMes:
                sheet_gastosmes.cell(row,1,line.fecha)
                sheet_gastosmes.cell(row,2,line.comprobante)
                sheet_gastosmes.cell(row,3,line.cuenta)
                sheet_gastosmes.cell(row,4,line.descrip_cuenta)
                sheet_gastosmes.cell(row,5,line.tercero)
                sheet_gastosmes.cell(row,6,line.descrip_gasto)
                sheet_gastosmes.cell(row,7,line.item_asociado)
                sheet_gastosmes.cell(row,8,line.valor)
                sheet_gastosmes.cell(row,9,line.empresa)
                row+=1
            row=2
            sheet_gastoshist=book.get_sheet_by_name('GastosHistorico')
            for line in obj_gastosAcumulados:
                sheet_gastoshist.cell(row,1,line.fecha)
                sheet_gastoshist.cell(row,2,line.comprobante)
                sheet_gastoshist.cell(row,3,line.cuenta)
                sheet_gastoshist.cell(row,4,line.descrip_cuenta)
                sheet_gastoshist.cell(row,5,line.tercero)
                sheet_gastoshist.cell(row,6,line.descrip_gasto)
                sheet_gastoshist.cell(row,7,line.item_asociado)
                sheet_gastoshist.cell(row,8,line.valor)
                sheet_gastoshist.cell(row,9,line.empresa)
                row+=1
            filename=f'Informe_{alttum}_periodo_{periodo}.xlsx'
            ruta=settings.DIR_EXPORT+filename
            _save_workbook_with_dirs(book, ruta)
            return FileResponse(open(ruta,'rb'),as_attachment=True,filename=filename)
        
        if request.POST.get('flujoingresos'):
            obj_flujo=saldos_adj.objects.using(proyecto).raw(f'CALL ver_presupuesto("{hasta}","")')
            book=openpyxl.Workbook()
            sheet=book.active
            encabezados=['Adj','Cliente','Tipo Cta','Nro Cta','Tipo Cartera','Fecha Cta','Capital','Interes Cte','Cuota',
                            'Dias Mora','Interes Mora','Asesor','Edad','Proyecto','Vencido','Origen']
            sheet.append(encabezados)
            i=2
            for fila in obj_flujo:
                sheet.cell(i,1,fila.adj)
                sheet.cell(i,2,fila.cliente)
                sheet.cell(i,3,fila.tipocta)
                sheet.cell(i,4,fila.nrocta)
                sheet.cell(i,5,fila.tipocartera)
                sheet.cell(i,6,fila.fechacta)
                sheet.cell(i,7,fila.saldocapital)
                sheet.cell(i,8,fila.saldointcte)
                sheet.cell(i,9,fila.saldocuota)
                sheet.cell(i,10,fila.diasmora)
                sheet.cell(i,11,fila.saldomora)
                sheet.cell(i,12,fila.asesor)
                sheet.cell(i,13,fila.edad)
                sheet.cell(i,14,proyecto)
                sheet.cell(i,16,fila.origenventa)
                
                mes_corte=datetime.date.today().month
                año_corte=datetime.date.today().year
                if fila.fechacta<datetime.date(año_corte,mes_corte,1):
                    vencido='SI'
                else: vencido='NO'
                sheet.cell(i,15,vencido)
                i+=1
            filename=f'FlujoIngresos_{proyecto}_hasta_{periodo}.xlsx'
            ruta=settings.DIR_EXPORT+filename
            _save_workbook_with_dirs(book, ruta)
            return FileResponse(open(ruta,'rb'),as_attachment=True,filename=filename)
    
    return render(request,'informe_mes.html',context)

@group_perm_required(('andinasoft.add_gastosinforme',),raise_exception=True)
def informe_gastos(request):
    context={
        
    }
    if request.is_ajax():
        if request.method == 'POST':
            data={}
            empresa=request.POST.get('Empresa')
            archivo_xlsx=request.FILES['archivocarga']
            ultima_fila=int(request.POST.get('ultimafila'))
            try:
                gastos_cargados=cargar_gastos_informe(empresa=empresa,file=archivo_xlsx,ultima_linea=ultima_fila,
                                      object_gastos=GastosInforme.objects,object_cc=CentrosCostos.objects,
                                      object_cuentas=CuentasAsociadas.objects)
                data['instance']=True
                if gastos_cargados>0:
                    data['errores']=True
                else:
                    data['errores']=False
                return JsonResponse(data)
            except:
                data['instance']=False
                file = open(settings.DIR_EXPORT+'console_error.txt','w')
                file.write(traceback.format_exc())
                file.close()
                return JsonResponse(data)
            
    return render(request,'gastos_proyectos.html',context)

@group_perm_required(('andinasoft.view_gastosinforme',),raise_exception=True)
def detalle_gastos(request,empresa,año,mes):
    if request.is_ajax():
        if request.method == 'GET':
            typecall=request.GET.get('typecall')
            if typecall == 'selectOptions':
                grupo_gastos=request.GET.get('grupo')
                obj_items = ItemsInforme.objects.filter(grupo=grupo_gastos)
                data = {
                    'items':serializers.serialize('json',obj_items),
                }
                return JsonResponse(data)
            elif typecall == 'UpdateData':
                check_perms(request,('andinasoft.change_gastosinforme',),raise_exception=True)
                comprobante = request.GET.get('comprobante')
                proyecto = request.GET.get('proyecto')
                item = request.GET.get('item')
                try:
                    obj_gasto = GastosInforme.objects.get(empresa=empresa,comprobante=comprobante)
                    obj_gasto.proyecto=proyecto
                    obj_gasto.item_asociado=item
                    obj_gasto.save()
                    return JsonResponse({'passed':True})
                except:
                    return JsonResponse({'passed':False})
            elif typecall == 'searchItem':
                item = request.GET.get('item')
                if item != '':
                    try:
                        obj_item=ItemsInforme.objects.get(pk=item)
                        nombre_item=obj_item.temas
                        grupo_item=obj_item.grupo
                        data={
                            'passed':True,
                            'descrip_item':nombre_item,
                            'grupo_item':grupo_item,
                        }
                        return JsonResponse(data)
                    except:
                        return JsonResponse({'passed',False})
            elif typecall == 'DeleteData':
                if check_perms(request,('andinasoft.change_gastosinforme',),raise_exception=False):
                    comprobante = request.GET.get('comprobante')
                    proyecto = request.GET.get('proyecto')
                    item = request.GET.get('item')
                    try:
                        obj_gasto = GastosInforme.objects.get(comprobante=comprobante,proyecto=proyecto)
                        obj_gasto.delete()
                        data={
                            'passed':True
                        }
                    except:
                        data={
                            'passed':False
                        }
                else:
                    data={
                        'passed':False,
                        'message':'No tienes permisos para realizar este cambio'
                    }
                return JsonResponse(data)
    
    obj_proyectos = proyectos.objects.all()
    if mes=='Todos':
        fecha_desde=datetime.date(int(año),1,1)
        fecha_hasta=datetime.date(int(año),12,31)
    else:
        fecha_desde=datetime.date(int(año),int(mes),1)
        last_day=calendar.monthrange(int(año),int(mes))[1]
        fecha_hasta=datetime.date(int(año),int(mes),last_day)
    if request.method=='POST' and not request.is_ajax():
        obj_gastos = GastosInforme.objects.filter(empresa=empresa,fecha__gte=fecha_desde,fecha__lte=fecha_hasta)
        obj_gastos.delete()
    obj_gastos=GastosInforme.objects.filter(empresa=empresa,fecha__gte=fecha_desde,fecha__lte=fecha_hasta)
    obj_sinproyecto=obj_gastos.filter(proyecto="").count()
    obj_sinitem=obj_gastos.filter(item_asociado="").count()
    context={
        'gastos':obj_gastos,
        'año':año,
        'mes':mes,
        'empresa':empresa,
        'proyectos':obj_proyectos,
        'obj_sinproyecto':obj_sinproyecto,
        'obj_sinitem':obj_sinitem,
    }
    return render(request,'clasificar_gastos.html',context)

@group_perm_required(('andinasoft.view_gastosinforme',),raise_exception=True)
def general_gastos(request):
    context={
        
    }
    return render(request,'general_gastos.html',context)

@group_perm_required(('andinasoft.change_parametros_operaciones',),raise_exception=True)
def parametros(request,proyecto):
    obj_parametro=Parametros_Operaciones.objects.using(proyecto)
    jsonparametros=serializers.serialize('json',obj_parametro.all())
    context={
        'proyecto':proyecto,
        'parameters':jsonparametros,
    }
    
    if request.is_ajax():
        if request.method == 'GET':
            parametro=request.GET.get('parametro')
            status=request.GET.get('status')
            if status=="true":status=True
            else: status=False
            obj_parametro=obj_parametro.get(descripcion=parametro)
            obj_parametro.estado=status
            obj_parametro.save()
            if status:
                mensaje='Parametro Activado'
            else:
                mensaje='Parametro Desactivado'
            
            data={'passed':True,'msj':mensaje}
            return JsonResponse(data)
    
    return render(request,'parametros.html',context)

def asociar_cuentas(request,empresa):
    
    if request.is_ajax():
        if request.method == 'GET':
            typecall = request.GET.get('typecall')
            if typecall == 'selectOptions':
                grupo_gastos=request.GET.get('grupo')
                obj_items = ItemsInforme.objects.filter(grupo=grupo_gastos)
                data = {
                    'items':serializers.serialize('json',obj_items),
                }
                return JsonResponse(data)
            elif typecall == 'ChangeAsociation':
                cuenta = request.GET.get('cuenta')
                item = request.GET.get('item')
                try:
                    obj_cuenta = CuentasAsociadas.objects.get(cuenta=cuenta,empresa=empresa)
                    obj_cuenta.item_asociado=item
                    obj_cuenta.save()
                    obj_item=ItemsInforme.objects.get(pk=item)
                    data={
                        'passed':True,
                        'descrip_item':obj_item.temas,
                        'grupo_item':obj_item.grupo,
                    }
                except:
                    data={
                        'passed':False,
                    }
                return JsonResponse(data)
    if request.method == 'POST':
        cuenta = request.POST.get('idCuenta')
        item = request.POST.get('NuevaCtaItem')
        descripcion = request.POST.get('DescripCuenta')
        CuentasAsociadas.objects.create(cuenta=cuenta,empresa=empresa,descrip_cuenta=descripcion,item_asociado=item)
    cuentas = CuentasAsociadas.objects.raw(f'CALL gastos_asociados("{empresa}")')
    context={
        'cuentas':cuentas,
        'empresa':empresa,
    }
    return render(request,'asociar_cuentas.html',context)

def asociar_cc(request,empresa):
    
    if request.is_ajax():
        if request.method == 'GET':
            typecall = request.GET.get('typecall')
            if typecall == 'ChangeCC':
                cc = request.GET.get('cc')
                proyecto = request.GET.get('proyecto')
                obj_changecc = CentrosCostos.objects.get(empresa=empresa,idcentrocosto=cc)
                obj_changecc.proyecto = proyecto
                obj_changecc.save()
                data={'passed':True}
                return JsonResponse(data)
    if request.method == 'POST':
        centrocosto = request.POST.get('idCC')
        centrocosto = f'{centrocosto}   '
        proyecto = request.POST.get('ProyectoCC')
        CentrosCostos.objects.create(proyecto=proyecto,empresa=empresa,idcentrocosto=centrocosto,
                                     centro=centrocosto[:4],subcentro=centrocosto[5:],detalle='')
            
    
    obj_proyectos=proyectos.objects.all()
    obj_cc=CentrosCostos.objects.filter(empresa=empresa)
    context={
        'proyectos':obj_proyectos,
        'centroscosto':obj_cc,
        'empresa':empresa,
    }
    return render(request,'asociar_centroscosto.html',context)

@group_perm_required(('andinasoft.view_clientes',),raise_exception=True)
def buscar_cliente(request):
    context={
        
    }
    
    if request.is_ajax():
        if request.method == 'GET':
            q=request.GET.get('q')
            todo = request.GET.get('todo')
            
            
            if todo == 'search':
                
                cliente=clientes.objects.exclude(pk__icontains='alt'
                        ).exclude(nombrecompleto__exact="").order_by('nombrecompleto')
                
                total = cliente.count()
                
                draw = request.GET.get('draw')
                start = request.GET.get('start')
                length = request.GET.get('length')
                search_val = request.GET.get('search[value]')
                
                if search_val != "":
                    cliente = cliente.filter(Q(pk__icontains=search_val)|
                                            Q(nombrecompleto__icontains=search_val))
                
                
                defered_list = cliente[int(start):int(start)+int(length)]
                
                data={
                    "draw": int(draw),
                    "recordsTotal": total,
                    "recordsFiltered": cliente.count(),
                    'data':JSONRender(defered_list).render()
                    }
                
                return JsonResponse(data)
            
            
            elif todo == 'clientdata':
                
                pk=request.GET.get('pk')
                cliente=clientes.objects.get(pk=pk)
                nombre=cliente.nombrecompleto
                    
                lista_proyectos=proyectos.objects.all()
                contratos_asociados={}
                i=0
                for proyecto in lista_proyectos:
                    proyecto=str(proyecto)
                    if proyecto in settings.DATABASES:
                        titular1=Adjudicacion.objects.using(proyecto).filter(idtercero1=pk)
                        
                        if titular1.exists():
                            for adj in titular1:
                                info = Vista_Adjudicacion.objects.using(proyecto).get(IdAdjudicacion=adj.pk)
                                contratos_asociados[i]={
                                    'proyecto':proyecto,
                                    'adj':adj.pk,
                                    'inmueble':adj.idinmueble,
                                    'cartera':info.tipo_cartera,
                                    'estado':adj.estado,
                                    'nrotitular':'1er Titular'
                                }
                                
                                i+=1
                        titular2=Adjudicacion.objects.using(proyecto).filter(idtercero2=pk)
                        if titular2.exists():
                            for adj in titular2:
                                info = Vista_Adjudicacion.objects.using(proyecto).get(IdAdjudicacion=adj.pk)
                                contratos_asociados[i]={
                                    'proyecto':proyecto,
                                    'adj':adj.pk,
                                    'inmueble':adj.idinmueble,
                                    'cartera':info.tipo_cartera,
                                    'estado':adj.estado,
                                    'nrotitular':'2do Titular'
                                }
                                i+=1
                        titular3=Adjudicacion.objects.using(proyecto).filter(idtercero3=pk)
                        if titular3.exists():
                            for adj in titular3:
                                info = Vista_Adjudicacion.objects.using(proyecto).get(IdAdjudicacion=adj.pk)
                                contratos_asociados[i]={
                                    'proyecto':proyecto,
                                    'adj':adj.pk,
                                    'inmueble':adj.idinmueble,
                                    'cartera':info.tipo_cartera,
                                    'estado':adj.estado,
                                    'nrotitular':'3er Titular'
                                }
                                i+=1
                        titular4=Adjudicacion.objects.using(proyecto).filter(idtercero4=pk)
                        if titular4.exists():
                            for adj in titular4:
                                info = Vista_Adjudicacion.objects.using(proyecto).get(IdAdjudicacion=adj.pk)
                                contratos_asociados[i]={
                                    'proyecto':proyecto,
                                    'adj':adj.pk,
                                    'inmueble':adj.idinmueble,
                                    'cartera':info.tipo_cartera,
                                    'estado':adj.estado,
                                    'nrotitular':'4to Titular'
                                }
                                i+=1
                contratos_asociados['cantidad']=i
                json_data=json.dumps(contratos_asociados)
                data={'contratos':json_data,'cliente':nombre}
                return JsonResponse(data)        
        elif request.method == 'POST':
            todo = request.POST.get('todo')
            
            if todo == 'interfaz':
                check_groups(request,['Contabilidad',])
                lista_clientes = request.POST.getlist('clientes[]')
                book=openpyxl.load_workbook("resources/excel_formats/InterfazTerceros.xlsx")
                sheet=book.active
                
                row=6
                for i in lista_clientes:
                    cliente = clientes.objects.get(pk=i)
                    
                    try: dv = digito_verificacion(cliente.pk)
                    except: dv = ""
                    
                    names = dividir_nombre(cliente.nombres, cliente.apellidos)
                    
                    
                    sheet.cell(row,1,str(cliente.pk))
                    sheet.cell(row,2,0)
                    sheet.cell(row,3,dv)
                    sheet.cell(row,4,cliente.nombrecompleto.upper())
                    sheet.cell(row,5,"N")
                    sheet.cell(row,6,names[0])
                    sheet.cell(row,7,names[1])
                    sheet.cell(row,8,names[2])
                    sheet.cell(row,9,names[3])
                    sheet.cell(row,10,cliente.domicilio.upper())
                    sheet.cell(row,11,cliente.pais.pk)
                    sheet.cell(row,12,cliente.city.id_city)
                    sheet.cell(row,13,cliente.celular1)
                    sheet.cell(row,15,1)
                    sheet.cell(row,16,cliente.email.upper())
                    sheet.cell(row,17,"C")
                    row+=1
                
                nombre_doc='Interfaz_Terceros.xlsx'
                ruta=settings.MEDIA_ROOT+'/tmp/'+nombre_doc
                _save_workbook_with_dirs(book, ruta)
                
                data = {
                    'ruta': settings.MEDIA_URL + 'tmp/' + nombre_doc
                }
                print(data)
                return JsonResponse(data)
                    
                    
    
    return render(request,'buscar_cliente.html',context)

def dividir_nombre(nombres,apellidos):
    nombres_divididos = nombres.split(' ')
    if nombres_divididos.__len__() > 1:
        fn1 = nombres_divididos[0]
        nombres_divididos.pop(0)
        fn2 = ' '.join(nombres_divididos)
    else:
        fn1 = nombres
        fn2 = ''
    
    apellidos_divididos = apellidos.split(' ')
    if apellidos_divididos.__len__() > 1:
        ln1 = apellidos_divididos[0]
        apellidos_divididos.pop(0)
        ln2 = ' '.join(apellidos_divididos)
    else:
        ln1 = apellidos
        ln2 = ''
        
    return fn1, fn2, ln1, ln2

def digito_verificacion(nit):
    k,j = 0, 0
    digits = [i for i in str(nit)]
    s = [3,7,13,17,19,23,29,37,41,43,47,53,59,67,71]
    for n in reversed(digits):
        j += int(n) * s[k]
        k+=1
        
    r = j%11
    dv = 0 if r == 0 else 1 if r == 1 else 11-r
    return dv
    

def advisersarea(request):
    context={
        
    }
    return render(request,'asesores_area.html',context)

def graphs(request,proyecto):
    check_groups(request,('Gerente de area',))
    context={}
    if request.method =='POST':
        mes=int(request.POST.get('periodomes'))
        año=int(request.POST.get('periodoaño'))
        ultimodia=calendar.monthrange(año,mes)[1]
        fecha_inicial=f'{año}-{int(mes):02d}-01'
        fecha_final=f'{año}-{int(mes):02d}-{ultimodia}'
        obj_recaudos=Recaudos_general.objects.using(proyecto).raw(f'CALL recaudos_cartera("{fecha_inicial}","{fecha_final}")')
        obj_ventasmes=Recaudos_general.objects.using(proyecto).raw(f'CALL recaudos_ventas("{fecha_inicial}","{fecha_final}")')
        
        context={
            'recaudos':obj_recaudos,
            'rec_ventasmes':obj_ventasmes,
            'post':True
        }
    return render(request,'graphs/graph_recaudos.html',context)

def graph_rcdo_com_año(request,proyecto):
    check_groups(request,('Gerente de area',))
    context={}
    if request.method =='POST':
        año=int(request.POST.get('periodoaño'))
        datos_año=[]
        mes_actual=datetime.date.today().month
        for i in range(1,mes_actual+1):
            ultimodia=calendar.monthrange(año,i)[1]
            fecha_inicial=f'{año}-{i:02d}-01'
            fecha_final=f'{año}-{i:02d}-{ultimodia}'
            rcdo_cartera=Recaudos_general.objects.using(proyecto).raw(f'CALL recaudos_cartera("{fecha_inicial}","{fecha_final}")')
            rcdo_ventasmes=Recaudos_general.objects.using(proyecto).raw(f'CALL recaudos_ventas("{fecha_inicial}","{fecha_final}")')
            datos_año.append({'rcdo_cartera':rcdo_cartera,
                        'rcdo_ventasmes':rcdo_ventasmes,
                        'mes':i})
        context={
            'data_año':datos_año,
            'post':True
        }
    return render(request,'graphs/graph_year.html',context)

def graph_ventas_anuales(request,proyecto):
    check_groups(request,('Gerente de area',))
    context={}
    if request.method =='POST':
        año=request.POST.get('periodoaño')
        fecha_inicial=f'{año}-01-01'
        fecha_final=f'{año}-12-31'
        
        obj_ventas=Adjudicacion.objects.using(proyecto).raw(f'CALL detalle_ventas("{fecha_inicial}","{fecha_final}")')
        
        context={
            'ventas':obj_ventas,
            'año':año,
            'proyecto':proyecto,
            'post':True
        }
    
    return render(request,'graphs/graph_ventas_mes.html',context)

def graph_cartera_anual(request,proyecto):
    check_groups(request,('Gerente de area',))
    context={}
    if request.method =='POST':
        año=request.POST.get('periodoaño')
        agrupador={}
        for mes in range(1,13):
            contenido_ppto=Adjudicacion.objects.using(proyecto).raw(f'CALL informe_cartera("{año}{mes:02d}",NULL)')
            for fila in contenido_ppto:
                grupo=str(mes)+'-'+fila.tipocartera+'-'+fila.edad
                if grupo not in agrupador:
                    agrupador.update({
                        grupo:{
                            'año':int(año),
                            'mes':mes,
                            'edad':fila.edad,
                            'tipocartera':fila.tipocartera,
                            'nroclientes':1,
                            'pptomes':fila.ppto_mes,
                            'rcdomes':fila.recaudo_mes,
                            'pptovenc':fila.ppto_vencido,
                            'rcdovenc':fila.recaudo_vencido,
                            'totalppto':fila.presupuesto,
                            'rcdopptado':fila.recaudo_pptado,
                            'rcdonopptado':fila.recaudo_nopptado,
                            'rcdototal':fila.recaudo_total
                        }
                    })
                else:
                    agrupador[grupo]['pptomes']+=fila.ppto_mes
                    agrupador[grupo]['rcdomes']+=fila.recaudo_mes
                    agrupador[grupo]['pptovenc']+=fila.ppto_vencido
                    agrupador[grupo]['rcdovenc']+=fila.recaudo_vencido
                    agrupador[grupo]['totalppto']+=fila.presupuesto
                    agrupador[grupo]['rcdopptado']+=fila.recaudo_pptado
                    agrupador[grupo]['rcdonopptado']+=fila.recaudo_nopptado
                    agrupador[grupo]['rcdototal']+=fila.recaudo_total
                    agrupador[grupo]['nroclientes']+=1
        context={
            'cartera_año':agrupador.values(),
            'proyecto':proyecto,
            'año':año,
            'post':True
        }
    return render(request,'graphs/graph_cartera_year.html',context)

#ajax

def ajax_imprimir_promesa(request):
    if request.method == 'GET':
        
        todo=request.GET.get('todo')
        proyecto = request.GET.get('proyecto')
        adj = request.GET.get('adj')
        
        if todo == 'actual':
            obj_promesa = Promesas.objects.using(proyecto).get(idadjudicacion=adj)
            obj_adj = Adjudicacion.objects.using(proyecto).get(idadjudicacion=adj)
            titular1 = clientes.objects.get(pk=obj_adj.idtercero1)
            titular2 = clientes.objects.get(pk=obj_adj.idtercero2)
            titular3 = clientes.objects.get(pk=obj_adj.idtercero3)
            titular4 = clientes.objects.get(pk=obj_adj.idtercero4)
            obj_inmueble = Inmuebles.objects.using(proyecto).get(idinmueble=obj_adj.idinmueble)
            obj_planpagos=PlanPagos.objects.using(proyecto).filter(adj=adj)
            formapago = obj_promesa.formapago
            valor_letras = Utilidades().numeros_letras(obj_adj.valor)
            cuota_inicial=obj_planpagos.filter(tipocta='CI').aggregate(Sum('capital'))['capital__sum']
            if cuota_inicial==None: cuota_inicial=0
            saldo=obj_planpagos.exclude(tipocta='CI').aggregate(Sum('capital'))['capital__sum']
            if saldo==None: saldo=0
            if formapago=='Contado':
                forma=('x','','')
            elif formapago=='Credicontado':
                forma=('','x','')
            elif formapago=='Amortizacion':
                forma=('','','x')
            diacontrato = obj_promesa.fechapromesa.day
            mescontrato = obj_promesa.fechapromesa.month
            añocontrato = obj_promesa.fechapromesa.year
            
            filename = f'Promesa_{proyecto}_{adj}.pdf'
            ruta = settings.MEDIA_ROOT+f'/tmp/pdf/{filename}'
            GenerarPDF().ExportPromesaVegasVenecia(nro_contrato=obj_promesa.nropromesa,
                    nombre_t1=titular1.nombrecompleto,cc_t1=titular1.idTercero,tel_t1=titular1.telefono1,cel_t1=titular1.celular1,ofic_t1=titular1.oficina,cdof_t1=titular1.ciudad,
                    telof_t1=titular1.telefono2,resid_t1=titular1.domicilio,cdresid_t1=titular1.ciudad,telresid_t1=titular1.telefono1,email_t1=titular1.email,
                    nombre_t2=titular2.nombrecompleto,cc_t2=titular2.idTercero,tel_t2=titular2.telefono1,cel_t2=titular2.celular1,ofic_t2=titular2.oficina,cdof_t2=titular2.ciudad,
                    telof_t2=titular2.telefono2,resid_t2=titular2.domicilio,cdresid_t2=titular2.ciudad,telresid_t2=titular2.telefono1,email_t2=titular2.email,
                    nombre_t3=titular3.nombrecompleto,cc_t3=titular3.idTercero,tel_t3=titular3.telefono1,cel_t3=titular3.celular1,ofic_t3=titular3.oficina,cdof_t3=titular3.ciudad,
                    telof_t3=titular3.telefono2,resid_t3=titular3.domicilio,cdresid_t3=titular3.ciudad,telresid_t3=titular3.telefono1,email_t3=titular3.email,
                    nombre_t4=titular4.nombrecompleto,cc_t4=titular4.idTercero,tel_t4=titular4.telefono1,cel_t4=titular4.celular1,ofic_t4=titular4.oficina,cdof_t4=titular4.ciudad,
                    telof_t4=titular4.telefono2,resid_t4=titular4.domicilio,cdresid_t4=titular4.ciudad,telresid_t4=titular4.telefono1,email_t4=titular4.email,
                    lote=obj_inmueble.lotenumero,manzana=obj_inmueble.manzananumero,area=str(obj_inmueble.areaprivada),
                    mtsnorte=str(obj_inmueble.norte),colnorte=obj_inmueble.colindante_norte,mtseste=str(obj_inmueble.este),coleste=obj_inmueble.colidante_este,
                    mtssur=str(obj_inmueble.sur),colsur=obj_inmueble.colindante_sur,mtsoeste=str(obj_inmueble.oeste),coloeste=obj_inmueble.colindante_oeste,
                    valor=int(obj_adj.valor),valor_letras=valor_letras,ci=int(cuota_inicial),saldo=int(saldo),contado_x=forma[0],credic_x=forma[1],amort_x=forma[2],
                    formaCI=obj_promesa.formaci,formaFN=obj_promesa.formasaldo,obs=obj_promesa.observaciones,dia_contrato=str(diacontrato),mes_contrato=mescontrato,año_contrato=str(añocontrato),
                    fecha_entrega=obj_promesa.fechaentrega,fecha_escritura=obj_promesa.fechaescritura,ciudad_entrega=obj_promesa.ciudad,ruta=ruta)
            
            data = {
                'ruta':settings.MEDIA_URL+f'/tmp/pdf/{filename}'
            }
            
            return JsonResponse(data)
        elif todo == 'cambiolote':
            
            adj = request.GET.get('adj')
            nuevo_lote = request.GET.get('nuevolote')
            mesesentrega = request.GET.get('mesesentrega')
            fecha_escritura = request.GET.get('fechaescritura')
            override_date = request.GET.get('overridedate')
            dt_fmt = datetime.datetime.strptime(fecha_escritura,'%Y-%m-%d')
            
            
            prorroga = request.GET.get('prorroga')
            
            prorroga = int(prorroga) if prorroga else 365
            
            mesesentrega = int(mesesentrega) if mesesentrega else 0
            
            try: 
                promesa = Promesas.objects.using(proyecto).get(idadjudicacion=adj)
            except:
                #print(traceback.format_exc())
                data = {
                    'response': 'La adjudicación seleccionada no tiene ninguna promesa asociada en el modulo Estado de Promesas.'
                }
                return JsonResponse(data)
            
            if nuevo_lote: 
                try:
                    inmueble = Inmuebles.objects.using(proyecto).get(pk=nuevo_lote)
                except:
                    data = {
                        'response': 'El lote ingresado no está registrado en la base de datos.'
                    }
                    return JsonResponse(data)
            else:
                inmueble = promesa.general_info().get('inmueble')     
                is_adj = Adjudicacion.objects.using(proyecto).filter(
                    idinmueble = nuevo_lote
                ).exclude(estado__icointains='Desistido')
                is_rsv = ventas_nuevas.objects.using(proyecto).filter(
                    Q(estado='Pendiente')|Q(estado='Aprobado'),
                    inmueble = nuevo_lote
                )
                
                if is_adj.exists() or is_rsv.exists():
                    data = {
                        'response': 'El lote ingresado ya se encuentra asignado a una adjudicacion o a una venta nueva, revisalo'
                    }
                    return JsonResponse(data)
                
            
            new_date =  datetime.date.today() if override_date else None
                
            try:                
                parcela = int(inmueble.lotenumero)
                fraccion = None
                porcentaje = '100%'             
            except:  
                fraccion = inmueble.lotenumero[-1]
                parcela =  inmueble.lotenumero[:-1]
                porcentaje = '50%'     
            
            
            
            context = {
                'es_promesa':True,
                'proyecto':proyecto,
                'ctr':promesa,
                'inmueble': inmueble,
                'fecha_escritura':dt_fmt,
                'meses_entrega':mesesentrega,
                'oficina':'Medellín',
                'new_date':new_date,
                'parcela':parcela,
                'fraccion':fraccion,      
                'prorroga':prorroga,   
                'prorroga_letras': Utilidades().numeros_letras(prorroga, formato=None).replace('PESOS M/CTE',''),
            }
            
            filename = f'Contrato_bien_futuro_{adj}_{proyecto}.pdf'
            
                
            if proyecto == 'Oasis':
                pdf = pdf_gen_weasy(f'pdf/{proyecto}/contrato.html', context, filename)
            else:
                pdf = pdf_gen(f'pdf/{proyecto}/contrato.html', context, filename)
            
            ruta = pdf.get('root')
                                    
            ruta_link=pdf.get('url')
            
            print('llega')
            
            return FileResponse(open(ruta,'rb'),as_attachment=True,filename=filename)
            
def ajax_print_estado_cuenta(request):
    
    if request.method == 'GET':
        proyecto = request.GET.get('proyecto')
        adj = request.GET.get('adj')
        
        obj_adj = Adjudicacion.objects.using(proyecto).get(pk=adj)
        today = datetime.date.today()
        next_30_days = today+ datetime.timedelta(days=30)
        cuotas_a_la_fecha = PlanPagos.objects.using(proyecto).filter(adj=adj, fecha__lte=today).order_by('fecha')
        cuotas_futuras = PlanPagos.objects.using(proyecto).filter(
            adj=adj, fecha__gt=today,fecha__lte = next_30_days).order_by('fecha')
        
        cuotas_vencidas = []
        
        total_cuotas_vencidas = {
             'valor': 0,
             'intereses_mora':0,
             'total': 0
        }
        
        recaudador = {
            'Tesoro Escondido': 'STATUS COMERCIALIZADORA S.A.S. NIT: 901018375-4',
            'Vegas de Venecia': 'STATUS COMERCIALIZADORA S.A.S. NIT: 901018375-4',
            'Perla del Mar': 'ANDINA CONCEPTOS INMOBILIARIOS S.A.S. NIT: 900993044-9',
            'Sandville Beach': 'ANDINA CONCEPTOS INMOBILIARIOS S.A.S. NIT: 900993044-9',
            'Carmelo Reservado': 'ANDINA CONCEPTOS INMOBILIARIOS S.A.S. NIT: 900993044-9', 
        }
        
        for q in cuotas_a_la_fecha:
            pendiente = q.pendiente()
            if pendiente.get('total', 0) > 0:
                mora = q.mora()
                cuotas_vencidas.append({
                    'fecha':q.fecha,
                    'idcta': q.pk.split('ADJ')[0],
                    'pendiente':pendiente,
                    'mora':mora,
                })
                total_cuotas_vencidas['valor'] += pendiente.get('total')
                total_cuotas_vencidas['intereses_mora'] += mora.get('valor')
                total_cuotas_vencidas['total'] += pendiente.get('total') + mora.get('valor')


        context = {
            'adj':obj_adj,
            'cuotas_a_la_fecha': cuotas_vencidas,
            'cuotas_futuras':cuotas_futuras,
            'user': request.user,
            'now': datetime.datetime.now(),
            'totals': total_cuotas_vencidas,
            'recaudador': recaudador.get(proyecto)
        }
        
        
        filename = f'Estado_de_cuenta_{adj}_{proyecto}.pdf'
                
                    
        pdf = pdf_gen(f'pdf/statement_of_account.html',context,filename)
        
        ruta = pdf.get('root')
                                
        ruta_link=pdf.get('url')
        
        return FileResponse(open(ruta,'rb'),as_attachment=True,filename=filename)

def print_documents(request,proyecto):
    if request.method == 'GET':
        ctr = request.GET.get('ctr')
        template = request.GET.get('template')
        obj_ctr = ventas_nuevas.objects.using(proyecto).get(pk=ctr)
        if template == 'contrato_nuevo':
            context = {
                'proyecto':proyecto,
                'ctr':obj_ctr,
                'fecha_escritura':datetime.date.today(),
                'meses_entrega':24,
                'oficina':'Medellín',
                'plan_pagos':calcular_tabla_amortizacion(obj_ctr),
            }
            
            filename = 'contrato.pdf'
            path = f'pdf/{proyecto}/contrato.html'
        
        elif template == 'verificacion':
            context = {
                            'proyecto':proyecto,
                            'ctr':obj_ctr,
                            'meses_entrega':24,
                            'oficina':'Medellin'
                        }
            filename = 'verificacion.pdf'
            path = f'pdf/{proyecto}/verificacion.html'
            
        elif template == 'sagrilaft':
            pdf_files = []
            for titular in obj_ctr.titulares():
                context = {
                    'cliente':titular,
                    'today': datetime.date.today(),
                }
            
                filename = f'Vinculacion_PN_{titular.pk}_Sagrilaft_Ctr_{ctr}_{proyecto.replace(" ","_")}.pdf'
                path = f'pdf/sagrilaft.html'
                pdf = pdf_gen(path,context,filename)
                pdf_files.append({
                        'cliente':titular.nombrecompleto,
                        'url':pdf.get('url')
                    }
                )
                
            
            data = {
                'data': pdf_files
            }
            
            return JsonResponse(data)
                
                
                
        if proyecto == 'Oasis' and template == 'contrato_nuevo':
            pdf = pdf_gen_weasy(path, context, filename)
        else:
            pdf = pdf_gen(path, context, filename)
        
        file = pdf.get('root')
                
        return JsonResponse({'file':pdf.get('url')})
        #return FileResponse(open(file,'rb'),as_attachment=True,filename=filename)

def compare_siigo_andinasoft_clients(request):
    if request.method == 'POST' and request.FILES:
        file = request.FILES['filesiigo']
        projects = request.POST.getlist('proyecto')
        
        documento=openpyxl.load_workbook(file)
        hoja=documento.get_sheet_names()
        sheet_to_read=documento.get_sheet_by_name(hoja[0])
        
        blank_cells = 0
        row = 8
        
        font = Font(
            bold=True,
        )
        alignment = Alignment(horizontal='center')
        
        fill = PatternFill(fill_type = "solid", start_color="99CCFF", end_color="99CCFF")
        
        fecha = sheet_to_read['J1'].value
        fecha_dt = datetime.datetime.strptime(fecha,'%b/%d/%Y')
        
        k = 6
        j = 0
        
        
        
        #SIIGO vs ANDINASOFT
        for proyecto in projects:
            cel_1 = sheet_to_read.cell(6 , 11 + j*k )
            cel_1.value = proyecto.upper()
            cel_1.font = font
            cel_1.alignment = alignment
            cel_1.fill = fill
            cel_2 = sheet_to_read.cell(7,11 + j*k)
            cel_2.value = 'ADJ'
            cel_2.font = font
            cel_2.alignment = alignment
            cel_2.fill = fill
            
            
            
            cel_3 = sheet_to_read.cell(7,12 + j*k)
            cel_3.value = 'CAPITAL'
            cel_3.font = font
            cel_3.alignment = alignment
            cel_3.fill = fill
            cel_4 = sheet_to_read.cell(7,13 + j*k)
            cel_4.value = 'INT CTE'
            cel_4.font = font
            cel_4.alignment = alignment
            cel_4.fill = fill
            cel_5 = sheet_to_read.cell(7,14 + j*k)
            cel_5.value = 'INT MORA'
            cel_5.font = font
            cel_5.alignment = alignment
            cel_5.fill = fill
            cel_6 = sheet_to_read.cell(7,15 + j*k)
            cel_6.value = 'TOTAL'
            cel_6.font = font
            cel_6.alignment = alignment
            cel_6.fill = fill
            
            sc = 11+j*k
            ec = sc+ (k-1)
            
            sheet_to_read.merge_cells(start_row=6,end_row=6,start_column=sc,end_column=ec )

            j += 1
            
        while blank_cells < 2:
            tercero = sheet_to_read[f'C{row}'].value
            if tercero == None or tercero == "":
                blank_cells += 1
                row+=1
                continue
            else:
                blank_cells = 0
            
            j=0
            
            for proyecto in projects:
                adjs = Adjudicacion.objects.using(proyecto).filter(
                    Q(idtercero1 = tercero) | Q(idtercero2 = tercero) |
                    Q(idtercero3 = tercero) | Q(idtercero4 = tercero) 
                ).exclude(
                    estado__icontains='desistido'
                )
                
                if adjs.count() < 1: row += 1
                for adj in adjs:
                        
                    cantidad_t = adj.cantidad_titulares()
                    cantidad_t = decimal.Decimal(cantidad_t)                           
                    recaudo = adj.recaudo_detallado(date = fecha_dt)
                    
                    capital = f"=SUMIFS('Andinasoft_{proyecto}'!G:G, 'Andinasoft_{proyecto}'!A:A, K{row}, 'Andinasoft_{proyecto}'!E:E, C{row})"
                    intcte = f"=SUMIFS('Andinasoft_{proyecto}'!H:H, 'Andinasoft_{proyecto}'!A:A, K{row}, 'Andinasoft_{proyecto}'!E:E, C{row})"
                    intmora = f"=SUMIFS('Andinasoft_{proyecto}'!I:I, 'Andinasoft_{proyecto}'!A:A, K{row}, 'Andinasoft_{proyecto}'!E:E, C{row})"
                    total = f"=SUM(L{row}:N{row})"
                    
                    sheet_to_read.cell(row,11 + j*k, adj.pk)
                    sheet_to_read.cell(row,12 + j*k, capital).style='Comma'
                    sheet_to_read.cell(row,13 + j*k, intcte).style='Comma'
                    sheet_to_read.cell(row,14 + j*k, intmora).style='Comma'
                    sheet_to_read.cell(row,15 + j*k, total).style='Comma'
                    
                    if adjs.count() > 1 and adj != adjs.last(): sheet_to_read.insert_rows(row+1)
                    
                    row += 1
                
                j+=1
                
                
        #ANDINASOFT vs SIIGO
        
        for proyecto in projects:
        
            terceros = []
            
            sheet = documento.create_sheet('Andinasoft_'+proyecto)
            sheet.cell(1,1).value = 'Adj'
            sheet.cell(1,2).value = 'Estado'
            sheet.cell(1,3).value = 'Tipo'
            sheet.cell(1,4).value = 'Valor'
            sheet.cell(1,5).value = 'id Titulares'
            sheet.cell(1,6).value = 'Nombres'
            sheet.cell(1,7).value = 'Capital'
            sheet.cell(1,8).value = 'IntCte'
            sheet.cell(1,9).value = 'IntMora'
            sheet.cell(1,10).value = 'Total'
            
            adjs = Adjudicacion.objects.using(proyecto).exclude(estado__icontains='desistido')
            
            row = 2
            for adj in adjs:
                titulares = adj.titulares2()
                cantidad_t = decimal.Decimal(len(titulares))                           
                recaudo = adj.recaudo_detallado(date = fecha_dt)
                
                capital = recaudo.get('capital')/cantidad_t
                intcte = recaudo.get('interescte')/cantidad_t
                intmora = recaudo.get('interesmora')/cantidad_t
                
                sheet.cell(row,1).value = adj.pk
                sheet.cell(row,2).value = adj.estado
                sheet.cell(row,3).value = adj.origenventa
                sheet.cell(row,4).value = adj.valor
                #Titulares
                for titular in titulares:
                    sheet.cell(row,1).value = adj.pk
                    sheet.cell(row,5).value = titular.pk
                    sheet.cell(row,6).value = titular.nombrecompleto.upper()
                    sheet.cell(row,7).value = capital
                    sheet.cell(row,8).value = intcte   
                    sheet.cell(row,9).value = intmora     
                    sheet.cell(row,10).value = capital+intcte+intmora 
                    
                    row+=1
                    if len(titulares) > 1 and titular != titulares[-1]: 
                        sheet.insert_rows(row)

        
        
        #Generacion de doc
        nombre_doc=f'siigo_saldos_terceros.xlsx'
        ruta=settings.MEDIA_ROOT+'/tmp/xlsx/'+nombre_doc
        ruta_dw=settings.MEDIA_URL+'/tmp/xlsx/'+nombre_doc
        _save_workbook_with_dirs(documento, ruta)
        
        data = {
            'corpus':f'''<ul>
                <li>
                    Puedes descargar el archivo de egresos <strong><a href="{ruta_dw}" target="_blank">aquí</a></strong>
                </li>
            </ul>'''
        }
        
        return JsonResponse(data)

#Cartera
@login_required
def reestructuraciones_cartera(request):

    if request.is_ajax:
        if request.method == 'GET':
            todo = request.GET.get('todo')
            if todo == 'datatable':
                proyecto = request.GET.get('proyecto')
                if proyecto is not None:
                    rstr = reestructuraciones_otrosi.objects.filter(proyecto=proyecto
                                            ).order_by('-fecha')
                    
                    js = JSONRender(rstr, query_functions=('extra_info',))
                    
                    data = {
                        'data': js.render()
                    }
                else:
                    data = {
                        'data':[]
                    }
                return JsonResponse(data)
            
        if request.method == 'POST':
            if request.is_ajax:
                todo = request.POST.get('todo')
                
                if todo == 'add-reestructuration':
                    print(request.POST)
                    data = {
                        'status':'ok',
                        'id':0
                    }
                    
                    return JsonResponse(data)
                
        
    context = {
        'proyectos':proyectos.objects.all(),
        'formbuscar':form_buscar_reestr(),
        'form_reg_legaliz':form_nueva_reestr(),
    }
    
    return render(request, 'reestructuraciones_new.html',context)

#Fractal
@login_required
def landing_fractal(request):
    check_project(request,'Fractal')
    context = {
        
    }
    
    return render(request, 'landing_fractal.html',context)

@login_required
def nueva_venta_fractal(request):
    if request.method == 'POST':
        if request.is_ajax():   
                        
            proyecto = request.POST.get('proyecto')
            
            
            inversionista = request.POST.get('inversionista')
            inmueble = request.POST.get('inmueble')
            valor_venta = request.POST.get('valor_inversion').replace(',','')
            forma_pago = request.POST.get('tipo_saldo')
            cuota_inicial = request.POST.get('valor_ci').replace(',','')
            saldo = request.POST.get('valor_saldo').replace(',','')
            valor_ctas_fn = request.POST.get('valor_cuota_saldo').replace(',','')
            fecha_cta_saldo = request.POST.get('fecha_cta_saldo')
            nro_cuotas_fn = request.POST.get('nro_cuota_saldo')
            valor_cuota_extra = request.POST.get('valor_cuota_extra')
            fecha_cta_extra = request.POST.get('fecha_cta_extra')
            nro_cuota_extra = request.POST.get('nro_cuota_extra')
            period_ce = request.POST.get('periodo_extra')
            observaciones = request.POST.get('observaciones')
            tasa = request.POST.get('tasa')
            
            nro_ctas_ci = request.POST.getlist('nro_cuota_ci')
            fecha_ctas_ci = request.POST.getlist('fecha_cta_ci')
            valor_cuota_ci = request.POST.getlist('valor_cuota_ci')
            
            nro_fracciones = request.POST.get('numero_fracciones')
            valor_fraccion = int(valor_venta)/int(nro_fracciones)
            valor_lista_fraccion = request.POST.get('valor_lista').replace('.','').replace('$','')
            
            
            obj_inmueble = Inmuebles.objects.using(proyecto).get(pk=inmueble)
            
            if obj_inmueble.fractales_disponibles() < int(nro_fracciones):
                data = {
                    'status':'error',
                    'msj':'El inmueble escogido no tiene suficientes fracciones disponibles para la venta'
                }
            
                return JsonResponse(data)
            
            venta = ventas_nuevas.objects.using(proyecto).create(
                id_t1 = inversionista, inmueble = inmueble, valor_venta = valor_venta,
                forma_pago = forma_pago, cuota_inicial = cuota_inicial, 
                saldo = saldo, forma_saldo = forma_pago, valor_ctas_fn = valor_ctas_fn,
                inicio_fn = fecha_cta_saldo, nro_cuotas_fn = nro_cuotas_fn,
                observaciones = observaciones, fecha_contrato = datetime.date.today(),
                usuario = request.user.username, estado = 'Pendiente', tasa = tasa,
                tipo_venta = 'Fractal'
            )
            
            for i in range(1,len(nro_ctas_ci)+1):
                n = f'venta.cant_ci{i} = nro_ctas_ci[{i-1}]'
                exec(n)
                f = f'venta.fecha_ci{i} = fecha_ctas_ci[{i-1}]'
                exec(f)
                v = f'venta.valor_ci{i} = valor_cuota_ci[{i-1}].replace(",","")'
                exec(v)
                venta.save()
            
            if nro_cuota_extra != '':
                venta.valor_ctas_ce = valor_cuota_extra
                venta.inicio_ce = fecha_cta_extra
                venta.nro_cuotas_ce = nro_cuota_extra
                venta.period_ce = period_ce
                venta.save()
            
            
            fractal = fractales_ventas.objects.using(proyecto).create(
                contrato =venta, nro_fracciones = nro_fracciones,
                valor_fraccion = valor_fraccion, valor_lista_fraccion = valor_lista_fraccion,
                valor_venta = valor_venta
            )
            
            timeline.objects.using(proyecto).create(
                adj=venta.pk, fecha = datetime.date.today(),
                usuario = request.user.username,
                accion = 'Modificó el contrato en estado no radicado.'
            )
            
            data = {
                'status':'success',
                'data':{
                    'venta':venta.pk
                    },
                'message':{
                }
            }
            
            return JsonResponse(data)
    
@login_required
def acciones_venta_fractal(request):
    if request.method == 'GET':
        if request.is_ajax():
            todo = request.GET.get('todo')
            
            if todo == 'getsaleinfo':
                proyecto = request.GET.get('proyecto')
                venta = request.GET.get('venta')
                
                obj_venta = ventas_nuevas.objects.using(proyecto).filter(pk=venta)
                
                data = {
                 'status':'success',
                 'venta':JSONRender(obj_venta,query_functions=['fractal','documents','recaudos']).render()
                }
                
                return JsonResponse(data)
            
            elif todo == 'datatable':
                proyecto = request.GET.get('proyecto')
                show_commercial = request.GET.get('showcommercial')
                show_admin = request.GET.get('showadmin')
                
                
                draw = request.GET.get('draw')
                start = request.GET.get('start')
                length = request.GET.get('length')
                search_val = request.GET.get('search[value]')
                
                
                adjs = Vista_Adjudicacion.objects.using(proyecto).exclude(Q(Estado__icontains='desistido')|
                                                                    Q(Estado__icontains='pagado')
                                                                    ).order_by('-FechaContrato')
                rt=adjs.count()
                if show_commercial:
                    adjs = adjs.filter(tipo_cartera='Comercial')
                
                elif show_admin:
                    adjs = adjs.filter(tipo_cartera='Administrativa')
                    
                if search_val != "" and search_val != None:
                    adjs = adjs.filter(Q(IdAdjudicacion__icontains=search_val)|Q(Nombre__icontains=search_val))
                
                rf = adjs.count()
                
                if int(length) != -1:
                    defered_list = adjs[int(start):int(start)+int(length)]
                else:
                    defered_list = adjs
                
                jsondata = JSONRender(defered_list, query_functions=['fractal','rcdo_total'])
                    
                data={
                    "draw": int(draw),
                    "recordsTotal": rt,
                    "recordsFiltered": rf,
                    'data':jsondata.render(),
                }
                
                return JsonResponse(data)

            elif todo == 'printdoc':
                proyecto = request.GET.get('proyecto')
                tipodoc = request.GET.get('tipodoc')
                venta = request.GET.get('venta')
                
                                
                if tipodoc == 'contrato':
                    obj_venta = ventas_nuevas.objects.using(proyecto).get(pk=venta)
                
                    context = {
                        'proyecto':proyecto,
                        'ctr':obj_venta,
                    }
                    if proyecto == 'Oasis':
                        context.update({
                            'fecha_escritura': datetime.date.today(),
                            'meses_entrega': 24,
                            'oficina': 'Medellín',
                        })
                    filename = f'{venta}_acuerdo_vinculacion_{proyecto}.pdf'
                    if proyecto == 'Oasis':
                        pdf = pdf_gen_weasy(f'pdf/{proyecto}/contrato.html', context, filename)
                    else:
                        pdf = pdf_gen(f'pdf/{proyecto}/contrato.html', context, filename)
                
                elif tipodoc == 'verificacion':
                    obj_venta = ventas_nuevas.objects.using(proyecto).get(pk=venta)
                
                    context = {
                        'proyecto':proyecto,
                        'ctr':obj_venta,
                    }
                    filename = f'{venta}_verificacion_{proyecto}.pdf'
                    pdf = pdf_gen(f'pdf/{proyecto}/verificacion.html',context,filename)
                    
                elif tipodoc == 'sagrilaft':
                    obj_venta = ventas_nuevas.objects.using(proyecto).get(pk=venta)
                    context = {
                        'cliente':clientes.objects.get(pk = obj_venta.id_t1),
                        'today': datetime.date.today(),
                    }
                
                    filename = f'Vinculacion_PN_{obj_venta.id_t1}.pdf'
                    path = f'pdf/sagrilaft.html'
                    pdf = pdf_gen(path,context,filename)
                
                elif tipodoc == 'pagare':
                    obj_venta = ventas_nuevas.objects.using(proyecto).get(pk=venta)
                
                    context = {
                        'proyecto':proyecto,
                        'ctr':obj_venta,
                    }
                    filename = f'{venta}_pagare_{proyecto}.pdf'
                    pdf = pdf_gen(f'pdf/{proyecto}/pagare.html',context,filename)
                
                elif tipodoc == 'alttum':
                    obj_venta = ventas_nuevas.objects.using(proyecto).get(pk=venta)
                
                    context = {
                        'proyecto':proyecto,
                        'ctr':obj_venta,
                    }
                    filename = f'{venta}_alttum_{proyecto}.pdf'
                    pdf = pdf_gen(f'pdf/{proyecto}/alttum.html',context,filename)
                
                elif tipodoc == 'recibo':
                    numero_recibo = request.GET.get('recibo')
                    filename = f'Recibo_caja_NR_{numero_recibo}_{proyecto}.pdf'
                    obj_recibo = RecaudosNoradicados.objects.using(proyecto).get(recibo = numero_recibo)
                    context = {
                        'recibo':obj_recibo
                    }
                    pdf = pdf_gen(f'pdf/{proyecto}/recibo.html',context,filename)
                    
                elif tipodoc == 'estadocuenta':
                    obj_adj = Adjudicacion.objects.using(proyecto).get(pk=venta)
                    today = datetime.date.today()
                    next_30_days = today+ datetime.timedelta(days=30)
                    cuotas_a_la_fecha = PlanPagos.objects.using(proyecto).filter(adj=venta, fecha__lte=today).order_by('fecha')
                    cuotas_futuras = PlanPagos.objects.using(proyecto).filter(
                        adj=venta, fecha__gt=today,fecha__lte = next_30_days).order_by('fecha')
                    
                    cuotas_vencidas = []
                    
                    total_cuotas_vencidas = {
                        'valor': 0,
                        'intereses_mora':0,
                        'total': 0
                    }
                    
                    for q in cuotas_a_la_fecha:
                        pendiente = q.pendiente()
                        if pendiente.get('total', 0) > 0:
                            mora = q.mora()
                            cuotas_vencidas.append({
                                'fecha':q.fecha,
                                'idcta': q.pk.split('ADJ')[0],
                                'pendiente':pendiente,
                                'mora':mora,
                            })
                            total_cuotas_vencidas['valor'] += pendiente.get('total')
                            total_cuotas_vencidas['intereses_mora'] += mora.get('valor')
                            total_cuotas_vencidas['total'] += pendiente.get('total') + mora.get('valor')
                    
                    
                    context = {
                        'adj':obj_adj,
                        'cuotas_a_la_fecha': cuotas_vencidas,
                        'cuotas_futuras':cuotas_futuras,
                        'user': request.user,
                        'now': datetime.datetime.now(),
                        'totals': total_cuotas_vencidas,
                        'recaudador': 'ANDINA CONCEPTOS INMOBILIARIOS S.A.S. NIT: 900993044-9'
                    }
                    
                    
                    filename = f'Estado_de_cuenta_{venta}_{proyecto}.pdf'
                            
                                
                    pdf = pdf_gen(f'pdf/Fractal/statement_of_account.html',context,filename)
                
                
                ruta = pdf.get('root')
                                        
                ruta_link=pdf.get('url')

                msj = f'Puedes descargar el documento <a href="{ruta_link}" target="_blank">aquí</a>'
                
                data = {
                    'status': 'success',
                    'message':{
                        'class': 'success',
                        'text':msj
                    }
                }
                
                return JsonResponse(data)
            
            
            
    elif request.method == 'POST':
        if request.is_ajax():
            todo = request.POST.get('todo')
            
            if todo == 'modificar':
                proyecto = request.POST.get('proyecto')
                idventa = request.POST.get('idventa')
                venta = ventas_nuevas.objects.using(proyecto).get(pk=idventa)

                inversionista = request.POST.get('inversionista')
                inmueble = request.POST.get('inmueble')
                valor_venta = request.POST.get('valor_inversion').replace(',','')
                forma_pago = request.POST.get('tipo_saldo')
                cuota_inicial = request.POST.get('valor_ci').replace(',','')
                saldo = request.POST.get('valor_saldo').replace(',','')
                valor_ctas_fn = request.POST.get('valor_cuota_saldo').replace(',','')
                fecha_cta_saldo = request.POST.get('fecha_cta_saldo')
                nro_cuotas_fn = request.POST.get('nro_cuota_saldo')
                valor_cuota_extra = request.POST.get('valor_cuota_extra')
                fecha_cta_extra = request.POST.get('fecha_cta_extra')
                nro_cuota_extra = request.POST.get('nro_cuota_extra')
                period_ce = request.POST.get('periodo_extra')
                observaciones = request.POST.get('observaciones')
                tasa = request.POST.get('tasa')
                
                nro_ctas_ci = request.POST.getlist('nro_cuota_ci')
                fecha_ctas_ci = request.POST.getlist('fecha_cta_ci')
                valor_cuota_ci = request.POST.getlist('valor_cuota_ci')
                
                nro_fracciones = request.POST.get('numero_fracciones')
                
                
                obj_inmueble = Inmuebles.objects.using(proyecto).get(pk=inmueble)
                
                if obj_inmueble.fractales_disponibles() < int(nro_fracciones):
                    data = {
                        'status':'error',
                        'msj':'El inmueble escogido no tiene suficientes fracciones disponibles para la venta'
                    }
                
                    return JsonResponse(data)
                    
                ventas = ventas_nuevas.objects.using(proyecto).filter(pk=idventa)
                
                ventas.update(
                    id_t1 = inversionista, inmueble = inmueble, valor_venta = valor_venta,
                    forma_pago = forma_pago, cuota_inicial = cuota_inicial, 
                    saldo = saldo, forma_saldo = forma_pago, valor_ctas_fn = valor_ctas_fn,
                    inicio_fn = fecha_cta_saldo, nro_cuotas_fn = nro_cuotas_fn,
                    observaciones = observaciones, fecha_contrato = datetime.date.today(),
                    usuario = request.user.username, estado = 'Pendiente', tasa = tasa,
                    tipo_venta = 'Fractal'
                )
                
                venta = ventas[0]
                venta.save()
                    
                
                for i in range(1,len(nro_ctas_ci)+1):
                    n = f'venta.cant_ci{i} = nro_ctas_ci[{i-1}]'
                    exec(n)
                    f = f'venta.fecha_ci{i} = fecha_ctas_ci[{i-1}]'
                    exec(f)
                    v = f'venta.valor_ci{i} = valor_cuota_ci[{i-1}].replace(",","")'
                    exec(v)
                    venta.save()
                
                if nro_cuota_extra != '':
                    venta.valor_ctas_ce = valor_cuota_extra.replace(',','')
                    venta.inicio_ce = fecha_cta_extra
                    venta.nro_cuotas_ce = nro_cuota_extra
                    venta.period_ce = period_ce
                    venta.save()
                
                
                fractal = fractales_ventas.objects.using(proyecto).get(contrato =idventa) 
                fractal.nro_fracciones = nro_fracciones
                fractal.valor_venta = valor_venta
                fractal.save()
                
                data = {
                        'status':'success',
                        'data':{
                            'venta':venta.pk
                            },
                        'message':{
                            'class':'success',
                            'text':'El contrato fué modificado sin problemas.'
                        }
                    }
                
                return JsonResponse(data)
            
            elif todo == 'uploadfile':
                proyecto = request.POST.get('proyecto')
                idventa = request.POST.get('idventa')
                contrato = request.POST.get('idventa')
                descrip_doc = request.POST.get('name')
                
                input_name = descrip_doc.lower() + '_upload'
                documento = request.FILES.get(input_name)
                
                doc = documentos_contratos.objects.using(proyecto).filter(
                    adj=contrato,descripcion_doc=descrip_doc,
                )
                
                if doc.exists(): 
                    for i in doc: 
                        doc.delete()
                                
                upload_docs_contratos(documento,contrato,proyecto,descrip_doc)
                documentos_contratos.objects.using(proyecto).create(
                    adj=contrato,descripcion_doc=descrip_doc,
                    fecha_carga=datetime.date.today(),usuario_carga=request.user.username
                )
                doc_path = f"docs_andinasoft/doc_contratos/{proyecto}/{idventa}/{descrip_doc}.pdf"
                href = default_storage.url(doc_path)
                data = {
                        'status':'success',
                        'data':{
                            'href': href
                            },
                        'message':{
                            'class':'success',
                            'text':'El archivo fue cargado sin problema'
                        }
                    }
                
                return JsonResponse(data)
            
            elif todo == 'aprobar':
                if not check_perms(request,('andinasoft.delete_ventas_nuevas',),raise_exception=False):
                    data = {
                    'status':'success',
                    'data':{
                        'venta':''
                        },
                    'message':{
                        'message':'No tienes permisos para aprobar una venta',
                        'class':'error'
                        }
                    }
                    
                    return JsonResponse(data)
                
                proyecto = request.POST.get('proyecto')
                idventa = request.POST.get('idventa')
                
                venta = ventas_nuevas.objects.using(proyecto).get(pk=idventa)
                venta.estado = 'Aprobado'
                venta.usuarioaprueba = request.user.username
                venta.fecha_aprueba = datetime.date.today()
                venta.save()
                
                data = {
                    'status':'success',
                    'data':{
                        'venta':venta.pk
                        },
                    'message':{
                        'message':'Se aprobó la venta seleccionada, serás redirigido en breve',
                        'class':'success'
                    }
                }
                
                return JsonResponse(data)

            elif todo == 'newreceipt':
                proyecto = request.POST.get('proyecto')
                contrato = request.POST.get('idventa')
                valor = request.POST.get('valor_recibo')
                fp = request.POST.get('formapagorecibo')
                soporte = request.FILES.get('soporterecibo')
                
                obj_consecutivo=consecutivos.objects.using(proyecto).get(documento='RC')
                consecutivo = obj_consecutivo.consecutivo
                
                name_doc = f'soporte_rc_{consecutivo}_{proyecto}_CTR{contrato}'
                typedoc = soporte.name.split('.')[-1].lower()


                file_dir = f'{settings.MEDIA_ROOT}/soportes_recibos/ventas_nuevas/{proyecto}/'
                file_key = _to_storage_key(f'{file_dir}{name_doc}.{typedoc}')
                upload_docs(soporte,file_dir,name_doc,typedoc)
                file_dwnld = default_storage.url(file_key)
                
                
                
                recaudo = RecaudosNoradicados.objects.using(proyecto).create(
                    recibo = consecutivo, contrato = contrato,
                    fecha = datetime.date.today(), formapago = fp,
                    usuario = request.user.username, soportepago = file_key,
                    concepto = 'PAGO RECIBIDO DE CLIENTE',
                    valor = valor.replace(',','')
                )
                
                filename = f'Recibo_caja_NR_{consecutivo}_{proyecto}.pdf'
        
                obj_recibo = RecaudosNoradicados.objects.using(proyecto).get(recibo = consecutivo)
                
                obj_consecutivo.consecutivo = consecutivo + 1
                obj_consecutivo.save()
                
                context = {
                    'recibo':obj_recibo
                }
                                    
                pdf = pdf_gen(f'pdf/{proyecto}/recibo.html', context,filename)
                
                ruta_link=pdf.get('url')

                msj = f'Puedes descargar el recibo <a href="{ruta_link}" target="_blank">aquí</a>'
                
                data = {
                    'status': 'success',
                    'message':{
                        'class': 'success',
                        'text':msj
                    },
                    'data':{
                        'rc':{
                            'fecha':obj_recibo.fecha,
                            'recibo':obj_recibo.recibo,
                            'usuario':obj_recibo.usuario,
                            'valor':obj_recibo.valor,
                            'formapago':obj_recibo.formapago,
                            'soportepago':file_dwnld
                        }
                    }
                }
                
                return JsonResponse(data)
            
@login_required     
def comisiones_ajax(request):
    
    if request.method == 'GET':
        if request.is_ajax:
            idventa = request.GET.get('idventa')
            proyecto = request.GET.get('proyecto')
            
            comisiones = AsignacionComisiones.objects.using(proyecto).filter(idadjudicacion=idventa)
            
            data = {
                'data': JSONRender(comisiones).render()
            }
            
            return JsonResponse(data)
    
    if request.method == 'POST':
        if request.is_ajax:
            contrato = request.POST.get('idventa')
            proyecto = request.POST.get('proyecto')
            
            cargos = request.POST.getlist('cargo_comision')
            porcentajes = request.POST.getlist('porcentaje_comision')
            asesores = request.POST.getlist('asesor_comision')
            
            obj_comis = AsignacionComisiones.objects.using(proyecto)
            for i in range(0, len(asesores)):
                cargo = cargos[i]
                cc = asesores[i]
                porc = porcentajes[i]
                if porc != "" and porc != None:
                    obj_comis.create(id_comision=f'{cargo}-{contrato}',
                                                    idadjudicacion=contrato,
                                                    fecha=datetime.datetime.today(),
                                                    idgestor=cc,
                                                    idcargo=cargo,
                                                    comision=porc,
                                                    usuario='Activo',
                                                    )
                
            cargos_fijos = CargosFijos.objects.using(proyecto).all()
            
            for fijo in cargos_fijos:
                obj_comis.create(id_comision=f'{fijo.idcargo}-{contrato}',
                                                idadjudicacion=contrato,
                                                fecha=datetime.datetime.today(),
                                                idgestor=fijo.cc_fija,
                                                idcargo=fijo.idcargo,
                                                comision=fijo.porc_fijo,
                                                usuario='Activo',
                                                )
                
            data = {
                
            }
            
            return JsonResponse(data)
            
@login_required
@group_perm_required(perms=('andinasoft.add_presupuestocartera',),raise_exception=True)
def cartera_month_results(request):
    if request.method == 'GET' and request.GET:
        mes = request.GET.get('month')
        anio = request.GET.get('year')
        
        periodo = f'{anio}{mes}'
        
        usuario_administrador=check_perms(request,('andinasoft.change_presupuestocartera',))
        if not usuario_administrador:
            raise PermissionError
        
        lista_proyectos = proyectos.objects.exclude(
            Q(proyecto__startswith='Alttum')|
            Q(proyecto = 'Sandville del sol')|
            Q(proyecto = 'Sotavento')
        )
        
        book=openpyxl.load_workbook("resources/excel_formats/Bonos_cartera.xlsx")
        failed_projects = []
        for proyecto in lista_proyectos:
            try:
                contenido_ppto = list(
                    Adjudicacion.objects.using(proyecto.proyecto).raw(
                        f'CALL informe_cartera("{periodo}",NULL)'
                    )
                )
            except DBInternalError as exc:
                failed_projects.append((proyecto.proyecto, str(exc)))
                continue
            
            sheet=book[proyecto.proyecto]
            encabezados=['Adjudicacion','Cliente','Estado','Origen','Venta Mes','Tipo Cartera','Edad',
                            'Cuota Mes','Recaudo Mes','Cuotas Vencidas','Recaudo Vencido','Presupuesto Total','Recaudo Presupuestado',
                            'Recaudo No Pptado','Recaudo Total','Asesor','Cashout']
            j = 1
            for x in encabezados:
                sheet.cell(1,j,x)
                j+=1
            
            #sheet.append(encabezados)
            i=2
            for fila in contenido_ppto:
                sheet.cell(i,1,fila.pk)
                sheet.cell(i,2,fila.cliente)
                sheet.cell(i,3,fila.estado)
                sheet.cell(i,4,fila.origen)
                sheet.cell(i,5,fila.venta_mes)
                sheet.cell(i,6,fila.tipocartera)
                sheet.cell(i,7,fila.edad)
                sheet.cell(i,8,fila.ppto_mes)
                sheet.cell(i,9,fila.recaudo_mes)
                sheet.cell(i,10,fila.ppto_vencido)
                sheet.cell(i,11,fila.recaudo_vencido)
                sheet.cell(i,12,fila.presupuesto)
                sheet.cell(i,13,fila.recaudo_pptado)
                sheet.cell(i,14,fila.recaudo_nopptado)
                sheet.cell(i,15,fila.recaudo_total)
                sheet.cell(i,16,fila.asesor)
                i+=1

        if failed_projects:
            sheet_name = 'Errores'
            if sheet_name in book.sheetnames:
                error_sheet = book[sheet_name]
                error_sheet.delete_rows(1, error_sheet.max_row)
            else:
                error_sheet = book.create_sheet(sheet_name)
            error_sheet.append(['Proyecto', 'Error'])
            for project_name, error_message in failed_projects:
                error_sheet.append([project_name, error_message[:500]])
                
        filename = f'Informe_cartera_periodo_{periodo}.xlsx'
        ruta=settings.DIR_EXPORT+filename
        _save_workbook_with_dirs(book, ruta)
        return FileResponse(open(ruta,'rb'),as_attachment=True,filename=filename)
        
    return render(request,'reporte_ov.html')

@login_required
def ajustar_al_peso(request):
    if request.method == 'POST':
        if request.is_ajax:
            proyecto= request.POST.get('proyecto')
            adj = request.POST.get('adj')
            valor_ajuste = request.POST.get('valor_ajuste')
            
            fp = formas_pago.objects.using(proyecto).filter(descripcion_icontains='ajuste al peso')
            consecutivo=consecutivos.objects.using(proyecto).get(documento='RC')
            
            if fp.exists():
                fp = fp[0]
            else:
                data = {
                    'class': 'alert-danger',
                    'msj':'No existe forma de pago de ajuste al peso.'
                }
            
                return JsonResponse(data)
            
            obj_adj= Adjudicacion.objects.using(proyecto).get(pk=adj)
            saldo_cuotas=saldos_adj.objects.using(proyecto).filter(adj=adj,saldocuota__gt=0)
            
            for cuota in saldo_cuotas:
                Recaudos.objects.using(proyecto).create(recibo=consecutivo.consecutivo,
                                            fecha=datetime.date.today(),
                                            idcta=cuota.idcta,
                                            idadjudicacion=adj,
                                            capital=valor_ajuste,
                                            interescte=0,
                                            interesmora=0,
                                            moralqd=0,
                                            fechaoperacion=datetime.datetime.today(),
                                            usuario=request.user,
                                            estado='Aprobado')
            
            nro_recibo=f'{consecutivo.consecutivo}'
            Recaudos_general.objects.using(proyecto).create(idadjudicacion=adj,
                                                            fecha=datetime.date.today(),
                                                            fecha_pago = datetime.date.today(),
                                                            numrecibo=nro_recibo,
                                                            idtercero=obj_adj.titulares().get('titular_1').pk,
                                                            operacion='Ajuste al peso',
                                                            valor=valor_ajuste,
                                                            formapago=str(fp),
                                                            concepto='Ajuste al peso',
                                                            usuario=request.user)
            
            consecutivo.consecutivo=consecutivo.consecutivo+1
            consecutivo.save()
            obj_saldos=saldos_adj.objects.using(proyecto).filter(adj=adj)
            saldos=obj_saldos.aggregate(Sum('saldocuota'))
            saldos=saldos['saldocuota__sum']
            if saldos<=0:
                obj_adj.estado='Pagado'
                obj_adj.save()
                
            filename = f'Recibo_caja_{nro_recibo}_{proyecto}.pdf'
            ruta=settings.DIR_EXPORT+filename
            
            obj_recibo = Recaudos_general.objects.using(proyecto).get(numrecibo = nro_recibo)
                                
            context_rc = {
                'recibo':obj_recibo
            }
                                
            pdf = pdf_gen(f'pdf/{proyecto}/recibo.html', context_rc,filename)
            
            file = pdf.get('url')
            
            data = {
                'class': 'alert-success',
                'msj':f'Puedes descargar el recibo de ajuste <a href="{file}"><b>aquí</b></a>'
            }
            
            return JsonResponse(data)
            
def imprimir_recibos(request):
    if request.GET and request.method == 'GET':
        proyecto = request.GET.get('proyecto')
        recibo = request.GET.get('recibo')
        
        
        filename = f'Recibo_caja_{recibo}_{proyecto}.pdf'
        
        obj_recibo = Recaudos_general.objects.using(proyecto).get(numrecibo = recibo)
        
        
        context = {
            'recibo':obj_recibo
        }
                            
        pdf = pdf_gen(f'pdf/{proyecto}/recibo.html', context,filename)
        
        file = pdf.get('root')
                                
        return FileResponse(open(file,'rb'),as_attachment=True,filename=filename)

     
Ajax_URL = [
    path('ajax/printPromesa',ajax_imprimir_promesa),
    path('printdocuments/<proyecto>',print_documents),
    path('conciliartercerosiigo',compare_siigo_andinasoft_clients),
    path('informecartera',cartera_month_results),
    path('printincome',imprimir_recibos),
    path('ajustealpeso',ajustar_al_peso),
    path('estadodecuenta', ajax_print_estado_cuenta),
    path('ventafraccion',nueva_venta_fractal),
    path('detallesventa',acciones_venta_fractal),
    path('comisiones',comisiones_ajax),
]
#helper functions

def api_socrata(request):
    return render(request,'base.html') 

def blank_request(request):
    return render(request,'blank_request.html')

def handler403(request, *args, **argv):
    context={}
    return render(request,'403.html',context)

def handler500(request, *args, **argv):
    context={}
    return render(request,'500.html',context, status=500)

def check_perms(request,perms:list,raise_exception=True):
    user=request.user
    permissions=user.get_all_permissions()
    permissions_granted=0
    permissions_required=len(perms)
    for perm in perms:
        if perm in permissions:
            permissions_granted+=1
    if permissions_granted==permissions_required:
        return True
    if raise_exception:
        raise PermissionDenied
    return False

def check_groups(request,groups:list,raise_exception=True):
    if request.user.is_superuser:
        return True
    
    user_groups=request.user.groups.all()
    for user_group in user_groups:
        if user_group.name in groups:
            return True
    
    if raise_exception:
        raise PermissionDenied
    return False

def check_project(request,proyecto,raise_exception=True):
    if request.user.is_superuser:
        return True
    user=request.user.pk
    user_projects=Usuarios_Proyectos.objects.filter(usuario=user)
    if user_projects.exists():
        user_projects=user_projects[0].proyecto.all()
        for p in user_projects:
            if p.proyecto==proyecto:
                return True
    if raise_exception:
        raise PermissionDenied
    return False
